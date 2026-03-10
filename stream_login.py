# # # # # # # # """
# # # # # # # # ╔══════════════════════════════════════════════════════════╗
# # # # # # # # ║   SecureITLab Pipeline Dashboard — Streamlit             ║
# # # # # # # # ║   Reads from MongoDB → secureitlab_job_pipeline          ║
# # # # # # # # ╠══════════════════════════════════════════════════════════╣
# # # # # # # # ║  Install: pip install streamlit pymongo python-dotenv    ║
# # # # # # # # ║  Run:     streamlit run streamlit_dashboard.py           ║
# # # # # # # # ╚══════════════════════════════════════════════════════════╝
# # # # # # # # """

# # # # # # # # import streamlit as st
# # # # # # # # from pymongo import MongoClient
# # # # # # # # import json
# # # # # # # # import threading
# # # # # # # # import time
# # # # # # # # from datetime import datetime, timezone

# # # # # # # # # ── Thread-safe shared state ──────────────────────────────────────────────────
# # # # # # # # _thread_logs   = []
# # # # # # # # _thread_result = [None]
# # # # # # # # _thread_done   = [False]
# # # # # # # # _thread_lock   = threading.Lock()

# # # # # # # # # ── Page config ───────────────────────────────────────────────────────────────
# # # # # # # # st.set_page_config(
# # # # # # # #     page_title="SecureITLab Pipeline",
# # # # # # # #     page_icon="🛡️",
# # # # # # # #     layout="wide",
# # # # # # # #     initial_sidebar_state="expanded",
# # # # # # # # )

# # # # # # # # # ── LOGIN CREDENTIALS (change these) ─────────────────────────────────────────
# # # # # # # # LOGIN_USERNAME = "admin"
# # # # # # # # LOGIN_PASSWORD = "neha123"

# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # #  GLOBAL CSS (login + dashboard share this)
# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # st.markdown("""
# # # # # # # # <style>

# # # # # # # # @import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');

# # # # # # # # :root {
# # # # # # # #     --bg:        #f4f7fb;
# # # # # # # #     --surface:   #ffffff;
# # # # # # # #     --surface2:  #eef2f7;
# # # # # # # #     --border:    #d9e2ec;
# # # # # # # #     --accent:    #2563eb;
# # # # # # # #     --accent2:   #7c3aed;
# # # # # # # #     --green:     #16a34a;
# # # # # # # #     --yellow:    #f59e0b;
# # # # # # # #     --red:       #dc2626;
# # # # # # # #     --text:      #0f172a;
# # # # # # # #     --muted:     #64748b;
# # # # # # # # }

# # # # # # # # html, body, [class*="css"] {
# # # # # # # #     background-color: var(--bg) !important;
# # # # # # # #     color: var(--text) !important;
# # # # # # # #     font-family: 'DM Sans', sans-serif !important;
# # # # # # # # }

# # # # # # # # /* ── LOGIN PAGE ─────────────────────────────────────────── */
# # # # # # # # .login-wrap {
# # # # # # # #     display: flex;
# # # # # # # #     align-items: center;
# # # # # # # #     justify-content: center;
# # # # # # # #     min-height: 80vh;
# # # # # # # # }
# # # # # # # # .login-card {
# # # # # # # #     background: var(--surface);
# # # # # # # #     border: 1px solid var(--border);
# # # # # # # #     border-radius: 20px;
# # # # # # # #     padding: 3rem 3.5rem;
# # # # # # # #     width: 100%;
# # # # # # # #     max-width: 420px;
# # # # # # # #     box-shadow: 0 20px 60px rgba(37,99,235,0.08);
# # # # # # # #     text-align: center;
# # # # # # # # }
# # # # # # # # .login-logo {
# # # # # # # #     font-family: 'Syne', sans-serif;
# # # # # # # #     font-size: 1.6rem;
# # # # # # # #     font-weight: 800;
# # # # # # # #     color: var(--accent);
# # # # # # # #     margin-bottom: .25rem;
# # # # # # # # }
# # # # # # # # .login-subtitle {
# # # # # # # #     font-size: .75rem;
# # # # # # # #     color: var(--muted);
# # # # # # # #     letter-spacing: .12em;
# # # # # # # #     text-transform: uppercase;
# # # # # # # #     margin-bottom: 2.5rem;
# # # # # # # # }
# # # # # # # # .login-error {
# # # # # # # #     background: #fef2f2;
# # # # # # # #     border: 1px solid #fecaca;
# # # # # # # #     border-radius: 8px;
# # # # # # # #     padding: .65rem 1rem;
# # # # # # # #     color: #b91c1c;
# # # # # # # #     font-size: .85rem;
# # # # # # # #     margin-top: 1rem;
# # # # # # # # }
# # # # # # # # .login-divider {
# # # # # # # #     width: 40px;
# # # # # # # #     height: 3px;
# # # # # # # #     background: linear-gradient(90deg, var(--accent), var(--accent2));
# # # # # # # #     border-radius: 2px;
# # # # # # # #     margin: 0 auto 2rem;
# # # # # # # # }

# # # # # # # # /* ── DASHBOARD ──────────────────────────────────────────── */
# # # # # # # # [data-testid="stSidebar"] {
# # # # # # # #     background: var(--surface) !important;
# # # # # # # #     border-right: 1px solid var(--border) !important;
# # # # # # # # }
# # # # # # # # [data-testid="stSidebar"] * { color: var(--text) !important; }

# # # # # # # # .main .block-container { padding: 2rem 2rem 3rem !important; }

# # # # # # # # h1, h2, h3, h4 {
# # # # # # # #     font-family: 'Syne', sans-serif !important;
# # # # # # # #     color: var(--text) !important;
# # # # # # # # }

# # # # # # # # .sil-card {
# # # # # # # #     background: var(--surface);
# # # # # # # #     border: 1px solid var(--border);
# # # # # # # #     border-radius: 14px;
# # # # # # # #     padding: 1.25rem 1.5rem;
# # # # # # # #     margin-bottom: 1rem;
# # # # # # # #     transition: all 0.25s ease;
# # # # # # # # }
# # # # # # # # .sil-card:hover {
# # # # # # # #     transform: translateY(-2px);
# # # # # # # #     box-shadow: 0 8px 22px rgba(0,0,0,0.05);
# # # # # # # # }
# # # # # # # # .sil-card-accent { border-left: 4px solid var(--accent); }

# # # # # # # # .metric-row { display:flex; gap:1rem; flex-wrap:wrap; margin-bottom:1.5rem; }
# # # # # # # # .metric-tile {
# # # # # # # #     flex:1; min-width:140px;
# # # # # # # #     background:var(--surface2);
# # # # # # # #     border:1px solid var(--border);
# # # # # # # #     border-radius:12px;
# # # # # # # #     padding:1rem; text-align:center;
# # # # # # # #     transition:all .25s ease;
# # # # # # # # }
# # # # # # # # .metric-tile:hover { transform:translateY(-3px); box-shadow:0 10px 24px rgba(0,0,0,0.06); }
# # # # # # # # .metric-tile .val { font-family:'Syne',sans-serif; font-size:2rem; font-weight:800; color:var(--accent); }
# # # # # # # # .metric-tile .lbl { font-size:.72rem; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; }

# # # # # # # # .badge { padding:.25rem .7rem; border-radius:20px; font-size:.72rem; font-weight:600; font-family:'DM Mono',monospace; }
# # # # # # # # .badge-green  { background:#ecfdf5; color:#15803d; }
# # # # # # # # .badge-yellow { background:#fffbeb; color:#b45309; }
# # # # # # # # .badge-red    { background:#fef2f2; color:#b91c1c; }
# # # # # # # # .badge-blue   { background:#eff6ff; color:#1d4ed8; }
# # # # # # # # .badge-purple { background:#f5f3ff; color:#6d28d9; }

# # # # # # # # .contact-card {
# # # # # # # #     background:var(--surface2); border:1px solid var(--border);
# # # # # # # #     border-radius:10px; padding:1rem; margin-bottom:.8rem;
# # # # # # # # }
# # # # # # # # .contact-name  { font-family:'Syne',sans-serif; font-weight:700; color:var(--text); }
# # # # # # # # .contact-title { color:var(--muted); font-size:.85rem; }
# # # # # # # # .contact-email { font-family:'DM Mono',monospace; color:var(--accent); font-size:.8rem; }

# # # # # # # # .email-box {
# # # # # # # #     background: #f8fafc;
# # # # # # # #     border: 1px solid var(--border);
# # # # # # # #     border-radius: 10px;
# # # # # # # #     padding: 1rem 1.25rem;
# # # # # # # #     font-size: .9rem;
# # # # # # # #     line-height: 1.65;
# # # # # # # #     white-space: pre-wrap;
# # # # # # # #     color: var(--text);
# # # # # # # # }
# # # # # # # # .email-subject { font-family:'Syne',sans-serif; font-weight:700; color:var(--accent); margin-bottom:.5rem; }

# # # # # # # # .section-label {
# # # # # # # #     font-family:'DM Mono',monospace; font-size:.72rem;
# # # # # # # #     text-transform:uppercase; letter-spacing:.12em;
# # # # # # # #     color:var(--accent); margin-bottom:.6rem;
# # # # # # # # }
# # # # # # # # .sil-divider { border-top:1px solid var(--border); margin:1rem 0; }

# # # # # # # # [data-testid="stExpander"] {
# # # # # # # #     background: var(--surface) !important;
# # # # # # # #     border: 1px solid var(--border) !important;
# # # # # # # #     border-radius: 10px !important;
# # # # # # # # }
# # # # # # # # [data-testid="stSelectbox"] > div,
# # # # # # # # [data-testid="stMultiSelect"] > div {
# # # # # # # #     background: var(--surface2) !important;
# # # # # # # #     border-color: var(--border) !important;
# # # # # # # # }
# # # # # # # # [data-testid="stTabs"] button {
# # # # # # # #     font-family:'Syne',sans-serif !important;
# # # # # # # #     font-weight:600 !important;
# # # # # # # # }
# # # # # # # # ::-webkit-scrollbar { width:6px; }
# # # # # # # # ::-webkit-scrollbar-thumb { background:var(--border); border-radius:3px; }

# # # # # # # # .pipeline-log {
# # # # # # # #     background: #0f172a;
# # # # # # # #     color: #e2e8f0;
# # # # # # # #     border-radius: 10px;
# # # # # # # #     padding: 1rem 1.25rem;
# # # # # # # #     font-family: 'DM Mono', monospace;
# # # # # # # #     font-size: .8rem;
# # # # # # # #     line-height: 1.6;
# # # # # # # #     max-height: 380px;
# # # # # # # #     overflow-y: auto;
# # # # # # # #     white-space: pre-wrap;
# # # # # # # # }

# # # # # # # # .fit-box {
# # # # # # # #     border-radius: 8px;
# # # # # # # #     padding: .75rem;
# # # # # # # #     margin-bottom: .5rem;
# # # # # # # #     font-size: .85rem;
# # # # # # # # }

# # # # # # # # /* Hide sidebar on login page */
# # # # # # # # .hide-sidebar [data-testid="stSidebar"] { display: none !important; }
# # # # # # # # .hide-sidebar .main .block-container { max-width: 480px; margin: 0 auto; }

# # # # # # # # </style>
# # # # # # # # """, unsafe_allow_html=True)


# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # #  SESSION STATE INIT
# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # for _k, _v in [
# # # # # # # #     ("logged_in",        False),
# # # # # # # #     ("login_error",      ""),
# # # # # # # #     ("pipeline_running", False),
# # # # # # # #     ("pipeline_logs",    []),
# # # # # # # #     ("pipeline_result",  None),
# # # # # # # #     ("pipeline_done",    False),
# # # # # # # # ]:
# # # # # # # #     if _k not in st.session_state:
# # # # # # # #         st.session_state[_k] = _v


# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # #  LOGIN PAGE
# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # if not st.session_state.logged_in:

# # # # # # # #     # Hide sidebar on login page
# # # # # # # #     st.markdown("""
# # # # # # # #     <script>
# # # # # # # #     document.querySelector('[data-testid="stSidebar"]') &&
# # # # # # # #     (document.querySelector('[data-testid="stSidebar"]').style.display = 'none');
# # # # # # # #     </script>
# # # # # # # #     """, unsafe_allow_html=True)

# # # # # # # #     # Center the login card
# # # # # # # #     _, col, _ = st.columns([1, 1.2, 1])

# # # # # # # #     with col:
# # # # # # # #         st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)

# # # # # # # #         st.markdown("""
# # # # # # # #         <div class="login-card">
# # # # # # # #           <div class="login-logo">🛡️ SecureITLab</div>
# # # # # # # #           <div class="login-subtitle">Pipeline Intelligence</div>
# # # # # # # #           <div class="login-divider"></div>
# # # # # # # #         </div>
# # # # # # # #         """, unsafe_allow_html=True)

# # # # # # # #         st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # # # # #         with st.container():
# # # # # # # #             username = st.text_input(
# # # # # # # #                 "Username",
# # # # # # # #                 placeholder="Enter username",
# # # # # # # #                 key="login_username",
# # # # # # # #             )
# # # # # # # #             password = st.text_input(
# # # # # # # #                 "Password",
# # # # # # # #                 placeholder="Enter password",
# # # # # # # #                 type="password",
# # # # # # # #                 key="login_password",
# # # # # # # #             )

# # # # # # # #             st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # # # # #             login_btn = st.button(
# # # # # # # #                 "Sign In →",
# # # # # # # #                 use_container_width=True,
# # # # # # # #                 type="primary",
# # # # # # # #             )

# # # # # # # #             if login_btn:
# # # # # # # #                 if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
# # # # # # # #                     st.session_state.logged_in = True
# # # # # # # #                     st.session_state.login_error = ""
# # # # # # # #                     st.rerun()
# # # # # # # #                 else:
# # # # # # # #                     st.session_state.login_error = "Incorrect username or password."

# # # # # # # #             if st.session_state.login_error:
# # # # # # # #                 st.markdown(
# # # # # # # #                     f'<div class="login-error">⚠️ {st.session_state.login_error}</div>',
# # # # # # # #                     unsafe_allow_html=True,
# # # # # # # #                 )

# # # # # # # #         st.markdown(
# # # # # # # #             "<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>"
# # # # # # # #             "SecureITLab · Confidential</div>",
# # # # # # # #             unsafe_allow_html=True,
# # # # # # # #         )

# # # # # # # #     st.stop()  # ← Stop here — nothing below renders until logged in


# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # #  EVERYTHING BELOW ONLY RUNS WHEN LOGGED IN
# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # # # # # # ── MongoDB helpers ───────────────────────────────────────────────────────────
# # # # # # # # @st.cache_resource
# # # # # # # # def get_db():
# # # # # # # #     URI = st.secrets.get("MONGO_URI", "mongodb://localhost:27017")
# # # # # # # #     DB  = st.secrets.get("MONGO_DB",  "secureitlab_job_pipeline")
# # # # # # # #     client = MongoClient(URI, serverSelectionTimeoutMS=6000)
# # # # # # # #     return client[DB]

# # # # # # # # @st.cache_data(ttl=60)
# # # # # # # # def load_all_jobs():
# # # # # # # #     db = get_db()
# # # # # # # #     return list(db.jobs.find({}, {
# # # # # # # #         "_id": 1, "company": 1, "role": 1, "job_number": 1,
# # # # # # # #         "opp_score": 1, "contacts_found": 1, "pipeline_ok": 1,
# # # # # # # #         "coverage_score": 1, "run_at": 1, "contact_domain": 1,
# # # # # # # #     }))

# # # # # # # # @st.cache_data(ttl=60)
# # # # # # # # def load_job(job_id):
# # # # # # # #     from bson import ObjectId
# # # # # # # #     return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# # # # # # # # # ── Render helpers ────────────────────────────────────────────────────────────
# # # # # # # # def badge(text, color="blue"):
# # # # # # # #     return f'<span class="badge badge-{color}">{text}</span>'

# # # # # # # # def safe_str(val, limit=300):
# # # # # # # #     if val is None: return "—"
# # # # # # # #     s = str(val)
# # # # # # # #     return s[:limit] + "…" if len(s) > limit else s

# # # # # # # # def as_dict(raw):
# # # # # # # #     if isinstance(raw, dict): return raw
# # # # # # # #     if isinstance(raw, list): return next((x for x in raw if isinstance(x, dict)), {})
# # # # # # # #     return {}

# # # # # # # # def render_json_pretty(data, title=""):
# # # # # # # #     if not data: return
# # # # # # # #     with st.expander(f"📄 Raw JSON — {title}", expanded=False):
# # # # # # # #         st.code(json.dumps(data, indent=2, default=str), language="json")

# # # # # # # # def render_qa_block(data, label):
# # # # # # # #     if not data:
# # # # # # # #         st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>', unsafe_allow_html=True)
# # # # # # # #         return
# # # # # # # #     data = as_dict(data)
# # # # # # # #     if not data: return
# # # # # # # #     passed    = data.get("passed") or data.get("Passed") or False
# # # # # # # #     rec       = data.get("recommendation") or data.get("Recommendation", "")
# # # # # # # #     issues    = data.get("issues") or data.get("Issues") or []
# # # # # # # #     checklist = data.get("checklist") or data.get("Checklist") or []
# # # # # # # #     color     = "green" if passed else "yellow"
# # # # # # # #     status    = "✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
# # # # # # # #     html = f"""
# # # # # # # #     <div class="sil-card sil-card-accent">
# # # # # # # #       <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
# # # # # # # #         <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
# # # # # # # #         {badge(status, color)}
# # # # # # # #       </div>"""
# # # # # # # #     if rec:
# # # # # # # #         html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
# # # # # # # #     if checklist:
# # # # # # # #         html += '<div style="font-size:.82rem;margin-bottom:.5rem">'
# # # # # # # #         for item in (checklist if isinstance(checklist, list) else [checklist]):
# # # # # # # #             if isinstance(item, dict):
# # # # # # # #                 ip = item.get("pass") or item.get("passed") or item.get("status","") == "pass"
# # # # # # # #                 nm = item.get("item") or item.get("name") or item.get("check","")
# # # # # # # #                 nt = item.get("reason") or item.get("note") or item.get("issue","")
# # # # # # # #                 html += f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
# # # # # # # #                 if nt: html += f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
# # # # # # # #                 html += '</div>'
# # # # # # # #         html += '</div>'
# # # # # # # #     if issues:
# # # # # # # #         html += '<div style="margin-top:.5rem">'
# # # # # # # #         for iss in (issues if isinstance(issues, list) else [issues])[:4]:
# # # # # # # #             txt = iss if isinstance(iss, str) else json.dumps(iss)
# # # # # # # #             html += f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
# # # # # # # #         html += '</div>'
# # # # # # # #     st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # # # # def render_contacts(contacts, title="Contacts"):
# # # # # # # #     if not contacts: st.info("No contacts found for this job."); return
# # # # # # # #     pri_color = {"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
# # # # # # # #     st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>', unsafe_allow_html=True)
# # # # # # # #     cols = st.columns(2)
# # # # # # # #     for i, c in enumerate(contacts):
# # # # # # # #         col = cols[i % 2]
# # # # # # # #         prio = c.get("priority", "General")
# # # # # # # #         email = c.get("email", ""); li = c.get("linkedin_url", "")
# # # # # # # #         patterns = c.get("email_patterns", []); src = c.get("source", "")
# # # # # # # #         with col:
# # # # # # # #             html = f"""<div class="contact-card">
# # # # # # # #               <div style="display:flex;justify-content:space-between;align-items:flex-start">
# # # # # # # #                 <div>
# # # # # # # #                   <div class="contact-name">{c.get('name','—')}</div>
# # # # # # # #                   <div class="contact-title">{c.get('title','—')}</div>
# # # # # # # #                 </div>
# # # # # # # #                 {badge(prio, pri_color.get(prio,'blue'))}
# # # # # # # #               </div>"""
# # # # # # # #             if email:      html += f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
# # # # # # # #             elif patterns: html += f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
# # # # # # # #             if li:         html += f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
# # # # # # # #             if src:        html += f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
# # # # # # # #             st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # # # # def render_emails(emails_data):
# # # # # # # #     if not emails_data: st.info("No email data available."); return
# # # # # # # #     emails_data = as_dict(emails_data)
# # # # # # # #     if not emails_data: return
# # # # # # # #     variants = {}
# # # # # # # #     for k, v in emails_data.items():
# # # # # # # #         kl = k.lower().replace("_","").replace(" ","")
# # # # # # # #         if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl == "a":
# # # # # # # #             variants["Variant A — Hiring Manager"] = v
# # # # # # # #         elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl == "b":
# # # # # # # #             variants["Variant B — CISO / VP Level"] = v
# # # # # # # #         else:
# # # # # # # #             variants[k] = v
# # # # # # # #     for label, v in variants.items():
# # # # # # # #         st.markdown(f'<div class="section-label">✉️ {label}</div>', unsafe_allow_html=True)
# # # # # # # #         if isinstance(v, dict):
# # # # # # # #             subj = v.get("subject") or v.get("Subject","")
# # # # # # # #             body = v.get("body") or v.get("Body") or v.get("content","")
# # # # # # # #             if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>', unsafe_allow_html=True)
# # # # # # # #             if body: st.markdown(f'<div class="email-box">{body}</div>', unsafe_allow_html=True)
# # # # # # # #             else:    st.code(json.dumps(v, indent=2), language="json")
# # # # # # # #         elif isinstance(v, str):
# # # # # # # #             st.markdown(f'<div class="email-box">{v}</div>', unsafe_allow_html=True)
# # # # # # # #         st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

# # # # # # # # def render_service_mapping(data):
# # # # # # # #     if not data: st.info("No service mapping data."); return
# # # # # # # #     items = data if isinstance(data, list) else []
# # # # # # # #     if not items and isinstance(data, dict):
# # # # # # # #         for key in ("services","mappings","service_mapping","ServiceMapping","items"):
# # # # # # # #             if isinstance(data.get(key), list): items = data[key]; break
# # # # # # # #         if not items: items = [data]
# # # # # # # #     fit_colors = {"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
# # # # # # # #     for item in items:
# # # # # # # #         if not isinstance(item, dict): continue
# # # # # # # #         svc  = item.get("service") or item.get("Service") or item.get("name","")
# # # # # # # #         fit  = (item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
# # # # # # # #         why  = item.get("justification") or item.get("rationale") or item.get("why","")
# # # # # # # #         reqs = item.get("requirements_addressed") or item.get("requirements") or ""
# # # # # # # #         eng  = item.get("engagement_type") or item.get("engagement","")
# # # # # # # #         color = fit_colors.get(fit, "blue")
# # # # # # # #         html = f"""<div class="sil-card" style="margin-bottom:.75rem">
# # # # # # # #           <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
# # # # # # # #             <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
# # # # # # # #             {badge(fit or "MAPPED", color) if fit else ""}
# # # # # # # #           </div>"""
# # # # # # # #         if why:  html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
# # # # # # # #         if reqs:
# # # # # # # #             rs = ", ".join(reqs) if isinstance(reqs, list) else str(reqs)
# # # # # # # #             html += f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
# # # # # # # #         if eng:  html += f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
# # # # # # # #         st.markdown(html + '</div>', unsafe_allow_html=True)
# # # # # # # #     render_json_pretty(data, "Service Mapping")

# # # # # # # # def render_microplans(data):
# # # # # # # #     if not data: st.info("No micro-plan data."); return
# # # # # # # #     plans = data if isinstance(data, list) else []
# # # # # # # #     if not plans and isinstance(data, dict):
# # # # # # # #         for k in ("plans","micro_plans","microplans","top_3","improvements"):
# # # # # # # #             if isinstance(data.get(k), list): plans = data[k]; break
# # # # # # # #         if not plans: plans = [data]
# # # # # # # #     for i, plan in enumerate(plans[:3], 1):
# # # # # # # #         if not isinstance(plan, dict): continue
# # # # # # # #         title = plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
# # # # # # # #         weeks = plan.get("duration") or plan.get("timeline","")
# # # # # # # #         obj   = plan.get("objective") or plan.get("goal","")
# # # # # # # #         kpis  = plan.get("kpis") or plan.get("KPIs") or []
# # # # # # # #         tasks = plan.get("tasks") or plan.get("workstreams") or []
# # # # # # # #         with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}", expanded=(i==1)):
# # # # # # # #             if obj and obj != title: st.markdown(f"**Objective:** {obj}")
# # # # # # # #             if kpis:
# # # # # # # #                 st.markdown("**KPIs:**")
# # # # # # # #                 for kpi in (kpis if isinstance(kpis, list) else [kpis]): st.markdown(f"• {kpi}")
# # # # # # # #             if tasks:
# # # # # # # #                 st.markdown("**Tasks / Workstreams:**")
# # # # # # # #                 for t in (tasks if isinstance(tasks, list) else [tasks]):
# # # # # # # #                     if isinstance(t, dict):
# # # # # # # #                         tn = t.get("task") or t.get("name","")
# # # # # # # #                         te = t.get("effort") or t.get("duration","")
# # # # # # # #                         st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
# # # # # # # #                     else: st.markdown(f"• {t}")
# # # # # # # #             render_json_pretty(plan, f"Plan {i}")

# # # # # # # # def render_deal_assurance(data):
# # # # # # # #     if not data: st.info("No deal assurance data."); return
# # # # # # # #     if not isinstance(data, dict): render_json_pretty(data, "Deal Assurance Pack"); return
# # # # # # # #     evp = (data.get("executive_value_proposition") or
# # # # # # # #            data.get("value_proposition") or data.get("ExecutiveValueProposition",""))
# # # # # # # #     if evp:
# # # # # # # #         st.markdown('<div class="section-label">💼 Executive Value Proposition</div>', unsafe_allow_html=True)
# # # # # # # #         st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7;color:var(--text)">{evp}</div>', unsafe_allow_html=True)
# # # # # # # #     caps = data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
# # # # # # # #     if caps:
# # # # # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>', unsafe_allow_html=True)
# # # # # # # #         c1, c2 = st.columns(2)
# # # # # # # #         for i, cap in enumerate(caps if isinstance(caps, list) else [caps]):
# # # # # # # #             (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
# # # # # # # #     risk = data.get("risk_mitigation") or data.get("RiskMitigation","")
# # # # # # # #     if risk:
# # # # # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>', unsafe_allow_html=True)
# # # # # # # #         if isinstance(risk, dict):
# # # # # # # #             for k, v in risk.items(): st.markdown(f"**{k}:** {v}")
# # # # # # # #         else: st.markdown(str(risk))
# # # # # # # #     render_json_pretty(data, "Deal Assurance Pack")


# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # #  SIDEBAR
# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # with st.sidebar:
# # # # # # # #     st.markdown("""
# # # # # # # #     <div style="padding:.75rem 0 1.25rem">
# # # # # # # #       <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;
# # # # # # # #                   color:#2563eb">🛡️ SecureITLab</div>
# # # # # # # #       <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;
# # # # # # # #                   text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
# # # # # # # #     </div>
# # # # # # # #     """, unsafe_allow_html=True)

# # # # # # # #     # ── Logout button ─────────────────────────────────────────────────────────
# # # # # # # #     if st.button("🚪 Logout", use_container_width=True):
# # # # # # # #         st.session_state.logged_in = False
# # # # # # # #         st.session_state.login_error = ""
# # # # # # # #         st.rerun()

# # # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # # # #     # ── Find Jobs button ──────────────────────────────────────────────────────
# # # # # # # #     st.markdown("**🔍 Find New Jobs**")
# # # # # # # #     st.caption(
# # # # # # # #         "Runs scraper → checks MongoDB for duplicates → "
# # # # # # # #         "runs AI pipeline only on NEW jobs → stores in MongoDB"
# # # # # # # #     )

# # # # # # # #     find_jobs_clicked = st.button(
# # # # # # # #         "🔍  Find Jobs",
# # # # # # # #         disabled=st.session_state.pipeline_running,
# # # # # # # #         use_container_width=True,
# # # # # # # #         type="primary",
# # # # # # # #     )

# # # # # # # #     if st.session_state.pipeline_running:
# # # # # # # #         st.markdown("""
# # # # # # # #         <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
# # # # # # # #                     padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;
# # # # # # # #                     font-size:.8rem;color:#1d4ed8">
# # # # # # # #           ⏳ Pipeline is running… check log below
# # # # # # # #         </div>""", unsafe_allow_html=True)

# # # # # # # #     if st.session_state.pipeline_done and st.session_state.pipeline_result:
# # # # # # # #         res = st.session_state.pipeline_result
# # # # # # # #         ok  = res.get("success", False)
# # # # # # # #         bg  = "#ecfdf5" if ok else "#fef2f2"
# # # # # # # #         bc  = "#bbf7d0" if ok else "#fecaca"
# # # # # # # #         tc  = "#15803d" if ok else "#b91c1c"
# # # # # # # #         ic  = "✅" if ok else "❌"
# # # # # # # #         st.markdown(f"""
# # # # # # # #         <div style="background:{bg};border:1px solid {bc};border-radius:8px;
# # # # # # # #                     padding:.75rem;margin-top:.5rem;font-size:.82rem">
# # # # # # # #           <div style="font-weight:700;color:{tc};margin-bottom:.4rem">{ic} Last Run</div>
# # # # # # # #           <div>📦 Scraped: <b>{res.get('scraped',0)}</b></div>
# # # # # # # #           <div>🆕 New jobs: <b>{res.get('new_jobs',0)}</b></div>
# # # # # # # #           <div>⏭️ Already in DB (skipped): <b>{res.get('skipped_db',0)}</b></div>
# # # # # # # #           <div>🤖 Processed by AI: <b>{res.get('processed',0)}</b></div>
# # # # # # # #           {f'<div style="color:{tc};margin-top:.3rem">⚠️ {res.get("error","")}</div>' if res.get("error") else ""}
# # # # # # # #         </div>""", unsafe_allow_html=True)

# # # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # # # #     # ── Job list ──────────────────────────────────────────────────────────────
# # # # # # # #     try:
# # # # # # # #         all_jobs = load_all_jobs()
# # # # # # # #     except Exception as e:
# # # # # # # #         st.error(f"MongoDB connection failed: {e}")
# # # # # # # #         st.stop()

# # # # # # # #     if not all_jobs:
# # # # # # # #         st.warning("No jobs in MongoDB yet. Click **Find Jobs** to scrape and process.")
# # # # # # # #         st.stop()

# # # # # # # #     st.markdown(
# # # # # # # #         f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">'
# # # # # # # #         f'{len(all_jobs)} jobs in database</div>',
# # # # # # # #         unsafe_allow_html=True,
# # # # # # # #     )

# # # # # # # #     search   = st.text_input("🔍 Filter by company / role", placeholder="e.g. Bounteous")
# # # # # # # #     filtered = [j for j in all_jobs
# # # # # # # #                 if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

# # # # # # # #     def job_label(j):
# # # # # # # #         score = j.get("opp_score")
# # # # # # # #         s_str = f" [{score}/10]" if score else ""
# # # # # # # #         ok    = "✅" if j.get("pipeline_ok") else "❌"
# # # # # # # #         return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s_str}"

# # # # # # # #     if not filtered:
# # # # # # # #         st.warning("No matching jobs.")
# # # # # # # #         st.stop()

# # # # # # # #     sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
# # # # # # # #     sel_idx     = [job_label(j) for j in filtered].index(sel_label)
# # # # # # # #     selected_id = str(filtered[sel_idx]["_id"])

# # # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # # # #     ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
# # # # # # # #     total_c  = sum(j.get("contacts_found",0) for j in all_jobs)
# # # # # # # #     st.markdown(f"""
# # # # # # # #     <div style="font-size:.75rem;color:#64748b">
# # # # # # # #       <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
# # # # # # # #       <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
# # # # # # # #     </div>""", unsafe_allow_html=True)

# # # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # # # #     if st.button("🔄 Refresh Data", use_container_width=True):
# # # # # # # #         st.cache_data.clear()
# # # # # # # #         st.rerun()


# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # #  FIND JOBS — background thread
# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # if find_jobs_clicked and not st.session_state.pipeline_running:
# # # # # # # #     with _thread_lock:
# # # # # # # #         _thread_logs.clear()
# # # # # # # #         _thread_result[0] = None
# # # # # # # #         _thread_done[0]   = False

# # # # # # # #     st.session_state.pipeline_running = True
# # # # # # # #     st.session_state.pipeline_done    = False
# # # # # # # #     st.session_state.pipeline_logs    = []
# # # # # # # #     st.session_state.pipeline_result  = None
# # # # # # # #     st.cache_data.clear()

# # # # # # # #     def _run_pipeline_bg():
# # # # # # # #         try:
# # # # # # # #             import main as _main

# # # # # # # #             def _cb(msg: str):
# # # # # # # #                 with _thread_lock:
# # # # # # # #                     _thread_logs.append(f"{datetime.now().strftime('%H:%M:%S')} | {msg}")

# # # # # # # #             result = _main.run_pipeline(progress_callback=_cb)
# # # # # # # #             with _thread_lock:
# # # # # # # #                 _thread_result[0] = result
# # # # # # # #         except Exception as e:
# # # # # # # #             with _thread_lock:
# # # # # # # #                 _thread_logs.append(f"❌ Unexpected error: {e}")
# # # # # # # #                 _thread_result[0] = {
# # # # # # # #                     "success": False, "error": str(e),
# # # # # # # #                     "scraped": 0, "new_jobs": 0, "skipped_db": 0, "processed": 0,
# # # # # # # #                 }
# # # # # # # #         finally:
# # # # # # # #             with _thread_lock:
# # # # # # # #                 _thread_done[0] = True

# # # # # # # #     threading.Thread(target=_run_pipeline_bg, daemon=True).start()
# # # # # # # #     st.rerun()

# # # # # # # # # ── Sync thread state → session_state ────────────────────────────────────────
# # # # # # # # with _thread_lock:
# # # # # # # #     if _thread_logs:
# # # # # # # #         st.session_state.pipeline_logs = list(_thread_logs)
# # # # # # # #     if _thread_done[0] and _thread_result[0] is not None:
# # # # # # # #         st.session_state.pipeline_result  = _thread_result[0]
# # # # # # # #         st.session_state.pipeline_running = False
# # # # # # # #         st.session_state.pipeline_done    = True

# # # # # # # # # ── Live log panel ────────────────────────────────────────────────────────────
# # # # # # # # if st.session_state.pipeline_running or (
# # # # # # # #         st.session_state.pipeline_done and st.session_state.pipeline_logs):
# # # # # # # #     heading = "⏳ Pipeline running — live log…" if st.session_state.pipeline_running \
# # # # # # # #               else "📋 Last pipeline run log"
# # # # # # # #     with st.expander(heading, expanded=st.session_state.pipeline_running):
# # # # # # # #         log_text = "\n".join(st.session_state.pipeline_logs[-100:]) \
# # # # # # # #                    if st.session_state.pipeline_logs else "Starting…"
# # # # # # # #         st.markdown(f'<div class="pipeline-log">{log_text}</div>', unsafe_allow_html=True)
# # # # # # # #         if st.session_state.pipeline_running:
# # # # # # # #             time.sleep(2)
# # # # # # # #             st.rerun()


# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # #  MAIN CONTENT — selected job detail
# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # with st.spinner("Loading job from MongoDB…"):
# # # # # # # #     job = load_job(selected_id)

# # # # # # # # if not job:
# # # # # # # #     st.error("Could not load job document.")
# # # # # # # #     st.stop()

# # # # # # # # company   = job.get("company",  "Unknown")
# # # # # # # # role      = job.get("role",     "Unknown")
# # # # # # # # opp_score = job.get("opp_score")
# # # # # # # # p_ok      = job.get("pipeline_ok", False)
# # # # # # # # p_min     = job.get("pipeline_min", "?")
# # # # # # # # c_found   = job.get("contacts_found", 0)
# # # # # # # # c_cov     = job.get("coverage_score")
# # # # # # # # c_domain  = job.get("contact_domain","")
# # # # # # # # run_at    = job.get("run_at","")

# # # # # # # # # ── Header ────────────────────────────────────────────────────────────────────
# # # # # # # # st.markdown(f"""
# # # # # # # # <div style="margin-bottom:1.75rem">
# # # # # # # #   <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;
# # # # # # # #               letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
# # # # # # # #     Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
# # # # # # # #   </div>
# # # # # # # #   <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;
# # # # # # # #              color:#0f172a;margin:0;line-height:1.15">{role}</h1>
# # # # # # # #   <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
# # # # # # # #     @ <span style="color:#334155;font-weight:600">{company}</span>
# # # # # # # #     {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
# # # # # # # #   </div>
# # # # # # # # </div>
# # # # # # # # """, unsafe_allow_html=True)

# # # # # # # # # ── Metric tiles ──────────────────────────────────────────────────────────────
# # # # # # # # try:
# # # # # # # #     sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
# # # # # # # #     sc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # # # # # except Exception:
# # # # # # # #     sc = "#2563eb"

# # # # # # # # st.markdown(f"""
# # # # # # # # <div class="metric-row">
# # # # # # # #   <div class="metric-tile">
# # # # # # # #     <div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div>
# # # # # # # #     <div class="lbl">Opportunity Score</div>
# # # # # # # #   </div>
# # # # # # # #   <div class="metric-tile">
# # # # # # # #     <div class="val">{c_found}</div>
# # # # # # # #     <div class="lbl">Contacts Found</div>
# # # # # # # #   </div>
# # # # # # # #   <div class="metric-tile">
# # # # # # # #     <div class="val">{f"{c_cov}%" if c_cov else "—"}</div>
# # # # # # # #     <div class="lbl">Contact Coverage</div>
# # # # # # # #   </div>
# # # # # # # #   <div class="metric-tile">
# # # # # # # #     <div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">
# # # # # # # #       {'✅ OK' if p_ok else '❌ Failed'}
# # # # # # # #     </div>
# # # # # # # #     <div class="lbl">Pipeline ({p_min} min)</div>
# # # # # # # #   </div>
# # # # # # # # </div>
# # # # # # # # """, unsafe_allow_html=True)

# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # #  TABS
# # # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # # tabs = st.tabs([
# # # # # # # #     "📋 Job & Enrichment",
# # # # # # # #     "🎯 Service Mapping",
# # # # # # # #     "🔍 Fit / Gap",
# # # # # # # #     "🛠️ Capability & Plans",
# # # # # # # #     "📦 Deal Assurance",
# # # # # # # #     "✉️ Outreach Emails",
# # # # # # # #     "👥 Contacts",
# # # # # # # #     "✅ QA Gates",
# # # # # # # #     "🗄️ Raw Data",
# # # # # # # # ])

# # # # # # # # # ── Tab 1: Job Research + Enrichment ─────────────────────────────────────────
# # # # # # # # with tabs[0]:
# # # # # # # #     col1, col2 = st.columns([1, 1])
# # # # # # # #     with col1:
# # # # # # # #         st.markdown('<div class="section-label">📄 Job Research</div>', unsafe_allow_html=True)
# # # # # # # #         jr = as_dict(job.get("agent_job_research") or {})
# # # # # # # #         if jr:
# # # # # # # #             for label, keys in [
# # # # # # # #                 ("Job Role",        ["job_role","Job Role","role","title"]),
# # # # # # # #                 ("Company",         ["company_name","Company Name","company"]),
# # # # # # # #                 ("Location",        ["location","Location"]),
# # # # # # # #                 ("Organization URL",["organization_url","Organization URL","url"]),
# # # # # # # #             ]:
# # # # # # # #                 val = next((jr.get(k) for k in keys if jr.get(k)), None)
# # # # # # # #                 if val: st.markdown(f"**{label}:** {val}")
# # # # # # # #             desc = jr.get("job_description") or jr.get("Job Description","")
# # # # # # # #             if desc:
# # # # # # # #                 st.markdown("**Job Description:**")
# # # # # # # #                 st.markdown(
# # # # # # # #                     f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;'
# # # # # # # #                     f'max-height:300px;overflow-y:auto;color:var(--text)">{desc[:2000]}</div>',
# # # # # # # #                     unsafe_allow_html=True,
# # # # # # # #                 )
# # # # # # # #             render_json_pretty(jr, "Job Research")
# # # # # # # #         else:
# # # # # # # #             st.info("No job research data.")
# # # # # # # #     with col2:
# # # # # # # #         st.markdown('<div class="section-label">🏢 Company Enrichment</div>', unsafe_allow_html=True)
# # # # # # # #         enr = as_dict(job.get("agent_enrichment") or {})
# # # # # # # #         if enr:
# # # # # # # #             for label, keys in [
# # # # # # # #                 ("Industry",          ["industry","Industry"]),
# # # # # # # #                 ("Company Size",      ["company_size","size","Company Size"]),
# # # # # # # #                 ("Regulatory Env",    ["regulatory_environment","regulatory"]),
# # # # # # # #                 ("Certifications",    ["certifications","Certifications"]),
# # # # # # # #                 ("Tech Stack",        ["tech_stack","technology_stack"]),
# # # # # # # #                 ("Security Maturity", ["security_maturity","maturity"]),
# # # # # # # #             ]:
# # # # # # # #                 val = next((enr.get(k) for k in keys if enr.get(k)), None)
# # # # # # # #                 if val:
# # # # # # # #                     if isinstance(val, list): val = ", ".join(str(v) for v in val)
# # # # # # # #                     st.markdown(f"**{label}:** {safe_str(val, 200)}")
# # # # # # # #             render_json_pretty(enr, "Enrichment")
# # # # # # # #         else:
# # # # # # # #             st.info("No enrichment data.")

# # # # # # # # # ── Tab 2: Service Mapping ────────────────────────────────────────────────────
# # # # # # # # with tabs[1]:
# # # # # # # #     st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>', unsafe_allow_html=True)
# # # # # # # #     render_service_mapping(job.get("agent_service_mapping"))

# # # # # # # # # ── Tab 3: Fit / Gap Analysis ─────────────────────────────────────────────────
# # # # # # # # with tabs[2]:
# # # # # # # #     fg = as_dict(job.get("agent_fit_gap") or {})
# # # # # # # #     if opp_score:
# # # # # # # #         try:
# # # # # # # #             sn = float(str(opp_score).split("/")[0])
# # # # # # # #             bc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # # # # #             bp = int(sn / 10 * 100)
# # # # # # # #             st.markdown(f"""
# # # # # # # #             <div style="margin-bottom:1.5rem">
# # # # # # # #               <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
# # # # # # # #                 <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">Opportunity Score</span>
# # # # # # # #                 <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
# # # # # # # #               </div>
# # # # # # # #               <div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%">
# # # # # # # #                 <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
# # # # # # # #               </div>
# # # # # # # #             </div>""", unsafe_allow_html=True)
# # # # # # # #         except Exception:
# # # # # # # #             pass
# # # # # # # #     st.markdown('<div class="section-label">📊 Service Classifications</div>', unsafe_allow_html=True)
# # # # # # # #     services = []
# # # # # # # #     if isinstance(fg, dict):
# # # # # # # #         for k in ("services","classifications","service_classifications","items","fit_gap"):
# # # # # # # #             v = fg.get(k)
# # # # # # # #             if isinstance(v, list): services = v; break
# # # # # # # #         if not services and (fg.get("service") or fg.get("Service")):
# # # # # # # #             services = [fg]
# # # # # # # #     elif isinstance(fg, list):
# # # # # # # #         services = fg
# # # # # # # #     if services:
# # # # # # # #         buckets = {"STRONG FIT": [], "PARTIAL FIT": [], "GAP": []}
# # # # # # # #         other   = []
# # # # # # # #         for s in services:
# # # # # # # #             if not isinstance(s, dict): continue
# # # # # # # #             fit = (s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
# # # # # # # #             if "STRONG" in fit: buckets["STRONG FIT"].append(s)
# # # # # # # #             elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
# # # # # # # #             elif "GAP" in fit: buckets["GAP"].append(s)
# # # # # # # #             else: other.append(s)
# # # # # # # #         c1, c2, c3 = st.columns(3)
# # # # # # # #         cm  = {"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
# # # # # # # #         bgm = {"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
# # # # # # # #         bdm = {"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
# # # # # # # #         for col, (fl, items) in zip([c1, c2, c3], buckets.items()):
# # # # # # # #             col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>', unsafe_allow_html=True)
# # # # # # # #             for s in items:
# # # # # # # #                 svc  = s.get("service") or s.get("Service") or s.get("name","")
# # # # # # # #                 just = s.get("justification") or s.get("reason","")
# # # # # # # #                 col.markdown(
# # # # # # # #                     f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem">'
# # # # # # # #                     f'<div style="font-weight:600;color:#0f172a">{svc}</div>'
# # # # # # # #                     f'<div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>',
# # # # # # # #                     unsafe_allow_html=True,
# # # # # # # #                 )
# # # # # # # #     elif fg:
# # # # # # # #         st.json(fg)
# # # # # # # #     else:
# # # # # # # #         st.info("No fit/gap data.")
# # # # # # # #     render_json_pretty(job.get("agent_fit_gap"), "Fit/Gap Analysis")

# # # # # # # # # ── Tab 4: Capability + Micro-plans ──────────────────────────────────────────
# # # # # # # # with tabs[3]:
# # # # # # # #     col1, col2 = st.columns([1, 1])
# # # # # # # #     with col1:
# # # # # # # #         st.markdown('<div class="section-label">🔧 Capability Improvements</div>', unsafe_allow_html=True)
# # # # # # # #         cap = job.get("agent_capability") or {}
# # # # # # # #         items_cap = cap if isinstance(cap, list) else []
# # # # # # # #         if not items_cap and isinstance(cap, dict):
# # # # # # # #             for k in ("improvements","recommendations","capabilities","items"):
# # # # # # # #                 v = cap.get(k)
# # # # # # # #                 if isinstance(v, list): items_cap = v; break
# # # # # # # #             if not items_cap: items_cap = [cap]
# # # # # # # #         for item in items_cap:
# # # # # # # #             if not isinstance(item, dict): continue
# # # # # # # #             title  = item.get("title") or item.get("gap") or item.get("service","")
# # # # # # # #             rec    = item.get("recommendation") or item.get("steps","")
# # # # # # # #             effort = item.get("build_effort") or item.get("effort","")
# # # # # # # #             demand = item.get("market_demand") or item.get("priority","")
# # # # # # # #             st.markdown(f"""
# # # # # # # #             <div class="sil-card" style="margin-bottom:.6rem">
# # # # # # # #               <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem;color:var(--text)">{title}</div>
# # # # # # # #               <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec, 250)}</div>
# # # # # # # #               {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
# # # # # # # #             </div>""", unsafe_allow_html=True)
# # # # # # # #         if not items_cap:
# # # # # # # #             render_json_pretty(cap, "Capability Improvement")
# # # # # # # #     with col2:
# # # # # # # #         st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>', unsafe_allow_html=True)
# # # # # # # #         render_microplans(job.get("agent_microplans"))

# # # # # # # # # ── Tab 5: Deal Assurance Pack ────────────────────────────────────────────────
# # # # # # # # with tabs[4]:
# # # # # # # #     render_deal_assurance(job.get("agent_deal_assurance"))

# # # # # # # # # ── Tab 6: Outreach Emails ────────────────────────────────────────────────────
# # # # # # # # with tabs[5]:
# # # # # # # #     st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>', unsafe_allow_html=True)
# # # # # # # #     emails_src = job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
# # # # # # # #     oq = as_dict(job.get("agent_outreach_qa") or {})
# # # # # # # #     improved = (oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
# # # # # # # #     if improved:
# # # # # # # #         st.info("⚡ Showing QA-improved versions where available")
# # # # # # # #         render_emails(improved)
# # # # # # # #         with st.expander("📬 Original (pre-QA) versions"):
# # # # # # # #             render_emails(emails_src)
# # # # # # # #     else:
# # # # # # # #         render_emails(emails_src)

# # # # # # # # # ── Tab 7: Contacts ───────────────────────────────────────────────────────────
# # # # # # # # with tabs[6]:
# # # # # # # #     contacts        = job.get("contacts") or []
# # # # # # # #     contact_sources = job.get("contact_sources") or []
# # # # # # # #     pri = [c for c in contacts if c.get("priority") == "Primary"]
# # # # # # # #     sec = [c for c in contacts if c.get("priority") == "Secondary"]
# # # # # # # #     ter = [c for c in contacts if c.get("priority") == "Tertiary"]
# # # # # # # #     gen = [c for c in contacts if c.get("priority") == "General"]
# # # # # # # #     st.markdown(f"""
# # # # # # # #     <div class="metric-row" style="margin-bottom:1.5rem">
# # # # # # # #       <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
# # # # # # # #       <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
# # # # # # # #       <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
# # # # # # # #       <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
# # # # # # # #     </div>""", unsafe_allow_html=True)
# # # # # # # #     if contact_sources:
# # # # # # # #         st.markdown('Sources: ' + " ".join(badge(s,"blue") for s in contact_sources), unsafe_allow_html=True)
# # # # # # # #         st.markdown("")
# # # # # # # #     missing = job.get("missing_roles") or []
# # # # # # # #     if missing:
# # # # # # # #         st.markdown('⚠️ Missing roles: ' + " ".join(badge(r,"red") for r in missing), unsafe_allow_html=True)
# # # # # # # #         st.markdown("")
# # # # # # # #     if contacts:
# # # # # # # #         pri_filter = st.multiselect(
# # # # # # # #             "Filter by priority",
# # # # # # # #             ["Primary","Secondary","Tertiary","General"],
# # # # # # # #             default=["Primary","Secondary","Tertiary","General"],
# # # # # # # #         )
# # # # # # # #         shown = [c for c in contacts if c.get("priority","General") in pri_filter]
# # # # # # # #         render_contacts(shown, f"Contacts ({len(shown)} shown)")
# # # # # # # #         agent_contacts = job.get("agent_prospect_contacts") or {}
# # # # # # # #         if agent_contacts:
# # # # # # # #             with st.expander("🤖 CrewAI Agent's Contact Search"):
# # # # # # # #                 if isinstance(agent_contacts, dict):
# # # # # # # #                     ac_list = agent_contacts.get("contacts") or []
# # # # # # # #                     if ac_list: render_contacts(ac_list, "Agent Contacts")
# # # # # # # #                     else:       st.json(agent_contacts)
# # # # # # # #                 else:           st.json(agent_contacts)
# # # # # # # #     else:
# # # # # # # #         st.info("No contacts found for this job.")

# # # # # # # # # ── Tab 8: QA Gates ───────────────────────────────────────────────────────────
# # # # # # # # with tabs[7]:
# # # # # # # #     st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>', unsafe_allow_html=True)
# # # # # # # #     col1, col2 = st.columns(2)
# # # # # # # #     for i, (label, key) in enumerate([
# # # # # # # #         ("Research QA",       "agent_research_qa"),
# # # # # # # #         ("Service Mapping QA","agent_mapping_qa"),
# # # # # # # #         ("Deal Assurance QA", "agent_assurance_qa"),
# # # # # # # #         ("Outreach Email QA", "agent_outreach_qa"),
# # # # # # # #     ]):
# # # # # # # #         with (col1 if i % 2 == 0 else col2):
# # # # # # # #             render_qa_block(job.get(key), label)
# # # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # # # #     st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>', unsafe_allow_html=True)
# # # # # # # #     enf = as_dict(job.get("agent_prospect_enforcer") or {})
# # # # # # # #     if enf:
# # # # # # # #         cov  = enf.get("coverage_score","?")
# # # # # # # #         miss = enf.get("missing_roles",[])
# # # # # # # #         note = enf.get("note","")
# # # # # # # #         ec   = enf.get("contacts",[])
# # # # # # # #         x1, x2, x3 = st.columns(3)
# # # # # # # #         x1.metric("Coverage Score",    f"{cov}%")
# # # # # # # #         x2.metric("Missing Roles",     len(miss))
# # # # # # # #         x3.metric("Contacts Verified", len(ec))
# # # # # # # #         if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
# # # # # # # #         if note: st.caption(note)
# # # # # # # #     else:
# # # # # # # #         st.info("No enforcer data.")

# # # # # # # # # ── Tab 9: Raw Data ───────────────────────────────────────────────────────────
# # # # # # # # with tabs[8]:
# # # # # # # #     st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>', unsafe_allow_html=True)
# # # # # # # #     st.caption("All fields stored in the `jobs` collection for this document.")
# # # # # # # #     rows = []
# # # # # # # #     for k, v in job.items():
# # # # # # # #         if k == "_id": continue
# # # # # # # #         rows.append({"Field": k, "Type": type(v).__name__, "Len": len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
# # # # # # # #     hc1, hc2, hc3 = st.columns([3,1,1])
# # # # # # # #     hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
# # # # # # # #     for r in rows:
# # # # # # # #         rc1, rc2, rc3 = st.columns([3,1,1])
# # # # # # # #         rc1.code(r["Field"], language=None)
# # # # # # # #         rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>', unsafe_allow_html=True)
# # # # # # # #         rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>',  unsafe_allow_html=True)
# # # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # # # #     for label, key in [
# # # # # # # #         ("Job Research","agent_job_research"), ("Enrichment","agent_enrichment"),
# # # # # # # #         ("Service Mapping","agent_service_mapping"), ("Fit/Gap Analysis","agent_fit_gap"),
# # # # # # # #         ("Capability","agent_capability"), ("Micro-Plans","agent_microplans"),
# # # # # # # #         ("Deal Assurance","agent_deal_assurance"), ("Outreach Emails","agent_outreach_emails"),
# # # # # # # #         ("Prospect Contacts","agent_prospect_contacts"), ("Prospect Enforcer","agent_prospect_enforcer"),
# # # # # # # #         ("Research QA","agent_research_qa"), ("Mapping QA","agent_mapping_qa"),
# # # # # # # #         ("Assurance QA","agent_assurance_qa"), ("Outreach QA","agent_outreach_qa"),
# # # # # # # #         ("Contacts (5-source)","contacts"),
# # # # # # # #     ]:
# # # # # # # #         data = job.get(key)
# # # # # # # #         if data:
# # # # # # # #             with st.expander(f"📄 {label}"):
# # # # # # # #                 st.code(json.dumps(data, indent=2, default=str), language="json")































# # # # # # # """
# # # # # # # ╔══════════════════════════════════════════════════════════╗
# # # # # # # ║   SecureITLab Pipeline Dashboard — Streamlit             ║
# # # # # # # ║   Reads from MongoDB → secureitlab_job_pipeline          ║
# # # # # # # ╠══════════════════════════════════════════════════════════╣
# # # # # # # ║  Install: pip install streamlit pymongo python-dotenv    ║
# # # # # # # ║  Run:     streamlit run streamlit_dashboard.py           ║
# # # # # # # ╚══════════════════════════════════════════════════════════╝
# # # # # # # """

# # # # # # # import io
# # # # # # # import re
# # # # # # # import streamlit as st
# # # # # # # from pymongo import MongoClient
# # # # # # # import json
# # # # # # # import threading
# # # # # # # import time
# # # # # # # from datetime import datetime, timezone

# # # # # # # # ── Thread-safe shared state ──────────────────────────────────────────────────
# # # # # # # _thread_logs   = []
# # # # # # # _thread_result = [None]
# # # # # # # _thread_done   = [False]
# # # # # # # _thread_lock   = threading.Lock()

# # # # # # # # ── Page config ───────────────────────────────────────────────────────────────
# # # # # # # st.set_page_config(
# # # # # # #     page_title="SecureITLab Pipeline",
# # # # # # #     page_icon="🛡️",
# # # # # # #     layout="wide",
# # # # # # #     initial_sidebar_state="expanded",
# # # # # # # )

# # # # # # # # ── LOGIN CREDENTIALS (change these) ─────────────────────────────────────────
# # # # # # # LOGIN_USERNAME = "admin"
# # # # # # # LOGIN_PASSWORD = "secureitlab2024"

# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # #  GLOBAL CSS (login + dashboard share this)
# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # st.markdown("""
# # # # # # # <style>

# # # # # # # @import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');

# # # # # # # :root {
# # # # # # #     --bg:        #f4f7fb;
# # # # # # #     --surface:   #ffffff;
# # # # # # #     --surface2:  #eef2f7;
# # # # # # #     --border:    #d9e2ec;
# # # # # # #     --accent:    #2563eb;
# # # # # # #     --accent2:   #7c3aed;
# # # # # # #     --green:     #16a34a;
# # # # # # #     --yellow:    #f59e0b;
# # # # # # #     --red:       #dc2626;
# # # # # # #     --text:      #0f172a;
# # # # # # #     --muted:     #64748b;
# # # # # # # }

# # # # # # # html, body, [class*="css"] {
# # # # # # #     background-color: var(--bg) !important;
# # # # # # #     color: var(--text) !important;
# # # # # # #     font-family: 'DM Sans', sans-serif !important;
# # # # # # # }

# # # # # # # /* ── LOGIN PAGE ─────────────────────────────────────────── */
# # # # # # # .login-wrap {
# # # # # # #     display: flex;
# # # # # # #     align-items: center;
# # # # # # #     justify-content: center;
# # # # # # #     min-height: 80vh;
# # # # # # # }
# # # # # # # .login-card {
# # # # # # #     background: var(--surface);
# # # # # # #     border: 1px solid var(--border);
# # # # # # #     border-radius: 20px;
# # # # # # #     padding: 3rem 3.5rem;
# # # # # # #     width: 100%;
# # # # # # #     max-width: 420px;
# # # # # # #     box-shadow: 0 20px 60px rgba(37,99,235,0.08);
# # # # # # #     text-align: center;
# # # # # # # }
# # # # # # # .login-logo {
# # # # # # #     font-family: 'Syne', sans-serif;
# # # # # # #     font-size: 1.6rem;
# # # # # # #     font-weight: 800;
# # # # # # #     color: var(--accent);
# # # # # # #     margin-bottom: .25rem;
# # # # # # # }
# # # # # # # .login-subtitle {
# # # # # # #     font-size: .75rem;
# # # # # # #     color: var(--muted);
# # # # # # #     letter-spacing: .12em;
# # # # # # #     text-transform: uppercase;
# # # # # # #     margin-bottom: 2.5rem;
# # # # # # # }
# # # # # # # .login-error {
# # # # # # #     background: #fef2f2;
# # # # # # #     border: 1px solid #fecaca;
# # # # # # #     border-radius: 8px;
# # # # # # #     padding: .65rem 1rem;
# # # # # # #     color: #b91c1c;
# # # # # # #     font-size: .85rem;
# # # # # # #     margin-top: 1rem;
# # # # # # # }
# # # # # # # .login-divider {
# # # # # # #     width: 40px;
# # # # # # #     height: 3px;
# # # # # # #     background: linear-gradient(90deg, var(--accent), var(--accent2));
# # # # # # #     border-radius: 2px;
# # # # # # #     margin: 0 auto 2rem;
# # # # # # # }

# # # # # # # /* ── DASHBOARD ──────────────────────────────────────────── */
# # # # # # # [data-testid="stSidebar"] {
# # # # # # #     background: var(--surface) !important;
# # # # # # #     border-right: 1px solid var(--border) !important;
# # # # # # # }
# # # # # # # [data-testid="stSidebar"] * { color: var(--text) !important; }

# # # # # # # .main .block-container { padding: 2rem 2rem 3rem !important; }

# # # # # # # h1, h2, h3, h4 {
# # # # # # #     font-family: 'Syne', sans-serif !important;
# # # # # # #     color: var(--text) !important;
# # # # # # # }

# # # # # # # .sil-card {
# # # # # # #     background: var(--surface);
# # # # # # #     border: 1px solid var(--border);
# # # # # # #     border-radius: 14px;
# # # # # # #     padding: 1.25rem 1.5rem;
# # # # # # #     margin-bottom: 1rem;
# # # # # # #     transition: all 0.25s ease;
# # # # # # # }
# # # # # # # .sil-card:hover {
# # # # # # #     transform: translateY(-2px);
# # # # # # #     box-shadow: 0 8px 22px rgba(0,0,0,0.05);
# # # # # # # }
# # # # # # # .sil-card-accent { border-left: 4px solid var(--accent); }

# # # # # # # .metric-row { display:flex; gap:1rem; flex-wrap:wrap; margin-bottom:1.5rem; }
# # # # # # # .metric-tile {
# # # # # # #     flex:1; min-width:140px;
# # # # # # #     background:var(--surface2);
# # # # # # #     border:1px solid var(--border);
# # # # # # #     border-radius:12px;
# # # # # # #     padding:1rem; text-align:center;
# # # # # # #     transition:all .25s ease;
# # # # # # # }
# # # # # # # .metric-tile:hover { transform:translateY(-3px); box-shadow:0 10px 24px rgba(0,0,0,0.06); }
# # # # # # # .metric-tile .val { font-family:'Syne',sans-serif; font-size:2rem; font-weight:800; color:var(--accent); }
# # # # # # # .metric-tile .lbl { font-size:.72rem; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; }

# # # # # # # .badge { padding:.25rem .7rem; border-radius:20px; font-size:.72rem; font-weight:600; font-family:'DM Mono',monospace; }
# # # # # # # .badge-green  { background:#ecfdf5; color:#15803d; }
# # # # # # # .badge-yellow { background:#fffbeb; color:#b45309; }
# # # # # # # .badge-red    { background:#fef2f2; color:#b91c1c; }
# # # # # # # .badge-blue   { background:#eff6ff; color:#1d4ed8; }
# # # # # # # .badge-purple { background:#f5f3ff; color:#6d28d9; }

# # # # # # # .contact-card {
# # # # # # #     background:var(--surface2); border:1px solid var(--border);
# # # # # # #     border-radius:10px; padding:1rem; margin-bottom:.8rem;
# # # # # # # }
# # # # # # # .contact-name  { font-family:'Syne',sans-serif; font-weight:700; color:var(--text); }
# # # # # # # .contact-title { color:var(--muted); font-size:.85rem; }
# # # # # # # .contact-email { font-family:'DM Mono',monospace; color:var(--accent); font-size:.8rem; }

# # # # # # # .email-box {
# # # # # # #     background: #f8fafc;
# # # # # # #     border: 1px solid var(--border);
# # # # # # #     border-radius: 10px;
# # # # # # #     padding: 1rem 1.25rem;
# # # # # # #     font-size: .9rem;
# # # # # # #     line-height: 1.65;
# # # # # # #     white-space: pre-wrap;
# # # # # # #     color: var(--text);
# # # # # # # }
# # # # # # # .email-subject { font-family:'Syne',sans-serif; font-weight:700; color:var(--accent); margin-bottom:.5rem; }

# # # # # # # .section-label {
# # # # # # #     font-family:'DM Mono',monospace; font-size:.72rem;
# # # # # # #     text-transform:uppercase; letter-spacing:.12em;
# # # # # # #     color:var(--accent); margin-bottom:.6rem;
# # # # # # # }
# # # # # # # .sil-divider { border-top:1px solid var(--border); margin:1rem 0; }

# # # # # # # [data-testid="stExpander"] {
# # # # # # #     background: var(--surface) !important;
# # # # # # #     border: 1px solid var(--border) !important;
# # # # # # #     border-radius: 10px !important;
# # # # # # # }
# # # # # # # [data-testid="stSelectbox"] > div,
# # # # # # # [data-testid="stMultiSelect"] > div {
# # # # # # #     background: var(--surface2) !important;
# # # # # # #     border-color: var(--border) !important;
# # # # # # # }
# # # # # # # [data-testid="stTabs"] button {
# # # # # # #     font-family:'Syne',sans-serif !important;
# # # # # # #     font-weight:600 !important;
# # # # # # # }
# # # # # # # ::-webkit-scrollbar { width:6px; }
# # # # # # # ::-webkit-scrollbar-thumb { background:var(--border); border-radius:3px; }

# # # # # # # .pipeline-log {
# # # # # # #     background: #0f172a;
# # # # # # #     color: #e2e8f0;
# # # # # # #     border-radius: 10px;
# # # # # # #     padding: 1rem 1.25rem;
# # # # # # #     font-family: 'DM Mono', monospace;
# # # # # # #     font-size: .8rem;
# # # # # # #     line-height: 1.6;
# # # # # # #     max-height: 380px;
# # # # # # #     overflow-y: auto;
# # # # # # #     white-space: pre-wrap;
# # # # # # # }

# # # # # # # .fit-box {
# # # # # # #     border-radius: 8px;
# # # # # # #     padding: .75rem;
# # # # # # #     margin-bottom: .5rem;
# # # # # # #     font-size: .85rem;
# # # # # # # }

# # # # # # # /* Hide sidebar on login page */
# # # # # # # .hide-sidebar [data-testid="stSidebar"] { display: none !important; }
# # # # # # # .hide-sidebar .main .block-container { max-width: 480px; margin: 0 auto; }

# # # # # # # </style>
# # # # # # # """, unsafe_allow_html=True)


# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # #  SESSION STATE INIT
# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # for _k, _v in [
# # # # # # #     ("logged_in",        False),
# # # # # # #     ("login_error",      ""),
# # # # # # #     ("pipeline_running", False),
# # # # # # #     ("pipeline_logs",    []),
# # # # # # #     ("pipeline_result",  None),
# # # # # # #     ("pipeline_done",    False),
# # # # # # # ]:
# # # # # # #     if _k not in st.session_state:
# # # # # # #         st.session_state[_k] = _v


# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # #  LOGIN PAGE
# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # if not st.session_state.logged_in:

# # # # # # #     # Hide sidebar on login page
# # # # # # #     st.markdown("""
# # # # # # #     <script>
# # # # # # #     document.querySelector('[data-testid="stSidebar"]') &&
# # # # # # #     (document.querySelector('[data-testid="stSidebar"]').style.display = 'none');
# # # # # # #     </script>
# # # # # # #     """, unsafe_allow_html=True)

# # # # # # #     # Center the login card
# # # # # # #     _, col, _ = st.columns([1, 1.2, 1])

# # # # # # #     with col:
# # # # # # #         st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)

# # # # # # #         st.markdown("""
# # # # # # #         <div class="login-card">
# # # # # # #           <div class="login-logo">🛡️ SecureITLab</div>
# # # # # # #           <div class="login-subtitle">Pipeline Intelligence</div>
# # # # # # #           <div class="login-divider"></div>
# # # # # # #         </div>
# # # # # # #         """, unsafe_allow_html=True)

# # # # # # #         st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # # # #         with st.container():
# # # # # # #             username = st.text_input(
# # # # # # #                 "Username",
# # # # # # #                 placeholder="Enter username",
# # # # # # #                 key="login_username",
# # # # # # #             )
# # # # # # #             password = st.text_input(
# # # # # # #                 "Password",
# # # # # # #                 placeholder="Enter password",
# # # # # # #                 type="password",
# # # # # # #                 key="login_password",
# # # # # # #             )

# # # # # # #             st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # # # #             login_btn = st.button(
# # # # # # #                 "Sign In →",
# # # # # # #                 use_container_width=True,
# # # # # # #                 type="primary",
# # # # # # #             )

# # # # # # #             if login_btn:
# # # # # # #                 if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
# # # # # # #                     st.session_state.logged_in = True
# # # # # # #                     st.session_state.login_error = ""
# # # # # # #                     st.rerun()
# # # # # # #                 else:
# # # # # # #                     st.session_state.login_error = "Incorrect username or password."

# # # # # # #             if st.session_state.login_error:
# # # # # # #                 st.markdown(
# # # # # # #                     f'<div class="login-error">⚠️ {st.session_state.login_error}</div>',
# # # # # # #                     unsafe_allow_html=True,
# # # # # # #                 )

# # # # # # #         st.markdown(
# # # # # # #             "<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>"
# # # # # # #             "SecureITLab · Confidential</div>",
# # # # # # #             unsafe_allow_html=True,
# # # # # # #         )

# # # # # # #     st.stop()  # ← Stop here — nothing below renders until logged in


# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # #  EVERYTHING BELOW ONLY RUNS WHEN LOGGED IN
# # # # # # # # ══════════════════════════════════════════════════════════════════════════════


# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # #  EXCEL EXPORT HELPER
# # # # # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # # # # def build_contacts_excel(contacts: list, company: str, role: str):
# # # # # # #     """Build a formatted Excel workbook from contacts. Returns bytes or None."""
# # # # # # #     try:
# # # # # # #         from openpyxl import Workbook
# # # # # # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # # # # # #         from openpyxl.utils import get_column_letter
# # # # # # #     except ImportError:
# # # # # # #         return None

# # # # # # #     wb = Workbook()
# # # # # # #     ws = wb.active
# # # # # # #     ws.title = "Contacts"

# # # # # # #     NAVY  = "1E3A5F"
# # # # # # #     BLUE  = "2563EB"
# # # # # # #     LBLUE = "EFF6FF"
# # # # # # #     GREY  = "F8FAFC"
# # # # # # #     WHITE = "FFFFFF"

# # # # # # #     pri_colors = {
# # # # # # #         "Primary":   ("FEF2F2", "B91C1C"),
# # # # # # #         "Secondary": ("FFFBEB", "B45309"),
# # # # # # #         "Tertiary":  ("EFF6FF", "1D4ED8"),
# # # # # # #         "General":   ("F5F3FF", "6D28D9"),
# # # # # # #     }

# # # # # # #     thin   = Side(border_style="thin", color="D9E2EC")
# # # # # # #     border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # # # # # #     # Row 1: Title
# # # # # # #     ws.merge_cells("A1:H1")
# # # # # # #     c = ws["A1"]
# # # # # # #     c.value     = f"Contacts Export  —  {company}  |  {role}"
# # # # # # #     c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
# # # # # # #     c.fill      = PatternFill("solid", fgColor=NAVY)
# # # # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # # #     ws.row_dimensions[1].height = 30

# # # # # # #     # Row 2: Sub-title
# # # # # # #     ws.merge_cells("A2:H2")
# # # # # # #     c = ws["A2"]
# # # # # # #     c.value     = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(contacts)} contacts"
# # # # # # #     c.font      = Font(name="Arial", size=9, color="64748B")
# # # # # # #     c.fill      = PatternFill("solid", fgColor="F4F7FB")
# # # # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # # #     ws.row_dimensions[2].height = 18

# # # # # # #     # Row 3: Headers
# # # # # # #     headers    = ["#", "Priority", "Name", "Title / Role", "Company", "Email", "LinkedIn URL", "Source"]
# # # # # # #     col_widths = [4,    12,         24,     32,              22,        34,       42,              18]

# # # # # # #     for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
# # # # # # #         c = ws.cell(row=3, column=ci, value=h)
# # # # # # #         c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
# # # # # # #         c.fill      = PatternFill("solid", fgColor=BLUE)
# # # # # # #         c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # # #         c.border    = border
# # # # # # #         ws.column_dimensions[get_column_letter(ci)].width = w

# # # # # # #     ws.row_dimensions[3].height = 22

# # # # # # #     # Data rows
# # # # # # #     for ri, ct in enumerate(contacts, start=4):
# # # # # # #         prio           = ct.get("priority", "General")
# # # # # # #         bg_hex, fg_hex = pri_colors.get(prio, (WHITE, "0F172A"))
# # # # # # #         patterns       = ct.get("email_patterns", [])
# # # # # # #         email_val      = ct.get("email") or (patterns[0] + "  (pattern)" if patterns else "")
# # # # # # #         row_fill       = bg_hex if ri % 2 == 0 else GREY

# # # # # # #         for ci, val in enumerate([
# # # # # # #             ri - 3,
# # # # # # #             prio,
# # # # # # #             ct.get("name", ""),
# # # # # # #             ct.get("title", ""),
# # # # # # #             ct.get("company", ""),
# # # # # # #             email_val,
# # # # # # #             ct.get("linkedin_url", ""),
# # # # # # #             ct.get("source", ""),
# # # # # # #         ], 1):
# # # # # # #             cell = ws.cell(row=ri, column=ci, value=val)
# # # # # # #             cell.font      = Font(name="Arial", size=9,
# # # # # # #                                   bold=(ci == 2),
# # # # # # #                                   color=fg_hex if ci == 2 else "0F172A")
# # # # # # #             cell.fill      = PatternFill("solid", fgColor=row_fill)
# # # # # # #             cell.alignment = Alignment(vertical="center", wrap_text=(ci in [4, 7]))
# # # # # # #             cell.border    = border

# # # # # # #         ws.row_dimensions[ri].height = 18

# # # # # # #     ws.freeze_panes   = "A4"
# # # # # # #     ws.auto_filter.ref = f"A3:H{3 + len(contacts)}"

# # # # # # #     # Summary sheet
# # # # # # #     ws2 = wb.create_sheet("Summary")
# # # # # # #     ws2.merge_cells("A1:C1")
# # # # # # #     t = ws2["A1"]
# # # # # # #     t.value     = "Export Summary"
# # # # # # #     t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
# # # # # # #     t.fill      = PatternFill("solid", fgColor=NAVY)
# # # # # # #     t.alignment = Alignment(horizontal="center")

# # # # # # #     for r, (lbl, val) in enumerate([
# # # # # # #         ("Company",        company),
# # # # # # #         ("Role",           role),
# # # # # # #         ("Total Contacts", len(contacts)),
# # # # # # #         ("Primary",        sum(1 for x in contacts if x.get("priority") == "Primary")),
# # # # # # #         ("Secondary",      sum(1 for x in contacts if x.get("priority") == "Secondary")),
# # # # # # #         ("Tertiary",       sum(1 for x in contacts if x.get("priority") == "Tertiary")),
# # # # # # #         ("General",        sum(1 for x in contacts if x.get("priority") == "General")),
# # # # # # #         ("With Email",     sum(1 for x in contacts if x.get("email"))),
# # # # # # #         ("With LinkedIn",  sum(1 for x in contacts if x.get("linkedin_url"))),
# # # # # # #         ("Generated",      datetime.now().strftime("%d %b %Y  %H:%M")),
# # # # # # #     ], 2):
# # # # # # #         lc = ws2.cell(row=r, column=1, value=lbl)
# # # # # # #         vc = ws2.cell(row=r, column=2, value=val)
# # # # # # #         bg = LBLUE if r % 2 == 0 else WHITE
# # # # # # #         for cell in (lc, vc):
# # # # # # #             cell.font   = Font(name="Arial", bold=(cell.column == 1), size=10)
# # # # # # #             cell.fill   = PatternFill("solid", fgColor=bg)
# # # # # # #             cell.border = border

# # # # # # #     ws2.column_dimensions["A"].width = 20
# # # # # # #     ws2.column_dimensions["B"].width = 30

# # # # # # #     buf = io.BytesIO()
# # # # # # #     wb.save(buf)
# # # # # # #     buf.seek(0)
# # # # # # #     return buf.getvalue()


# # # # # # # # ── MongoDB helpers ───────────────────────────────────────────────────────────
# # # # # # # @st.cache_resource
# # # # # # # def get_db():
# # # # # # #     URI = st.secrets.get("MONGO_URI", "mongodb://localhost:27017")
# # # # # # #     DB  = st.secrets.get("MONGO_DB",  "secureitlab_job_pipeline")
# # # # # # #     client = MongoClient(URI, serverSelectionTimeoutMS=6000)
# # # # # # #     return client[DB]

# # # # # # # @st.cache_data(ttl=60)
# # # # # # # def load_all_jobs():
# # # # # # #     db = get_db()
# # # # # # #     return list(db.jobs.find({}, {
# # # # # # #         "_id": 1, "company": 1, "role": 1, "job_number": 1,
# # # # # # #         "opp_score": 1, "contacts_found": 1, "pipeline_ok": 1,
# # # # # # #         "coverage_score": 1, "run_at": 1, "contact_domain": 1,
# # # # # # #     }))

# # # # # # # @st.cache_data(ttl=60)
# # # # # # # def load_job(job_id):
# # # # # # #     from bson import ObjectId
# # # # # # #     return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# # # # # # # # ── Render helpers ────────────────────────────────────────────────────────────
# # # # # # # def badge(text, color="blue"):
# # # # # # #     return f'<span class="badge badge-{color}">{text}</span>'

# # # # # # # def safe_str(val, limit=300):
# # # # # # #     if val is None: return "—"
# # # # # # #     s = str(val)
# # # # # # #     return s[:limit] + "…" if len(s) > limit else s

# # # # # # # def as_dict(raw):
# # # # # # #     if isinstance(raw, dict): return raw
# # # # # # #     if isinstance(raw, list): return next((x for x in raw if isinstance(x, dict)), {})
# # # # # # #     return {}

# # # # # # # def render_json_pretty(data, title=""):
# # # # # # #     if not data: return
# # # # # # #     with st.expander(f"📄 Raw JSON — {title}", expanded=False):
# # # # # # #         st.code(json.dumps(data, indent=2, default=str), language="json")

# # # # # # # def render_qa_block(data, label):
# # # # # # #     if not data:
# # # # # # #         st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>', unsafe_allow_html=True)
# # # # # # #         return
# # # # # # #     data = as_dict(data)
# # # # # # #     if not data: return
# # # # # # #     passed    = data.get("passed") or data.get("Passed") or False
# # # # # # #     rec       = data.get("recommendation") or data.get("Recommendation", "")
# # # # # # #     issues    = data.get("issues") or data.get("Issues") or []
# # # # # # #     checklist = data.get("checklist") or data.get("Checklist") or []
# # # # # # #     color     = "green" if passed else "yellow"
# # # # # # #     status    = "✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
# # # # # # #     html = f"""
# # # # # # #     <div class="sil-card sil-card-accent">
# # # # # # #       <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
# # # # # # #         <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
# # # # # # #         {badge(status, color)}
# # # # # # #       </div>"""
# # # # # # #     if rec:
# # # # # # #         html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
# # # # # # #     if checklist:
# # # # # # #         html += '<div style="font-size:.82rem;margin-bottom:.5rem">'
# # # # # # #         for item in (checklist if isinstance(checklist, list) else [checklist]):
# # # # # # #             if isinstance(item, dict):
# # # # # # #                 ip = item.get("pass") or item.get("passed") or item.get("status","") == "pass"
# # # # # # #                 nm = item.get("item") or item.get("name") or item.get("check","")
# # # # # # #                 nt = item.get("reason") or item.get("note") or item.get("issue","")
# # # # # # #                 html += f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
# # # # # # #                 if nt: html += f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
# # # # # # #                 html += '</div>'
# # # # # # #         html += '</div>'
# # # # # # #     if issues:
# # # # # # #         html += '<div style="margin-top:.5rem">'
# # # # # # #         for iss in (issues if isinstance(issues, list) else [issues])[:4]:
# # # # # # #             txt = iss if isinstance(iss, str) else json.dumps(iss)
# # # # # # #             html += f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
# # # # # # #         html += '</div>'
# # # # # # #     st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # # # def render_contacts(contacts, title="Contacts"):
# # # # # # #     if not contacts: st.info("No contacts found for this job."); return
# # # # # # #     pri_color = {"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
# # # # # # #     st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>', unsafe_allow_html=True)
# # # # # # #     cols = st.columns(2)
# # # # # # #     for i, c in enumerate(contacts):
# # # # # # #         col = cols[i % 2]
# # # # # # #         prio = c.get("priority", "General")
# # # # # # #         email = c.get("email", ""); li = c.get("linkedin_url", "")
# # # # # # #         patterns = c.get("email_patterns", []); src = c.get("source", "")
# # # # # # #         with col:
# # # # # # #             html = f"""<div class="contact-card">
# # # # # # #               <div style="display:flex;justify-content:space-between;align-items:flex-start">
# # # # # # #                 <div>
# # # # # # #                   <div class="contact-name">{c.get('name','—')}</div>
# # # # # # #                   <div class="contact-title">{c.get('title','—')}</div>
# # # # # # #                 </div>
# # # # # # #                 {badge(prio, pri_color.get(prio,'blue'))}
# # # # # # #               </div>"""
# # # # # # #             if email:      html += f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
# # # # # # #             elif patterns: html += f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
# # # # # # #             if li:         html += f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
# # # # # # #             if src:        html += f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
# # # # # # #             st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # # # def render_emails(emails_data):
# # # # # # #     if not emails_data: st.info("No email data available."); return
# # # # # # #     emails_data = as_dict(emails_data)
# # # # # # #     if not emails_data: return
# # # # # # #     variants = {}
# # # # # # #     for k, v in emails_data.items():
# # # # # # #         kl = k.lower().replace("_","").replace(" ","")
# # # # # # #         if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl == "a":
# # # # # # #             variants["Variant A — Hiring Manager"] = v
# # # # # # #         elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl == "b":
# # # # # # #             variants["Variant B — CISO / VP Level"] = v
# # # # # # #         else:
# # # # # # #             variants[k] = v
# # # # # # #     for label, v in variants.items():
# # # # # # #         st.markdown(f'<div class="section-label">✉️ {label}</div>', unsafe_allow_html=True)
# # # # # # #         if isinstance(v, dict):
# # # # # # #             subj = v.get("subject") or v.get("Subject","")
# # # # # # #             body = v.get("body") or v.get("Body") or v.get("content","")
# # # # # # #             if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>', unsafe_allow_html=True)
# # # # # # #             if body: st.markdown(f'<div class="email-box">{body}</div>', unsafe_allow_html=True)
# # # # # # #             else:    st.code(json.dumps(v, indent=2), language="json")
# # # # # # #         elif isinstance(v, str):
# # # # # # #             st.markdown(f'<div class="email-box">{v}</div>', unsafe_allow_html=True)
# # # # # # #         st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

# # # # # # # def render_service_mapping(data):
# # # # # # #     if not data: st.info("No service mapping data."); return
# # # # # # #     items = data if isinstance(data, list) else []
# # # # # # #     if not items and isinstance(data, dict):
# # # # # # #         for key in ("services","mappings","service_mapping","ServiceMapping","items"):
# # # # # # #             if isinstance(data.get(key), list): items = data[key]; break
# # # # # # #         if not items: items = [data]
# # # # # # #     fit_colors = {"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
# # # # # # #     for item in items:
# # # # # # #         if not isinstance(item, dict): continue
# # # # # # #         svc  = item.get("service") or item.get("Service") or item.get("name","")
# # # # # # #         fit  = (item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
# # # # # # #         why  = item.get("justification") or item.get("rationale") or item.get("why","")
# # # # # # #         reqs = item.get("requirements_addressed") or item.get("requirements") or ""
# # # # # # #         eng  = item.get("engagement_type") or item.get("engagement","")
# # # # # # #         color = fit_colors.get(fit, "blue")
# # # # # # #         html = f"""<div class="sil-card" style="margin-bottom:.75rem">
# # # # # # #           <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
# # # # # # #             <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
# # # # # # #             {badge(fit or "MAPPED", color) if fit else ""}
# # # # # # #           </div>"""
# # # # # # #         if why:  html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
# # # # # # #         if reqs:
# # # # # # #             rs = ", ".join(reqs) if isinstance(reqs, list) else str(reqs)
# # # # # # #             html += f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
# # # # # # #         if eng:  html += f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
# # # # # # #         st.markdown(html + '</div>', unsafe_allow_html=True)
# # # # # # #     render_json_pretty(data, "Service Mapping")

# # # # # # # def render_microplans(data):
# # # # # # #     if not data: st.info("No micro-plan data."); return
# # # # # # #     plans = data if isinstance(data, list) else []
# # # # # # #     if not plans and isinstance(data, dict):
# # # # # # #         for k in ("plans","micro_plans","microplans","top_3","improvements"):
# # # # # # #             if isinstance(data.get(k), list): plans = data[k]; break
# # # # # # #         if not plans: plans = [data]
# # # # # # #     for i, plan in enumerate(plans[:3], 1):
# # # # # # #         if not isinstance(plan, dict): continue
# # # # # # #         title = plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
# # # # # # #         weeks = plan.get("duration") or plan.get("timeline","")
# # # # # # #         obj   = plan.get("objective") or plan.get("goal","")
# # # # # # #         kpis  = plan.get("kpis") or plan.get("KPIs") or []
# # # # # # #         tasks = plan.get("tasks") or plan.get("workstreams") or []
# # # # # # #         with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}", expanded=(i==1)):
# # # # # # #             if obj and obj != title: st.markdown(f"**Objective:** {obj}")
# # # # # # #             if kpis:
# # # # # # #                 st.markdown("**KPIs:**")
# # # # # # #                 for kpi in (kpis if isinstance(kpis, list) else [kpis]): st.markdown(f"• {kpi}")
# # # # # # #             if tasks:
# # # # # # #                 st.markdown("**Tasks / Workstreams:**")
# # # # # # #                 for t in (tasks if isinstance(tasks, list) else [tasks]):
# # # # # # #                     if isinstance(t, dict):
# # # # # # #                         tn = t.get("task") or t.get("name","")
# # # # # # #                         te = t.get("effort") or t.get("duration","")
# # # # # # #                         st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
# # # # # # #                     else: st.markdown(f"• {t}")
# # # # # # #             render_json_pretty(plan, f"Plan {i}")

# # # # # # # def render_deal_assurance(data):
# # # # # # #     if not data: st.info("No deal assurance data."); return
# # # # # # #     if not isinstance(data, dict): render_json_pretty(data, "Deal Assurance Pack"); return
# # # # # # #     evp = (data.get("executive_value_proposition") or
# # # # # # #            data.get("value_proposition") or data.get("ExecutiveValueProposition",""))
# # # # # # #     if evp:
# # # # # # #         st.markdown('<div class="section-label">💼 Executive Value Proposition</div>', unsafe_allow_html=True)
# # # # # # #         st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7;color:var(--text)">{evp}</div>', unsafe_allow_html=True)
# # # # # # #     caps = data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
# # # # # # #     if caps:
# # # # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>', unsafe_allow_html=True)
# # # # # # #         c1, c2 = st.columns(2)
# # # # # # #         for i, cap in enumerate(caps if isinstance(caps, list) else [caps]):
# # # # # # #             (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
# # # # # # #     risk = data.get("risk_mitigation") or data.get("RiskMitigation","")
# # # # # # #     if risk:
# # # # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>', unsafe_allow_html=True)
# # # # # # #         if isinstance(risk, dict):
# # # # # # #             for k, v in risk.items(): st.markdown(f"**{k}:** {v}")
# # # # # # #         else: st.markdown(str(risk))
# # # # # # #     render_json_pretty(data, "Deal Assurance Pack")


# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # #  SIDEBAR
# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # with st.sidebar:
# # # # # # #     st.markdown("""
# # # # # # #     <div style="padding:.75rem 0 1.25rem">
# # # # # # #       <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;
# # # # # # #                   color:#2563eb">🛡️ SecureITLab</div>
# # # # # # #       <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;
# # # # # # #                   text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
# # # # # # #     </div>
# # # # # # #     """, unsafe_allow_html=True)

# # # # # # #     # ── Logout button ─────────────────────────────────────────────────────────
# # # # # # #     if st.button("🚪 Logout", use_container_width=True):
# # # # # # #         st.session_state.logged_in = False
# # # # # # #         st.session_state.login_error = ""
# # # # # # #         st.rerun()

# # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # # #     # ── Find Jobs button ──────────────────────────────────────────────────────
# # # # # # #     st.markdown("**🔍 Find New Jobs**")
# # # # # # #     st.caption(
# # # # # # #         "Runs scraper → checks MongoDB for duplicates → "
# # # # # # #         "runs AI pipeline only on NEW jobs → stores in MongoDB"
# # # # # # #     )

# # # # # # #     find_jobs_clicked = st.button(
# # # # # # #         "🔍  Find Jobs",
# # # # # # #         disabled=st.session_state.pipeline_running,
# # # # # # #         use_container_width=True,
# # # # # # #         type="primary",
# # # # # # #     )

# # # # # # #     if st.session_state.pipeline_running:
# # # # # # #         st.markdown("""
# # # # # # #         <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
# # # # # # #                     padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;
# # # # # # #                     font-size:.8rem;color:#1d4ed8">
# # # # # # #           ⏳ Pipeline is running… check log below
# # # # # # #         </div>""", unsafe_allow_html=True)

# # # # # # #     if st.session_state.pipeline_done and st.session_state.pipeline_result:
# # # # # # #         res = st.session_state.pipeline_result
# # # # # # #         ok  = res.get("success", False)
# # # # # # #         bg  = "#ecfdf5" if ok else "#fef2f2"
# # # # # # #         bc  = "#bbf7d0" if ok else "#fecaca"
# # # # # # #         tc  = "#15803d" if ok else "#b91c1c"
# # # # # # #         ic  = "✅" if ok else "❌"
# # # # # # #         st.markdown(f"""
# # # # # # #         <div style="background:{bg};border:1px solid {bc};border-radius:8px;
# # # # # # #                     padding:.75rem;margin-top:.5rem;font-size:.82rem">
# # # # # # #           <div style="font-weight:700;color:{tc};margin-bottom:.4rem">{ic} Last Run</div>
# # # # # # #           <div>📦 Scraped: <b>{res.get('scraped',0)}</b></div>
# # # # # # #           <div>🆕 New jobs: <b>{res.get('new_jobs',0)}</b></div>
# # # # # # #           <div>⏭️ Already in DB (skipped): <b>{res.get('skipped_db',0)}</b></div>
# # # # # # #           <div>🤖 Processed by AI: <b>{res.get('processed',0)}</b></div>
# # # # # # #           {f'<div style="color:{tc};margin-top:.3rem">⚠️ {res.get("error","")}</div>' if res.get("error") else ""}
# # # # # # #         </div>""", unsafe_allow_html=True)

# # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # # #     # ── Job list ──────────────────────────────────────────────────────────────
# # # # # # #     try:
# # # # # # #         all_jobs = load_all_jobs()
# # # # # # #     except Exception as e:
# # # # # # #         st.error(f"MongoDB connection failed: {e}")
# # # # # # #         st.stop()

# # # # # # #     if not all_jobs:
# # # # # # #         st.warning("No jobs in MongoDB yet. Click **Find Jobs** to scrape and process.")
# # # # # # #         st.stop()

# # # # # # #     st.markdown(
# # # # # # #         f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">'
# # # # # # #         f'{len(all_jobs)} jobs in database</div>',
# # # # # # #         unsafe_allow_html=True,
# # # # # # #     )

# # # # # # #     search   = st.text_input("🔍 Filter by company / role", placeholder="e.g. Bounteous")
# # # # # # #     filtered = [j for j in all_jobs
# # # # # # #                 if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

# # # # # # #     def job_label(j):
# # # # # # #         score = j.get("opp_score")
# # # # # # #         s_str = f" [{score}/10]" if score else ""
# # # # # # #         ok    = "✅" if j.get("pipeline_ok") else "❌"
# # # # # # #         return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s_str}"

# # # # # # #     if not filtered:
# # # # # # #         st.warning("No matching jobs.")
# # # # # # #         st.stop()

# # # # # # #     sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
# # # # # # #     sel_idx     = [job_label(j) for j in filtered].index(sel_label)
# # # # # # #     selected_id = str(filtered[sel_idx]["_id"])

# # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # # #     ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
# # # # # # #     total_c  = sum(j.get("contacts_found",0) for j in all_jobs)
# # # # # # #     st.markdown(f"""
# # # # # # #     <div style="font-size:.75rem;color:#64748b">
# # # # # # #       <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
# # # # # # #       <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
# # # # # # #     </div>""", unsafe_allow_html=True)

# # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # # #     if st.button("🔄 Refresh Data", use_container_width=True):
# # # # # # #         st.cache_data.clear()
# # # # # # #         st.rerun()


# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # #  FIND JOBS — background thread
# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # if find_jobs_clicked and not st.session_state.pipeline_running:
# # # # # # #     with _thread_lock:
# # # # # # #         _thread_logs.clear()
# # # # # # #         _thread_result[0] = None
# # # # # # #         _thread_done[0]   = False

# # # # # # #     st.session_state.pipeline_running = True
# # # # # # #     st.session_state.pipeline_done    = False
# # # # # # #     st.session_state.pipeline_logs    = []
# # # # # # #     st.session_state.pipeline_result  = None
# # # # # # #     st.cache_data.clear()

# # # # # # #     def _run_pipeline_bg():
# # # # # # #         try:
# # # # # # #             import main as _main

# # # # # # #             def _cb(msg: str):
# # # # # # #                 with _thread_lock:
# # # # # # #                     _thread_logs.append(f"{datetime.now().strftime('%H:%M:%S')} | {msg}")

# # # # # # #             result = _main.run_pipeline(progress_callback=_cb)
# # # # # # #             with _thread_lock:
# # # # # # #                 _thread_result[0] = result
# # # # # # #         except Exception as e:
# # # # # # #             with _thread_lock:
# # # # # # #                 _thread_logs.append(f"❌ Unexpected error: {e}")
# # # # # # #                 _thread_result[0] = {
# # # # # # #                     "success": False, "error": str(e),
# # # # # # #                     "scraped": 0, "new_jobs": 0, "skipped_db": 0, "processed": 0,
# # # # # # #                 }
# # # # # # #         finally:
# # # # # # #             with _thread_lock:
# # # # # # #                 _thread_done[0] = True

# # # # # # #     threading.Thread(target=_run_pipeline_bg, daemon=True).start()
# # # # # # #     st.rerun()

# # # # # # # # ── Sync thread state → session_state ────────────────────────────────────────
# # # # # # # with _thread_lock:
# # # # # # #     if _thread_logs:
# # # # # # #         st.session_state.pipeline_logs = list(_thread_logs)
# # # # # # #     if _thread_done[0] and _thread_result[0] is not None:
# # # # # # #         st.session_state.pipeline_result  = _thread_result[0]
# # # # # # #         st.session_state.pipeline_running = False
# # # # # # #         st.session_state.pipeline_done    = True

# # # # # # # # ── Live log panel ────────────────────────────────────────────────────────────
# # # # # # # if st.session_state.pipeline_running or (
# # # # # # #         st.session_state.pipeline_done and st.session_state.pipeline_logs):
# # # # # # #     heading = "⏳ Pipeline running — live log…" if st.session_state.pipeline_running \
# # # # # # #               else "📋 Last pipeline run log"
# # # # # # #     with st.expander(heading, expanded=st.session_state.pipeline_running):
# # # # # # #         log_text = "\n".join(st.session_state.pipeline_logs[-100:]) \
# # # # # # #                    if st.session_state.pipeline_logs else "Starting…"
# # # # # # #         st.markdown(f'<div class="pipeline-log">{log_text}</div>', unsafe_allow_html=True)
# # # # # # #         if st.session_state.pipeline_running:
# # # # # # #             time.sleep(2)
# # # # # # #             st.rerun()


# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # #  MAIN CONTENT — selected job detail
# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # with st.spinner("Loading job from MongoDB…"):
# # # # # # #     job = load_job(selected_id)

# # # # # # # if not job:
# # # # # # #     st.error("Could not load job document.")
# # # # # # #     st.stop()

# # # # # # # company   = job.get("company",  "Unknown")
# # # # # # # role      = job.get("role",     "Unknown")
# # # # # # # opp_score = job.get("opp_score")
# # # # # # # p_ok      = job.get("pipeline_ok", False)
# # # # # # # p_min     = job.get("pipeline_min", "?")
# # # # # # # c_found   = job.get("contacts_found", 0)
# # # # # # # c_cov     = job.get("coverage_score")
# # # # # # # c_domain  = job.get("contact_domain","")
# # # # # # # run_at    = job.get("run_at","")

# # # # # # # # ── Header ────────────────────────────────────────────────────────────────────
# # # # # # # st.markdown(f"""
# # # # # # # <div style="margin-bottom:1.75rem">
# # # # # # #   <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;
# # # # # # #               letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
# # # # # # #     Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
# # # # # # #   </div>
# # # # # # #   <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;
# # # # # # #              color:#0f172a;margin:0;line-height:1.15">{role}</h1>
# # # # # # #   <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
# # # # # # #     @ <span style="color:#334155;font-weight:600">{company}</span>
# # # # # # #     {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
# # # # # # #   </div>
# # # # # # # </div>
# # # # # # # """, unsafe_allow_html=True)

# # # # # # # # ── Metric tiles ──────────────────────────────────────────────────────────────
# # # # # # # try:
# # # # # # #     sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
# # # # # # #     sc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # # # # except Exception:
# # # # # # #     sc = "#2563eb"

# # # # # # # st.markdown(f"""
# # # # # # # <div class="metric-row">
# # # # # # #   <div class="metric-tile">
# # # # # # #     <div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div>
# # # # # # #     <div class="lbl">Opportunity Score</div>
# # # # # # #   </div>
# # # # # # #   <div class="metric-tile">
# # # # # # #     <div class="val">{c_found}</div>
# # # # # # #     <div class="lbl">Contacts Found</div>
# # # # # # #   </div>
# # # # # # #   <div class="metric-tile">
# # # # # # #     <div class="val">{f"{c_cov}%" if c_cov else "—"}</div>
# # # # # # #     <div class="lbl">Contact Coverage</div>
# # # # # # #   </div>
# # # # # # #   <div class="metric-tile">
# # # # # # #     <div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">
# # # # # # #       {'✅ OK' if p_ok else '❌ Failed'}
# # # # # # #     </div>
# # # # # # #     <div class="lbl">Pipeline ({p_min} min)</div>
# # # # # # #   </div>
# # # # # # # </div>
# # # # # # # """, unsafe_allow_html=True)

# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # #  TABS
# # # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # # tabs = st.tabs([
# # # # # # #     "📋 Job & Enrichment",
# # # # # # #     "🎯 Service Mapping",
# # # # # # #     "🔍 Fit / Gap",
# # # # # # #     "🛠️ Capability & Plans",
# # # # # # #     "📦 Deal Assurance",
# # # # # # #     "✉️ Outreach Emails",
# # # # # # #     "👥 Contacts",
# # # # # # #     "✅ QA Gates",
# # # # # # #     "🗄️ Raw Data",
# # # # # # # ])

# # # # # # # # ── Tab 1: Job Research + Enrichment ─────────────────────────────────────────
# # # # # # # with tabs[0]:
# # # # # # #     col1, col2 = st.columns([1, 1])
# # # # # # #     with col1:
# # # # # # #         st.markdown('<div class="section-label">📄 Job Research</div>', unsafe_allow_html=True)
# # # # # # #         jr = as_dict(job.get("agent_job_research") or {})
# # # # # # #         if jr:
# # # # # # #             for label, keys in [
# # # # # # #                 ("Job Role",        ["job_role","Job Role","role","title"]),
# # # # # # #                 ("Company",         ["company_name","Company Name","company"]),
# # # # # # #                 ("Location",        ["location","Location"]),
# # # # # # #                 ("Organization URL",["organization_url","Organization URL","url"]),
# # # # # # #             ]:
# # # # # # #                 val = next((jr.get(k) for k in keys if jr.get(k)), None)
# # # # # # #                 if val: st.markdown(f"**{label}:** {val}")
# # # # # # #             desc = jr.get("job_description") or jr.get("Job Description","")
# # # # # # #             if desc:
# # # # # # #                 st.markdown("**Job Description:**")
# # # # # # #                 st.markdown(
# # # # # # #                     f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;'
# # # # # # #                     f'max-height:300px;overflow-y:auto;color:var(--text)">{desc[:2000]}</div>',
# # # # # # #                     unsafe_allow_html=True,
# # # # # # #                 )
# # # # # # #             render_json_pretty(jr, "Job Research")
# # # # # # #         else:
# # # # # # #             st.info("No job research data.")
# # # # # # #     with col2:
# # # # # # #         st.markdown('<div class="section-label">🏢 Company Enrichment</div>', unsafe_allow_html=True)
# # # # # # #         enr = as_dict(job.get("agent_enrichment") or {})
# # # # # # #         if enr:
# # # # # # #             for label, keys in [
# # # # # # #                 ("Industry",          ["industry","Industry"]),
# # # # # # #                 ("Company Size",      ["company_size","size","Company Size"]),
# # # # # # #                 ("Regulatory Env",    ["regulatory_environment","regulatory"]),
# # # # # # #                 ("Certifications",    ["certifications","Certifications"]),
# # # # # # #                 ("Tech Stack",        ["tech_stack","technology_stack"]),
# # # # # # #                 ("Security Maturity", ["security_maturity","maturity"]),
# # # # # # #             ]:
# # # # # # #                 val = next((enr.get(k) for k in keys if enr.get(k)), None)
# # # # # # #                 if val:
# # # # # # #                     if isinstance(val, list): val = ", ".join(str(v) for v in val)
# # # # # # #                     st.markdown(f"**{label}:** {safe_str(val, 200)}")
# # # # # # #             render_json_pretty(enr, "Enrichment")
# # # # # # #         else:
# # # # # # #             st.info("No enrichment data.")

# # # # # # # # ── Tab 2: Service Mapping ────────────────────────────────────────────────────
# # # # # # # with tabs[1]:
# # # # # # #     st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>', unsafe_allow_html=True)
# # # # # # #     render_service_mapping(job.get("agent_service_mapping"))

# # # # # # # # ── Tab 3: Fit / Gap Analysis ─────────────────────────────────────────────────
# # # # # # # with tabs[2]:
# # # # # # #     fg = as_dict(job.get("agent_fit_gap") or {})
# # # # # # #     if opp_score:
# # # # # # #         try:
# # # # # # #             sn = float(str(opp_score).split("/")[0])
# # # # # # #             bc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # # # #             bp = int(sn / 10 * 100)
# # # # # # #             st.markdown(f"""
# # # # # # #             <div style="margin-bottom:1.5rem">
# # # # # # #               <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
# # # # # # #                 <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">Opportunity Score</span>
# # # # # # #                 <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
# # # # # # #               </div>
# # # # # # #               <div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%">
# # # # # # #                 <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
# # # # # # #               </div>
# # # # # # #             </div>""", unsafe_allow_html=True)
# # # # # # #         except Exception:
# # # # # # #             pass
# # # # # # #     st.markdown('<div class="section-label">📊 Service Classifications</div>', unsafe_allow_html=True)
# # # # # # #     services = []
# # # # # # #     if isinstance(fg, dict):
# # # # # # #         for k in ("services","classifications","service_classifications","items","fit_gap"):
# # # # # # #             v = fg.get(k)
# # # # # # #             if isinstance(v, list): services = v; break
# # # # # # #         if not services and (fg.get("service") or fg.get("Service")):
# # # # # # #             services = [fg]
# # # # # # #     elif isinstance(fg, list):
# # # # # # #         services = fg
# # # # # # #     if services:
# # # # # # #         buckets = {"STRONG FIT": [], "PARTIAL FIT": [], "GAP": []}
# # # # # # #         other   = []
# # # # # # #         for s in services:
# # # # # # #             if not isinstance(s, dict): continue
# # # # # # #             fit = (s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
# # # # # # #             if "STRONG" in fit: buckets["STRONG FIT"].append(s)
# # # # # # #             elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
# # # # # # #             elif "GAP" in fit: buckets["GAP"].append(s)
# # # # # # #             else: other.append(s)
# # # # # # #         c1, c2, c3 = st.columns(3)
# # # # # # #         cm  = {"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
# # # # # # #         bgm = {"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
# # # # # # #         bdm = {"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
# # # # # # #         for col, (fl, items) in zip([c1, c2, c3], buckets.items()):
# # # # # # #             col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>', unsafe_allow_html=True)
# # # # # # #             for s in items:
# # # # # # #                 svc  = s.get("service") or s.get("Service") or s.get("name","")
# # # # # # #                 just = s.get("justification") or s.get("reason","")
# # # # # # #                 col.markdown(
# # # # # # #                     f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem">'
# # # # # # #                     f'<div style="font-weight:600;color:#0f172a">{svc}</div>'
# # # # # # #                     f'<div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>',
# # # # # # #                     unsafe_allow_html=True,
# # # # # # #                 )
# # # # # # #     elif fg:
# # # # # # #         st.json(fg)
# # # # # # #     else:
# # # # # # #         st.info("No fit/gap data.")
# # # # # # #     render_json_pretty(job.get("agent_fit_gap"), "Fit/Gap Analysis")

# # # # # # # # ── Tab 4: Capability + Micro-plans ──────────────────────────────────────────
# # # # # # # with tabs[3]:
# # # # # # #     col1, col2 = st.columns([1, 1])
# # # # # # #     with col1:
# # # # # # #         st.markdown('<div class="section-label">🔧 Capability Improvements</div>', unsafe_allow_html=True)
# # # # # # #         cap = job.get("agent_capability") or {}
# # # # # # #         items_cap = cap if isinstance(cap, list) else []
# # # # # # #         if not items_cap and isinstance(cap, dict):
# # # # # # #             for k in ("improvements","recommendations","capabilities","items"):
# # # # # # #                 v = cap.get(k)
# # # # # # #                 if isinstance(v, list): items_cap = v; break
# # # # # # #             if not items_cap: items_cap = [cap]
# # # # # # #         for item in items_cap:
# # # # # # #             if not isinstance(item, dict): continue
# # # # # # #             title  = item.get("title") or item.get("gap") or item.get("service","")
# # # # # # #             rec    = item.get("recommendation") or item.get("steps","")
# # # # # # #             effort = item.get("build_effort") or item.get("effort","")
# # # # # # #             demand = item.get("market_demand") or item.get("priority","")
# # # # # # #             st.markdown(f"""
# # # # # # #             <div class="sil-card" style="margin-bottom:.6rem">
# # # # # # #               <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem;color:var(--text)">{title}</div>
# # # # # # #               <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec, 250)}</div>
# # # # # # #               {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
# # # # # # #             </div>""", unsafe_allow_html=True)
# # # # # # #         if not items_cap:
# # # # # # #             render_json_pretty(cap, "Capability Improvement")
# # # # # # #     with col2:
# # # # # # #         st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>', unsafe_allow_html=True)
# # # # # # #         render_microplans(job.get("agent_microplans"))

# # # # # # # # ── Tab 5: Deal Assurance Pack ────────────────────────────────────────────────
# # # # # # # with tabs[4]:
# # # # # # #     render_deal_assurance(job.get("agent_deal_assurance"))

# # # # # # # # ── Tab 6: Outreach Emails ────────────────────────────────────────────────────
# # # # # # # with tabs[5]:
# # # # # # #     st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>', unsafe_allow_html=True)
# # # # # # #     emails_src = job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
# # # # # # #     oq = as_dict(job.get("agent_outreach_qa") or {})
# # # # # # #     improved = (oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
# # # # # # #     if improved:
# # # # # # #         st.info("⚡ Showing QA-improved versions where available")
# # # # # # #         render_emails(improved)
# # # # # # #         with st.expander("📬 Original (pre-QA) versions"):
# # # # # # #             render_emails(emails_src)
# # # # # # #     else:
# # # # # # #         render_emails(emails_src)

# # # # # # # # ── Tab 7: Contacts ───────────────────────────────────────────────────────────
# # # # # # # with tabs[6]:
# # # # # # #     contacts        = job.get("contacts") or []
# # # # # # #     contact_sources = job.get("contact_sources") or []
# # # # # # #     pri = [c for c in contacts if c.get("priority") == "Primary"]
# # # # # # #     sec = [c for c in contacts if c.get("priority") == "Secondary"]
# # # # # # #     ter = [c for c in contacts if c.get("priority") == "Tertiary"]
# # # # # # #     gen = [c for c in contacts if c.get("priority") == "General"]
# # # # # # #     st.markdown(f"""
# # # # # # #     <div class="metric-row" style="margin-bottom:1.5rem">
# # # # # # #       <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
# # # # # # #       <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
# # # # # # #       <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
# # # # # # #       <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
# # # # # # #     </div>""", unsafe_allow_html=True)
# # # # # # #     if contact_sources:
# # # # # # #         st.markdown('Sources: ' + " ".join(badge(s,"blue") for s in contact_sources), unsafe_allow_html=True)
# # # # # # #         st.markdown("")
# # # # # # #     missing = job.get("missing_roles") or []
# # # # # # #     if missing:
# # # # # # #         st.markdown('⚠️ Missing roles: ' + " ".join(badge(r,"red") for r in missing), unsafe_allow_html=True)
# # # # # # #         st.markdown("")
# # # # # # #     if contacts:
# # # # # # #         # ── Excel export button ───────────────────────────────────────────────
# # # # # # #         excel_bytes = build_contacts_excel(contacts, company, role)
# # # # # # #         if excel_bytes:
# # # # # # #             safe_co = re.sub(r'[^a-z0-9]', '_', company.lower())[:20]
# # # # # # #             fname   = f"contacts_{safe_co}_{datetime.now().strftime('%Y%m%d')}.xlsx"
# # # # # # #             btn_col, _ = st.columns([1, 3])
# # # # # # #             with btn_col:
# # # # # # #                 st.download_button(
# # # # # # #                     label="📥  Download Contacts (.xlsx)",
# # # # # # #                     data=excel_bytes,
# # # # # # #                     file_name=fname,
# # # # # # #                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# # # # # # #                     use_container_width=True,
# # # # # # #                     type="primary",
# # # # # # #                 )
# # # # # # #         else:
# # # # # # #             st.warning("Run `pip install openpyxl` to enable Excel export.")

# # # # # # #         st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # # #         pri_filter = st.multiselect(
# # # # # # #             "Filter by priority",
# # # # # # #             ["Primary","Secondary","Tertiary","General"],
# # # # # # #             default=["Primary","Secondary","Tertiary","General"],
# # # # # # #         )
# # # # # # #         shown = [c for c in contacts if c.get("priority","General") in pri_filter]
# # # # # # #         render_contacts(shown, f"Contacts ({len(shown)} shown)")
# # # # # # #         agent_contacts = job.get("agent_prospect_contacts") or {}
# # # # # # #         if agent_contacts:
# # # # # # #             with st.expander("🤖 CrewAI Agent's Contact Search"):
# # # # # # #                 if isinstance(agent_contacts, dict):
# # # # # # #                     ac_list = agent_contacts.get("contacts") or []
# # # # # # #                     if ac_list: render_contacts(ac_list, "Agent Contacts")
# # # # # # #                     else:       st.json(agent_contacts)
# # # # # # #                 else:           st.json(agent_contacts)
# # # # # # #     else:
# # # # # # #         st.info("No contacts found for this job.")

# # # # # # # # ── Tab 8: QA Gates ───────────────────────────────────────────────────────────
# # # # # # # with tabs[7]:
# # # # # # #     st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>', unsafe_allow_html=True)
# # # # # # #     col1, col2 = st.columns(2)
# # # # # # #     for i, (label, key) in enumerate([
# # # # # # #         ("Research QA",       "agent_research_qa"),
# # # # # # #         ("Service Mapping QA","agent_mapping_qa"),
# # # # # # #         ("Deal Assurance QA", "agent_assurance_qa"),
# # # # # # #         ("Outreach Email QA", "agent_outreach_qa"),
# # # # # # #     ]):
# # # # # # #         with (col1 if i % 2 == 0 else col2):
# # # # # # #             render_qa_block(job.get(key), label)
# # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # # #     st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>', unsafe_allow_html=True)
# # # # # # #     enf = as_dict(job.get("agent_prospect_enforcer") or {})
# # # # # # #     if enf:
# # # # # # #         cov  = enf.get("coverage_score","?")
# # # # # # #         miss = enf.get("missing_roles",[])
# # # # # # #         note = enf.get("note","")
# # # # # # #         ec   = enf.get("contacts",[])
# # # # # # #         x1, x2, x3 = st.columns(3)
# # # # # # #         x1.metric("Coverage Score",    f"{cov}%")
# # # # # # #         x2.metric("Missing Roles",     len(miss))
# # # # # # #         x3.metric("Contacts Verified", len(ec))
# # # # # # #         if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
# # # # # # #         if note: st.caption(note)
# # # # # # #     else:
# # # # # # #         st.info("No enforcer data.")

# # # # # # # # ── Tab 9: Raw Data ───────────────────────────────────────────────────────────
# # # # # # # with tabs[8]:
# # # # # # #     st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>', unsafe_allow_html=True)
# # # # # # #     st.caption("All fields stored in the `jobs` collection for this document.")
# # # # # # #     rows = []
# # # # # # #     for k, v in job.items():
# # # # # # #         if k == "_id": continue
# # # # # # #         rows.append({"Field": k, "Type": type(v).__name__, "Len": len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
# # # # # # #     hc1, hc2, hc3 = st.columns([3,1,1])
# # # # # # #     hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
# # # # # # #     for r in rows:
# # # # # # #         rc1, rc2, rc3 = st.columns([3,1,1])
# # # # # # #         rc1.code(r["Field"], language=None)
# # # # # # #         rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>', unsafe_allow_html=True)
# # # # # # #         rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>',  unsafe_allow_html=True)
# # # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # # #     for label, key in [
# # # # # # #         ("Job Research","agent_job_research"), ("Enrichment","agent_enrichment"),
# # # # # # #         ("Service Mapping","agent_service_mapping"), ("Fit/Gap Analysis","agent_fit_gap"),
# # # # # # #         ("Capability","agent_capability"), ("Micro-Plans","agent_microplans"),
# # # # # # #         ("Deal Assurance","agent_deal_assurance"), ("Outreach Emails","agent_outreach_emails"),
# # # # # # #         ("Prospect Contacts","agent_prospect_contacts"), ("Prospect Enforcer","agent_prospect_enforcer"),
# # # # # # #         ("Research QA","agent_research_qa"), ("Mapping QA","agent_mapping_qa"),
# # # # # # #         ("Assurance QA","agent_assurance_qa"), ("Outreach QA","agent_outreach_qa"),
# # # # # # #         ("Contacts (5-source)","contacts"),
# # # # # # #     ]:
# # # # # # #         data = job.get(key)
# # # # # # #         if data:
# # # # # # #             with st.expander(f"📄 {label}"):
# # # # # # #                 st.code(json.dumps(data, indent=2, default=str), language="json")








































# # # # # # """
# # # # # # ╔══════════════════════════════════════════════════════════╗
# # # # # # ║   SecureITLab Pipeline Dashboard — Streamlit             ║
# # # # # # ║   Reads from MongoDB → secureitlab_job_pipeline          ║
# # # # # # ╠══════════════════════════════════════════════════════════╣
# # # # # # ║  Install: pip install streamlit pymongo python-dotenv    ║
# # # # # # ║  Run:     streamlit run streamlit_dashboard.py           ║
# # # # # # ╚══════════════════════════════════════════════════════════╝
# # # # # # """

# # # # # # import io
# # # # # # import re
# # # # # # import streamlit as st
# # # # # # from pymongo import MongoClient
# # # # # # import json
# # # # # # import threading
# # # # # # import time
# # # # # # from datetime import datetime, timezone

# # # # # # # ── Thread-safe shared state ──────────────────────────────────────────────────
# # # # # # _thread_logs   = []
# # # # # # _thread_result = [None]
# # # # # # _thread_done   = [False]
# # # # # # _thread_lock   = threading.Lock()

# # # # # # # ── Page config ───────────────────────────────────────────────────────────────
# # # # # # st.set_page_config(
# # # # # #     page_title="SecureITLab Pipeline",
# # # # # #     page_icon="🛡️",
# # # # # #     layout="wide",
# # # # # #     initial_sidebar_state="expanded",
# # # # # # )

# # # # # # # ── LOGIN CREDENTIALS (change these) ─────────────────────────────────────────
# # # # # # LOGIN_USERNAME = "admin"
# # # # # # LOGIN_PASSWORD = "secureitlab2024"

# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  GLOBAL CSS (login + dashboard share this)
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # st.markdown("""
# # # # # # <style>

# # # # # # @import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');

# # # # # # :root {
# # # # # #     --bg:        #f4f7fb;
# # # # # #     --surface:   #ffffff;
# # # # # #     --surface2:  #eef2f7;
# # # # # #     --border:    #d9e2ec;
# # # # # #     --accent:    #2563eb;
# # # # # #     --accent2:   #7c3aed;
# # # # # #     --green:     #16a34a;
# # # # # #     --yellow:    #f59e0b;
# # # # # #     --red:       #dc2626;
# # # # # #     --text:      #0f172a;
# # # # # #     --muted:     #64748b;
# # # # # # }

# # # # # # html, body, [class*="css"] {
# # # # # #     background-color: var(--bg) !important;
# # # # # #     color: var(--text) !important;
# # # # # #     font-family: 'DM Sans', sans-serif !important;
# # # # # # }

# # # # # # /* ── LOGIN PAGE ─────────────────────────────────────────── */
# # # # # # .login-wrap {
# # # # # #     display: flex;
# # # # # #     align-items: center;
# # # # # #     justify-content: center;
# # # # # #     min-height: 80vh;
# # # # # # }
# # # # # # .login-card {
# # # # # #     background: var(--surface);
# # # # # #     border: 1px solid var(--border);
# # # # # #     border-radius: 20px;
# # # # # #     padding: 3rem 3.5rem;
# # # # # #     width: 100%;
# # # # # #     max-width: 420px;
# # # # # #     box-shadow: 0 20px 60px rgba(37,99,235,0.08);
# # # # # #     text-align: center;
# # # # # # }
# # # # # # .login-logo {
# # # # # #     font-family: 'Syne', sans-serif;
# # # # # #     font-size: 1.6rem;
# # # # # #     font-weight: 800;
# # # # # #     color: var(--accent);
# # # # # #     margin-bottom: .25rem;
# # # # # # }
# # # # # # .login-subtitle {
# # # # # #     font-size: .75rem;
# # # # # #     color: var(--muted);
# # # # # #     letter-spacing: .12em;
# # # # # #     text-transform: uppercase;
# # # # # #     margin-bottom: 2.5rem;
# # # # # # }
# # # # # # .login-error {
# # # # # #     background: #fef2f2;
# # # # # #     border: 1px solid #fecaca;
# # # # # #     border-radius: 8px;
# # # # # #     padding: .65rem 1rem;
# # # # # #     color: #b91c1c;
# # # # # #     font-size: .85rem;
# # # # # #     margin-top: 1rem;
# # # # # # }
# # # # # # .login-divider {
# # # # # #     width: 40px;
# # # # # #     height: 3px;
# # # # # #     background: linear-gradient(90deg, var(--accent), var(--accent2));
# # # # # #     border-radius: 2px;
# # # # # #     margin: 0 auto 2rem;
# # # # # # }

# # # # # # /* ── DASHBOARD ──────────────────────────────────────────── */
# # # # # # [data-testid="stSidebar"] {
# # # # # #     background: var(--surface) !important;
# # # # # #     border-right: 1px solid var(--border) !important;
# # # # # # }
# # # # # # [data-testid="stSidebar"] * { color: var(--text) !important; }

# # # # # # .main .block-container { padding: 2rem 2rem 3rem !important; }

# # # # # # h1, h2, h3, h4 {
# # # # # #     font-family: 'Syne', sans-serif !important;
# # # # # #     color: var(--text) !important;
# # # # # # }

# # # # # # .sil-card {
# # # # # #     background: var(--surface);
# # # # # #     border: 1px solid var(--border);
# # # # # #     border-radius: 14px;
# # # # # #     padding: 1.25rem 1.5rem;
# # # # # #     margin-bottom: 1rem;
# # # # # #     transition: all 0.25s ease;
# # # # # # }
# # # # # # .sil-card:hover {
# # # # # #     transform: translateY(-2px);
# # # # # #     box-shadow: 0 8px 22px rgba(0,0,0,0.05);
# # # # # # }
# # # # # # .sil-card-accent { border-left: 4px solid var(--accent); }

# # # # # # .metric-row { display:flex; gap:1rem; flex-wrap:wrap; margin-bottom:1.5rem; }
# # # # # # .metric-tile {
# # # # # #     flex:1; min-width:140px;
# # # # # #     background:var(--surface2);
# # # # # #     border:1px solid var(--border);
# # # # # #     border-radius:12px;
# # # # # #     padding:1rem; text-align:center;
# # # # # #     transition:all .25s ease;
# # # # # # }
# # # # # # .metric-tile:hover { transform:translateY(-3px); box-shadow:0 10px 24px rgba(0,0,0,0.06); }
# # # # # # .metric-tile .val { font-family:'Syne',sans-serif; font-size:2rem; font-weight:800; color:var(--accent); }
# # # # # # .metric-tile .lbl { font-size:.72rem; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; }

# # # # # # .badge { padding:.25rem .7rem; border-radius:20px; font-size:.72rem; font-weight:600; font-family:'DM Mono',monospace; }
# # # # # # .badge-green  { background:#ecfdf5; color:#15803d; }
# # # # # # .badge-yellow { background:#fffbeb; color:#b45309; }
# # # # # # .badge-red    { background:#fef2f2; color:#b91c1c; }
# # # # # # .badge-blue   { background:#eff6ff; color:#1d4ed8; }
# # # # # # .badge-purple { background:#f5f3ff; color:#6d28d9; }

# # # # # # .contact-card {
# # # # # #     background:var(--surface2); border:1px solid var(--border);
# # # # # #     border-radius:10px; padding:1rem; margin-bottom:.8rem;
# # # # # # }
# # # # # # .contact-name  { font-family:'Syne',sans-serif; font-weight:700; color:var(--text); }
# # # # # # .contact-title { color:var(--muted); font-size:.85rem; }
# # # # # # .contact-email { font-family:'DM Mono',monospace; color:var(--accent); font-size:.8rem; }

# # # # # # .email-box {
# # # # # #     background: #f8fafc;
# # # # # #     border: 1px solid var(--border);
# # # # # #     border-radius: 10px;
# # # # # #     padding: 1rem 1.25rem;
# # # # # #     font-size: .9rem;
# # # # # #     line-height: 1.65;
# # # # # #     white-space: pre-wrap;
# # # # # #     color: var(--text);
# # # # # # }
# # # # # # .email-subject { font-family:'Syne',sans-serif; font-weight:700; color:var(--accent); margin-bottom:.5rem; }

# # # # # # .section-label {
# # # # # #     font-family:'DM Mono',monospace; font-size:.72rem;
# # # # # #     text-transform:uppercase; letter-spacing:.12em;
# # # # # #     color:var(--accent); margin-bottom:.6rem;
# # # # # # }
# # # # # # .sil-divider { border-top:1px solid var(--border); margin:1rem 0; }

# # # # # # [data-testid="stExpander"] {
# # # # # #     background: var(--surface) !important;
# # # # # #     border: 1px solid var(--border) !important;
# # # # # #     border-radius: 10px !important;
# # # # # # }
# # # # # # [data-testid="stSelectbox"] > div,
# # # # # # [data-testid="stMultiSelect"] > div {
# # # # # #     background: var(--surface2) !important;
# # # # # #     border-color: var(--border) !important;
# # # # # # }
# # # # # # [data-testid="stTabs"] button {
# # # # # #     font-family:'Syne',sans-serif !important;
# # # # # #     font-weight:600 !important;
# # # # # # }
# # # # # # ::-webkit-scrollbar { width:6px; }
# # # # # # ::-webkit-scrollbar-thumb { background:var(--border); border-radius:3px; }

# # # # # # .pipeline-log {
# # # # # #     background: #0f172a;
# # # # # #     color: #e2e8f0;
# # # # # #     border-radius: 10px;
# # # # # #     padding: 1rem 1.25rem;
# # # # # #     font-family: 'DM Mono', monospace;
# # # # # #     font-size: .8rem;
# # # # # #     line-height: 1.6;
# # # # # #     max-height: 380px;
# # # # # #     overflow-y: auto;
# # # # # #     white-space: pre-wrap;
# # # # # # }

# # # # # # .fit-box {
# # # # # #     border-radius: 8px;
# # # # # #     padding: .75rem;
# # # # # #     margin-bottom: .5rem;
# # # # # #     font-size: .85rem;
# # # # # # }

# # # # # # /* Hide sidebar on login page */
# # # # # # .hide-sidebar [data-testid="stSidebar"] { display: none !important; }
# # # # # # .hide-sidebar .main .block-container { max-width: 480px; margin: 0 auto; }

# # # # # # </style>
# # # # # # """, unsafe_allow_html=True)


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  SESSION STATE INIT
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # for _k, _v in [
# # # # # #     ("logged_in",        False),
# # # # # #     ("login_error",      ""),
# # # # # #     ("pipeline_running", False),
# # # # # #     ("pipeline_logs",    []),
# # # # # #     ("pipeline_result",  None),
# # # # # #     ("pipeline_done",    False),
# # # # # # ]:
# # # # # #     if _k not in st.session_state:
# # # # # #         st.session_state[_k] = _v


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  LOGIN PAGE
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # if not st.session_state.logged_in:

# # # # # #     # Hide sidebar on login page
# # # # # #     st.markdown("""
# # # # # #     <script>
# # # # # #     document.querySelector('[data-testid="stSidebar"]') &&
# # # # # #     (document.querySelector('[data-testid="stSidebar"]').style.display = 'none');
# # # # # #     </script>
# # # # # #     """, unsafe_allow_html=True)

# # # # # #     # Center the login card
# # # # # #     _, col, _ = st.columns([1, 1.2, 1])

# # # # # #     with col:
# # # # # #         st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)

# # # # # #         st.markdown("""
# # # # # #         <div class="login-card">
# # # # # #           <div class="login-logo">🛡️ SecureITLab</div>
# # # # # #           <div class="login-subtitle">Pipeline Intelligence</div>
# # # # # #           <div class="login-divider"></div>
# # # # # #         </div>
# # # # # #         """, unsafe_allow_html=True)

# # # # # #         st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # # #         with st.container():
# # # # # #             username = st.text_input(
# # # # # #                 "Username",
# # # # # #                 placeholder="Enter username",
# # # # # #                 key="login_username",
# # # # # #             )
# # # # # #             password = st.text_input(
# # # # # #                 "Password",
# # # # # #                 placeholder="Enter password",
# # # # # #                 type="password",
# # # # # #                 key="login_password",
# # # # # #             )

# # # # # #             st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # # #             login_btn = st.button(
# # # # # #                 "Sign In →",
# # # # # #                 use_container_width=True,
# # # # # #                 type="primary",
# # # # # #             )

# # # # # #             if login_btn:
# # # # # #                 if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
# # # # # #                     st.session_state.logged_in = True
# # # # # #                     st.session_state.login_error = ""
# # # # # #                     st.rerun()
# # # # # #                 else:
# # # # # #                     st.session_state.login_error = "Incorrect username or password."

# # # # # #             if st.session_state.login_error:
# # # # # #                 st.markdown(
# # # # # #                     f'<div class="login-error">⚠️ {st.session_state.login_error}</div>',
# # # # # #                     unsafe_allow_html=True,
# # # # # #                 )

# # # # # #         st.markdown(
# # # # # #             "<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>"
# # # # # #             "SecureITLab · Confidential</div>",
# # # # # #             unsafe_allow_html=True,
# # # # # #         )

# # # # # #     st.stop()  # ← Stop here — nothing below renders until logged in


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  EVERYTHING BELOW ONLY RUNS WHEN LOGGED IN
# # # # # # # ══════════════════════════════════════════════════════════════════════════════


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  EXCEL EXPORT HELPER
# # # # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # # # def build_contacts_excel(contacts: list, company: str, role: str):
# # # # # #     """Build a formatted Excel workbook from contacts. Returns bytes or None."""
# # # # # #     try:
# # # # # #         from openpyxl import Workbook
# # # # # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # # # # #         from openpyxl.utils import get_column_letter
# # # # # #     except ImportError:
# # # # # #         return None

# # # # # #     wb = Workbook()
# # # # # #     ws = wb.active
# # # # # #     ws.title = "Contacts"

# # # # # #     NAVY  = "1E3A5F"
# # # # # #     BLUE  = "2563EB"
# # # # # #     LBLUE = "EFF6FF"
# # # # # #     GREY  = "F8FAFC"
# # # # # #     WHITE = "FFFFFF"

# # # # # #     pri_colors = {
# # # # # #         "Primary":   ("FEF2F2", "B91C1C"),
# # # # # #         "Secondary": ("FFFBEB", "B45309"),
# # # # # #         "Tertiary":  ("EFF6FF", "1D4ED8"),
# # # # # #         "General":   ("F5F3FF", "6D28D9"),
# # # # # #     }

# # # # # #     thin   = Side(border_style="thin", color="D9E2EC")
# # # # # #     border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # # # # #     # Row 1: Title
# # # # # #     ws.merge_cells("A1:H1")
# # # # # #     c = ws["A1"]
# # # # # #     c.value     = f"Contacts Export  —  {company}  |  {role}"
# # # # # #     c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
# # # # # #     c.fill      = PatternFill("solid", fgColor=NAVY)
# # # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # #     ws.row_dimensions[1].height = 30

# # # # # #     # Row 2: Sub-title
# # # # # #     ws.merge_cells("A2:H2")
# # # # # #     c = ws["A2"]
# # # # # #     c.value     = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(contacts)} contacts"
# # # # # #     c.font      = Font(name="Arial", size=9, color="64748B")
# # # # # #     c.fill      = PatternFill("solid", fgColor="F4F7FB")
# # # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # #     ws.row_dimensions[2].height = 18

# # # # # #     # Row 3: Headers
# # # # # #     headers    = ["#", "Priority", "Name", "Title / Role", "Company", "Email", "LinkedIn URL", "Source"]
# # # # # #     col_widths = [4,    12,         24,     32,              22,        34,       42,              18]

# # # # # #     for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
# # # # # #         c = ws.cell(row=3, column=ci, value=h)
# # # # # #         c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
# # # # # #         c.fill      = PatternFill("solid", fgColor=BLUE)
# # # # # #         c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # #         c.border    = border
# # # # # #         ws.column_dimensions[get_column_letter(ci)].width = w

# # # # # #     ws.row_dimensions[3].height = 22

# # # # # #     # Data rows
# # # # # #     for ri, ct in enumerate(contacts, start=4):
# # # # # #         prio           = ct.get("priority", "General")
# # # # # #         bg_hex, fg_hex = pri_colors.get(prio, (WHITE, "0F172A"))
# # # # # #         patterns       = ct.get("email_patterns", [])
# # # # # #         email_val      = ct.get("email") or (patterns[0] + "  (pattern)" if patterns else "")
# # # # # #         row_fill       = bg_hex if ri % 2 == 0 else GREY

# # # # # #         for ci, val in enumerate([
# # # # # #             ri - 3,
# # # # # #             prio,
# # # # # #             ct.get("name", ""),
# # # # # #             ct.get("title", ""),
# # # # # #             ct.get("company", ""),
# # # # # #             email_val,
# # # # # #             ct.get("linkedin_url", ""),
# # # # # #             ct.get("source", ""),
# # # # # #         ], 1):
# # # # # #             cell = ws.cell(row=ri, column=ci, value=val)
# # # # # #             cell.font      = Font(name="Arial", size=9,
# # # # # #                                   bold=(ci == 2),
# # # # # #                                   color=fg_hex if ci == 2 else "0F172A")
# # # # # #             cell.fill      = PatternFill("solid", fgColor=row_fill)
# # # # # #             cell.alignment = Alignment(vertical="center", wrap_text=(ci in [4, 7]))
# # # # # #             cell.border    = border

# # # # # #         ws.row_dimensions[ri].height = 18

# # # # # #     ws.freeze_panes   = "A4"
# # # # # #     ws.auto_filter.ref = f"A3:H{3 + len(contacts)}"

# # # # # #     # Summary sheet
# # # # # #     ws2 = wb.create_sheet("Summary")
# # # # # #     ws2.merge_cells("A1:C1")
# # # # # #     t = ws2["A1"]
# # # # # #     t.value     = "Export Summary"
# # # # # #     t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
# # # # # #     t.fill      = PatternFill("solid", fgColor=NAVY)
# # # # # #     t.alignment = Alignment(horizontal="center")

# # # # # #     for r, (lbl, val) in enumerate([
# # # # # #         ("Company",        company),
# # # # # #         ("Role",           role),
# # # # # #         ("Total Contacts", len(contacts)),
# # # # # #         ("Primary",        sum(1 for x in contacts if x.get("priority") == "Primary")),
# # # # # #         ("Secondary",      sum(1 for x in contacts if x.get("priority") == "Secondary")),
# # # # # #         ("Tertiary",       sum(1 for x in contacts if x.get("priority") == "Tertiary")),
# # # # # #         ("General",        sum(1 for x in contacts if x.get("priority") == "General")),
# # # # # #         ("With Email",     sum(1 for x in contacts if x.get("email"))),
# # # # # #         ("With LinkedIn",  sum(1 for x in contacts if x.get("linkedin_url"))),
# # # # # #         ("Generated",      datetime.now().strftime("%d %b %Y  %H:%M")),
# # # # # #     ], 2):
# # # # # #         lc = ws2.cell(row=r, column=1, value=lbl)
# # # # # #         vc = ws2.cell(row=r, column=2, value=val)
# # # # # #         bg = LBLUE if r % 2 == 0 else WHITE
# # # # # #         for cell in (lc, vc):
# # # # # #             cell.font   = Font(name="Arial", bold=(cell.column == 1), size=10)
# # # # # #             cell.fill   = PatternFill("solid", fgColor=bg)
# # # # # #             cell.border = border

# # # # # #     ws2.column_dimensions["A"].width = 20
# # # # # #     ws2.column_dimensions["B"].width = 30

# # # # # #     buf = io.BytesIO()
# # # # # #     wb.save(buf)
# # # # # #     buf.seek(0)
# # # # # #     return buf.getvalue()


# # # # # # # ── MongoDB helpers ───────────────────────────────────────────────────────────
# # # # # # @st.cache_resource
# # # # # # def get_db():
# # # # # #     URI = st.secrets.get("MONGO_URI", "mongodb://localhost:27017")
# # # # # #     DB  = st.secrets.get("MONGO_DB",  "secureitlab_job_pipeline")
# # # # # #     client = MongoClient(URI, serverSelectionTimeoutMS=6000)
# # # # # #     return client[DB]

# # # # # # @st.cache_data(ttl=60)
# # # # # # def load_all_jobs():
# # # # # #     db = get_db()
# # # # # #     return list(db.jobs.find({}, {
# # # # # #         "_id": 1, "company": 1, "role": 1, "job_number": 1,
# # # # # #         "opp_score": 1, "contacts_found": 1, "pipeline_ok": 1,
# # # # # #         "coverage_score": 1, "run_at": 1, "contact_domain": 1,
# # # # # #     }))

# # # # # # @st.cache_data(ttl=60)
# # # # # # def load_job(job_id):
# # # # # #     from bson import ObjectId
# # # # # #     return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# # # # # # # ── Render helpers ────────────────────────────────────────────────────────────
# # # # # # def badge(text, color="blue"):
# # # # # #     return f'<span class="badge badge-{color}">{text}</span>'

# # # # # # def safe_str(val, limit=300):
# # # # # #     if val is None: return "—"
# # # # # #     s = str(val)
# # # # # #     return s[:limit] + "…" if len(s) > limit else s

# # # # # # def as_dict(raw):
# # # # # #     if isinstance(raw, dict): return raw
# # # # # #     if isinstance(raw, list): return next((x for x in raw if isinstance(x, dict)), {})
# # # # # #     return {}

# # # # # # def render_json_pretty(data, title=""):
# # # # # #     if not data: return
# # # # # #     with st.expander(f"📄 Raw JSON — {title}", expanded=False):
# # # # # #         st.code(json.dumps(data, indent=2, default=str), language="json")

# # # # # # def render_qa_block(data, label):
# # # # # #     if not data:
# # # # # #         st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>', unsafe_allow_html=True)
# # # # # #         return
# # # # # #     data = as_dict(data)
# # # # # #     if not data: return
# # # # # #     passed    = data.get("passed") or data.get("Passed") or False
# # # # # #     rec       = data.get("recommendation") or data.get("Recommendation", "")
# # # # # #     issues    = data.get("issues") or data.get("Issues") or []
# # # # # #     checklist = data.get("checklist") or data.get("Checklist") or []
# # # # # #     color     = "green" if passed else "yellow"
# # # # # #     status    = "✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
# # # # # #     html = f"""
# # # # # #     <div class="sil-card sil-card-accent">
# # # # # #       <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
# # # # # #         <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
# # # # # #         {badge(status, color)}
# # # # # #       </div>"""
# # # # # #     if rec:
# # # # # #         html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
# # # # # #     if checklist:
# # # # # #         html += '<div style="font-size:.82rem;margin-bottom:.5rem">'
# # # # # #         for item in (checklist if isinstance(checklist, list) else [checklist]):
# # # # # #             if isinstance(item, dict):
# # # # # #                 ip = item.get("pass") or item.get("passed") or item.get("status","") == "pass"
# # # # # #                 nm = item.get("item") or item.get("name") or item.get("check","")
# # # # # #                 nt = item.get("reason") or item.get("note") or item.get("issue","")
# # # # # #                 html += f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
# # # # # #                 if nt: html += f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
# # # # # #                 html += '</div>'
# # # # # #         html += '</div>'
# # # # # #     if issues:
# # # # # #         html += '<div style="margin-top:.5rem">'
# # # # # #         for iss in (issues if isinstance(issues, list) else [issues])[:4]:
# # # # # #             txt = iss if isinstance(iss, str) else json.dumps(iss)
# # # # # #             html += f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
# # # # # #         html += '</div>'
# # # # # #     st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # # def render_contacts(contacts, title="Contacts"):
# # # # # #     if not contacts: st.info("No contacts found for this job."); return
# # # # # #     pri_color = {"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
# # # # # #     st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>', unsafe_allow_html=True)
# # # # # #     cols = st.columns(2)
# # # # # #     for i, c in enumerate(contacts):
# # # # # #         col = cols[i % 2]
# # # # # #         prio = c.get("priority", "General")
# # # # # #         email = c.get("email", ""); li = c.get("linkedin_url", "")
# # # # # #         patterns = c.get("email_patterns", []); src = c.get("source", "")
# # # # # #         with col:
# # # # # #             html = f"""<div class="contact-card">
# # # # # #               <div style="display:flex;justify-content:space-between;align-items:flex-start">
# # # # # #                 <div>
# # # # # #                   <div class="contact-name">{c.get('name','—')}</div>
# # # # # #                   <div class="contact-title">{c.get('title','—')}</div>
# # # # # #                 </div>
# # # # # #                 {badge(prio, pri_color.get(prio,'blue'))}
# # # # # #               </div>"""
# # # # # #             if email:      html += f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
# # # # # #             elif patterns: html += f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
# # # # # #             if li:         html += f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
# # # # # #             if src:        html += f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
# # # # # #             st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # # def render_emails(emails_data):
# # # # # #     if not emails_data: st.info("No email data available."); return
# # # # # #     emails_data = as_dict(emails_data)
# # # # # #     if not emails_data: return
# # # # # #     variants = {}
# # # # # #     for k, v in emails_data.items():
# # # # # #         kl = k.lower().replace("_","").replace(" ","")
# # # # # #         if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl == "a":
# # # # # #             variants["Variant A — Hiring Manager"] = v
# # # # # #         elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl == "b":
# # # # # #             variants["Variant B — CISO / VP Level"] = v
# # # # # #         else:
# # # # # #             variants[k] = v
# # # # # #     for label, v in variants.items():
# # # # # #         st.markdown(f'<div class="section-label">✉️ {label}</div>', unsafe_allow_html=True)
# # # # # #         if isinstance(v, dict):
# # # # # #             subj = v.get("subject") or v.get("Subject","")
# # # # # #             body = v.get("body") or v.get("Body") or v.get("content","")
# # # # # #             if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>', unsafe_allow_html=True)
# # # # # #             if body: st.markdown(f'<div class="email-box">{body}</div>', unsafe_allow_html=True)
# # # # # #             else:    st.code(json.dumps(v, indent=2), language="json")
# # # # # #         elif isinstance(v, str):
# # # # # #             st.markdown(f'<div class="email-box">{v}</div>', unsafe_allow_html=True)
# # # # # #         st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

# # # # # # def render_service_mapping(data):
# # # # # #     if not data: st.info("No service mapping data."); return
# # # # # #     items = data if isinstance(data, list) else []
# # # # # #     if not items and isinstance(data, dict):
# # # # # #         for key in ("services","mappings","service_mapping","ServiceMapping","items"):
# # # # # #             if isinstance(data.get(key), list): items = data[key]; break
# # # # # #         if not items: items = [data]
# # # # # #     fit_colors = {"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
# # # # # #     for item in items:
# # # # # #         if not isinstance(item, dict): continue
# # # # # #         svc  = item.get("service") or item.get("Service") or item.get("name","")
# # # # # #         fit  = (item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
# # # # # #         why  = item.get("justification") or item.get("rationale") or item.get("why","")
# # # # # #         reqs = item.get("requirements_addressed") or item.get("requirements") or ""
# # # # # #         eng  = item.get("engagement_type") or item.get("engagement","")
# # # # # #         color = fit_colors.get(fit, "blue")
# # # # # #         html = f"""<div class="sil-card" style="margin-bottom:.75rem">
# # # # # #           <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
# # # # # #             <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
# # # # # #             {badge(fit or "MAPPED", color) if fit else ""}
# # # # # #           </div>"""
# # # # # #         if why:  html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
# # # # # #         if reqs:
# # # # # #             rs = ", ".join(reqs) if isinstance(reqs, list) else str(reqs)
# # # # # #             html += f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
# # # # # #         if eng:  html += f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
# # # # # #         st.markdown(html + '</div>', unsafe_allow_html=True)
# # # # # #     render_json_pretty(data, "Service Mapping")

# # # # # # def render_microplans(data):
# # # # # #     if not data: st.info("No micro-plan data."); return
# # # # # #     plans = data if isinstance(data, list) else []
# # # # # #     if not plans and isinstance(data, dict):
# # # # # #         for k in ("plans","micro_plans","microplans","top_3","improvements"):
# # # # # #             if isinstance(data.get(k), list): plans = data[k]; break
# # # # # #         if not plans: plans = [data]
# # # # # #     for i, plan in enumerate(plans[:3], 1):
# # # # # #         if not isinstance(plan, dict): continue
# # # # # #         title = plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
# # # # # #         weeks = plan.get("duration") or plan.get("timeline","")
# # # # # #         obj   = plan.get("objective") or plan.get("goal","")
# # # # # #         kpis  = plan.get("kpis") or plan.get("KPIs") or []
# # # # # #         tasks = plan.get("tasks") or plan.get("workstreams") or []
# # # # # #         with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}", expanded=(i==1)):
# # # # # #             if obj and obj != title: st.markdown(f"**Objective:** {obj}")
# # # # # #             if kpis:
# # # # # #                 st.markdown("**KPIs:**")
# # # # # #                 for kpi in (kpis if isinstance(kpis, list) else [kpis]): st.markdown(f"• {kpi}")
# # # # # #             if tasks:
# # # # # #                 st.markdown("**Tasks / Workstreams:**")
# # # # # #                 for t in (tasks if isinstance(tasks, list) else [tasks]):
# # # # # #                     if isinstance(t, dict):
# # # # # #                         tn = t.get("task") or t.get("name","")
# # # # # #                         te = t.get("effort") or t.get("duration","")
# # # # # #                         st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
# # # # # #                     else: st.markdown(f"• {t}")
# # # # # #             if plan:
# # # # # #                 st.code(json.dumps(plan, indent=2, default=str), language="json")

# # # # # # def render_deal_assurance(data):
# # # # # #     if not data: st.info("No deal assurance data."); return
# # # # # #     if not isinstance(data, dict): render_json_pretty(data, "Deal Assurance Pack"); return
# # # # # #     evp = (data.get("executive_value_proposition") or
# # # # # #            data.get("value_proposition") or data.get("ExecutiveValueProposition",""))
# # # # # #     if evp:
# # # # # #         st.markdown('<div class="section-label">💼 Executive Value Proposition</div>', unsafe_allow_html=True)
# # # # # #         st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7;color:var(--text)">{evp}</div>', unsafe_allow_html=True)
# # # # # #     caps = data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
# # # # # #     if caps:
# # # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>', unsafe_allow_html=True)
# # # # # #         c1, c2 = st.columns(2)
# # # # # #         for i, cap in enumerate(caps if isinstance(caps, list) else [caps]):
# # # # # #             (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
# # # # # #     risk = data.get("risk_mitigation") or data.get("RiskMitigation","")
# # # # # #     if risk:
# # # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>', unsafe_allow_html=True)
# # # # # #         if isinstance(risk, dict):
# # # # # #             for k, v in risk.items(): st.markdown(f"**{k}:** {v}")
# # # # # #         else: st.markdown(str(risk))
# # # # # #     render_json_pretty(data, "Deal Assurance Pack")


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  SIDEBAR
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # with st.sidebar:
# # # # # #     st.markdown("""
# # # # # #     <div style="padding:.75rem 0 1.25rem">
# # # # # #       <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;
# # # # # #                   color:#2563eb">🛡️ SecureITLab</div>
# # # # # #       <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;
# # # # # #                   text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
# # # # # #     </div>
# # # # # #     """, unsafe_allow_html=True)

# # # # # #     # ── Logout button ─────────────────────────────────────────────────────────
# # # # # #     if st.button("🚪 Logout", use_container_width=True):
# # # # # #         st.session_state.logged_in = False
# # # # # #         st.session_state.login_error = ""
# # # # # #         st.rerun()

# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # #     # ── Find Jobs button ──────────────────────────────────────────────────────
# # # # # #     st.markdown("**🔍 Find New Jobs**")
# # # # # #     st.caption(
# # # # # #         "Runs scraper → checks MongoDB for duplicates → "
# # # # # #         "runs AI pipeline only on NEW jobs → stores in MongoDB"
# # # # # #     )

# # # # # #     find_jobs_clicked = st.button(
# # # # # #         "🔍  Find Jobs",
# # # # # #         disabled=st.session_state.pipeline_running,
# # # # # #         use_container_width=True,
# # # # # #         type="primary",
# # # # # #     )

# # # # # #     if st.session_state.pipeline_running:
# # # # # #         st.markdown("""
# # # # # #         <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
# # # # # #                     padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;
# # # # # #                     font-size:.8rem;color:#1d4ed8">
# # # # # #           ⏳ Pipeline is running… check log below
# # # # # #         </div>""", unsafe_allow_html=True)

# # # # # #     if st.session_state.pipeline_done and st.session_state.pipeline_result:
# # # # # #         res = st.session_state.pipeline_result
# # # # # #         ok  = res.get("success", False)
# # # # # #         bg  = "#ecfdf5" if ok else "#fef2f2"
# # # # # #         bc  = "#bbf7d0" if ok else "#fecaca"
# # # # # #         tc  = "#15803d" if ok else "#b91c1c"
# # # # # #         ic  = "✅" if ok else "❌"
# # # # # #         st.markdown(f"""
# # # # # #         <div style="background:{bg};border:1px solid {bc};border-radius:8px;
# # # # # #                     padding:.75rem;margin-top:.5rem;font-size:.82rem">
# # # # # #           <div style="font-weight:700;color:{tc};margin-bottom:.4rem">{ic} Last Run</div>
# # # # # #           <div>📦 Scraped: <b>{res.get('scraped',0)}</b></div>
# # # # # #           <div>🆕 New jobs: <b>{res.get('new_jobs',0)}</b></div>
# # # # # #           <div>⏭️ Already in DB (skipped): <b>{res.get('skipped_db',0)}</b></div>
# # # # # #           <div>🤖 Processed by AI: <b>{res.get('processed',0)}</b></div>
# # # # # #           {f'<div style="color:{tc};margin-top:.3rem">⚠️ {res.get("error","")}</div>' if res.get("error") else ""}
# # # # # #         </div>""", unsafe_allow_html=True)

# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # #     # ── Job list ──────────────────────────────────────────────────────────────
# # # # # #     try:
# # # # # #         all_jobs = load_all_jobs()
# # # # # #     except Exception as e:
# # # # # #         st.error(f"MongoDB connection failed: {e}")
# # # # # #         st.stop()

# # # # # #     if not all_jobs:
# # # # # #         st.warning("No jobs in MongoDB yet. Click **Find Jobs** to scrape and process.")
# # # # # #         st.stop()

# # # # # #     st.markdown(
# # # # # #         f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">'
# # # # # #         f'{len(all_jobs)} jobs in database</div>',
# # # # # #         unsafe_allow_html=True,
# # # # # #     )

# # # # # #     search   = st.text_input("🔍 Filter by company / role", placeholder="e.g. Bounteous")
# # # # # #     filtered = [j for j in all_jobs
# # # # # #                 if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

# # # # # #     def job_label(j):
# # # # # #         score = j.get("opp_score")
# # # # # #         s_str = f" [{score}/10]" if score else ""
# # # # # #         ok    = "✅" if j.get("pipeline_ok") else "❌"
# # # # # #         return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s_str}"

# # # # # #     if not filtered:
# # # # # #         st.warning("No matching jobs.")
# # # # # #         st.stop()

# # # # # #     sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
# # # # # #     sel_idx     = [job_label(j) for j in filtered].index(sel_label)
# # # # # #     selected_id = str(filtered[sel_idx]["_id"])

# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # #     ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
# # # # # #     total_c  = sum(j.get("contacts_found",0) for j in all_jobs)
# # # # # #     st.markdown(f"""
# # # # # #     <div style="font-size:.75rem;color:#64748b">
# # # # # #       <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
# # # # # #       <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
# # # # # #     </div>""", unsafe_allow_html=True)

# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # #     if st.button("🔄 Refresh Data", use_container_width=True):
# # # # # #         st.cache_data.clear()
# # # # # #         st.rerun()


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  FIND JOBS — background thread
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # if find_jobs_clicked and not st.session_state.pipeline_running:
# # # # # #     with _thread_lock:
# # # # # #         _thread_logs.clear()
# # # # # #         _thread_result[0] = None
# # # # # #         _thread_done[0]   = False

# # # # # #     st.session_state.pipeline_running = True
# # # # # #     st.session_state.pipeline_done    = False
# # # # # #     st.session_state.pipeline_logs    = []
# # # # # #     st.session_state.pipeline_result  = None
# # # # # #     st.cache_data.clear()

# # # # # #     def _run_pipeline_bg():
# # # # # #         try:
# # # # # #             import main as _main

# # # # # #             def _cb(msg: str):
# # # # # #                 with _thread_lock:
# # # # # #                     _thread_logs.append(f"{datetime.now().strftime('%H:%M:%S')} | {msg}")

# # # # # #             result = _main.run_pipeline(progress_callback=_cb)
# # # # # #             with _thread_lock:
# # # # # #                 _thread_result[0] = result
# # # # # #         except Exception as e:
# # # # # #             with _thread_lock:
# # # # # #                 _thread_logs.append(f"❌ Unexpected error: {e}")
# # # # # #                 _thread_result[0] = {
# # # # # #                     "success": False, "error": str(e),
# # # # # #                     "scraped": 0, "new_jobs": 0, "skipped_db": 0, "processed": 0,
# # # # # #                 }
# # # # # #         finally:
# # # # # #             with _thread_lock:
# # # # # #                 _thread_done[0] = True

# # # # # #     threading.Thread(target=_run_pipeline_bg, daemon=True).start()
# # # # # #     st.rerun()

# # # # # # # ── Sync thread state → session_state ────────────────────────────────────────
# # # # # # with _thread_lock:
# # # # # #     if _thread_logs:
# # # # # #         st.session_state.pipeline_logs = list(_thread_logs)
# # # # # #     if _thread_done[0] and _thread_result[0] is not None:
# # # # # #         st.session_state.pipeline_result  = _thread_result[0]
# # # # # #         st.session_state.pipeline_running = False
# # # # # #         st.session_state.pipeline_done    = True

# # # # # # # ── Live log panel ────────────────────────────────────────────────────────────
# # # # # # if st.session_state.pipeline_running or (
# # # # # #         st.session_state.pipeline_done and st.session_state.pipeline_logs):
# # # # # #     heading = "⏳ Pipeline running — live log…" if st.session_state.pipeline_running \
# # # # # #               else "📋 Last pipeline run log"
# # # # # #     with st.expander(heading, expanded=st.session_state.pipeline_running):
# # # # # #         log_text = "\n".join(st.session_state.pipeline_logs[-100:]) \
# # # # # #                    if st.session_state.pipeline_logs else "Starting…"
# # # # # #         st.markdown(f'<div class="pipeline-log">{log_text}</div>', unsafe_allow_html=True)
# # # # # #         if st.session_state.pipeline_running:
# # # # # #             time.sleep(2)
# # # # # #             st.rerun()


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  MAIN CONTENT — selected job detail
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # with st.spinner("Loading job from MongoDB…"):
# # # # # #     job = load_job(selected_id)

# # # # # # if not job:
# # # # # #     st.error("Could not load job document.")
# # # # # #     st.stop()

# # # # # # company   = job.get("company",  "Unknown")
# # # # # # role      = job.get("role",     "Unknown")
# # # # # # opp_score = job.get("opp_score")
# # # # # # p_ok      = job.get("pipeline_ok", False)
# # # # # # p_min     = job.get("pipeline_min", "?")
# # # # # # c_found   = job.get("contacts_found", 0)
# # # # # # c_cov     = job.get("coverage_score")
# # # # # # c_domain  = job.get("contact_domain","")
# # # # # # run_at    = job.get("run_at","")

# # # # # # # ── Header ────────────────────────────────────────────────────────────────────
# # # # # # st.markdown(f"""
# # # # # # <div style="margin-bottom:1.75rem">
# # # # # #   <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;
# # # # # #               letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
# # # # # #     Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
# # # # # #   </div>
# # # # # #   <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;
# # # # # #              color:#0f172a;margin:0;line-height:1.15">{role}</h1>
# # # # # #   <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
# # # # # #     @ <span style="color:#334155;font-weight:600">{company}</span>
# # # # # #     {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
# # # # # #   </div>
# # # # # # </div>
# # # # # # """, unsafe_allow_html=True)

# # # # # # # ── Metric tiles ──────────────────────────────────────────────────────────────
# # # # # # try:
# # # # # #     sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
# # # # # #     sc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # # # except Exception:
# # # # # #     sc = "#2563eb"

# # # # # # st.markdown(f"""
# # # # # # <div class="metric-row">
# # # # # #   <div class="metric-tile">
# # # # # #     <div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div>
# # # # # #     <div class="lbl">Opportunity Score</div>
# # # # # #   </div>
# # # # # #   <div class="metric-tile">
# # # # # #     <div class="val">{c_found}</div>
# # # # # #     <div class="lbl">Contacts Found</div>
# # # # # #   </div>
# # # # # #   <div class="metric-tile">
# # # # # #     <div class="val">{f"{c_cov}%" if c_cov else "—"}</div>
# # # # # #     <div class="lbl">Contact Coverage</div>
# # # # # #   </div>
# # # # # #   <div class="metric-tile">
# # # # # #     <div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">
# # # # # #       {'✅ OK' if p_ok else '❌ Failed'}
# # # # # #     </div>
# # # # # #     <div class="lbl">Pipeline ({p_min} min)</div>
# # # # # #   </div>
# # # # # # </div>
# # # # # # """, unsafe_allow_html=True)

# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  TABS
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # tabs = st.tabs([
# # # # # #     "📋 Job & Enrichment",
# # # # # #     "🎯 Service Mapping",
# # # # # #     "🔍 Fit / Gap",
# # # # # #     "🛠️ Capability & Plans",
# # # # # #     "📦 Deal Assurance",
# # # # # #     "✉️ Outreach Emails",
# # # # # #     "👥 Contacts",
# # # # # #     "✅ QA Gates",
# # # # # #     "🗄️ Raw Data",
# # # # # # ])

# # # # # # # ── Tab 1: Job Research + Enrichment ─────────────────────────────────────────
# # # # # # with tabs[0]:
# # # # # #     col1, col2 = st.columns([1, 1])
# # # # # #     with col1:
# # # # # #         st.markdown('<div class="section-label">📄 Job Research</div>', unsafe_allow_html=True)
# # # # # #         jr = as_dict(job.get("agent_job_research") or {})
# # # # # #         if jr:
# # # # # #             for label, keys in [
# # # # # #                 ("Job Role",        ["job_role","Job Role","role","title"]),
# # # # # #                 ("Company",         ["company_name","Company Name","company"]),
# # # # # #                 ("Location",        ["location","Location"]),
# # # # # #                 ("Organization URL",["organization_url","Organization URL","url"]),
# # # # # #             ]:
# # # # # #                 val = next((jr.get(k) for k in keys if jr.get(k)), None)
# # # # # #                 if val: st.markdown(f"**{label}:** {val}")
# # # # # #             desc = jr.get("job_description") or jr.get("Job Description","")
# # # # # #             if desc:
# # # # # #                 st.markdown("**Job Description:**")
# # # # # #                 st.markdown(
# # # # # #                     f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;'
# # # # # #                     f'max-height:300px;overflow-y:auto;color:var(--text)">{desc[:2000]}</div>',
# # # # # #                     unsafe_allow_html=True,
# # # # # #                 )
# # # # # #             render_json_pretty(jr, "Job Research")
# # # # # #         else:
# # # # # #             st.info("No job research data.")
# # # # # #     with col2:
# # # # # #         st.markdown('<div class="section-label">🏢 Company Enrichment</div>', unsafe_allow_html=True)
# # # # # #         enr = as_dict(job.get("agent_enrichment") or {})
# # # # # #         if enr:
# # # # # #             for label, keys in [
# # # # # #                 ("Industry",          ["industry","Industry"]),
# # # # # #                 ("Company Size",      ["company_size","size","Company Size"]),
# # # # # #                 ("Regulatory Env",    ["regulatory_environment","regulatory"]),
# # # # # #                 ("Certifications",    ["certifications","Certifications"]),
# # # # # #                 ("Tech Stack",        ["tech_stack","technology_stack"]),
# # # # # #                 ("Security Maturity", ["security_maturity","maturity"]),
# # # # # #             ]:
# # # # # #                 val = next((enr.get(k) for k in keys if enr.get(k)), None)
# # # # # #                 if val:
# # # # # #                     if isinstance(val, list): val = ", ".join(str(v) for v in val)
# # # # # #                     st.markdown(f"**{label}:** {safe_str(val, 200)}")
# # # # # #             render_json_pretty(enr, "Enrichment")
# # # # # #         else:
# # # # # #             st.info("No enrichment data.")

# # # # # # # ── Tab 2: Service Mapping ────────────────────────────────────────────────────
# # # # # # with tabs[1]:
# # # # # #     st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>', unsafe_allow_html=True)
# # # # # #     render_service_mapping(job.get("agent_service_mapping"))

# # # # # # # ── Tab 3: Fit / Gap Analysis ─────────────────────────────────────────────────
# # # # # # with tabs[2]:
# # # # # #     fg = as_dict(job.get("agent_fit_gap") or {})
# # # # # #     if opp_score:
# # # # # #         try:
# # # # # #             sn = float(str(opp_score).split("/")[0])
# # # # # #             bc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # # #             bp = int(sn / 10 * 100)
# # # # # #             st.markdown(f"""
# # # # # #             <div style="margin-bottom:1.5rem">
# # # # # #               <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
# # # # # #                 <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">Opportunity Score</span>
# # # # # #                 <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
# # # # # #               </div>
# # # # # #               <div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%">
# # # # # #                 <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
# # # # # #               </div>
# # # # # #             </div>""", unsafe_allow_html=True)
# # # # # #         except Exception:
# # # # # #             pass
# # # # # #     st.markdown('<div class="section-label">📊 Service Classifications</div>', unsafe_allow_html=True)
# # # # # #     services = []
# # # # # #     if isinstance(fg, dict):
# # # # # #         for k in ("services","classifications","service_classifications","items","fit_gap"):
# # # # # #             v = fg.get(k)
# # # # # #             if isinstance(v, list): services = v; break
# # # # # #         if not services and (fg.get("service") or fg.get("Service")):
# # # # # #             services = [fg]
# # # # # #     elif isinstance(fg, list):
# # # # # #         services = fg
# # # # # #     if services:
# # # # # #         buckets = {"STRONG FIT": [], "PARTIAL FIT": [], "GAP": []}
# # # # # #         other   = []
# # # # # #         for s in services:
# # # # # #             if not isinstance(s, dict): continue
# # # # # #             fit = (s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
# # # # # #             if "STRONG" in fit: buckets["STRONG FIT"].append(s)
# # # # # #             elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
# # # # # #             elif "GAP" in fit: buckets["GAP"].append(s)
# # # # # #             else: other.append(s)
# # # # # #         c1, c2, c3 = st.columns(3)
# # # # # #         cm  = {"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
# # # # # #         bgm = {"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
# # # # # #         bdm = {"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
# # # # # #         for col, (fl, items) in zip([c1, c2, c3], buckets.items()):
# # # # # #             col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>', unsafe_allow_html=True)
# # # # # #             for s in items:
# # # # # #                 svc  = s.get("service") or s.get("Service") or s.get("name","")
# # # # # #                 just = s.get("justification") or s.get("reason","")
# # # # # #                 col.markdown(
# # # # # #                     f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem">'
# # # # # #                     f'<div style="font-weight:600;color:#0f172a">{svc}</div>'
# # # # # #                     f'<div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>',
# # # # # #                     unsafe_allow_html=True,
# # # # # #                 )
# # # # # #     elif fg:
# # # # # #         st.json(fg)
# # # # # #     else:
# # # # # #         st.info("No fit/gap data.")
# # # # # #     render_json_pretty(job.get("agent_fit_gap"), "Fit/Gap Analysis")

# # # # # # # ── Tab 4: Capability + Micro-plans ──────────────────────────────────────────
# # # # # # with tabs[3]:
# # # # # #     col1, col2 = st.columns([1, 1])
# # # # # #     with col1:
# # # # # #         st.markdown('<div class="section-label">🔧 Capability Improvements</div>', unsafe_allow_html=True)
# # # # # #         cap = job.get("agent_capability") or {}
# # # # # #         items_cap = cap if isinstance(cap, list) else []
# # # # # #         if not items_cap and isinstance(cap, dict):
# # # # # #             for k in ("improvements","recommendations","capabilities","items"):
# # # # # #                 v = cap.get(k)
# # # # # #                 if isinstance(v, list): items_cap = v; break
# # # # # #             if not items_cap: items_cap = [cap]
# # # # # #         for item in items_cap:
# # # # # #             if not isinstance(item, dict): continue
# # # # # #             title  = item.get("title") or item.get("gap") or item.get("service","")
# # # # # #             rec    = item.get("recommendation") or item.get("steps","")
# # # # # #             effort = item.get("build_effort") or item.get("effort","")
# # # # # #             demand = item.get("market_demand") or item.get("priority","")
# # # # # #             st.markdown(f"""
# # # # # #             <div class="sil-card" style="margin-bottom:.6rem">
# # # # # #               <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem;color:var(--text)">{title}</div>
# # # # # #               <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec, 250)}</div>
# # # # # #               {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
# # # # # #             </div>""", unsafe_allow_html=True)
# # # # # #         if not items_cap:
# # # # # #             render_json_pretty(cap, "Capability Improvement")
# # # # # #     with col2:
# # # # # #         st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>', unsafe_allow_html=True)
# # # # # #         render_microplans(job.get("agent_microplans"))

# # # # # # # ── Tab 5: Deal Assurance Pack ────────────────────────────────────────────────
# # # # # # with tabs[4]:
# # # # # #     render_deal_assurance(job.get("agent_deal_assurance"))

# # # # # # # ── Tab 6: Outreach Emails ────────────────────────────────────────────────────
# # # # # # with tabs[5]:
# # # # # #     st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>', unsafe_allow_html=True)
# # # # # #     emails_src = job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
# # # # # #     oq = as_dict(job.get("agent_outreach_qa") or {})
# # # # # #     improved = (oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
# # # # # #     if improved:
# # # # # #         st.info("⚡ Showing QA-improved versions where available")
# # # # # #         render_emails(improved)
# # # # # #         with st.expander("📬 Original (pre-QA) versions"):
# # # # # #             render_emails(emails_src)
# # # # # #     else:
# # # # # #         render_emails(emails_src)

# # # # # # # ── Tab 7: Contacts ───────────────────────────────────────────────────────────
# # # # # # with tabs[6]:
# # # # # #     contacts        = job.get("contacts") or []
# # # # # #     contact_sources = job.get("contact_sources") or []
# # # # # #     pri = [c for c in contacts if c.get("priority") == "Primary"]
# # # # # #     sec = [c for c in contacts if c.get("priority") == "Secondary"]
# # # # # #     ter = [c for c in contacts if c.get("priority") == "Tertiary"]
# # # # # #     gen = [c for c in contacts if c.get("priority") == "General"]
# # # # # #     st.markdown(f"""
# # # # # #     <div class="metric-row" style="margin-bottom:1.5rem">
# # # # # #       <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
# # # # # #       <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
# # # # # #       <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
# # # # # #       <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
# # # # # #     </div>""", unsafe_allow_html=True)
# # # # # #     if contact_sources:
# # # # # #         st.markdown('Sources: ' + " ".join(badge(s,"blue") for s in contact_sources), unsafe_allow_html=True)
# # # # # #         st.markdown("")
# # # # # #     missing = job.get("missing_roles") or []
# # # # # #     if missing:
# # # # # #         st.markdown('⚠️ Missing roles: ' + " ".join(badge(r,"red") for r in missing), unsafe_allow_html=True)
# # # # # #         st.markdown("")
# # # # # #     if contacts:
# # # # # #         # ── Excel export button ───────────────────────────────────────────────
# # # # # #         excel_bytes = build_contacts_excel(contacts, company, role)
# # # # # #         if excel_bytes:
# # # # # #             safe_co = re.sub(r'[^a-z0-9]', '_', company.lower())[:20]
# # # # # #             fname   = f"contacts_{safe_co}_{datetime.now().strftime('%Y%m%d')}.xlsx"
# # # # # #             btn_col, _ = st.columns([1, 3])
# # # # # #             with btn_col:
# # # # # #                 st.download_button(
# # # # # #                     label="📥  Download Contacts (.xlsx)",
# # # # # #                     data=excel_bytes,
# # # # # #                     file_name=fname,
# # # # # #                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# # # # # #                     use_container_width=True,
# # # # # #                     type="primary",
# # # # # #                 )
# # # # # #         else:
# # # # # #             st.warning("Run `pip install openpyxl` to enable Excel export.")

# # # # # #         st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # #         pri_filter = st.multiselect(
# # # # # #             "Filter by priority",
# # # # # #             ["Primary","Secondary","Tertiary","General"],
# # # # # #             default=["Primary","Secondary","Tertiary","General"],
# # # # # #         )
# # # # # #         shown = [c for c in contacts if c.get("priority","General") in pri_filter]
# # # # # #         render_contacts(shown, f"Contacts ({len(shown)} shown)")
# # # # # #         agent_contacts = job.get("agent_prospect_contacts") or {}
# # # # # #         if agent_contacts:
# # # # # #             with st.expander("🤖 CrewAI Agent's Contact Search"):
# # # # # #                 if isinstance(agent_contacts, dict):
# # # # # #                     ac_list = agent_contacts.get("contacts") or []
# # # # # #                     if ac_list: render_contacts(ac_list, "Agent Contacts")
# # # # # #                     else:       st.json(agent_contacts)
# # # # # #                 else:           st.json(agent_contacts)
# # # # # #     else:
# # # # # #         st.info("No contacts found for this job.")

# # # # # # # ── Tab 8: QA Gates ───────────────────────────────────────────────────────────
# # # # # # with tabs[7]:
# # # # # #     st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>', unsafe_allow_html=True)
# # # # # #     col1, col2 = st.columns(2)
# # # # # #     for i, (label, key) in enumerate([
# # # # # #         ("Research QA",       "agent_research_qa"),
# # # # # #         ("Service Mapping QA","agent_mapping_qa"),
# # # # # #         ("Deal Assurance QA", "agent_assurance_qa"),
# # # # # #         ("Outreach Email QA", "agent_outreach_qa"),
# # # # # #     ]):
# # # # # #         with (col1 if i % 2 == 0 else col2):
# # # # # #             render_qa_block(job.get(key), label)
# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # #     st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>', unsafe_allow_html=True)
# # # # # #     enf = as_dict(job.get("agent_prospect_enforcer") or {})
# # # # # #     if enf:
# # # # # #         cov  = enf.get("coverage_score","?")
# # # # # #         miss = enf.get("missing_roles",[])
# # # # # #         note = enf.get("note","")
# # # # # #         ec   = enf.get("contacts",[])
# # # # # #         x1, x2, x3 = st.columns(3)
# # # # # #         x1.metric("Coverage Score",    f"{cov}%")
# # # # # #         x2.metric("Missing Roles",     len(miss))
# # # # # #         x3.metric("Contacts Verified", len(ec))
# # # # # #         if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
# # # # # #         if note: st.caption(note)
# # # # # #     else:
# # # # # #         st.info("No enforcer data.")

# # # # # # # ── Tab 9: Raw Data ───────────────────────────────────────────────────────────
# # # # # # with tabs[8]:
# # # # # #     st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>', unsafe_allow_html=True)
# # # # # #     st.caption("All fields stored in the `jobs` collection for this document.")
# # # # # #     rows = []
# # # # # #     for k, v in job.items():
# # # # # #         if k == "_id": continue
# # # # # #         rows.append({"Field": k, "Type": type(v).__name__, "Len": len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
# # # # # #     hc1, hc2, hc3 = st.columns([3,1,1])
# # # # # #     hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
# # # # # #     for r in rows:
# # # # # #         rc1, rc2, rc3 = st.columns([3,1,1])
# # # # # #         rc1.code(r["Field"], language=None)
# # # # # #         rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>', unsafe_allow_html=True)
# # # # # #         rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>',  unsafe_allow_html=True)
# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # #     for label, key in [
# # # # # #         ("Job Research","agent_job_research"), ("Enrichment","agent_enrichment"),
# # # # # #         ("Service Mapping","agent_service_mapping"), ("Fit/Gap Analysis","agent_fit_gap"),
# # # # # #         ("Capability","agent_capability"), ("Micro-Plans","agent_microplans"),
# # # # # #         ("Deal Assurance","agent_deal_assurance"), ("Outreach Emails","agent_outreach_emails"),
# # # # # #         ("Prospect Contacts","agent_prospect_contacts"), ("Prospect Enforcer","agent_prospect_enforcer"),
# # # # # #         ("Research QA","agent_research_qa"), ("Mapping QA","agent_mapping_qa"),
# # # # # #         ("Assurance QA","agent_assurance_qa"), ("Outreach QA","agent_outreach_qa"),
# # # # # #         ("Contacts (5-source)","contacts"),
# # # # # #     ]:
# # # # # #         data = job.get(key)
# # # # # #         if data:
# # # # # #             with st.expander(f"📄 {label}"):
# # # # # #                 st.code(json.dumps(data, indent=2, default=str), language="json")


























# # # # # """
# # # # # ╔══════════════════════════════════════════════════════════╗
# # # # # ║   SecureITLab Pipeline Dashboard — Streamlit             ║
# # # # # ║   WITH MASTER CONTACTS AGGREGATION & AUTO-UPDATE         ║
# # # # # ║   Reads from MongoDB → secureitlab_job_pipeline          ║
# # # # # ╠══════════════════════════════════════════════════════════╣
# # # # # ║  Install: pip install streamlit pymongo python-dotenv    ║
# # # # # ║  Run:     streamlit run streamlit_dashboard.py           ║
# # # # # ╚══════════════════════════════════════════════════════════╝
# # # # # """

# # # # # import io
# # # # # import re
# # # # # import streamlit as st
# # # # # from pymongo import MongoClient
# # # # # import json
# # # # # import threading
# # # # # import time
# # # # # from datetime import datetime, timezone

# # # # # # ── Thread-safe shared state ──────────────────────────────────────────────────
# # # # # _thread_logs   = []
# # # # # _thread_result = [None]
# # # # # _thread_done   = [False]
# # # # # _thread_lock   = threading.Lock()

# # # # # # ── Page config ───────────────────────────────────────────────────────────────
# # # # # st.set_page_config(
# # # # #     page_title="SecureITLab Pipeline",
# # # # #     page_icon="🛡️",
# # # # #     layout="wide",
# # # # #     initial_sidebar_state="expanded",
# # # # # )

# # # # # # ── LOGIN CREDENTIALS (change these) ─────────────────────────────────────────
# # # # # LOGIN_USERNAME = "admin"
# # # # # LOGIN_PASSWORD = "secureitlab2024"

# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  GLOBAL CSS (login + dashboard share this)
# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # st.markdown("""
# # # # # <style>

# # # # # @import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');

# # # # # :root {
# # # # #     --bg:        #f4f7fb;
# # # # #     --surface:   #ffffff;
# # # # #     --surface2:  #eef2f7;
# # # # #     --border:    #d9e2ec;
# # # # #     --accent:    #2563eb;
# # # # #     --accent2:   #7c3aed;
# # # # #     --green:     #16a34a;
# # # # #     --yellow:    #f59e0b;
# # # # #     --red:       #dc2626;
# # # # #     --text:      #0f172a;
# # # # #     --muted:     #64748b;
# # # # # }

# # # # # html, body, [class*="css"] {
# # # # #     background-color: var(--bg) !important;
# # # # #     color: var(--text) !important;
# # # # #     font-family: 'DM Sans', sans-serif !important;
# # # # # }

# # # # # /* ── LOGIN PAGE ─────────────────────────────────────────── */
# # # # # .login-wrap {
# # # # #     display: flex;
# # # # #     align-items: center;
# # # # #     justify-content: center;
# # # # #     min-height: 80vh;
# # # # # }
# # # # # .login-card {
# # # # #     background: var(--surface);
# # # # #     border: 1px solid var(--border);
# # # # #     border-radius: 20px;
# # # # #     padding: 3rem 3.5rem;
# # # # #     width: 100%;
# # # # #     max-width: 420px;
# # # # #     box-shadow: 0 20px 60px rgba(37,99,235,0.08);
# # # # #     text-align: center;
# # # # # }
# # # # # .login-logo {
# # # # #     font-family: 'Syne', sans-serif;
# # # # #     font-size: 1.6rem;
# # # # #     font-weight: 800;
# # # # #     color: var(--accent);
# # # # #     margin-bottom: .25rem;
# # # # # }
# # # # # .login-subtitle {
# # # # #     font-size: .75rem;
# # # # #     color: var(--muted);
# # # # #     letter-spacing: .12em;
# # # # #     text-transform: uppercase;
# # # # #     margin-bottom: 2.5rem;
# # # # # }
# # # # # .login-error {
# # # # #     background: #fef2f2;
# # # # #     border: 1px solid #fecaca;
# # # # #     border-radius: 8px;
# # # # #     padding: .65rem 1rem;
# # # # #     color: #b91c1c;
# # # # #     font-size: .85rem;
# # # # #     margin-top: 1rem;
# # # # # }
# # # # # .login-divider {
# # # # #     width: 40px;
# # # # #     height: 3px;
# # # # #     background: linear-gradient(90deg, var(--accent), var(--accent2));
# # # # #     border-radius: 2px;
# # # # #     margin: 0 auto 2rem;
# # # # # }

# # # # # /* ── DASHBOARD ──────────────────────────────────────────── */
# # # # # [data-testid="stSidebar"] {
# # # # #     background: var(--surface) !important;
# # # # #     border-right: 1px solid var(--border) !important;
# # # # # }
# # # # # [data-testid="stSidebar"] * { color: var(--text) !important; }

# # # # # .main .block-container { padding: 2rem 2rem 3rem !important; }

# # # # # h1, h2, h3, h4 {
# # # # #     font-family: 'Syne', sans-serif !important;
# # # # #     color: var(--text) !important;
# # # # # }

# # # # # .sil-card {
# # # # #     background: var(--surface);
# # # # #     border: 1px solid var(--border);
# # # # #     border-radius: 14px;
# # # # #     padding: 1.25rem 1.5rem;
# # # # #     margin-bottom: 1rem;
# # # # #     transition: all 0.25s ease;
# # # # # }
# # # # # .sil-card:hover {
# # # # #     transform: translateY(-2px);
# # # # #     box-shadow: 0 8px 22px rgba(0,0,0,0.05);
# # # # # }
# # # # # .sil-card-accent { border-left: 4px solid var(--accent); }

# # # # # .metric-row { display:flex; gap:1rem; flex-wrap:wrap; margin-bottom:1.5rem; }
# # # # # .metric-tile {
# # # # #     flex:1; min-width:140px;
# # # # #     background:var(--surface2);
# # # # #     border:1px solid var(--border);
# # # # #     border-radius:12px;
# # # # #     padding:1rem; text-align:center;
# # # # #     transition:all .25s ease;
# # # # # }
# # # # # .metric-tile:hover { transform:translateY(-3px); box-shadow:0 10px 24px rgba(0,0,0,0.06); }
# # # # # .metric-tile .val { font-family:'Syne',sans-serif; font-size:2rem; font-weight:800; color:var(--accent); }
# # # # # .metric-tile .lbl { font-size:.72rem; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; }

# # # # # .badge { padding:.25rem .7rem; border-radius:20px; font-size:.72rem; font-weight:600; font-family:'DM Mono',monospace; }
# # # # # .badge-green  { background:#ecfdf5; color:#15803d; }
# # # # # .badge-yellow { background:#fffbeb; color:#b45309; }
# # # # # .badge-red    { background:#fef2f2; color:#b91c1c; }
# # # # # .badge-blue   { background:#eff6ff; color:#1d4ed8; }
# # # # # .badge-purple { background:#f5f3ff; color:#6d28d9; }

# # # # # .contact-card {
# # # # #     background:var(--surface2); border:1px solid var(--border);
# # # # #     border-radius:10px; padding:1rem; margin-bottom:.8rem;
# # # # # }
# # # # # .contact-name  { font-family:'Syne',sans-serif; font-weight:700; color:var(--text); }
# # # # # .contact-title { color:var(--muted); font-size:.85rem; }
# # # # # .contact-email { font-family:'DM Mono',monospace; color:var(--accent); font-size:.8rem; }

# # # # # .email-box {
# # # # #     background: #f8fafc;
# # # # #     border: 1px solid var(--border);
# # # # #     border-radius: 10px;
# # # # #     padding: 1rem 1.25rem;
# # # # #     font-size: .9rem;
# # # # #     line-height: 1.65;
# # # # #     white-space: pre-wrap;
# # # # #     color: var(--text);
# # # # # }
# # # # # .email-subject { font-family:'Syne',sans-serif; font-weight:700; color:var(--accent); margin-bottom:.5rem; }

# # # # # .section-label {
# # # # #     font-family:'DM Mono',monospace; font-size:.72rem;
# # # # #     text-transform:uppercase; letter-spacing:.12em;
# # # # #     color:var(--accent); margin-bottom:.6rem;
# # # # # }
# # # # # .sil-divider { border-top:1px solid var(--border); margin:1rem 0; }

# # # # # [data-testid="stExpander"] {
# # # # #     background: var(--surface) !important;
# # # # #     border: 1px solid var(--border) !important;
# # # # #     border-radius: 10px !important;
# # # # # }
# # # # # [data-testid="stSelectbox"] > div,
# # # # # [data-testid="stMultiSelect"] > div {
# # # # #     background: var(--surface2) !important;
# # # # #     border-color: var(--border) !important;
# # # # # }
# # # # # [data-testid="stTabs"] button {
# # # # #     font-family:'Syne',sans-serif !important;
# # # # #     font-weight:600 !important;
# # # # # }
# # # # # ::-webkit-scrollbar { width:6px; }
# # # # # ::-webkit-scrollbar-thumb { background:var(--border); border-radius:3px; }

# # # # # .pipeline-log {
# # # # #     background: #0f172a;
# # # # #     color: #e2e8f0;
# # # # #     border-radius: 10px;
# # # # #     padding: 1rem 1.25rem;
# # # # #     font-family: 'DM Mono', monospace;
# # # # #     font-size: .8rem;
# # # # #     line-height: 1.6;
# # # # #     max-height: 380px;
# # # # #     overflow-y: auto;
# # # # #     white-space: pre-wrap;
# # # # # }

# # # # # .fit-box {
# # # # #     border-radius: 8px;
# # # # #     padding: .75rem;
# # # # #     margin-bottom: .5rem;
# # # # #     font-size: .85rem;
# # # # # }

# # # # # /* Hide sidebar on login page */
# # # # # .hide-sidebar [data-testid="stSidebar"] { display: none !important; }
# # # # # .hide-sidebar .main .block-container { max-width: 480px; margin: 0 auto; }

# # # # # </style>
# # # # # """, unsafe_allow_html=True)


# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  SESSION STATE INIT
# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # for _k, _v in [
# # # # #     ("logged_in",        False),
# # # # #     ("login_error",      ""),
# # # # #     ("pipeline_running", False),
# # # # #     ("pipeline_logs",    []),
# # # # #     ("pipeline_result",  None),
# # # # #     ("pipeline_done",    False),
# # # # # ]:
# # # # #     if _k not in st.session_state:
# # # # #         st.session_state[_k] = _v


# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  LOGIN PAGE
# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # if not st.session_state.logged_in:

# # # # #     # Hide sidebar on login page
# # # # #     st.markdown("""
# # # # #     <script>
# # # # #     document.querySelector('[data-testid="stSidebar"]') &&
# # # # #     (document.querySelector('[data-testid="stSidebar"]').style.display = 'none');
# # # # #     </script>
# # # # #     """, unsafe_allow_html=True)

# # # # #     # Center the login card
# # # # #     _, col, _ = st.columns([1, 1.2, 1])

# # # # #     with col:
# # # # #         st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)

# # # # #         st.markdown("""
# # # # #         <div class="login-card">
# # # # #           <div class="login-logo">🛡️ SecureITLab</div>
# # # # #           <div class="login-subtitle">Pipeline Intelligence</div>
# # # # #           <div class="login-divider"></div>
# # # # #         </div>
# # # # #         """, unsafe_allow_html=True)

# # # # #         st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # #         with st.container():
# # # # #             username = st.text_input(
# # # # #                 "Username",
# # # # #                 placeholder="Enter username",
# # # # #                 key="login_username",
# # # # #             )
# # # # #             password = st.text_input(
# # # # #                 "Password",
# # # # #                 placeholder="Enter password",
# # # # #                 type="password",
# # # # #                 key="login_password",
# # # # #             )

# # # # #             st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # #             login_btn = st.button(
# # # # #                 "Sign In →",
# # # # #                 use_container_width=True,
# # # # #                 type="primary",
# # # # #             )

# # # # #             if login_btn:
# # # # #                 if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
# # # # #                     st.session_state.logged_in = True
# # # # #                     st.session_state.login_error = ""
# # # # #                     st.rerun()
# # # # #                 else:
# # # # #                     st.session_state.login_error = "Incorrect username or password."

# # # # #             if st.session_state.login_error:
# # # # #                 st.markdown(
# # # # #                     f'<div class="login-error">⚠️ {st.session_state.login_error}</div>',
# # # # #                     unsafe_allow_html=True,
# # # # #                 )

# # # # #         st.markdown(
# # # # #             "<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>"
# # # # #             "SecureITLab · Confidential</div>",
# # # # #             unsafe_allow_html=True,
# # # # #         )

# # # # #     st.stop()  # ← Stop here — nothing below renders until logged in


# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  EVERYTHING BELOW ONLY RUNS WHEN LOGGED IN
# # # # # # ══════════════════════════════════════════════════════════════════════════════


# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  MASTER CONTACTS EXCEL BUILDER (NEW FEATURE)
# # # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # # def build_master_contacts_excel(all_jobs: list):
# # # # #     """Build master Excel with ALL contacts from ALL jobs with job details."""
# # # # #     try:
# # # # #         from openpyxl import Workbook
# # # # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # # # #         from openpyxl.utils import get_column_letter
# # # # #     except ImportError:
# # # # #         return None

# # # # #     wb = Workbook()
# # # # #     ws = wb.active
# # # # #     ws.title = "All Contacts"

# # # # #     NAVY  = "1E3A5F"
# # # # #     BLUE  = "2563EB"
# # # # #     LBLUE = "EFF6FF"
# # # # #     GREY  = "F8FAFC"
# # # # #     WHITE = "FFFFFF"

# # # # #     pri_colors = {
# # # # #         "Primary":   ("FEF2F2", "B91C1C"),
# # # # #         "Secondary": ("FFFBEB", "B45309"),
# # # # #         "Tertiary":  ("EFF6FF", "1D4ED8"),
# # # # #         "General":   ("F5F3FF", "6D28D9"),
# # # # #     }

# # # # #     thin   = Side(border_style="thin", color="D9E2EC")
# # # # #     border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # # # #     # Row 1: Title
# # # # #     ws.merge_cells("A1:K1")
# # # # #     c = ws["A1"]
# # # # #     c.value     = f"Master Contacts Export — All Jobs"
# # # # #     c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
# # # # #     c.fill      = PatternFill("solid", fgColor=NAVY)
# # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # #     ws.row_dimensions[1].height = 30

# # # # #     # Row 2: Sub-title
# # # # #     ws.merge_cells("A2:K2")
# # # # #     c = ws["A2"]
# # # # #     total_contacts = sum(len(j.get("contacts", [])) for j in all_jobs)
# # # # #     c.value     = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(all_jobs)} jobs   ·   {total_contacts} total contacts"
# # # # #     c.font      = Font(name="Arial", size=9, color="64748B")
# # # # #     c.fill      = PatternFill("solid", fgColor="F4F7FB")
# # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # #     ws.row_dimensions[2].height = 18

# # # # #     # Row 3: Headers
# # # # #     headers    = ["#", "Job", "Company", "Country", "Priority", "Name", "Title / Role", "Email", "LinkedIn URL", "Source", "Job Score"]
# # # # #     col_widths = [4,   28,   18,        12,       12,        24,     32,              34,      42,             18,       10]

# # # # #     for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
# # # # #         c = ws.cell(row=3, column=ci, value=h)
# # # # #         c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
# # # # #         c.fill      = PatternFill("solid", fgColor=BLUE)
# # # # #         c.alignment = Alignment(horizontal="center", vertical="center")
# # # # #         c.border    = border
# # # # #         ws.column_dimensions[get_column_letter(ci)].width = w

# # # # #     ws.row_dimensions[3].height = 22

# # # # #     # Data rows
# # # # #     ri = 4
# # # # #     for job in all_jobs:
# # # # #         company    = job.get("company", "")
# # # # #         role       = job.get("role", "")
# # # # #         country    = job.get("country", "?")
# # # # #         contacts   = job.get("contacts", [])
# # # # #         job_score  = job.get("opp_score", "—")
        
# # # # #         for ci, contact in enumerate(contacts):
# # # # #             prio           = contact.get("priority", "General")
# # # # #             bg_hex, fg_hex = pri_colors.get(prio, (WHITE, "0F172A"))
# # # # #             patterns       = contact.get("email_patterns", [])
# # # # #             email_val      = contact.get("email") or (patterns[0] + "  (pattern)" if patterns else "")
# # # # #             row_fill       = bg_hex if ri % 2 == 0 else GREY

# # # # #             for col_idx, val in enumerate([
# # # # #                 ri - 3,
# # # # #                 role if ci == 0 else "",
# # # # #                 company if ci == 0 else "",
# # # # #                 country if ci == 0 else "",
# # # # #                 prio,
# # # # #                 contact.get("name", ""),
# # # # #                 contact.get("title", ""),
# # # # #                 email_val,
# # # # #                 contact.get("linkedin_url", ""),
# # # # #                 contact.get("source", ""),
# # # # #                 str(job_score) if ci == 0 else "",
# # # # #             ], 1):
# # # # #                 cell = ws.cell(row=ri, column=col_idx, value=val)
# # # # #                 cell.font      = Font(name="Arial", size=9,
# # # # #                                       bold=(col_idx == 5),
# # # # #                                       color=fg_hex if col_idx == 5 else "0F172A")
# # # # #                 cell.fill      = PatternFill("solid", fgColor=row_fill)
# # # # #                 cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [4, 7]))
# # # # #                 cell.border    = border

# # # # #             ws.row_dimensions[ri].height = 18
# # # # #             ri += 1

# # # # #     ws.freeze_panes   = "A4"
# # # # #     ws.auto_filter.ref = f"A3:K{ri-1}"

# # # # #     # Summary sheet
# # # # #     ws2 = wb.create_sheet("Summary")
# # # # #     ws2.merge_cells("A1:C1")
# # # # #     t = ws2["A1"]
# # # # #     t.value     = "Master Export Summary"
# # # # #     t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
# # # # #     t.fill      = PatternFill("solid", fgColor=NAVY)
# # # # #     t.alignment = Alignment(horizontal="center")

# # # # #     summary_stats = [
# # # # #         ("Total Jobs",        len(all_jobs)),
# # # # #         ("Total Contacts",    total_contacts),
# # # # #         ("Primary Contacts",  sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "Primary")),
# # # # #         ("Secondary",         sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "Secondary")),
# # # # #         ("Tertiary",          sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "Tertiary")),
# # # # #         ("General",           sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "General")),
# # # # #         ("With Email",        sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("email"))),
# # # # #         ("With LinkedIn",     sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("linkedin_url"))),
# # # # #         ("Generated",         datetime.now().strftime("%d %b %Y  %H:%M")),
# # # # #     ]

# # # # #     for r, (lbl, val) in enumerate(summary_stats, 2):
# # # # #         lc = ws2.cell(row=r, column=1, value=lbl)
# # # # #         vc = ws2.cell(row=r, column=2, value=val)
# # # # #         bg = LBLUE if r % 2 == 0 else WHITE
# # # # #         for cell in (lc, vc):
# # # # #             cell.font   = Font(name="Arial", bold=(cell.column == 1), size=10)
# # # # #             cell.fill   = PatternFill("solid", fgColor=bg)
# # # # #             cell.border = border

# # # # #     ws2.column_dimensions["A"].width = 20
# # # # #     ws2.column_dimensions["B"].width = 30

# # # # #     buf = io.BytesIO()
# # # # #     wb.save(buf)
# # # # #     buf.seek(0)
# # # # #     return buf.getvalue()


# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  INDIVIDUAL JOB CONTACTS EXCEL (ORIGINAL FUNCTION)
# # # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # # def build_contacts_excel(contacts: list, company: str, role: str):
# # # # #     """Build a formatted Excel workbook from contacts. Returns bytes or None."""
# # # # #     try:
# # # # #         from openpyxl import Workbook
# # # # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # # # #         from openpyxl.utils import get_column_letter
# # # # #     except ImportError:
# # # # #         return None

# # # # #     wb = Workbook()
# # # # #     ws = wb.active
# # # # #     ws.title = "Contacts"

# # # # #     NAVY  = "1E3A5F"
# # # # #     BLUE  = "2563EB"
# # # # #     LBLUE = "EFF6FF"
# # # # #     GREY  = "F8FAFC"
# # # # #     WHITE = "FFFFFF"

# # # # #     pri_colors = {
# # # # #         "Primary":   ("FEF2F2", "B91C1C"),
# # # # #         "Secondary": ("FFFBEB", "B45309"),
# # # # #         "Tertiary":  ("EFF6FF", "1D4ED8"),
# # # # #         "General":   ("F5F3FF", "6D28D9"),
# # # # #     }

# # # # #     thin   = Side(border_style="thin", color="D9E2EC")
# # # # #     border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # # # #     # Row 1: Title
# # # # #     ws.merge_cells("A1:H1")
# # # # #     c = ws["A1"]
# # # # #     c.value     = f"Contacts Export  —  {company}  |  {role}"
# # # # #     c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
# # # # #     c.fill      = PatternFill("solid", fgColor=NAVY)
# # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # #     ws.row_dimensions[1].height = 30

# # # # #     # Row 2: Sub-title
# # # # #     ws.merge_cells("A2:H2")
# # # # #     c = ws["A2"]
# # # # #     c.value     = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(contacts)} contacts"
# # # # #     c.font      = Font(name="Arial", size=9, color="64748B")
# # # # #     c.fill      = PatternFill("solid", fgColor="F4F7FB")
# # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # #     ws.row_dimensions[2].height = 18

# # # # #     # Row 3: Headers
# # # # #     headers    = ["#", "Priority", "Name", "Title / Role", "Company", "Email", "LinkedIn URL", "Source"]
# # # # #     col_widths = [4,    12,         24,     32,              22,        34,       42,              18]

# # # # #     for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
# # # # #         c = ws.cell(row=3, column=ci, value=h)
# # # # #         c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
# # # # #         c.fill      = PatternFill("solid", fgColor=BLUE)
# # # # #         c.alignment = Alignment(horizontal="center", vertical="center")
# # # # #         c.border    = border
# # # # #         ws.column_dimensions[get_column_letter(ci)].width = w

# # # # #     ws.row_dimensions[3].height = 22

# # # # #     # Data rows
# # # # #     for ri, ct in enumerate(contacts, start=4):
# # # # #         prio           = ct.get("priority", "General")
# # # # #         bg_hex, fg_hex = pri_colors.get(prio, (WHITE, "0F172A"))
# # # # #         patterns       = ct.get("email_patterns", [])
# # # # #         email_val      = ct.get("email") or (patterns[0] + "  (pattern)" if patterns else "")
# # # # #         row_fill       = bg_hex if ri % 2 == 0 else GREY

# # # # #         for ci, val in enumerate([
# # # # #             ri - 3,
# # # # #             prio,
# # # # #             ct.get("name", ""),
# # # # #             ct.get("title", ""),
# # # # #             ct.get("company", ""),
# # # # #             email_val,
# # # # #             ct.get("linkedin_url", ""),
# # # # #             ct.get("source", ""),
# # # # #         ], 1):
# # # # #             cell = ws.cell(row=ri, column=ci, value=val)
# # # # #             cell.font      = Font(name="Arial", size=9,
# # # # #                                   bold=(ci == 2),
# # # # #                                   color=fg_hex if ci == 2 else "0F172A")
# # # # #             cell.fill      = PatternFill("solid", fgColor=row_fill)
# # # # #             cell.alignment = Alignment(vertical="center", wrap_text=(ci in [4, 7]))
# # # # #             cell.border    = border

# # # # #         ws.row_dimensions[ri].height = 18

# # # # #     ws.freeze_panes   = "A4"
# # # # #     ws.auto_filter.ref = f"A3:H{3 + len(contacts)}"

# # # # #     # Summary sheet
# # # # #     ws2 = wb.create_sheet("Summary")
# # # # #     ws2.merge_cells("A1:C1")
# # # # #     t = ws2["A1"]
# # # # #     t.value     = "Export Summary"
# # # # #     t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
# # # # #     t.fill      = PatternFill("solid", fgColor=NAVY)
# # # # #     t.alignment = Alignment(horizontal="center")

# # # # #     for r, (lbl, val) in enumerate([
# # # # #         ("Company",        company),
# # # # #         ("Role",           role),
# # # # #         ("Total Contacts", len(contacts)),
# # # # #         ("Primary",        sum(1 for x in contacts if x.get("priority") == "Primary")),
# # # # #         ("Secondary",      sum(1 for x in contacts if x.get("priority") == "Secondary")),
# # # # #         ("Tertiary",       sum(1 for x in contacts if x.get("priority") == "Tertiary")),
# # # # #         ("General",        sum(1 for x in contacts if x.get("priority") == "General")),
# # # # #         ("With Email",     sum(1 for x in contacts if x.get("email"))),
# # # # #         ("With LinkedIn",  sum(1 for x in contacts if x.get("linkedin_url"))),
# # # # #         ("Generated",      datetime.now().strftime("%d %b %Y  %H:%M")),
# # # # #     ], 2):
# # # # #         lc = ws2.cell(row=r, column=1, value=lbl)
# # # # #         vc = ws2.cell(row=r, column=2, value=val)
# # # # #         bg = LBLUE if r % 2 == 0 else WHITE
# # # # #         for cell in (lc, vc):
# # # # #             cell.font   = Font(name="Arial", bold=(cell.column == 1), size=10)
# # # # #             cell.fill   = PatternFill("solid", fgColor=bg)
# # # # #             cell.border = border

# # # # #     ws2.column_dimensions["A"].width = 20
# # # # #     ws2.column_dimensions["B"].width = 30

# # # # #     buf = io.BytesIO()
# # # # #     wb.save(buf)
# # # # #     buf.seek(0)
# # # # #     return buf.getvalue()


# # # # # # ── MongoDB helpers ───────────────────────────────────────────────────────────
# # # # # @st.cache_resource
# # # # # def get_db():
# # # # #     URI = st.secrets.get("MONGO_URI", "mongodb://localhost:27017")
# # # # #     DB  = st.secrets.get("MONGO_DB",  "secureitlab_job_pipeline")
# # # # #     client = MongoClient(URI, serverSelectionTimeoutMS=6000)
# # # # #     return client[DB]

# # # # # @st.cache_data(ttl=60)
# # # # # def load_all_jobs():
# # # # #     db = get_db()
# # # # #     return list(db.jobs.find({}, {
# # # # #         "_id": 1, "company": 1, "role": 1, "job_number": 1,
# # # # #         "opp_score": 1, "contacts_found": 1, "pipeline_ok": 1,
# # # # #         "coverage_score": 1, "run_at": 1, "contact_domain": 1,
# # # # #         "contacts": 1, "country": 1,
# # # # #     }))

# # # # # @st.cache_data(ttl=60)
# # # # # def load_job(job_id):
# # # # #     from bson import ObjectId
# # # # #     return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# # # # # # ── Render helpers ────────────────────────────────────────────────────────────
# # # # # def badge(text, color="blue"):
# # # # #     return f'<span class="badge badge-{color}">{text}</span>'

# # # # # def safe_str(val, limit=300):
# # # # #     if val is None: return "—"
# # # # #     s = str(val)
# # # # #     return s[:limit] + "…" if len(s) > limit else s

# # # # # def as_dict(raw):
# # # # #     if isinstance(raw, dict): return raw
# # # # #     if isinstance(raw, list): return next((x for x in raw if isinstance(x, dict)), {})
# # # # #     return {}

# # # # # def render_json_pretty(data, title=""):
# # # # #     if not data: return
# # # # #     with st.expander(f"📄 Raw JSON — {title}", expanded=False):
# # # # #         st.code(json.dumps(data, indent=2, default=str), language="json")

# # # # # def render_qa_block(data, label):
# # # # #     if not data:
# # # # #         st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>', unsafe_allow_html=True)
# # # # #         return
# # # # #     data = as_dict(data)
# # # # #     if not data: return
# # # # #     passed    = data.get("passed") or data.get("Passed") or False
# # # # #     rec       = data.get("recommendation") or data.get("Recommendation", "")
# # # # #     issues    = data.get("issues") or data.get("Issues") or []
# # # # #     checklist = data.get("checklist") or data.get("Checklist") or []
# # # # #     color     = "green" if passed else "yellow"
# # # # #     status    = "✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
# # # # #     html = f"""
# # # # #     <div class="sil-card sil-card-accent">
# # # # #       <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
# # # # #         <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
# # # # #         {badge(status, color)}
# # # # #       </div>"""
# # # # #     if rec:
# # # # #         html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
# # # # #     if checklist:
# # # # #         html += '<div style="font-size:.82rem;margin-bottom:.5rem">'
# # # # #         for item in (checklist if isinstance(checklist, list) else [checklist]):
# # # # #             if isinstance(item, dict):
# # # # #                 ip = item.get("pass") or item.get("passed") or item.get("status","") == "pass"
# # # # #                 nm = item.get("item") or item.get("name") or item.get("check","")
# # # # #                 nt = item.get("reason") or item.get("note") or item.get("issue","")
# # # # #                 html += f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
# # # # #                 if nt: html += f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
# # # # #                 html += '</div>'
# # # # #         html += '</div>'
# # # # #     if issues:
# # # # #         html += '<div style="margin-top:.5rem">'
# # # # #         for iss in (issues if isinstance(issues, list) else [issues])[:4]:
# # # # #             txt = iss if isinstance(iss, str) else json.dumps(iss)
# # # # #             html += f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
# # # # #         html += '</div>'
# # # # #     st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # def render_contacts(contacts, title="Contacts"):
# # # # #     if not contacts: st.info("No contacts found for this job."); return
# # # # #     pri_color = {"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
# # # # #     st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>', unsafe_allow_html=True)
# # # # #     cols = st.columns(2)
# # # # #     for i, c in enumerate(contacts):
# # # # #         col = cols[i % 2]
# # # # #         prio = c.get("priority", "General")
# # # # #         email = c.get("email", ""); li = c.get("linkedin_url", "")
# # # # #         patterns = c.get("email_patterns", []); src = c.get("source", "")
# # # # #         with col:
# # # # #             html = f"""<div class="contact-card">
# # # # #               <div style="display:flex;justify-content:space-between;align-items:flex-start">
# # # # #                 <div>
# # # # #                   <div class="contact-name">{c.get('name','—')}</div>
# # # # #                   <div class="contact-title">{c.get('title','—')}</div>
# # # # #                 </div>
# # # # #                 {badge(prio, pri_color.get(prio,'blue'))}
# # # # #               </div>"""
# # # # #             if email:      html += f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
# # # # #             elif patterns: html += f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
# # # # #             if li:         html += f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
# # # # #             if src:        html += f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
# # # # #             st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # def render_emails(emails_data):
# # # # #     if not emails_data: st.info("No email data available."); return
# # # # #     emails_data = as_dict(emails_data)
# # # # #     if not emails_data: return
# # # # #     variants = {}
# # # # #     for k, v in emails_data.items():
# # # # #         kl = k.lower().replace("_","").replace(" ","")
# # # # #         if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl == "a":
# # # # #             variants["Variant A — Hiring Manager"] = v
# # # # #         elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl == "b":
# # # # #             variants["Variant B — CISO / VP Level"] = v
# # # # #         else:
# # # # #             variants[k] = v
# # # # #     for label, v in variants.items():
# # # # #         st.markdown(f'<div class="section-label">✉️ {label}</div>', unsafe_allow_html=True)
# # # # #         if isinstance(v, dict):
# # # # #             subj = v.get("subject") or v.get("Subject","")
# # # # #             body = v.get("body") or v.get("Body") or v.get("content","")
# # # # #             if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>', unsafe_allow_html=True)
# # # # #             if body: st.markdown(f'<div class="email-box">{body}</div>', unsafe_allow_html=True)
# # # # #             else:    st.code(json.dumps(v, indent=2), language="json")
# # # # #         elif isinstance(v, str):
# # # # #             st.markdown(f'<div class="email-box">{v}</div>', unsafe_allow_html=True)
# # # # #         st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

# # # # # def render_service_mapping(data):
# # # # #     if not data: st.info("No service mapping data."); return
# # # # #     items = data if isinstance(data, list) else []
# # # # #     if not items and isinstance(data, dict):
# # # # #         for key in ("services","mappings","service_mapping","ServiceMapping","items"):
# # # # #             if isinstance(data.get(key), list): items = data[key]; break
# # # # #         if not items: items = [data]
# # # # #     fit_colors = {"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
# # # # #     for item in items:
# # # # #         if not isinstance(item, dict): continue
# # # # #         svc  = item.get("service") or item.get("Service") or item.get("name","")
# # # # #         fit  = (item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
# # # # #         why  = item.get("justification") or item.get("rationale") or item.get("why","")
# # # # #         reqs = item.get("requirements_addressed") or item.get("requirements") or ""
# # # # #         eng  = item.get("engagement_type") or item.get("engagement","")
# # # # #         color = fit_colors.get(fit, "blue")
# # # # #         html = f"""<div class="sil-card" style="margin-bottom:.75rem">
# # # # #           <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
# # # # #             <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
# # # # #             {badge(fit or "MAPPED", color) if fit else ""}
# # # # #           </div>"""
# # # # #         if why:  html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
# # # # #         if reqs:
# # # # #             rs = ", ".join(reqs) if isinstance(reqs, list) else str(reqs)
# # # # #             html += f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
# # # # #         if eng:  html += f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
# # # # #         st.markdown(html + '</div>', unsafe_allow_html=True)
# # # # #     render_json_pretty(data, "Service Mapping")

# # # # # def render_microplans(data):
# # # # #     if not data: st.info("No micro-plan data."); return
# # # # #     plans = data if isinstance(data, list) else []
# # # # #     if not plans and isinstance(data, dict):
# # # # #         for k in ("plans","micro_plans","microplans","top_3","improvements"):
# # # # #             if isinstance(data.get(k), list): plans = data[k]; break
# # # # #         if not plans: plans = [data]
# # # # #     for i, plan in enumerate(plans[:3], 1):
# # # # #         if not isinstance(plan, dict): continue
# # # # #         title = plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
# # # # #         weeks = plan.get("duration") or plan.get("timeline","")
# # # # #         obj   = plan.get("objective") or plan.get("goal","")
# # # # #         kpis  = plan.get("kpis") or plan.get("KPIs") or []
# # # # #         tasks = plan.get("tasks") or plan.get("workstreams") or []
# # # # #         with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}", expanded=(i==1)):
# # # # #             if obj and obj != title: st.markdown(f"**Objective:** {obj}")
# # # # #             if kpis:
# # # # #                 st.markdown("**KPIs:**")
# # # # #                 for kpi in (kpis if isinstance(kpis, list) else [kpis]): st.markdown(f"• {kpi}")
# # # # #             if tasks:
# # # # #                 st.markdown("**Tasks / Workstreams:**")
# # # # #                 for t in (tasks if isinstance(tasks, list) else [tasks]):
# # # # #                     if isinstance(t, dict):
# # # # #                         tn = t.get("task") or t.get("name","")
# # # # #                         te = t.get("effort") or t.get("duration","")
# # # # #                         st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
# # # # #                     else: st.markdown(f"• {t}")
# # # # #             if plan:
# # # # #                 st.code(json.dumps(plan, indent=2, default=str), language="json")

# # # # # def render_deal_assurance(data):
# # # # #     if not data: st.info("No deal assurance data."); return
# # # # #     if not isinstance(data, dict): render_json_pretty(data, "Deal Assurance Pack"); return
# # # # #     evp = (data.get("executive_value_proposition") or
# # # # #            data.get("value_proposition") or data.get("ExecutiveValueProposition",""))
# # # # #     if evp:
# # # # #         st.markdown('<div class="section-label">💼 Executive Value Proposition</div>', unsafe_allow_html=True)
# # # # #         st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7;color:var(--text)">{evp}</div>', unsafe_allow_html=True)
# # # # #     caps = data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
# # # # #     if caps:
# # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>', unsafe_allow_html=True)
# # # # #         c1, c2 = st.columns(2)
# # # # #         for i, cap in enumerate(caps if isinstance(caps, list) else [caps]):
# # # # #             (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
# # # # #     risk = data.get("risk_mitigation") or data.get("RiskMitigation","")
# # # # #     if risk:
# # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>', unsafe_allow_html=True)
# # # # #         if isinstance(risk, dict):
# # # # #             for k, v in risk.items(): st.markdown(f"**{k}:** {v}")
# # # # #         else: st.markdown(str(risk))
# # # # #     render_json_pretty(data, "Deal Assurance Pack")


# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  SIDEBAR
# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # with st.sidebar:
# # # # #     st.markdown("""
# # # # #     <div style="padding:.75rem 0 1.25rem">
# # # # #       <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;
# # # # #                   color:#2563eb">🛡️ SecureITLab</div>
# # # # #       <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;
# # # # #                   text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
# # # # #     </div>
# # # # #     """, unsafe_allow_html=True)

# # # # #     # ── Logout button ─────────────────────────────────────────────────────────
# # # # #     if st.button("🚪 Logout", use_container_width=True):
# # # # #         st.session_state.logged_in = False
# # # # #         st.session_state.login_error = ""
# # # # #         st.rerun()

# # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # #     # ── Find Jobs button ──────────────────────────────────────────────────────
# # # # #     st.markdown("**🔍 Find New Jobs**")
# # # # #     st.caption(
# # # # #         "Runs scraper → checks MongoDB for duplicates → "
# # # # #         "runs AI pipeline only on NEW jobs → stores in MongoDB"
# # # # #     )

# # # # #     find_jobs_clicked = st.button(
# # # # #         "🔍  Find Jobs",
# # # # #         disabled=st.session_state.pipeline_running,
# # # # #         use_container_width=True,
# # # # #         type="primary",
# # # # #     )

# # # # #     if st.session_state.pipeline_running:
# # # # #         st.markdown("""
# # # # #         <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
# # # # #                     padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;
# # # # #                     font-size:.8rem;color:#1d4ed8">
# # # # #           ⏳ Pipeline is running… check log below
# # # # #         </div>""", unsafe_allow_html=True)

# # # # #     if st.session_state.pipeline_done and st.session_state.pipeline_result:
# # # # #         res = st.session_state.pipeline_result
# # # # #         ok  = res.get("success", False)
# # # # #         bg  = "#ecfdf5" if ok else "#fef2f2"
# # # # #         bc  = "#bbf7d0" if ok else "#fecaca"
# # # # #         tc  = "#15803d" if ok else "#b91c1c"
# # # # #         ic  = "✅" if ok else "❌"
# # # # #         st.markdown(f"""
# # # # #         <div style="background:{bg};border:1px solid {bc};border-radius:8px;
# # # # #                     padding:.75rem;margin-top:.5rem;font-size:.82rem">
# # # # #           <div style="font-weight:700;color:{tc};margin-bottom:.4rem">{ic} Last Run</div>
# # # # #           <div>📦 Scraped: <b>{res.get('scraped',0)}</b></div>
# # # # #           <div>🆕 New jobs: <b>{res.get('new_jobs',0)}</b></div>
# # # # #           <div>⏭️ Already in DB (skipped): <b>{res.get('skipped_db',0)}</b></div>
# # # # #           <div>🤖 Processed by AI: <b>{res.get('processed',0)}</b></div>
# # # # #           {f'<div style="color:{tc};margin-top:.3rem">⚠️ {res.get("error","")}</div>' if res.get("error") else ""}
# # # # #         </div>""", unsafe_allow_html=True)

# # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # #     # ── Job list ──────────────────────────────────────────────────────────────
# # # # #     try:
# # # # #         all_jobs = load_all_jobs()
# # # # #     except Exception as e:
# # # # #         st.error(f"MongoDB connection failed: {e}")
# # # # #         st.stop()

# # # # #     if not all_jobs:
# # # # #         st.warning("No jobs in MongoDB yet. Click **Find Jobs** to scrape and process.")
# # # # #         st.stop()

# # # # #     st.markdown(
# # # # #         f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">'
# # # # #         f'{len(all_jobs)} jobs in database</div>',
# # # # #         unsafe_allow_html=True,
# # # # #     )

# # # # #     # ── MASTER CONTACTS EXPORT (NEW) ────────────────────────────────────────────
# # # # #     st.markdown("**📋 Master Contacts Export**")
# # # # #     st.caption("All contacts from ALL jobs in one Excel file with auto-updates")
    
# # # # #     master_excel = build_master_contacts_excel(all_jobs)
# # # # #     if master_excel:
# # # # #         st.download_button(
# # # # #             label="📥  Download All Contacts",
# # # # #             data=master_excel,
# # # # #             file_name=f"master_contacts_{datetime.now().strftime('%Y%m%d')}.xlsx",
# # # # #             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# # # # #             use_container_width=True,
# # # # #         )
# # # # #     else:
# # # # #         st.warning("Run `pip install openpyxl` to enable Excel export.")

# # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # #     search   = st.text_input("🔍 Filter by company / role", placeholder="e.g. Bounteous")
# # # # #     filtered = [j for j in all_jobs
# # # # #                 if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

# # # # #     def job_label(j):
# # # # #         score = j.get("opp_score")
# # # # #         s_str = f" [{score}/10]" if score else ""
# # # # #         ok    = "✅" if j.get("pipeline_ok") else "❌"
# # # # #         return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s_str}"

# # # # #     if not filtered:
# # # # #         st.warning("No matching jobs.")
# # # # #         st.stop()

# # # # #     sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
# # # # #     sel_idx     = [job_label(j) for j in filtered].index(sel_label)
# # # # #     selected_id = str(filtered[sel_idx]["_id"])

# # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # #     ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
# # # # #     total_c  = sum(j.get("contacts_found",0) for j in all_jobs)
# # # # #     st.markdown(f"""
# # # # #     <div style="font-size:.75rem;color:#64748b">
# # # # #       <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
# # # # #       <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
# # # # #     </div>""", unsafe_allow_html=True)

# # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # #     if st.button("🔄 Refresh Data", use_container_width=True):
# # # # #         st.cache_data.clear()
# # # # #         st.rerun()


# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  FIND JOBS — background thread
# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # if find_jobs_clicked and not st.session_state.pipeline_running:
# # # # #     with _thread_lock:
# # # # #         _thread_logs.clear()
# # # # #         _thread_result[0] = None
# # # # #         _thread_done[0]   = False

# # # # #     st.session_state.pipeline_running = True
# # # # #     st.session_state.pipeline_done    = False
# # # # #     st.session_state.pipeline_logs    = []
# # # # #     st.session_state.pipeline_result  = None
# # # # #     st.cache_data.clear()

# # # # #     def _run_pipeline_bg():
# # # # #         try:
# # # # #             import main as _main

# # # # #             def _cb(msg: str):
# # # # #                 with _thread_lock:
# # # # #                     _thread_logs.append(f"{datetime.now().strftime('%H:%M:%S')} | {msg}")

# # # # #             result = _main.run_pipeline(progress_callback=_cb)
# # # # #             with _thread_lock:
# # # # #                 _thread_result[0] = result
# # # # #         except Exception as e:
# # # # #             with _thread_lock:
# # # # #                 _thread_logs.append(f"❌ Unexpected error: {e}")
# # # # #                 _thread_result[0] = {
# # # # #                     "success": False, "error": str(e),
# # # # #                     "scraped": 0, "new_jobs": 0, "skipped_db": 0, "processed": 0,
# # # # #                 }
# # # # #         finally:
# # # # #             with _thread_lock:
# # # # #                 _thread_done[0] = True

# # # # #     threading.Thread(target=_run_pipeline_bg, daemon=True).start()
# # # # #     st.rerun()

# # # # # # ── Sync thread state → session_state ────────────────────────────────────────
# # # # # with _thread_lock:
# # # # #     if _thread_logs:
# # # # #         st.session_state.pipeline_logs = list(_thread_logs)
# # # # #     if _thread_done[0] and _thread_result[0] is not None:
# # # # #         st.session_state.pipeline_result  = _thread_result[0]
# # # # #         st.session_state.pipeline_running = False
# # # # #         st.session_state.pipeline_done    = True

# # # # # # ── Live log panel ────────────────────────────────────────────────────────────
# # # # # if st.session_state.pipeline_running or (
# # # # #         st.session_state.pipeline_done and st.session_state.pipeline_logs):
# # # # #     heading = "⏳ Pipeline running — live log…" if st.session_state.pipeline_running \
# # # # #               else "📋 Last pipeline run log"
# # # # #     with st.expander(heading, expanded=st.session_state.pipeline_running):
# # # # #         log_text = "\n".join(st.session_state.pipeline_logs[-100:]) \
# # # # #                    if st.session_state.pipeline_logs else "Starting…"
# # # # #         st.markdown(f'<div class="pipeline-log">{log_text}</div>', unsafe_allow_html=True)
# # # # #         if st.session_state.pipeline_running:
# # # # #             time.sleep(2)
# # # # #             st.rerun()


# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  MAIN CONTENT — selected job detail (ORIGINAL CODE)
# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # with st.spinner("Loading job from MongoDB…"):
# # # # #     job = load_job(selected_id)

# # # # # if not job:
# # # # #     st.error("Could not load job document.")
# # # # #     st.stop()

# # # # # company   = job.get("company",  "Unknown")
# # # # # role      = job.get("role",     "Unknown")
# # # # # opp_score = job.get("opp_score")
# # # # # p_ok      = job.get("pipeline_ok", False)
# # # # # p_min     = job.get("pipeline_min", "?")
# # # # # c_found   = job.get("contacts_found", 0)
# # # # # c_cov     = job.get("coverage_score")
# # # # # c_domain  = job.get("contact_domain","")
# # # # # run_at    = job.get("run_at","")

# # # # # # ── Header ────────────────────────────────────────────────────────────────────
# # # # # st.markdown(f"""
# # # # # <div style="margin-bottom:1.75rem">
# # # # #   <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;
# # # # #               letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
# # # # #     Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
# # # # #   </div>
# # # # #   <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;
# # # # #              color:#0f172a;margin:0;line-height:1.15">{role}</h1>
# # # # #   <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
# # # # #     @ <span style="color:#334155;font-weight:600">{company}</span>
# # # # #     {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
# # # # #   </div>
# # # # # </div>
# # # # # """, unsafe_allow_html=True)

# # # # # # ── Metric tiles ──────────────────────────────────────────────────────────────
# # # # # try:
# # # # #     sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
# # # # #     sc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # # except Exception:
# # # # #     sc = "#2563eb"

# # # # # st.markdown(f"""
# # # # # <div class="metric-row">
# # # # #   <div class="metric-tile">
# # # # #     <div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div>
# # # # #     <div class="lbl">Opportunity Score</div>
# # # # #   </div>
# # # # #   <div class="metric-tile">
# # # # #     <div class="val">{c_found}</div>
# # # # #     <div class="lbl">Contacts Found</div>
# # # # #   </div>
# # # # #   <div class="metric-tile">
# # # # #     <div class="val">{f"{c_cov}%" if c_cov else "—"}</div>
# # # # #     <div class="lbl">Contact Coverage</div>
# # # # #   </div>
# # # # #   <div class="metric-tile">
# # # # #     <div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">
# # # # #       {'✅ OK' if p_ok else '❌ Failed'}
# # # # #     </div>
# # # # #     <div class="lbl">Pipeline ({p_min} min)</div>
# # # # #   </div>
# # # # # </div>
# # # # # """, unsafe_allow_html=True)

# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # #  TABS
# # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # tabs = st.tabs([
# # # # #     "📋 Job & Enrichment",
# # # # #     "🎯 Service Mapping",
# # # # #     "🔍 Fit / Gap",
# # # # #     "🛠️ Capability & Plans",
# # # # #     "📦 Deal Assurance",
# # # # #     "✉️ Outreach Emails",
# # # # #     "👥 Contacts",
# # # # #     "✅ QA Gates",
# # # # #     "🗄️ Raw Data",
# # # # # ])

# # # # # # ── Tab 1: Job Research + Enrichment ─────────────────────────────────────────
# # # # # with tabs[0]:
# # # # #     col1, col2 = st.columns([1, 1])
# # # # #     with col1:
# # # # #         st.markdown('<div class="section-label">📄 Job Research</div>', unsafe_allow_html=True)
# # # # #         jr = as_dict(job.get("agent_job_research") or {})
# # # # #         if jr:
# # # # #             for label, keys in [
# # # # #                 ("Job Role",        ["job_role","Job Role","role","title"]),
# # # # #                 ("Company",         ["company_name","Company Name","company"]),
# # # # #                 ("Location",        ["location","Location"]),
# # # # #                 ("Organization URL",["organization_url","Organization URL","url"]),
# # # # #             ]:
# # # # #                 val = next((jr.get(k) for k in keys if jr.get(k)), None)
# # # # #                 if val: st.markdown(f"**{label}:** {val}")
# # # # #             desc = jr.get("job_description") or jr.get("Job Description","")
# # # # #             if desc:
# # # # #                 st.markdown("**Job Description:**")
# # # # #                 st.markdown(
# # # # #                     f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;'
# # # # #                     f'max-height:300px;overflow-y:auto;color:var(--text)">{desc[:2000]}</div>',
# # # # #                     unsafe_allow_html=True,
# # # # #                 )
# # # # #             render_json_pretty(jr, "Job Research")
# # # # #         else:
# # # # #             st.info("No job research data.")
# # # # #     with col2:
# # # # #         st.markdown('<div class="section-label">🏢 Company Enrichment</div>', unsafe_allow_html=True)
# # # # #         enr = as_dict(job.get("agent_enrichment") or {})
# # # # #         if enr:
# # # # #             for label, keys in [
# # # # #                 ("Industry",          ["industry","Industry"]),
# # # # #                 ("Company Size",      ["company_size","size","Company Size"]),
# # # # #                 ("Regulatory Env",    ["regulatory_environment","regulatory"]),
# # # # #                 ("Certifications",    ["certifications","Certifications"]),
# # # # #                 ("Tech Stack",        ["tech_stack","technology_stack"]),
# # # # #                 ("Security Maturity", ["security_maturity","maturity"]),
# # # # #             ]:
# # # # #                 val = next((enr.get(k) for k in keys if enr.get(k)), None)
# # # # #                 if val:
# # # # #                     if isinstance(val, list): val = ", ".join(str(v) for v in val)
# # # # #                     st.markdown(f"**{label}:** {safe_str(val, 200)}")
# # # # #             render_json_pretty(enr, "Enrichment")
# # # # #         else:
# # # # #             st.info("No enrichment data.")

# # # # # # ── Tab 2: Service Mapping ────────────────────────────────────────────────────
# # # # # with tabs[1]:
# # # # #     st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>', unsafe_allow_html=True)
# # # # #     render_service_mapping(job.get("agent_service_mapping"))

# # # # # # ── Tab 3: Fit / Gap Analysis ─────────────────────────────────────────────────
# # # # # with tabs[2]:
# # # # #     fg = as_dict(job.get("agent_fit_gap") or {})
# # # # #     if opp_score:
# # # # #         try:
# # # # #             sn = float(str(opp_score).split("/")[0])
# # # # #             bc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # #             bp = int(sn / 10 * 100)
# # # # #             st.markdown(f"""
# # # # #             <div style="margin-bottom:1.5rem">
# # # # #               <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
# # # # #                 <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">Opportunity Score</span>
# # # # #                 <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
# # # # #               </div>
# # # # #               <div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%">
# # # # #                 <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
# # # # #               </div>
# # # # #             </div>""", unsafe_allow_html=True)
# # # # #         except Exception:
# # # # #             pass
# # # # #     st.markdown('<div class="section-label">📊 Service Classifications</div>', unsafe_allow_html=True)
# # # # #     services = []
# # # # #     if isinstance(fg, dict):
# # # # #         for k in ("services","classifications","service_classifications","items","fit_gap"):
# # # # #             v = fg.get(k)
# # # # #             if isinstance(v, list): services = v; break
# # # # #         if not services and (fg.get("service") or fg.get("Service")):
# # # # #             services = [fg]
# # # # #     elif isinstance(fg, list):
# # # # #         services = fg
# # # # #     if services:
# # # # #         buckets = {"STRONG FIT": [], "PARTIAL FIT": [], "GAP": []}
# # # # #         other   = []
# # # # #         for s in services:
# # # # #             if not isinstance(s, dict): continue
# # # # #             fit = (s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
# # # # #             if "STRONG" in fit: buckets["STRONG FIT"].append(s)
# # # # #             elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
# # # # #             elif "GAP" in fit: buckets["GAP"].append(s)
# # # # #             else: other.append(s)
# # # # #         c1, c2, c3 = st.columns(3)
# # # # #         cm  = {"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
# # # # #         bgm = {"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
# # # # #         bdm = {"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
# # # # #         for col, (fl, items) in zip([c1, c2, c3], buckets.items()):
# # # # #             col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>', unsafe_allow_html=True)
# # # # #             for s in items:
# # # # #                 svc  = s.get("service") or s.get("Service") or s.get("name","")
# # # # #                 just = s.get("justification") or s.get("reason","")
# # # # #                 col.markdown(
# # # # #                     f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem">'
# # # # #                     f'<div style="font-weight:600;color:#0f172a">{svc}</div>'
# # # # #                     f'<div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>',
# # # # #                     unsafe_allow_html=True,
# # # # #                 )
# # # # #     elif fg:
# # # # #         st.json(fg)
# # # # #     else:
# # # # #         st.info("No fit/gap data.")
# # # # #     render_json_pretty(job.get("agent_fit_gap"), "Fit/Gap Analysis")

# # # # # # ── Tab 4: Capability + Micro-plans ──────────────────────────────────────────
# # # # # with tabs[3]:
# # # # #     col1, col2 = st.columns([1, 1])
# # # # #     with col1:
# # # # #         st.markdown('<div class="section-label">🔧 Capability Improvements</div>', unsafe_allow_html=True)
# # # # #         cap = job.get("agent_capability") or {}
# # # # #         items_cap = cap if isinstance(cap, list) else []
# # # # #         if not items_cap and isinstance(cap, dict):
# # # # #             for k in ("improvements","recommendations","capabilities","items"):
# # # # #                 v = cap.get(k)
# # # # #                 if isinstance(v, list): items_cap = v; break
# # # # #             if not items_cap: items_cap = [cap]
# # # # #         for item in items_cap:
# # # # #             if not isinstance(item, dict): continue
# # # # #             title  = item.get("title") or item.get("gap") or item.get("service","")
# # # # #             rec    = item.get("recommendation") or item.get("steps","")
# # # # #             effort = item.get("build_effort") or item.get("effort","")
# # # # #             demand = item.get("market_demand") or item.get("priority","")
# # # # #             st.markdown(f"""
# # # # #             <div class="sil-card" style="margin-bottom:.6rem">
# # # # #               <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem;color:var(--text)">{title}</div>
# # # # #               <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec, 250)}</div>
# # # # #               {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
# # # # #             </div>""", unsafe_allow_html=True)
# # # # #         if not items_cap:
# # # # #             render_json_pretty(cap, "Capability Improvement")
# # # # #     with col2:
# # # # #         st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>', unsafe_allow_html=True)
# # # # #         render_microplans(job.get("agent_microplans"))

# # # # # # ── Tab 5: Deal Assurance Pack ────────────────────────────────────────────────
# # # # # with tabs[4]:
# # # # #     render_deal_assurance(job.get("agent_deal_assurance"))

# # # # # # ── Tab 6: Outreach Emails ────────────────────────────────────────────────────
# # # # # with tabs[5]:
# # # # #     st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>', unsafe_allow_html=True)
# # # # #     emails_src = job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
# # # # #     oq = as_dict(job.get("agent_outreach_qa") or {})
# # # # #     improved = (oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
# # # # #     if improved:
# # # # #         st.info("⚡ Showing QA-improved versions where available")
# # # # #         render_emails(improved)
# # # # #         with st.expander("📬 Original (pre-QA) versions"):
# # # # #             render_emails(emails_src)
# # # # #     else:
# # # # #         render_emails(emails_src)

# # # # # # ── Tab 7: Contacts ───────────────────────────────────────────────────────────
# # # # # with tabs[6]:
# # # # #     contacts        = job.get("contacts") or []
# # # # #     contact_sources = job.get("contact_sources") or []
# # # # #     pri = [c for c in contacts if c.get("priority") == "Primary"]
# # # # #     sec = [c for c in contacts if c.get("priority") == "Secondary"]
# # # # #     ter = [c for c in contacts if c.get("priority") == "Tertiary"]
# # # # #     gen = [c for c in contacts if c.get("priority") == "General"]
# # # # #     st.markdown(f"""
# # # # #     <div class="metric-row" style="margin-bottom:1.5rem">
# # # # #       <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
# # # # #       <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
# # # # #       <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
# # # # #       <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
# # # # #     </div>""", unsafe_allow_html=True)
# # # # #     if contact_sources:
# # # # #         st.markdown('Sources: ' + " ".join(badge(s,"blue") for s in contact_sources), unsafe_allow_html=True)
# # # # #         st.markdown("")
# # # # #     missing = job.get("missing_roles") or []
# # # # #     if missing:
# # # # #         st.markdown('⚠️ Missing roles: ' + " ".join(badge(r,"red") for r in missing), unsafe_allow_html=True)
# # # # #         st.markdown("")
# # # # #     if contacts:
# # # # #         # ── Excel export button (Individual Job) ───────────────────────────────
# # # # #         excel_bytes = build_contacts_excel(contacts, company, role)
# # # # #         if excel_bytes:
# # # # #             safe_co = re.sub(r'[^a-z0-9]', '_', company.lower())[:20]
# # # # #             fname   = f"contacts_{safe_co}_{datetime.now().strftime('%Y%m%d')}.xlsx"
# # # # #             btn_col, _ = st.columns([1, 3])
# # # # #             with btn_col:
# # # # #                 st.download_button(
# # # # #                     label="📥  Download Contacts (.xlsx)",
# # # # #                     data=excel_bytes,
# # # # #                     file_name=fname,
# # # # #                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# # # # #                     use_container_width=True,
# # # # #                     type="primary",
# # # # #                 )
# # # # #         else:
# # # # #             st.warning("Run `pip install openpyxl` to enable Excel export.")

# # # # #         st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # #         pri_filter = st.multiselect(
# # # # #             "Filter by priority",
# # # # #             ["Primary","Secondary","Tertiary","General"],
# # # # #             default=["Primary","Secondary","Tertiary","General"],
# # # # #         )
# # # # #         shown = [c for c in contacts if c.get("priority","General") in pri_filter]
# # # # #         render_contacts(shown, f"Contacts ({len(shown)} shown)")
# # # # #         agent_contacts = job.get("agent_prospect_contacts") or {}
# # # # #         if agent_contacts:
# # # # #             with st.expander("🤖 CrewAI Agent's Contact Search"):
# # # # #                 if isinstance(agent_contacts, dict):
# # # # #                     ac_list = agent_contacts.get("contacts") or []
# # # # #                     if ac_list: render_contacts(ac_list, "Agent Contacts")
# # # # #                     else:       st.json(agent_contacts)
# # # # #                 else:           st.json(agent_contacts)
# # # # #     else:
# # # # #         st.info("No contacts found for this job.")

# # # # # # ── Tab 8: QA Gates ───────────────────────────────────────────────────────────
# # # # # with tabs[7]:
# # # # #     st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>', unsafe_allow_html=True)
# # # # #     col1, col2 = st.columns(2)
# # # # #     for i, (label, key) in enumerate([
# # # # #         ("Research QA",       "agent_research_qa"),
# # # # #         ("Service Mapping QA","agent_mapping_qa"),
# # # # #         ("Deal Assurance QA", "agent_assurance_qa"),
# # # # #         ("Outreach Email QA", "agent_outreach_qa"),
# # # # #     ]):
# # # # #         with (col1 if i % 2 == 0 else col2):
# # # # #             render_qa_block(job.get(key), label)
# # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # #     st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>', unsafe_allow_html=True)
# # # # #     enf = as_dict(job.get("agent_prospect_enforcer") or {})
# # # # #     if enf:
# # # # #         cov  = enf.get("coverage_score","?")
# # # # #         miss = enf.get("missing_roles",[])
# # # # #         note = enf.get("note","")
# # # # #         ec   = enf.get("contacts",[])
# # # # #         x1, x2, x3 = st.columns(3)
# # # # #         x1.metric("Coverage Score",    f"{cov}%")
# # # # #         x2.metric("Missing Roles",     len(miss))
# # # # #         x3.metric("Contacts Verified", len(ec))
# # # # #         if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
# # # # #         if note: st.caption(note)
# # # # #     else:
# # # # #         st.info("No enforcer data.")

# # # # # # ── Tab 9: Raw Data ───────────────────────────────────────────────────────────
# # # # # with tabs[8]:
# # # # #     st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>', unsafe_allow_html=True)
# # # # #     st.caption("All fields stored in the `jobs` collection for this document.")
# # # # #     rows = []
# # # # #     for k, v in job.items():
# # # # #         if k == "_id": continue
# # # # #         rows.append({"Field": k, "Type": type(v).__name__, "Len": len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
# # # # #     hc1, hc2, hc3 = st.columns([3,1,1])
# # # # #     hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
# # # # #     for r in rows:
# # # # #         rc1, rc2, rc3 = st.columns([3,1,1])
# # # # #         rc1.code(r["Field"], language=None)
# # # # #         rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>', unsafe_allow_html=True)
# # # # #         rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>',  unsafe_allow_html=True)
# # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # #     for label, key in [
# # # # #         ("Job Research","agent_job_research"), ("Enrichment","agent_enrichment"),
# # # # #         ("Service Mapping","agent_service_mapping"), ("Fit/Gap Analysis","agent_fit_gap"),
# # # # #         ("Capability","agent_capability"), ("Micro-Plans","agent_microplans"),
# # # # #         ("Deal Assurance","agent_deal_assurance"), ("Outreach Emails","agent_outreach_emails"),
# # # # #         ("Prospect Contacts","agent_prospect_contacts"), ("Prospect Enforcer","agent_prospect_enforcer"),
# # # # #         ("Research QA","agent_research_qa"), ("Mapping QA","agent_mapping_qa"),
# # # # #         ("Assurance QA","agent_assurance_qa"), ("Outreach QA","agent_outreach_qa"),
# # # # #         ("Contacts (5-source)","contacts"),
# # # # #     ]:
# # # # #         data = job.get(key)
# # # # #         if data:
# # # # #             with st.expander(f"📄 {label}"):
# # # # #                 st.code(json.dumps(data, indent=2, default=str), language="json")



















# # # # # # """
# # # # # # ╔══════════════════════════════════════════════════════════╗
# # # # # # ║   SecureITLab Pipeline Dashboard — Streamlit             ║
# # # # # # ║   WITH MASTER CONTACTS AGGREGATION & AUTO-UPDATE         ║
# # # # # # ║   Reads from MongoDB → secureitlab_job_pipeline          ║
# # # # # # ╠══════════════════════════════════════════════════════════╣
# # # # # # ║  Install: pip install streamlit pymongo python-dotenv    ║
# # # # # # ║  Run:     streamlit run streamlit_dashboard.py           ║
# # # # # # ╚══════════════════════════════════════════════════════════╝
# # # # # # """

# # # # # # import io
# # # # # # import re
# # # # # # import streamlit as st
# # # # # # from pymongo import MongoClient
# # # # # # import json
# # # # # # import threading
# # # # # # import time
# # # # # # from datetime import datetime, timezone

# # # # # # # ── Thread-safe shared state ──────────────────────────────────────────────────
# # # # # # _thread_logs   = []
# # # # # # _thread_result = [None]
# # # # # # _thread_done   = [False]
# # # # # # _thread_lock   = threading.Lock()

# # # # # # # ── Page config ───────────────────────────────────────────────────────────────
# # # # # # st.set_page_config(
# # # # # #     page_title="SecureITLab Pipeline",
# # # # # #     page_icon="🛡️",
# # # # # #     layout="wide",
# # # # # #     initial_sidebar_state="expanded",
# # # # # # )

# # # # # # # ── LOGIN CREDENTIALS (change these) ─────────────────────────────────────────
# # # # # # LOGIN_USERNAME = "admin"
# # # # # # LOGIN_PASSWORD = "secureitlab2024"

# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  GLOBAL CSS (login + dashboard share this)
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # st.markdown("""
# # # # # # <style>

# # # # # # @import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');

# # # # # # :root {
# # # # # #     --bg:        #f4f7fb;
# # # # # #     --surface:   #ffffff;
# # # # # #     --surface2:  #eef2f7;
# # # # # #     --border:    #d9e2ec;
# # # # # #     --accent:    #2563eb;
# # # # # #     --accent2:   #7c3aed;
# # # # # #     --green:     #16a34a;
# # # # # #     --yellow:    #f59e0b;
# # # # # #     --red:       #dc2626;
# # # # # #     --text:      #0f172a;
# # # # # #     --muted:     #64748b;
# # # # # # }

# # # # # # html, body, [class*="css"] {
# # # # # #     background-color: var(--bg) !important;
# # # # # #     color: var(--text) !important;
# # # # # #     font-family: 'DM Sans', sans-serif !important;
# # # # # # }

# # # # # # /* ── LOGIN PAGE ─────────────────────────────────────────── */
# # # # # # .login-wrap {
# # # # # #     display: flex;
# # # # # #     align-items: center;
# # # # # #     justify-content: center;
# # # # # #     min-height: 80vh;
# # # # # # }
# # # # # # .login-card {
# # # # # #     background: var(--surface);
# # # # # #     border: 1px solid var(--border);
# # # # # #     border-radius: 20px;
# # # # # #     padding: 3rem 3.5rem;
# # # # # #     width: 100%;
# # # # # #     max-width: 420px;
# # # # # #     box-shadow: 0 20px 60px rgba(37,99,235,0.08);
# # # # # #     text-align: center;
# # # # # # }
# # # # # # .login-logo {
# # # # # #     font-family: 'Syne', sans-serif;
# # # # # #     font-size: 1.6rem;
# # # # # #     font-weight: 800;
# # # # # #     color: var(--accent);
# # # # # #     margin-bottom: .25rem;
# # # # # # }
# # # # # # .login-subtitle {
# # # # # #     font-size: .75rem;
# # # # # #     color: var(--muted);
# # # # # #     letter-spacing: .12em;
# # # # # #     text-transform: uppercase;
# # # # # #     margin-bottom: 2.5rem;
# # # # # # }
# # # # # # .login-error {
# # # # # #     background: #fef2f2;
# # # # # #     border: 1px solid #fecaca;
# # # # # #     border-radius: 8px;
# # # # # #     padding: .65rem 1rem;
# # # # # #     color: #b91c1c;
# # # # # #     font-size: .85rem;
# # # # # #     margin-top: 1rem;
# # # # # # }
# # # # # # .login-divider {
# # # # # #     width: 40px;
# # # # # #     height: 3px;
# # # # # #     background: linear-gradient(90deg, var(--accent), var(--accent2));
# # # # # #     border-radius: 2px;
# # # # # #     margin: 0 auto 2rem;
# # # # # # }

# # # # # # /* ── DASHBOARD ──────────────────────────────────────────── */
# # # # # # [data-testid="stSidebar"] {
# # # # # #     background: var(--surface) !important;
# # # # # #     border-right: 1px solid var(--border) !important;
# # # # # # }
# # # # # # [data-testid="stSidebar"] * { color: var(--text) !important; }

# # # # # # .main .block-container { padding: 2rem 2rem 3rem !important; }

# # # # # # h1, h2, h3, h4 {
# # # # # #     font-family: 'Syne', sans-serif !important;
# # # # # #     color: var(--text) !important;
# # # # # # }

# # # # # # .sil-card {
# # # # # #     background: var(--surface);
# # # # # #     border: 1px solid var(--border);
# # # # # #     border-radius: 14px;
# # # # # #     padding: 1.25rem 1.5rem;
# # # # # #     margin-bottom: 1rem;
# # # # # #     transition: all 0.25s ease;
# # # # # # }
# # # # # # .sil-card:hover {
# # # # # #     transform: translateY(-2px);
# # # # # #     box-shadow: 0 8px 22px rgba(0,0,0,0.05);
# # # # # # }
# # # # # # .sil-card-accent { border-left: 4px solid var(--accent); }

# # # # # # .metric-row { display:flex; gap:1rem; flex-wrap:wrap; margin-bottom:1.5rem; }
# # # # # # .metric-tile {
# # # # # #     flex:1; min-width:140px;
# # # # # #     background:var(--surface2);
# # # # # #     border:1px solid var(--border);
# # # # # #     border-radius:12px;
# # # # # #     padding:1rem; text-align:center;
# # # # # #     transition:all .25s ease;
# # # # # # }
# # # # # # .metric-tile:hover { transform:translateY(-3px); box-shadow:0 10px 24px rgba(0,0,0,0.06); }
# # # # # # .metric-tile .val { font-family:'Syne',sans-serif; font-size:2rem; font-weight:800; color:var(--accent); }
# # # # # # .metric-tile .lbl { font-size:.72rem; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; }

# # # # # # .badge { padding:.25rem .7rem; border-radius:20px; font-size:.72rem; font-weight:600; font-family:'DM Mono',monospace; }
# # # # # # .badge-green  { background:#ecfdf5; color:#15803d; }
# # # # # # .badge-yellow { background:#fffbeb; color:#b45309; }
# # # # # # .badge-red    { background:#fef2f2; color:#b91c1c; }
# # # # # # .badge-blue   { background:#eff6ff; color:#1d4ed8; }
# # # # # # .badge-purple { background:#f5f3ff; color:#6d28d9; }

# # # # # # .contact-card {
# # # # # #     background:var(--surface2); border:1px solid var(--border);
# # # # # #     border-radius:10px; padding:1rem; margin-bottom:.8rem;
# # # # # # }
# # # # # # .contact-name  { font-family:'Syne',sans-serif; font-weight:700; color:var(--text); }
# # # # # # .contact-title { color:var(--muted); font-size:.85rem; }
# # # # # # .contact-email { font-family:'DM Mono',monospace; color:var(--accent); font-size:.8rem; }

# # # # # # .email-box {
# # # # # #     background: #f8fafc;
# # # # # #     border: 1px solid var(--border);
# # # # # #     border-radius: 10px;
# # # # # #     padding: 1rem 1.25rem;
# # # # # #     font-size: .9rem;
# # # # # #     line-height: 1.65;
# # # # # #     white-space: pre-wrap;
# # # # # #     color: var(--text);
# # # # # # }
# # # # # # .email-subject { font-family:'Syne',sans-serif; font-weight:700; color:var(--accent); margin-bottom:.5rem; }

# # # # # # .section-label {
# # # # # #     font-family:'DM Mono',monospace; font-size:.72rem;
# # # # # #     text-transform:uppercase; letter-spacing:.12em;
# # # # # #     color:var(--accent); margin-bottom:.6rem;
# # # # # # }
# # # # # # .sil-divider { border-top:1px solid var(--border); margin:1rem 0; }

# # # # # # [data-testid="stExpander"] {
# # # # # #     background: var(--surface) !important;
# # # # # #     border: 1px solid var(--border) !important;
# # # # # #     border-radius: 10px !important;
# # # # # # }
# # # # # # [data-testid="stSelectbox"] > div,
# # # # # # [data-testid="stMultiSelect"] > div {
# # # # # #     background: var(--surface2) !important;
# # # # # #     border-color: var(--border) !important;
# # # # # # }
# # # # # # [data-testid="stTabs"] button {
# # # # # #     font-family:'Syne',sans-serif !important;
# # # # # #     font-weight:600 !important;
# # # # # # }
# # # # # # ::-webkit-scrollbar { width:6px; }
# # # # # # ::-webkit-scrollbar-thumb { background:var(--border); border-radius:3px; }

# # # # # # .pipeline-log {
# # # # # #     background: #0f172a;
# # # # # #     color: #e2e8f0;
# # # # # #     border-radius: 10px;
# # # # # #     padding: 1rem 1.25rem;
# # # # # #     font-family: 'DM Mono', monospace;
# # # # # #     font-size: .8rem;
# # # # # #     line-height: 1.6;
# # # # # #     max-height: 380px;
# # # # # #     overflow-y: auto;
# # # # # #     white-space: pre-wrap;
# # # # # # }

# # # # # # .fit-box {
# # # # # #     border-radius: 8px;
# # # # # #     padding: .75rem;
# # # # # #     margin-bottom: .5rem;
# # # # # #     font-size: .85rem;
# # # # # # }

# # # # # # /* Hide sidebar on login page */
# # # # # # .hide-sidebar [data-testid="stSidebar"] { display: none !important; }
# # # # # # .hide-sidebar .main .block-container { max-width: 480px; margin: 0 auto; }

# # # # # # </style>
# # # # # # """, unsafe_allow_html=True)


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  SESSION STATE INIT
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # for _k, _v in [
# # # # # #     ("logged_in",        False),
# # # # # #     ("login_error",      ""),
# # # # # #     ("pipeline_running", False),
# # # # # #     ("pipeline_logs",    []),
# # # # # #     ("pipeline_result",  None),
# # # # # #     ("pipeline_done",    False),
# # # # # # ]:
# # # # # #     if _k not in st.session_state:
# # # # # #         st.session_state[_k] = _v


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  LOGIN PAGE
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # if not st.session_state.logged_in:

# # # # # #     # Hide sidebar on login page
# # # # # #     st.markdown("""
# # # # # #     <script>
# # # # # #     document.querySelector('[data-testid="stSidebar"]') &&
# # # # # #     (document.querySelector('[data-testid="stSidebar"]').style.display = 'none');
# # # # # #     </script>
# # # # # #     """, unsafe_allow_html=True)

# # # # # #     # Center the login card
# # # # # #     _, col, _ = st.columns([1, 1.2, 1])

# # # # # #     with col:
# # # # # #         st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)

# # # # # #         st.markdown("""
# # # # # #         <div class="login-card">
# # # # # #           <div class="login-logo">🛡️ SecureITLab</div>
# # # # # #           <div class="login-subtitle">Pipeline Intelligence</div>
# # # # # #           <div class="login-divider"></div>
# # # # # #         </div>
# # # # # #         """, unsafe_allow_html=True)

# # # # # #         st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # # #         with st.container():
# # # # # #             username = st.text_input(
# # # # # #                 "Username",
# # # # # #                 placeholder="Enter username",
# # # # # #                 key="login_username",
# # # # # #             )
# # # # # #             password = st.text_input(
# # # # # #                 "Password",
# # # # # #                 placeholder="Enter password",
# # # # # #                 type="password",
# # # # # #                 key="login_password",
# # # # # #             )

# # # # # #             st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # # # #             login_btn = st.button(
# # # # # #                 "Sign In →",
# # # # # #                 use_container_width=True,
# # # # # #                 type="primary",
# # # # # #             )

# # # # # #             if login_btn:
# # # # # #                 if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
# # # # # #                     st.session_state.logged_in = True
# # # # # #                     st.session_state.login_error = ""
# # # # # #                     st.rerun()
# # # # # #                 else:
# # # # # #                     st.session_state.login_error = "Incorrect username or password."

# # # # # #             if st.session_state.login_error:
# # # # # #                 st.markdown(
# # # # # #                     f'<div class="login-error">⚠️ {st.session_state.login_error}</div>',
# # # # # #                     unsafe_allow_html=True,
# # # # # #                 )

# # # # # #         st.markdown(
# # # # # #             "<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>"
# # # # # #             "SecureITLab · Confidential</div>",
# # # # # #             unsafe_allow_html=True,
# # # # # #         )

# # # # # #     st.stop()  # ← Stop here — nothing below renders until logged in


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  EVERYTHING BELOW ONLY RUNS WHEN LOGGED IN
# # # # # # # ══════════════════════════════════════════════════════════════════════════════


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  MASTER CONTACTS EXCEL BUILDER (NEW FEATURE)
# # # # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # # # def build_master_contacts_excel(all_jobs: list):
# # # # # #     """Build master Excel with ALL contacts from ALL jobs with job details."""
# # # # # #     try:
# # # # # #         from openpyxl import Workbook
# # # # # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # # # # #         from openpyxl.utils import get_column_letter
# # # # # #     except ImportError:
# # # # # #         return None

# # # # # #     wb = Workbook()
# # # # # #     ws = wb.active
# # # # # #     ws.title = "All Contacts"

# # # # # #     NAVY  = "1E3A5F"
# # # # # #     BLUE  = "2563EB"
# # # # # #     LBLUE = "EFF6FF"
# # # # # #     GREY  = "F8FAFC"
# # # # # #     WHITE = "FFFFFF"

# # # # # #     pri_colors = {
# # # # # #         "Primary":   ("FEF2F2", "B91C1C"),
# # # # # #         "Secondary": ("FFFBEB", "B45309"),
# # # # # #         "Tertiary":  ("EFF6FF", "1D4ED8"),
# # # # # #         "General":   ("F5F3FF", "6D28D9"),
# # # # # #     }

# # # # # #     thin   = Side(border_style="thin", color="D9E2EC")
# # # # # #     border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # # # # #     # Row 1: Title
# # # # # #     ws.merge_cells("A1:M1")
# # # # # #     c = ws["A1"]
# # # # # #     c.value     = f"Master Contacts Export — All Jobs (with Outreach Templates)"
# # # # # #     c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
# # # # # #     c.fill      = PatternFill("solid", fgColor=NAVY)
# # # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # #     ws.row_dimensions[1].height = 30

# # # # # #     # Row 2: Sub-title
# # # # # #     ws.merge_cells("A2:M2")
# # # # # #     c = ws["A2"]
# # # # # #     total_contacts = sum(len(j.get("contacts", [])) for j in all_jobs)
# # # # # #     c.value     = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(all_jobs)} jobs   ·   {total_contacts} total contacts   ·   Includes Email & LinkedIn Messages"
# # # # # #     c.font      = Font(name="Arial", size=9, color="64748B")
# # # # # #     c.fill      = PatternFill("solid", fgColor="F4F7FB")
# # # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # #     ws.row_dimensions[2].height = 18

# # # # # #     # Row 3: Headers
# # # # # #     headers    = ["#", "Job", "Company", "Country", "Priority", "Name", "Title / Role", "Email", "LinkedIn URL", 
# # # # # #                   "Outreach Email (Body)", "LinkedIn Message", "Source", "Job Score"]
# # # # # #     col_widths = [4,   28,   18,        12,       12,        24,     32,              34,      42,
# # # # # #                   120,                    80,                  18,       10]
    
# # # # # #     # Key columns get darker blue highlighting
# # # # # #     HIGHLIGHT_BLUE = "1E5FA8"

# # # # # #     for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
# # # # # #         c = ws.cell(row=3, column=ci, value=h)
# # # # # #         c.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
# # # # # #         # Highlight contact/outreach columns with darker blue
# # # # # #         c.fill      = PatternFill("solid", fgColor=HIGHLIGHT_BLUE if ci in [8, 9, 10, 11] else BLUE)
# # # # # #         c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
# # # # # #         c.border    = border
# # # # # #         ws.column_dimensions[get_column_letter(ci)].width = w

# # # # # #     ws.row_dimensions[3].height = 28

# # # # # #     # Data rows
# # # # # #     ri = 4
# # # # # #     for job in all_jobs:
# # # # # #         company    = job.get("company", "")
# # # # # #         role       = job.get("role", "")
# # # # # #         country    = job.get("country", "?")
# # # # # #         contacts   = job.get("contacts", [])
# # # # # #         job_score  = job.get("opp_score", "—")
        
# # # # # #         # Get outreach emails from agent_outreach_emails in jobs collection
# # # # # #         emails_data = job.get("agent_outreach_emails", {})
        
# # # # # #         # Extract email body from VariantA (Variant B as fallback)
# # # # # #         email_body = ""
# # # # # #         if isinstance(emails_data, dict):
# # # # # #             # Try VariantA first
# # # # # #             if "VariantA" in emails_data and isinstance(emails_data["VariantA"], dict):
# # # # # #                 email_body = emails_data["VariantA"].get("body", "")
# # # # # #             # Fallback to VariantB
# # # # # #             elif "VariantB" in emails_data and isinstance(emails_data["VariantB"], dict):
# # # # # #                 email_body = emails_data["VariantB"].get("body", "")
        
# # # # # #         # Show full email body in Excel (will be wrapped in cell)
# # # # # #         email_body_display = str(email_body) if email_body else ""
        
# # # # # #         for ci, contact in enumerate(contacts):
# # # # # #             prio           = contact.get("priority", "General")
# # # # # #             bg_hex, fg_hex = pri_colors.get(prio, (WHITE, "0F172A"))
# # # # # #             patterns       = contact.get("email_patterns", [])
# # # # # #             email_val      = contact.get("email") or (patterns[0] + "  (pattern)" if patterns else "")
# # # # # #             row_fill       = bg_hex if ri % 2 == 0 else GREY

# # # # # #             for col_idx, val in enumerate([
# # # # # #                 ri - 3,
# # # # # #                 role if ci == 0 else "",
# # # # # #                 company if ci == 0 else "",
# # # # # #                 country if ci == 0 else "",
# # # # # #                 prio,
# # # # # #                 contact.get("name", ""),
# # # # # #                 contact.get("title", ""),
# # # # # #                 email_val,
# # # # # #                 contact.get("linkedin_url", ""),
# # # # # #                 email_body_display,  # Show email for EVERY contact
# # # # # #                 "",  # LinkedIn message column (placeholder for future)
# # # # # #                 contact.get("source", ""),
# # # # # #                 str(job_score) if ci == 0 else "",
# # # # # #             ], 1):
# # # # # #                 cell = ws.cell(row=ri, column=col_idx, value=val)
# # # # # #                 cell.font      = Font(name="Arial", size=8,
# # # # # #                                       bold=(col_idx == 6),
# # # # # #                                       color=fg_hex if col_idx == 6 else "0F172A")
# # # # # #                 cell.fill      = PatternFill("solid", fgColor=row_fill)
# # # # # #                 cell.alignment = Alignment(vertical="top", wrap_text=True)
# # # # # #                 cell.border    = border

# # # # # #             # Adjust row height for text-heavy cells
# # # # # #             if col_idx == 10 and val:
# # # # # #                 ws.row_dimensions[ri].height = 35
# # # # # #             else:
# # # # # #                 ws.row_dimensions[ri].height = 18
# # # # # #             ri += 1

# # # # # #     ws.freeze_panes   = "A4"
# # # # # #     ws.auto_filter.ref = f"A3:M{ri-1}"

# # # # # #     # Summary sheet
# # # # # #     ws2 = wb.create_sheet("Summary")
# # # # # #     ws2.merge_cells("A1:C1")
# # # # # #     t = ws2["A1"]
# # # # # #     t.value     = "Master Export Summary"
# # # # # #     t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
# # # # # #     t.fill      = PatternFill("solid", fgColor=NAVY)
# # # # # #     t.alignment = Alignment(horizontal="center")

# # # # # #     summary_stats = [
# # # # # #         ("Total Jobs",        len(all_jobs)),
# # # # # #         ("Total Contacts",    total_contacts),
# # # # # #         ("Primary Contacts",  sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "Primary")),
# # # # # #         ("Secondary",         sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "Secondary")),
# # # # # #         ("Tertiary",          sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "Tertiary")),
# # # # # #         ("General",           sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "General")),
# # # # # #         ("With Email",        sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("email"))),
# # # # # #         ("With LinkedIn",     sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("linkedin_url"))),
# # # # # #         ("Generated",         datetime.now().strftime("%d %b %Y  %H:%M")),
# # # # # #     ]

# # # # # #     for r, (lbl, val) in enumerate(summary_stats, 2):
# # # # # #         lc = ws2.cell(row=r, column=1, value=lbl)
# # # # # #         vc = ws2.cell(row=r, column=2, value=val)
# # # # # #         bg = LBLUE if r % 2 == 0 else WHITE
# # # # # #         for cell in (lc, vc):
# # # # # #             cell.font   = Font(name="Arial", bold=(cell.column == 1), size=10)
# # # # # #             cell.fill   = PatternFill("solid", fgColor=bg)
# # # # # #             cell.border = border

# # # # # #     ws2.column_dimensions["A"].width = 20
# # # # # #     ws2.column_dimensions["B"].width = 30

# # # # # #     buf = io.BytesIO()
# # # # # #     wb.save(buf)
# # # # # #     buf.seek(0)
# # # # # #     return buf.getvalue()


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  INDIVIDUAL JOB CONTACTS EXCEL (ORIGINAL FUNCTION)
# # # # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # # # def build_contacts_excel(contacts: list, company: str, role: str):
# # # # # #     """Build a formatted Excel workbook from contacts. Returns bytes or None."""
# # # # # #     try:
# # # # # #         from openpyxl import Workbook
# # # # # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # # # # #         from openpyxl.utils import get_column_letter
# # # # # #     except ImportError:
# # # # # #         return None

# # # # # #     wb = Workbook()
# # # # # #     ws = wb.active
# # # # # #     ws.title = "Contacts"

# # # # # #     NAVY  = "1E3A5F"
# # # # # #     BLUE  = "2563EB"
# # # # # #     LBLUE = "EFF6FF"
# # # # # #     GREY  = "F8FAFC"
# # # # # #     WHITE = "FFFFFF"

# # # # # #     pri_colors = {
# # # # # #         "Primary":   ("FEF2F2", "B91C1C"),
# # # # # #         "Secondary": ("FFFBEB", "B45309"),
# # # # # #         "Tertiary":  ("EFF6FF", "1D4ED8"),
# # # # # #         "General":   ("F5F3FF", "6D28D9"),
# # # # # #     }

# # # # # #     thin   = Side(border_style="thin", color="D9E2EC")
# # # # # #     border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # # # # #     # Row 1: Title
# # # # # #     ws.merge_cells("A1:H1")
# # # # # #     c = ws["A1"]
# # # # # #     c.value     = f"Contacts Export  —  {company}  |  {role}"
# # # # # #     c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
# # # # # #     c.fill      = PatternFill("solid", fgColor=NAVY)
# # # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # #     ws.row_dimensions[1].height = 30

# # # # # #     # Row 2: Sub-title
# # # # # #     ws.merge_cells("A2:H2")
# # # # # #     c = ws["A2"]
# # # # # #     c.value     = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(contacts)} contacts"
# # # # # #     c.font      = Font(name="Arial", size=9, color="64748B")
# # # # # #     c.fill      = PatternFill("solid", fgColor="F4F7FB")
# # # # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # #     ws.row_dimensions[2].height = 18

# # # # # #     # Row 3: Headers
# # # # # #     headers    = ["#", "Priority", "Name", "Title / Role", "Company", "Email", "LinkedIn URL", "Source"]
# # # # # #     col_widths = [4,    12,         24,     32,              22,        34,       42,              18]

# # # # # #     for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
# # # # # #         c = ws.cell(row=3, column=ci, value=h)
# # # # # #         c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
# # # # # #         c.fill      = PatternFill("solid", fgColor=BLUE)
# # # # # #         c.alignment = Alignment(horizontal="center", vertical="center")
# # # # # #         c.border    = border
# # # # # #         ws.column_dimensions[get_column_letter(ci)].width = w

# # # # # #     ws.row_dimensions[3].height = 22

# # # # # #     # Data rows
# # # # # #     for ri, ct in enumerate(contacts, start=4):
# # # # # #         prio           = ct.get("priority", "General")
# # # # # #         bg_hex, fg_hex = pri_colors.get(prio, (WHITE, "0F172A"))
# # # # # #         patterns       = ct.get("email_patterns", [])
# # # # # #         email_val      = ct.get("email") or (patterns[0] + "  (pattern)" if patterns else "")
# # # # # #         row_fill       = bg_hex if ri % 2 == 0 else GREY

# # # # # #         for ci, val in enumerate([
# # # # # #             ri - 3,
# # # # # #             prio,
# # # # # #             ct.get("name", ""),
# # # # # #             ct.get("title", ""),
# # # # # #             ct.get("company", ""),
# # # # # #             email_val,
# # # # # #             ct.get("linkedin_url", ""),
# # # # # #             ct.get("source", ""),
# # # # # #         ], 1):
# # # # # #             cell = ws.cell(row=ri, column=ci, value=val)
# # # # # #             cell.font      = Font(name="Arial", size=9,
# # # # # #                                   bold=(ci == 2),
# # # # # #                                   color=fg_hex if ci == 2 else "0F172A")
# # # # # #             cell.fill      = PatternFill("solid", fgColor=row_fill)
# # # # # #             cell.alignment = Alignment(vertical="center", wrap_text=(ci in [4, 7]))
# # # # # #             cell.border    = border

# # # # # #         ws.row_dimensions[ri].height = 18

# # # # # #     ws.freeze_panes   = "A4"
# # # # # #     ws.auto_filter.ref = f"A3:H{3 + len(contacts)}"

# # # # # #     # Summary sheet
# # # # # #     ws2 = wb.create_sheet("Summary")
# # # # # #     ws2.merge_cells("A1:C1")
# # # # # #     t = ws2["A1"]
# # # # # #     t.value     = "Export Summary"
# # # # # #     t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
# # # # # #     t.fill      = PatternFill("solid", fgColor=NAVY)
# # # # # #     t.alignment = Alignment(horizontal="center")

# # # # # #     for r, (lbl, val) in enumerate([
# # # # # #         ("Company",        company),
# # # # # #         ("Role",           role),
# # # # # #         ("Total Contacts", len(contacts)),
# # # # # #         ("Primary",        sum(1 for x in contacts if x.get("priority") == "Primary")),
# # # # # #         ("Secondary",      sum(1 for x in contacts if x.get("priority") == "Secondary")),
# # # # # #         ("Tertiary",       sum(1 for x in contacts if x.get("priority") == "Tertiary")),
# # # # # #         ("General",        sum(1 for x in contacts if x.get("priority") == "General")),
# # # # # #         ("With Email",     sum(1 for x in contacts if x.get("email"))),
# # # # # #         ("With LinkedIn",  sum(1 for x in contacts if x.get("linkedin_url"))),
# # # # # #         ("Generated",      datetime.now().strftime("%d %b %Y  %H:%M")),
# # # # # #     ], 2):
# # # # # #         lc = ws2.cell(row=r, column=1, value=lbl)
# # # # # #         vc = ws2.cell(row=r, column=2, value=val)
# # # # # #         bg = LBLUE if r % 2 == 0 else WHITE
# # # # # #         for cell in (lc, vc):
# # # # # #             cell.font   = Font(name="Arial", bold=(cell.column == 1), size=10)
# # # # # #             cell.fill   = PatternFill("solid", fgColor=bg)
# # # # # #             cell.border = border

# # # # # #     ws2.column_dimensions["A"].width = 20
# # # # # #     ws2.column_dimensions["B"].width = 30

# # # # # #     buf = io.BytesIO()
# # # # # #     wb.save(buf)
# # # # # #     buf.seek(0)
# # # # # #     return buf.getvalue()


# # # # # # # ── MongoDB helpers ───────────────────────────────────────────────────────────
# # # # # # @st.cache_resource
# # # # # # def get_db():
# # # # # #     URI = st.secrets.get("MONGO_URI", "mongodb://localhost:27017")
# # # # # #     DB  = st.secrets.get("MONGO_DB",  "secureitlab_job_pipeline")
# # # # # #     client = MongoClient(URI, serverSelectionTimeoutMS=6000)
# # # # # #     return client[DB]

# # # # # # @st.cache_data(ttl=60)
# # # # # # def load_all_jobs():
# # # # # #     db = get_db()
# # # # # #     return list(db.jobs.find({}, {
# # # # # #         "_id": 1, "company": 1, "role": 1, "job_number": 1,
# # # # # #         "opp_score": 1, "contacts_found": 1, "pipeline_ok": 1,
# # # # # #         "coverage_score": 1, "run_at": 1, "contact_domain": 1,
# # # # # #         "contacts": 1, "country": 1,
# # # # # #     }))

# # # # # # @st.cache_data(ttl=60)
# # # # # # def load_job(job_id):
# # # # # #     from bson import ObjectId
# # # # # #     return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# # # # # # # ── Render helpers ────────────────────────────────────────────────────────────
# # # # # # def badge(text, color="blue"):
# # # # # #     return f'<span class="badge badge-{color}">{text}</span>'

# # # # # # def safe_str(val, limit=300):
# # # # # #     if val is None: return "—"
# # # # # #     s = str(val)
# # # # # #     return s[:limit] + "…" if len(s) > limit else s

# # # # # # def as_dict(raw):
# # # # # #     if isinstance(raw, dict): return raw
# # # # # #     if isinstance(raw, list): return next((x for x in raw if isinstance(x, dict)), {})
# # # # # #     return {}

# # # # # # def render_json_pretty(data, title=""):
# # # # # #     if not data: return
# # # # # #     with st.expander(f"📄 Raw JSON — {title}", expanded=False):
# # # # # #         st.code(json.dumps(data, indent=2, default=str), language="json")

# # # # # # def render_qa_block(data, label):
# # # # # #     if not data:
# # # # # #         st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>', unsafe_allow_html=True)
# # # # # #         return
# # # # # #     data = as_dict(data)
# # # # # #     if not data: return
# # # # # #     passed    = data.get("passed") or data.get("Passed") or False
# # # # # #     rec       = data.get("recommendation") or data.get("Recommendation", "")
# # # # # #     issues    = data.get("issues") or data.get("Issues") or []
# # # # # #     checklist = data.get("checklist") or data.get("Checklist") or []
# # # # # #     color     = "green" if passed else "yellow"
# # # # # #     status    = "✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
# # # # # #     html = f"""
# # # # # #     <div class="sil-card sil-card-accent">
# # # # # #       <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
# # # # # #         <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
# # # # # #         {badge(status, color)}
# # # # # #       </div>"""
# # # # # #     if rec:
# # # # # #         html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
# # # # # #     if checklist:
# # # # # #         html += '<div style="font-size:.82rem;margin-bottom:.5rem">'
# # # # # #         for item in (checklist if isinstance(checklist, list) else [checklist]):
# # # # # #             if isinstance(item, dict):
# # # # # #                 ip = item.get("pass") or item.get("passed") or item.get("status","") == "pass"
# # # # # #                 nm = item.get("item") or item.get("name") or item.get("check","")
# # # # # #                 nt = item.get("reason") or item.get("note") or item.get("issue","")
# # # # # #                 html += f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
# # # # # #                 if nt: html += f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
# # # # # #                 html += '</div>'
# # # # # #         html += '</div>'
# # # # # #     if issues:
# # # # # #         html += '<div style="margin-top:.5rem">'
# # # # # #         for iss in (issues if isinstance(issues, list) else [issues])[:4]:
# # # # # #             txt = iss if isinstance(iss, str) else json.dumps(iss)
# # # # # #             html += f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
# # # # # #         html += '</div>'
# # # # # #     st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # # def render_contacts(contacts, title="Contacts"):
# # # # # #     if not contacts: st.info("No contacts found for this job."); return
# # # # # #     pri_color = {"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
# # # # # #     st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>', unsafe_allow_html=True)
# # # # # #     cols = st.columns(2)
# # # # # #     for i, c in enumerate(contacts):
# # # # # #         col = cols[i % 2]
# # # # # #         prio = c.get("priority", "General")
# # # # # #         email = c.get("email", ""); li = c.get("linkedin_url", "")
# # # # # #         patterns = c.get("email_patterns", []); src = c.get("source", "")
# # # # # #         with col:
# # # # # #             html = f"""<div class="contact-card">
# # # # # #               <div style="display:flex;justify-content:space-between;align-items:flex-start">
# # # # # #                 <div>
# # # # # #                   <div class="contact-name">{c.get('name','—')}</div>
# # # # # #                   <div class="contact-title">{c.get('title','—')}</div>
# # # # # #                 </div>
# # # # # #                 {badge(prio, pri_color.get(prio,'blue'))}
# # # # # #               </div>"""
# # # # # #             if email:      html += f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
# # # # # #             elif patterns: html += f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
# # # # # #             if li:         html += f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
# # # # # #             if src:        html += f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
# # # # # #             st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # # # def render_emails(emails_data):
# # # # # #     if not emails_data: st.info("No email data available."); return
# # # # # #     emails_data = as_dict(emails_data)
# # # # # #     if not emails_data: return
# # # # # #     variants = {}
# # # # # #     for k, v in emails_data.items():
# # # # # #         kl = k.lower().replace("_","").replace(" ","")
# # # # # #         if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl == "a":
# # # # # #             variants["Variant A — Hiring Manager"] = v
# # # # # #         elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl == "b":
# # # # # #             variants["Variant B — CISO / VP Level"] = v
# # # # # #         else:
# # # # # #             variants[k] = v
# # # # # #     for label, v in variants.items():
# # # # # #         st.markdown(f'<div class="section-label">✉️ {label}</div>', unsafe_allow_html=True)
# # # # # #         if isinstance(v, dict):
# # # # # #             subj = v.get("subject") or v.get("Subject","")
# # # # # #             body = v.get("body") or v.get("Body") or v.get("content","")
# # # # # #             if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>', unsafe_allow_html=True)
# # # # # #             if body: st.markdown(f'<div class="email-box">{body}</div>', unsafe_allow_html=True)
# # # # # #             else:    st.code(json.dumps(v, indent=2), language="json")
# # # # # #         elif isinstance(v, str):
# # # # # #             st.markdown(f'<div class="email-box">{v}</div>', unsafe_allow_html=True)
# # # # # #         st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

# # # # # # def render_service_mapping(data):
# # # # # #     if not data: st.info("No service mapping data."); return
# # # # # #     items = data if isinstance(data, list) else []
# # # # # #     if not items and isinstance(data, dict):
# # # # # #         for key in ("services","mappings","service_mapping","ServiceMapping","items"):
# # # # # #             if isinstance(data.get(key), list): items = data[key]; break
# # # # # #         if not items: items = [data]
# # # # # #     fit_colors = {"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
# # # # # #     for item in items:
# # # # # #         if not isinstance(item, dict): continue
# # # # # #         svc  = item.get("service") or item.get("Service") or item.get("name","")
# # # # # #         fit  = (item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
# # # # # #         why  = item.get("justification") or item.get("rationale") or item.get("why","")
# # # # # #         reqs = item.get("requirements_addressed") or item.get("requirements") or ""
# # # # # #         eng  = item.get("engagement_type") or item.get("engagement","")
# # # # # #         color = fit_colors.get(fit, "blue")
# # # # # #         html = f"""<div class="sil-card" style="margin-bottom:.75rem">
# # # # # #           <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
# # # # # #             <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
# # # # # #             {badge(fit or "MAPPED", color) if fit else ""}
# # # # # #           </div>"""
# # # # # #         if why:  html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
# # # # # #         if reqs:
# # # # # #             rs = ", ".join(reqs) if isinstance(reqs, list) else str(reqs)
# # # # # #             html += f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
# # # # # #         if eng:  html += f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
# # # # # #         st.markdown(html + '</div>', unsafe_allow_html=True)
# # # # # #     render_json_pretty(data, "Service Mapping")

# # # # # # def render_microplans(data):
# # # # # #     if not data: st.info("No micro-plan data."); return
# # # # # #     plans = data if isinstance(data, list) else []
# # # # # #     if not plans and isinstance(data, dict):
# # # # # #         for k in ("plans","micro_plans","microplans","top_3","improvements"):
# # # # # #             if isinstance(data.get(k), list): plans = data[k]; break
# # # # # #         if not plans: plans = [data]
# # # # # #     for i, plan in enumerate(plans[:3], 1):
# # # # # #         if not isinstance(plan, dict): continue
# # # # # #         title = plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
# # # # # #         weeks = plan.get("duration") or plan.get("timeline","")
# # # # # #         obj   = plan.get("objective") or plan.get("goal","")
# # # # # #         kpis  = plan.get("kpis") or plan.get("KPIs") or []
# # # # # #         tasks = plan.get("tasks") or plan.get("workstreams") or []
# # # # # #         with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}", expanded=(i==1)):
# # # # # #             if obj and obj != title: st.markdown(f"**Objective:** {obj}")
# # # # # #             if kpis:
# # # # # #                 st.markdown("**KPIs:**")
# # # # # #                 for kpi in (kpis if isinstance(kpis, list) else [kpis]): st.markdown(f"• {kpi}")
# # # # # #             if tasks:
# # # # # #                 st.markdown("**Tasks / Workstreams:**")
# # # # # #                 for t in (tasks if isinstance(tasks, list) else [tasks]):
# # # # # #                     if isinstance(t, dict):
# # # # # #                         tn = t.get("task") or t.get("name","")
# # # # # #                         te = t.get("effort") or t.get("duration","")
# # # # # #                         st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
# # # # # #                     else: st.markdown(f"• {t}")
# # # # # #             if plan:
# # # # # #                 st.code(json.dumps(plan, indent=2, default=str), language="json")

# # # # # # def render_deal_assurance(data):
# # # # # #     if not data: st.info("No deal assurance data."); return
# # # # # #     if not isinstance(data, dict): render_json_pretty(data, "Deal Assurance Pack"); return
# # # # # #     evp = (data.get("executive_value_proposition") or
# # # # # #            data.get("value_proposition") or data.get("ExecutiveValueProposition",""))
# # # # # #     if evp:
# # # # # #         st.markdown('<div class="section-label">💼 Executive Value Proposition</div>', unsafe_allow_html=True)
# # # # # #         st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7;color:var(--text)">{evp}</div>', unsafe_allow_html=True)
# # # # # #     caps = data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
# # # # # #     if caps:
# # # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>', unsafe_allow_html=True)
# # # # # #         c1, c2 = st.columns(2)
# # # # # #         for i, cap in enumerate(caps if isinstance(caps, list) else [caps]):
# # # # # #             (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
# # # # # #     risk = data.get("risk_mitigation") or data.get("RiskMitigation","")
# # # # # #     if risk:
# # # # # #         st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>', unsafe_allow_html=True)
# # # # # #         if isinstance(risk, dict):
# # # # # #             for k, v in risk.items(): st.markdown(f"**{k}:** {v}")
# # # # # #         else: st.markdown(str(risk))
# # # # # #     render_json_pretty(data, "Deal Assurance Pack")


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  SIDEBAR
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # with st.sidebar:
# # # # # #     st.markdown("""
# # # # # #     <div style="padding:.75rem 0 1.25rem">
# # # # # #       <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;
# # # # # #                   color:#2563eb">🛡️ SecureITLab</div>
# # # # # #       <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;
# # # # # #                   text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
# # # # # #     </div>
# # # # # #     """, unsafe_allow_html=True)

# # # # # #     # ── Logout button ─────────────────────────────────────────────────────────
# # # # # #     if st.button("🚪 Logout", use_container_width=True):
# # # # # #         st.session_state.logged_in = False
# # # # # #         st.session_state.login_error = ""
# # # # # #         st.rerun()

# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # #     # ── Find Jobs button ──────────────────────────────────────────────────────
# # # # # #     st.markdown("**🔍 Find New Jobs**")
# # # # # #     st.caption(
# # # # # #         "Runs scraper → checks MongoDB for duplicates → "
# # # # # #         "runs AI pipeline only on NEW jobs → stores in MongoDB"
# # # # # #     )

# # # # # #     find_jobs_clicked = st.button(
# # # # # #         "🔍  Find Jobs",
# # # # # #         disabled=st.session_state.pipeline_running,
# # # # # #         use_container_width=True,
# # # # # #         type="primary",
# # # # # #     )

# # # # # #     if st.session_state.pipeline_running:
# # # # # #         st.markdown("""
# # # # # #         <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
# # # # # #                     padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;
# # # # # #                     font-size:.8rem;color:#1d4ed8">
# # # # # #           ⏳ Pipeline is running… check log below
# # # # # #         </div>""", unsafe_allow_html=True)

# # # # # #     if st.session_state.pipeline_done and st.session_state.pipeline_result:
# # # # # #         res = st.session_state.pipeline_result
# # # # # #         ok  = res.get("success", False)
# # # # # #         bg  = "#ecfdf5" if ok else "#fef2f2"
# # # # # #         bc  = "#bbf7d0" if ok else "#fecaca"
# # # # # #         tc  = "#15803d" if ok else "#b91c1c"
# # # # # #         ic  = "✅" if ok else "❌"
# # # # # #         st.markdown(f"""
# # # # # #         <div style="background:{bg};border:1px solid {bc};border-radius:8px;
# # # # # #                     padding:.75rem;margin-top:.5rem;font-size:.82rem">
# # # # # #           <div style="font-weight:700;color:{tc};margin-bottom:.4rem">{ic} Last Run</div>
# # # # # #           <div>📦 Scraped: <b>{res.get('scraped',0)}</b></div>
# # # # # #           <div>🆕 New jobs: <b>{res.get('new_jobs',0)}</b></div>
# # # # # #           <div>⏭️ Already in DB (skipped): <b>{res.get('skipped_db',0)}</b></div>
# # # # # #           <div>🤖 Processed by AI: <b>{res.get('processed',0)}</b></div>
# # # # # #           {f'<div style="color:{tc};margin-top:.3rem">⚠️ {res.get("error","")}</div>' if res.get("error") else ""}
# # # # # #         </div>""", unsafe_allow_html=True)

# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # #     # ── Job list ──────────────────────────────────────────────────────────────
# # # # # #     try:
# # # # # #         all_jobs = load_all_jobs()
# # # # # #     except Exception as e:
# # # # # #         st.error(f"MongoDB connection failed: {e}")
# # # # # #         st.stop()

# # # # # #     if not all_jobs:
# # # # # #         st.warning("No jobs in MongoDB yet. Click **Find Jobs** to scrape and process.")
# # # # # #         st.stop()

# # # # # #     st.markdown(
# # # # # #         f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">'
# # # # # #         f'{len(all_jobs)} jobs in database</div>',
# # # # # #         unsafe_allow_html=True,
# # # # # #     )

# # # # # #     # ── MASTER CONTACTS EXPORT (NEW) ────────────────────────────────────────────
# # # # # #     st.markdown("**📋 Master Contacts Export**")
# # # # # #     st.caption("All contacts from ALL jobs in one Excel file with auto-updates")
    
# # # # # #     master_excel = build_master_contacts_excel(all_jobs)
# # # # # #     if master_excel:
# # # # # #         st.download_button(
# # # # # #             label="📥  Download All Contacts",
# # # # # #             data=master_excel,
# # # # # #             file_name=f"master_contacts_{datetime.now().strftime('%Y%m%d')}.xlsx",
# # # # # #             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# # # # # #             use_container_width=True,
# # # # # #         )
# # # # # #     else:
# # # # # #         st.warning("Run `pip install openpyxl` to enable Excel export.")

# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # #     search   = st.text_input("🔍 Filter by company / role", placeholder="e.g. Bounteous")
# # # # # #     filtered = [j for j in all_jobs
# # # # # #                 if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

# # # # # #     def job_label(j):
# # # # # #         score = j.get("opp_score")
# # # # # #         s_str = f" [{score}/10]" if score else ""
# # # # # #         ok    = "✅" if j.get("pipeline_ok") else "❌"
# # # # # #         return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s_str}"

# # # # # #     if not filtered:
# # # # # #         st.warning("No matching jobs.")
# # # # # #         st.stop()

# # # # # #     sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
# # # # # #     sel_idx     = [job_label(j) for j in filtered].index(sel_label)
# # # # # #     selected_id = str(filtered[sel_idx]["_id"])

# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # #     ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
# # # # # #     total_c  = sum(j.get("contacts_found",0) for j in all_jobs)
# # # # # #     st.markdown(f"""
# # # # # #     <div style="font-size:.75rem;color:#64748b">
# # # # # #       <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
# # # # # #       <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
# # # # # #     </div>""", unsafe_allow_html=True)

# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # #     if st.button("🔄 Refresh Data", use_container_width=True):
# # # # # #         st.cache_data.clear()
# # # # # #         st.rerun()


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  FIND JOBS — background thread
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # if find_jobs_clicked and not st.session_state.pipeline_running:
# # # # # #     with _thread_lock:
# # # # # #         _thread_logs.clear()
# # # # # #         _thread_result[0] = None
# # # # # #         _thread_done[0]   = False

# # # # # #     st.session_state.pipeline_running = True
# # # # # #     st.session_state.pipeline_done    = False
# # # # # #     st.session_state.pipeline_logs    = []
# # # # # #     st.session_state.pipeline_result  = None
# # # # # #     st.cache_data.clear()

# # # # # #     def _run_pipeline_bg():
# # # # # #         try:
# # # # # #             import main as _main

# # # # # #             def _cb(msg: str):
# # # # # #                 with _thread_lock:
# # # # # #                     _thread_logs.append(f"{datetime.now().strftime('%H:%M:%S')} | {msg}")

# # # # # #             result = _main.run_pipeline(progress_callback=_cb)
# # # # # #             with _thread_lock:
# # # # # #                 _thread_result[0] = result
# # # # # #         except Exception as e:
# # # # # #             with _thread_lock:
# # # # # #                 _thread_logs.append(f"❌ Unexpected error: {e}")
# # # # # #                 _thread_result[0] = {
# # # # # #                     "success": False, "error": str(e),
# # # # # #                     "scraped": 0, "new_jobs": 0, "skipped_db": 0, "processed": 0,
# # # # # #                 }
# # # # # #         finally:
# # # # # #             with _thread_lock:
# # # # # #                 _thread_done[0] = True

# # # # # #     threading.Thread(target=_run_pipeline_bg, daemon=True).start()
# # # # # #     st.rerun()

# # # # # # # ── Sync thread state → session_state ────────────────────────────────────────
# # # # # # with _thread_lock:
# # # # # #     if _thread_logs:
# # # # # #         st.session_state.pipeline_logs = list(_thread_logs)
# # # # # #     if _thread_done[0] and _thread_result[0] is not None:
# # # # # #         st.session_state.pipeline_result  = _thread_result[0]
# # # # # #         st.session_state.pipeline_running = False
# # # # # #         st.session_state.pipeline_done    = True

# # # # # # # ── Live log panel ────────────────────────────────────────────────────────────
# # # # # # if st.session_state.pipeline_running or (
# # # # # #         st.session_state.pipeline_done and st.session_state.pipeline_logs):
# # # # # #     heading = "⏳ Pipeline running — live log…" if st.session_state.pipeline_running \
# # # # # #               else "📋 Last pipeline run log"
# # # # # #     with st.expander(heading, expanded=st.session_state.pipeline_running):
# # # # # #         log_text = "\n".join(st.session_state.pipeline_logs[-100:]) \
# # # # # #                    if st.session_state.pipeline_logs else "Starting…"
# # # # # #         st.markdown(f'<div class="pipeline-log">{log_text}</div>', unsafe_allow_html=True)
# # # # # #         if st.session_state.pipeline_running:
# # # # # #             time.sleep(2)
# # # # # #             st.rerun()


# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  MAIN CONTENT — selected job detail (ORIGINAL CODE)
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # with st.spinner("Loading job from MongoDB…"):
# # # # # #     job = load_job(selected_id)

# # # # # # if not job:
# # # # # #     st.error("Could not load job document.")
# # # # # #     st.stop()

# # # # # # company   = job.get("company",  "Unknown")
# # # # # # role      = job.get("role",     "Unknown")
# # # # # # opp_score = job.get("opp_score")
# # # # # # p_ok      = job.get("pipeline_ok", False)
# # # # # # p_min     = job.get("pipeline_min", "?")
# # # # # # c_found   = job.get("contacts_found", 0)
# # # # # # c_cov     = job.get("coverage_score")
# # # # # # c_domain  = job.get("contact_domain","")
# # # # # # run_at    = job.get("run_at","")

# # # # # # # ── Header ────────────────────────────────────────────────────────────────────
# # # # # # st.markdown(f"""
# # # # # # <div style="margin-bottom:1.75rem">
# # # # # #   <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;
# # # # # #               letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
# # # # # #     Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
# # # # # #   </div>
# # # # # #   <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;
# # # # # #              color:#0f172a;margin:0;line-height:1.15">{role}</h1>
# # # # # #   <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
# # # # # #     @ <span style="color:#334155;font-weight:600">{company}</span>
# # # # # #     {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
# # # # # #   </div>
# # # # # # </div>
# # # # # # """, unsafe_allow_html=True)

# # # # # # # ── Metric tiles ──────────────────────────────────────────────────────────────
# # # # # # try:
# # # # # #     sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
# # # # # #     sc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # # # except Exception:
# # # # # #     sc = "#2563eb"

# # # # # # st.markdown(f"""
# # # # # # <div class="metric-row">
# # # # # #   <div class="metric-tile">
# # # # # #     <div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div>
# # # # # #     <div class="lbl">Opportunity Score</div>
# # # # # #   </div>
# # # # # #   <div class="metric-tile">
# # # # # #     <div class="val">{c_found}</div>
# # # # # #     <div class="lbl">Contacts Found</div>
# # # # # #   </div>
# # # # # #   <div class="metric-tile">
# # # # # #     <div class="val">{f"{c_cov}%" if c_cov else "—"}</div>
# # # # # #     <div class="lbl">Contact Coverage</div>
# # # # # #   </div>
# # # # # #   <div class="metric-tile">
# # # # # #     <div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">
# # # # # #       {'✅ OK' if p_ok else '❌ Failed'}
# # # # # #     </div>
# # # # # #     <div class="lbl">Pipeline ({p_min} min)</div>
# # # # # #   </div>
# # # # # # </div>
# # # # # # """, unsafe_allow_html=True)

# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # #  TABS
# # # # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # # # tabs = st.tabs([
# # # # # #     "📋 Job & Enrichment",
# # # # # #     "🎯 Service Mapping",
# # # # # #     "🔍 Fit / Gap",
# # # # # #     "🛠️ Capability & Plans",
# # # # # #     "📦 Deal Assurance",
# # # # # #     "✉️ Outreach Emails",
# # # # # #     "👥 Contacts",
# # # # # #     "✅ QA Gates",
# # # # # #     "🗄️ Raw Data",
# # # # # # ])

# # # # # # # ── Tab 1: Job Research + Enrichment ─────────────────────────────────────────
# # # # # # with tabs[0]:
# # # # # #     col1, col2 = st.columns([1, 1])
# # # # # #     with col1:
# # # # # #         st.markdown('<div class="section-label">📄 Job Research</div>', unsafe_allow_html=True)
# # # # # #         jr = as_dict(job.get("agent_job_research") or {})
# # # # # #         if jr:
# # # # # #             for label, keys in [
# # # # # #                 ("Job Role",        ["job_role","Job Role","role","title"]),
# # # # # #                 ("Company",         ["company_name","Company Name","company"]),
# # # # # #                 ("Location",        ["location","Location"]),
# # # # # #                 ("Organization URL",["organization_url","Organization URL","url"]),
# # # # # #             ]:
# # # # # #                 val = next((jr.get(k) for k in keys if jr.get(k)), None)
# # # # # #                 if val: st.markdown(f"**{label}:** {val}")
# # # # # #             desc = jr.get("job_description") or jr.get("Job Description","")
# # # # # #             if desc:
# # # # # #                 st.markdown("**Job Description:**")
# # # # # #                 st.markdown(
# # # # # #                     f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;'
# # # # # #                     f'max-height:300px;overflow-y:auto;color:var(--text)">{desc[:2000]}</div>',
# # # # # #                     unsafe_allow_html=True,
# # # # # #                 )
# # # # # #             render_json_pretty(jr, "Job Research")
# # # # # #         else:
# # # # # #             st.info("No job research data.")
# # # # # #     with col2:
# # # # # #         st.markdown('<div class="section-label">🏢 Company Enrichment</div>', unsafe_allow_html=True)
# # # # # #         enr = as_dict(job.get("agent_enrichment") or {})
# # # # # #         if enr:
# # # # # #             for label, keys in [
# # # # # #                 ("Industry",          ["industry","Industry"]),
# # # # # #                 ("Company Size",      ["company_size","size","Company Size"]),
# # # # # #                 ("Regulatory Env",    ["regulatory_environment","regulatory"]),
# # # # # #                 ("Certifications",    ["certifications","Certifications"]),
# # # # # #                 ("Tech Stack",        ["tech_stack","technology_stack"]),
# # # # # #                 ("Security Maturity", ["security_maturity","maturity"]),
# # # # # #             ]:
# # # # # #                 val = next((enr.get(k) for k in keys if enr.get(k)), None)
# # # # # #                 if val:
# # # # # #                     if isinstance(val, list): val = ", ".join(str(v) for v in val)
# # # # # #                     st.markdown(f"**{label}:** {safe_str(val, 200)}")
# # # # # #             render_json_pretty(enr, "Enrichment")
# # # # # #         else:
# # # # # #             st.info("No enrichment data.")

# # # # # # # ── Tab 2: Service Mapping ────────────────────────────────────────────────────
# # # # # # with tabs[1]:
# # # # # #     st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>', unsafe_allow_html=True)
# # # # # #     render_service_mapping(job.get("agent_service_mapping"))

# # # # # # # ── Tab 3: Fit / Gap Analysis ─────────────────────────────────────────────────
# # # # # # with tabs[2]:
# # # # # #     fg = as_dict(job.get("agent_fit_gap") or {})
# # # # # #     if opp_score:
# # # # # #         try:
# # # # # #             sn = float(str(opp_score).split("/")[0])
# # # # # #             bc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # # # #             bp = int(sn / 10 * 100)
# # # # # #             st.markdown(f"""
# # # # # #             <div style="margin-bottom:1.5rem">
# # # # # #               <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
# # # # # #                 <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">Opportunity Score</span>
# # # # # #                 <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
# # # # # #               </div>
# # # # # #               <div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%">
# # # # # #                 <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
# # # # # #               </div>
# # # # # #             </div>""", unsafe_allow_html=True)
# # # # # #         except Exception:
# # # # # #             pass
# # # # # #     st.markdown('<div class="section-label">📊 Service Classifications</div>', unsafe_allow_html=True)
# # # # # #     services = []
# # # # # #     if isinstance(fg, dict):
# # # # # #         for k in ("services","classifications","service_classifications","items","fit_gap"):
# # # # # #             v = fg.get(k)
# # # # # #             if isinstance(v, list): services = v; break
# # # # # #         if not services and (fg.get("service") or fg.get("Service")):
# # # # # #             services = [fg]
# # # # # #     elif isinstance(fg, list):
# # # # # #         services = fg
# # # # # #     if services:
# # # # # #         buckets = {"STRONG FIT": [], "PARTIAL FIT": [], "GAP": []}
# # # # # #         other   = []
# # # # # #         for s in services:
# # # # # #             if not isinstance(s, dict): continue
# # # # # #             fit = (s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
# # # # # #             if "STRONG" in fit: buckets["STRONG FIT"].append(s)
# # # # # #             elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
# # # # # #             elif "GAP" in fit: buckets["GAP"].append(s)
# # # # # #             else: other.append(s)
# # # # # #         c1, c2, c3 = st.columns(3)
# # # # # #         cm  = {"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
# # # # # #         bgm = {"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
# # # # # #         bdm = {"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
# # # # # #         for col, (fl, items) in zip([c1, c2, c3], buckets.items()):
# # # # # #             col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>', unsafe_allow_html=True)
# # # # # #             for s in items:
# # # # # #                 svc  = s.get("service") or s.get("Service") or s.get("name","")
# # # # # #                 just = s.get("justification") or s.get("reason","")
# # # # # #                 col.markdown(
# # # # # #                     f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem">'
# # # # # #                     f'<div style="font-weight:600;color:#0f172a">{svc}</div>'
# # # # # #                     f'<div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>',
# # # # # #                     unsafe_allow_html=True,
# # # # # #                 )
# # # # # #     elif fg:
# # # # # #         st.json(fg)
# # # # # #     else:
# # # # # #         st.info("No fit/gap data.")
# # # # # #     render_json_pretty(job.get("agent_fit_gap"), "Fit/Gap Analysis")

# # # # # # # ── Tab 4: Capability + Micro-plans ──────────────────────────────────────────
# # # # # # with tabs[3]:
# # # # # #     col1, col2 = st.columns([1, 1])
# # # # # #     with col1:
# # # # # #         st.markdown('<div class="section-label">🔧 Capability Improvements</div>', unsafe_allow_html=True)
# # # # # #         cap = job.get("agent_capability") or {}
# # # # # #         items_cap = cap if isinstance(cap, list) else []
# # # # # #         if not items_cap and isinstance(cap, dict):
# # # # # #             for k in ("improvements","recommendations","capabilities","items"):
# # # # # #                 v = cap.get(k)
# # # # # #                 if isinstance(v, list): items_cap = v; break
# # # # # #             if not items_cap: items_cap = [cap]
# # # # # #         for item in items_cap:
# # # # # #             if not isinstance(item, dict): continue
# # # # # #             title  = item.get("title") or item.get("gap") or item.get("service","")
# # # # # #             rec    = item.get("recommendation") or item.get("steps","")
# # # # # #             effort = item.get("build_effort") or item.get("effort","")
# # # # # #             demand = item.get("market_demand") or item.get("priority","")
# # # # # #             st.markdown(f"""
# # # # # #             <div class="sil-card" style="margin-bottom:.6rem">
# # # # # #               <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem;color:var(--text)">{title}</div>
# # # # # #               <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec, 250)}</div>
# # # # # #               {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
# # # # # #             </div>""", unsafe_allow_html=True)
# # # # # #         if not items_cap:
# # # # # #             render_json_pretty(cap, "Capability Improvement")
# # # # # #     with col2:
# # # # # #         st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>', unsafe_allow_html=True)
# # # # # #         render_microplans(job.get("agent_microplans"))

# # # # # # # ── Tab 5: Deal Assurance Pack ────────────────────────────────────────────────
# # # # # # with tabs[4]:
# # # # # #     render_deal_assurance(job.get("agent_deal_assurance"))

# # # # # # # ── Tab 6: Outreach Emails ────────────────────────────────────────────────────
# # # # # # with tabs[5]:
# # # # # #     st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>', unsafe_allow_html=True)
# # # # # #     emails_src = job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
# # # # # #     oq = as_dict(job.get("agent_outreach_qa") or {})
# # # # # #     improved = (oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
# # # # # #     if improved:
# # # # # #         st.info("⚡ Showing QA-improved versions where available")
# # # # # #         render_emails(improved)
# # # # # #         with st.expander("📬 Original (pre-QA) versions"):
# # # # # #             render_emails(emails_src)
# # # # # #     else:
# # # # # #         render_emails(emails_src)

# # # # # # # ── Tab 7: Contacts ───────────────────────────────────────────────────────────
# # # # # # with tabs[6]:
# # # # # #     contacts        = job.get("contacts") or []
# # # # # #     contact_sources = job.get("contact_sources") or []
# # # # # #     pri = [c for c in contacts if c.get("priority") == "Primary"]
# # # # # #     sec = [c for c in contacts if c.get("priority") == "Secondary"]
# # # # # #     ter = [c for c in contacts if c.get("priority") == "Tertiary"]
# # # # # #     gen = [c for c in contacts if c.get("priority") == "General"]
# # # # # #     st.markdown(f"""
# # # # # #     <div class="metric-row" style="margin-bottom:1.5rem">
# # # # # #       <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
# # # # # #       <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
# # # # # #       <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
# # # # # #       <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
# # # # # #     </div>""", unsafe_allow_html=True)
# # # # # #     if contact_sources:
# # # # # #         st.markdown('Sources: ' + " ".join(badge(s,"blue") for s in contact_sources), unsafe_allow_html=True)
# # # # # #         st.markdown("")
# # # # # #     missing = job.get("missing_roles") or []
# # # # # #     if missing:
# # # # # #         st.markdown('⚠️ Missing roles: ' + " ".join(badge(r,"red") for r in missing), unsafe_allow_html=True)
# # # # # #         st.markdown("")
# # # # # #     if contacts:
# # # # # #         # ── Excel export button (Individual Job) ───────────────────────────────
# # # # # #         excel_bytes = build_contacts_excel(contacts, company, role)
# # # # # #         if excel_bytes:
# # # # # #             safe_co = re.sub(r'[^a-z0-9]', '_', company.lower())[:20]
# # # # # #             fname   = f"contacts_{safe_co}_{datetime.now().strftime('%Y%m%d')}.xlsx"
# # # # # #             btn_col, _ = st.columns([1, 3])
# # # # # #             with btn_col:
# # # # # #                 st.download_button(
# # # # # #                     label="📥  Download Contacts (.xlsx)",
# # # # # #                     data=excel_bytes,
# # # # # #                     file_name=fname,
# # # # # #                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# # # # # #                     use_container_width=True,
# # # # # #                     type="primary",
# # # # # #                 )
# # # # # #         else:
# # # # # #             st.warning("Run `pip install openpyxl` to enable Excel export.")

# # # # # #         st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # # # #         pri_filter = st.multiselect(
# # # # # #             "Filter by priority",
# # # # # #             ["Primary","Secondary","Tertiary","General"],
# # # # # #             default=["Primary","Secondary","Tertiary","General"],
# # # # # #         )
# # # # # #         shown = [c for c in contacts if c.get("priority","General") in pri_filter]
# # # # # #         render_contacts(shown, f"Contacts ({len(shown)} shown)")
# # # # # #         agent_contacts = job.get("agent_prospect_contacts") or {}
# # # # # #         if agent_contacts:
# # # # # #             with st.expander("🤖 CrewAI Agent's Contact Search"):
# # # # # #                 if isinstance(agent_contacts, dict):
# # # # # #                     ac_list = agent_contacts.get("contacts") or []
# # # # # #                     if ac_list: render_contacts(ac_list, "Agent Contacts")
# # # # # #                     else:       st.json(agent_contacts)
# # # # # #                 else:           st.json(agent_contacts)
# # # # # #     else:
# # # # # #         st.info("No contacts found for this job.")

# # # # # # # ── Tab 8: QA Gates ───────────────────────────────────────────────────────────
# # # # # # with tabs[7]:
# # # # # #     st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>', unsafe_allow_html=True)
# # # # # #     col1, col2 = st.columns(2)
# # # # # #     for i, (label, key) in enumerate([
# # # # # #         ("Research QA",       "agent_research_qa"),
# # # # # #         ("Service Mapping QA","agent_mapping_qa"),
# # # # # #         ("Deal Assurance QA", "agent_assurance_qa"),
# # # # # #         ("Outreach Email QA", "agent_outreach_qa"),
# # # # # #     ]):
# # # # # #         with (col1 if i % 2 == 0 else col2):
# # # # # #             render_qa_block(job.get(key), label)
# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # #     st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>', unsafe_allow_html=True)
# # # # # #     enf = as_dict(job.get("agent_prospect_enforcer") or {})
# # # # # #     if enf:
# # # # # #         cov  = enf.get("coverage_score","?")
# # # # # #         miss = enf.get("missing_roles",[])
# # # # # #         note = enf.get("note","")
# # # # # #         ec   = enf.get("contacts",[])
# # # # # #         x1, x2, x3 = st.columns(3)
# # # # # #         x1.metric("Coverage Score",    f"{cov}%")
# # # # # #         x2.metric("Missing Roles",     len(miss))
# # # # # #         x3.metric("Contacts Verified", len(ec))
# # # # # #         if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
# # # # # #         if note: st.caption(note)
# # # # # #     else:
# # # # # #         st.info("No enforcer data.")

# # # # # # # ── Tab 9: Raw Data ───────────────────────────────────────────────────────────
# # # # # # with tabs[8]:
# # # # # #     st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>', unsafe_allow_html=True)
# # # # # #     st.caption("All fields stored in the `jobs` collection for this document.")
# # # # # #     rows = []
# # # # # #     for k, v in job.items():
# # # # # #         if k == "_id": continue
# # # # # #         rows.append({"Field": k, "Type": type(v).__name__, "Len": len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
# # # # # #     hc1, hc2, hc3 = st.columns([3,1,1])
# # # # # #     hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
# # # # # #     for r in rows:
# # # # # #         rc1, rc2, rc3 = st.columns([3,1,1])
# # # # # #         rc1.code(r["Field"], language=None)
# # # # # #         rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>', unsafe_allow_html=True)
# # # # # #         rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>',  unsafe_allow_html=True)
# # # # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # # # #     for label, key in [
# # # # # #         ("Job Research","agent_job_research"), ("Enrichment","agent_enrichment"),
# # # # # #         ("Service Mapping","agent_service_mapping"), ("Fit/Gap Analysis","agent_fit_gap"),
# # # # # #         ("Capability","agent_capability"), ("Micro-Plans","agent_microplans"),
# # # # # #         ("Deal Assurance","agent_deal_assurance"), ("Outreach Emails","agent_outreach_emails"),
# # # # # #         ("Prospect Contacts","agent_prospect_contacts"), ("Prospect Enforcer","agent_prospect_enforcer"),
# # # # # #         ("Research QA","agent_research_qa"), ("Mapping QA","agent_mapping_qa"),
# # # # # #         ("Assurance QA","agent_assurance_qa"), ("Outreach QA","agent_outreach_qa"),
# # # # # #         ("Contacts (5-source)","contacts"),
# # # # # #     ]:
# # # # # #         data = job.get(key)
# # # # # #         if data:
# # # # # #             with st.expander(f"📄 {label}"):
# # # # # #                 st.code(json.dumps(data, indent=2, default=str), language="json")























# # # # """
# # # # ╔══════════════════════════════════════════════════════════╗
# # # # ║   SecureITLab Pipeline Dashboard — Streamlit             ║
# # # # ║   WITH MASTER CONTACTS AGGREGATION & AUTO-UPDATE         ║
# # # # ║   Reads from MongoDB → secureitlab_job_pipeline          ║
# # # # ║   FIXED: Captures terminal logs properly                 ║
# # # # ╠══════════════════════════════════════════════════════════╣
# # # # ║  Install: pip install streamlit pymongo python-dotenv    ║
# # # # ║  Run:     streamlit run streamlit_dashboard.py           ║
# # # # ╚══════════════════════════════════════════════════════════╝
# # # # """

# # # # import io
# # # # import re
# # # # import streamlit as st
# # # # from pymongo import MongoClient
# # # # import json
# # # # import threading
# # # # import time
# # # # import logging
# # # # from datetime import datetime, timezone
# # # # from io import StringIO

# # # # # ── Thread-safe shared state ──────────────────────────────────────────────────
# # # # _thread_logs   = []
# # # # _thread_result = [None]
# # # # _thread_done   = [False]
# # # # _thread_lock   = threading.Lock()

# # # # # Capture all logs
# # # # _log_capture = StringIO()
# # # # _log_handler = logging.StreamHandler(_log_capture)
# # # # _log_handler.setLevel(logging.INFO)
# # # # _log_formatter = logging.Formatter('%(asctime)s | %(levelname)s | %(message)s')
# # # # _log_handler.setFormatter(_log_formatter)

# # # # # ── Page config ───────────────────────────────────────────────────────────────
# # # # st.set_page_config(
# # # #     page_title="SecureITLab Pipeline",
# # # #     page_icon="🛡️",
# # # #     layout="wide",
# # # #     initial_sidebar_state="expanded",
# # # # )

# # # # # ── LOGIN CREDENTIALS (change these) ─────────────────────────────────────────
# # # # LOGIN_USERNAME = "admin"
# # # # LOGIN_PASSWORD = "secureitlab2024"

# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # #  GLOBAL CSS (login + dashboard share this)
# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # st.markdown("""
# # # # <style>

# # # # @import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');

# # # # :root {
# # # #     --bg:        #f4f7fb;
# # # #     --surface:   #ffffff;
# # # #     --surface2:  #eef2f7;
# # # #     --border:    #d9e2ec;
# # # #     --accent:    #2563eb;
# # # #     --accent2:   #7c3aed;
# # # #     --green:     #16a34a;
# # # #     --yellow:    #f59e0b;
# # # #     --red:       #dc2626;
# # # #     --text:      #0f172a;
# # # #     --muted:     #64748b;
# # # # }

# # # # html, body, [class*="css"] {
# # # #     background-color: var(--bg) !important;
# # # #     color: var(--text) !important;
# # # #     font-family: 'DM Sans', sans-serif !important;
# # # # }

# # # # /* ── LOGIN PAGE ─────────────────────────────────────────── */
# # # # .login-wrap {
# # # #     display: flex;
# # # #     align-items: center;
# # # #     justify-content: center;
# # # #     min-height: 80vh;
# # # # }
# # # # .login-card {
# # # #     background: var(--surface);
# # # #     border: 1px solid var(--border);
# # # #     border-radius: 20px;
# # # #     padding: 3rem 3.5rem;
# # # #     width: 100%;
# # # #     max-width: 420px;
# # # #     box-shadow: 0 20px 60px rgba(37,99,235,0.08);
# # # #     text-align: center;
# # # # }
# # # # .login-logo {
# # # #     font-family: 'Syne', sans-serif;
# # # #     font-size: 1.6rem;
# # # #     font-weight: 800;
# # # #     color: var(--accent);
# # # #     margin-bottom: .25rem;
# # # # }
# # # # .login-subtitle {
# # # #     font-size: .75rem;
# # # #     color: var(--muted);
# # # #     letter-spacing: .12em;
# # # #     text-transform: uppercase;
# # # #     margin-bottom: 2.5rem;
# # # # }
# # # # .login-error {
# # # #     background: #fef2f2;
# # # #     border: 1px solid #fecaca;
# # # #     border-radius: 8px;
# # # #     padding: .65rem 1rem;
# # # #     color: #b91c1c;
# # # #     font-size: .85rem;
# # # #     margin-top: 1rem;
# # # # }
# # # # .login-divider {
# # # #     width: 40px;
# # # #     height: 3px;
# # # #     background: linear-gradient(90deg, var(--accent), var(--accent2));
# # # #     border-radius: 2px;
# # # #     margin: 0 auto 2rem;
# # # # }

# # # # /* ── DASHBOARD ──────────────────────────────────────────── */
# # # # [data-testid="stSidebar"] {
# # # #     background: var(--surface) !important;
# # # #     border-right: 1px solid var(--border) !important;
# # # # }
# # # # [data-testid="stSidebar"] * { color: var(--text) !important; }

# # # # .main .block-container { padding: 2rem 2rem 3rem !important; }

# # # # h1, h2, h3, h4 {
# # # #     font-family: 'Syne', sans-serif !important;
# # # #     color: var(--text) !important;
# # # # }

# # # # .sil-card {
# # # #     background: var(--surface);
# # # #     border: 1px solid var(--border);
# # # #     border-radius: 14px;
# # # #     padding: 1.25rem 1.5rem;
# # # #     margin-bottom: 1rem;
# # # #     transition: all 0.25s ease;
# # # # }
# # # # .sil-card:hover {
# # # #     transform: translateY(-2px);
# # # #     box-shadow: 0 8px 22px rgba(0,0,0,0.05);
# # # # }
# # # # .sil-card-accent { border-left: 4px solid var(--accent); }

# # # # .metric-row { display:flex; gap:1rem; flex-wrap:wrap; margin-bottom:1.5rem; }
# # # # .metric-tile {
# # # #     flex:1; min-width:140px;
# # # #     background:var(--surface2);
# # # #     border:1px solid var(--border);
# # # #     border-radius:12px;
# # # #     padding:1rem; text-align:center;
# # # #     transition:all .25s ease;
# # # # }
# # # # .metric-tile:hover { transform:translateY(-3px); box-shadow:0 10px 24px rgba(0,0,0,0.06); }
# # # # .metric-tile .val { font-family:'Syne',sans-serif; font-size:2rem; font-weight:800; color:var(--accent); }
# # # # .metric-tile .lbl { font-size:.72rem; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; }

# # # # .badge { padding:.25rem .7rem; border-radius:20px; font-size:.72rem; font-weight:600; font-family:'DM Mono',monospace; }
# # # # .badge-green  { background:#ecfdf5; color:#15803d; }
# # # # .badge-yellow { background:#fffbeb; color:#b45309; }
# # # # .badge-red    { background:#fef2f2; color:#b91c1c; }
# # # # .badge-blue   { background:#eff6ff; color:#1d4ed8; }
# # # # .badge-purple { background:#f5f3ff; color:#6d28d9; }

# # # # .contact-card {
# # # #     background:var(--surface2); border:1px solid var(--border);
# # # #     border-radius:10px; padding:1rem; margin-bottom:.8rem;
# # # # }
# # # # .contact-name  { font-family:'Syne',sans-serif; font-weight:700; color:var(--text); }
# # # # .contact-title { color:var(--muted); font-size:.85rem; }
# # # # .contact-email { font-family:'DM Mono',monospace; color:var(--accent); font-size:.8rem; }

# # # # .email-box {
# # # #     background: #f8fafc;
# # # #     border: 1px solid var(--border);
# # # #     border-radius: 10px;
# # # #     padding: 1rem 1.25rem;
# # # #     font-size: .9rem;
# # # #     line-height: 1.65;
# # # #     white-space: pre-wrap;
# # # #     color: var(--text);
# # # # }
# # # # .email-subject { font-family:'Syne',sans-serif; font-weight:700; color:var(--accent); margin-bottom:.5rem; }

# # # # .section-label {
# # # #     font-family:'DM Mono',monospace; font-size:.72rem;
# # # #     text-transform:uppercase; letter-spacing:.12em;
# # # #     color:var(--accent); margin-bottom:.6rem;
# # # # }
# # # # .sil-divider { border-top:1px solid var(--border); margin:1rem 0; }

# # # # [data-testid="stExpander"] {
# # # #     background: var(--surface) !important;
# # # #     border: 1px solid var(--border) !important;
# # # #     border-radius: 10px !important;
# # # # }
# # # # [data-testid="stSelectbox"] > div,
# # # # [data-testid="stMultiSelect"] > div {
# # # #     background: var(--surface2) !important;
# # # #     border-color: var(--border) !important;
# # # # }
# # # # [data-testid="stTabs"] button {
# # # #     font-family:'Syne',sans-serif !important;
# # # #     font-weight:600 !important;
# # # # }
# # # # ::-webkit-scrollbar { width:6px; }
# # # # ::-webkit-scrollbar-thumb { background:var(--border); border-radius:3px; }

# # # # .pipeline-log {
# # # #     background: #0f172a;
# # # #     color: #10b981;
# # # #     border-radius: 10px;
# # # #     padding: 1.5rem;
# # # #     font-family: 'DM Mono', monospace;
# # # #     font-size: .8rem;
# # # #     line-height: 1.8;
# # # #     max-height: 700px;
# # # #     overflow-y: auto;
# # # #     white-space: pre-wrap;
# # # #     word-break: break-word;
# # # #     border: 1px solid #1e293b;
# # # # }

# # # # .fit-box {
# # # #     border-radius: 8px;
# # # #     padding: .75rem;
# # # #     margin-bottom: .5rem;
# # # #     font-size: .85rem;
# # # # }

# # # # /* Hide sidebar on login page */
# # # # .hide-sidebar [data-testid="stSidebar"] { display: none !important; }
# # # # .hide-sidebar .main .block-container { max-width: 480px; margin: 0 auto; }

# # # # /* Logs page styles */
# # # # .logs-container {
# # # #     display: flex;
# # # #     flex-direction: column;
# # # #     gap: 1rem;
# # # # }

# # # # .logs-status {
# # # #     display: flex;
# # # #     gap: 1rem;
# # # #     justify-content: space-between;
# # # #     align-items: center;
# # # #     margin-bottom: 1.5rem;
# # # #     padding: 1rem;
# # # #     background: var(--surface2);
# # # #     border-radius: 10px;
# # # #     border: 1px solid var(--border);
# # # # }

# # # # .logs-status.running {
# # # #     background: #eff6ff;
# # # #     border-color: #bfdbfe;
# # # # }

# # # # .logs-status.success {
# # # #     background: #f0fdf4;
# # # #     border-color: #bbf7d0;
# # # # }

# # # # .logs-status.error {
# # # #     background: #fef2f2;
# # # #     border-color: #fecaca;
# # # # }

# # # # @keyframes pulse {
# # # #     0%, 100% { opacity: 1; }
# # # #     50% { opacity: 0.5; }
# # # # }

# # # # .pulse-dot {
# # # #     display: inline-block;
# # # #     width: 10px;
# # # #     height: 10px;
# # # #     background: #2563eb;
# # # #     border-radius: 50%;
# # # #     animation: pulse 2s infinite;
# # # #     margin-right: 8px;
# # # # }

# # # # </style>
# # # # """, unsafe_allow_html=True)


# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # #  SESSION STATE INIT
# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # for _k, _v in [
# # # #     ("logged_in",        False),
# # # #     ("login_error",      ""),
# # # #     ("pipeline_running", False),
# # # #     ("pipeline_logs",    []),
# # # #     ("pipeline_result",  None),
# # # #     ("pipeline_done",    False),
# # # #     ("current_page",     "dashboard"),
# # # # ]:
# # # #     if _k not in st.session_state:
# # # #         st.session_state[_k] = _v


# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # #  LOGIN PAGE
# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # if not st.session_state.logged_in:

# # # #     # Hide sidebar on login page
# # # #     st.markdown("""
# # # #     <script>
# # # #     document.querySelector('[data-testid="stSidebar"]') &&
# # # #     (document.querySelector('[data-testid="stSidebar"]').style.display = 'none');
# # # #     </script>
# # # #     """, unsafe_allow_html=True)

# # # #     # Center the login card
# # # #     _, col, _ = st.columns([1, 1.2, 1])

# # # #     with col:
# # # #         st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)

# # # #         st.markdown("""
# # # #         <div class="login-card">
# # # #           <div class="login-logo">🛡️ SecureITLab</div>
# # # #           <div class="login-subtitle">Pipeline Intelligence</div>
# # # #           <div class="login-divider"></div>
# # # #         </div>
# # # #         """, unsafe_allow_html=True)

# # # #         st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # #         with st.container():
# # # #             username = st.text_input(
# # # #                 "Username",
# # # #                 placeholder="Enter username",
# # # #                 key="login_username",
# # # #             )
# # # #             password = st.text_input(
# # # #                 "Password",
# # # #                 placeholder="Enter password",
# # # #                 type="password",
# # # #                 key="login_password",
# # # #             )

# # # #             st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # # #             login_btn = st.button(
# # # #                 "Sign In →",
# # # #                 use_container_width=True,
# # # #                 type="primary",
# # # #             )

# # # #             if login_btn:
# # # #                 if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
# # # #                     st.session_state.logged_in = True
# # # #                     st.session_state.login_error = ""
# # # #                     st.rerun()
# # # #                 else:
# # # #                     st.session_state.login_error = "Incorrect username or password."

# # # #             if st.session_state.login_error:
# # # #                 st.markdown(
# # # #                     f'<div class="login-error">⚠️ {st.session_state.login_error}</div>',
# # # #                     unsafe_allow_html=True,
# # # #                 )

# # # #         st.markdown(
# # # #             "<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>"
# # # #             "SecureITLab · Confidential</div>",
# # # #             unsafe_allow_html=True,
# # # #         )

# # # #     st.stop()


# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # #  MASTER CONTACTS EXCEL BUILDER
# # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # def build_master_contacts_excel(all_jobs: list):
# # # #     """Build master Excel with ALL contacts from ALL jobs with job details."""
# # # #     try:
# # # #         from openpyxl import Workbook
# # # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # # #         from openpyxl.utils import get_column_letter
# # # #     except ImportError:
# # # #         return None

# # # #     wb = Workbook()
# # # #     ws = wb.active
# # # #     ws.title = "All Contacts"

# # # #     NAVY  = "1E3A5F"
# # # #     BLUE  = "2563EB"
# # # #     LBLUE = "EFF6FF"
# # # #     GREY  = "F8FAFC"
# # # #     WHITE = "FFFFFF"

# # # #     pri_colors = {
# # # #         "Primary":   ("FEF2F2", "B91C1C"),
# # # #         "Secondary": ("FFFBEB", "B45309"),
# # # #         "Tertiary":  ("EFF6FF", "1D4ED8"),
# # # #         "General":   ("F5F3FF", "6D28D9"),
# # # #     }

# # # #     thin   = Side(border_style="thin", color="D9E2EC")
# # # #     border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # # #     # Row 1: Title
# # # #     ws.merge_cells("A1:K1")
# # # #     c = ws["A1"]
# # # #     c.value     = f"Master Contacts Export — All Jobs"
# # # #     c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
# # # #     c.fill      = PatternFill("solid", fgColor=NAVY)
# # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # #     ws.row_dimensions[1].height = 30

# # # #     # Row 2: Sub-title
# # # #     ws.merge_cells("A2:K2")
# # # #     c = ws["A2"]
# # # #     total_contacts = sum(len(j.get("contacts", [])) for j in all_jobs)
# # # #     c.value     = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(all_jobs)} jobs   ·   {total_contacts} total contacts"
# # # #     c.font      = Font(name="Arial", size=9, color="64748B")
# # # #     c.fill      = PatternFill("solid", fgColor="F4F7FB")
# # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # #     ws.row_dimensions[2].height = 18

# # # #     # Row 3: Headers
# # # #     headers    = ["#", "Job", "Company", "Country", "Priority", "Name", "Title / Role", "Email", "LinkedIn URL", "Source", "Job Score"]
# # # #     col_widths = [4,   28,   18,        12,       12,        24,     32,              34,      42,             18,       10]

# # # #     for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
# # # #         c = ws.cell(row=3, column=ci, value=h)
# # # #         c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
# # # #         c.fill      = PatternFill("solid", fgColor=BLUE)
# # # #         c.alignment = Alignment(horizontal="center", vertical="center")
# # # #         c.border    = border
# # # #         ws.column_dimensions[get_column_letter(ci)].width = w

# # # #     ws.row_dimensions[3].height = 22

# # # #     # Data rows
# # # #     ri = 4
# # # #     for job in all_jobs:
# # # #         company    = job.get("company", "")
# # # #         role       = job.get("role", "")
# # # #         country    = job.get("country", "?")
# # # #         contacts   = job.get("contacts", [])
# # # #         job_score  = job.get("opp_score", "—")
        
# # # #         for ci, contact in enumerate(contacts):
# # # #             prio           = contact.get("priority", "General")
# # # #             bg_hex, fg_hex = pri_colors.get(prio, (WHITE, "0F172A"))
# # # #             patterns       = contact.get("email_patterns", [])
# # # #             email_val      = contact.get("email") or (patterns[0] + "  (pattern)" if patterns else "")
# # # #             row_fill       = bg_hex if ri % 2 == 0 else GREY

# # # #             for col_idx, val in enumerate([
# # # #                 ri - 3,
# # # #                 role if ci == 0 else "",
# # # #                 company if ci == 0 else "",
# # # #                 country if ci == 0 else "",
# # # #                 prio,
# # # #                 contact.get("name", ""),
# # # #                 contact.get("title", ""),
# # # #                 email_val,
# # # #                 contact.get("linkedin_url", ""),
# # # #                 contact.get("source", ""),
# # # #                 str(job_score) if ci == 0 else "",
# # # #             ], 1):
# # # #                 cell = ws.cell(row=ri, column=col_idx, value=val)
# # # #                 cell.font      = Font(name="Arial", size=9,
# # # #                                       bold=(col_idx == 5),
# # # #                                       color=fg_hex if col_idx == 5 else "0F172A")
# # # #                 cell.fill      = PatternFill("solid", fgColor=row_fill)
# # # #                 cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [4, 7]))
# # # #                 cell.border    = border

# # # #             ws.row_dimensions[ri].height = 18
# # # #             ri += 1

# # # #     ws.freeze_panes   = "A4"
# # # #     ws.auto_filter.ref = f"A3:K{ri-1}"

# # # #     # Summary sheet
# # # #     ws2 = wb.create_sheet("Summary")
# # # #     ws2.merge_cells("A1:C1")
# # # #     t = ws2["A1"]
# # # #     t.value     = "Master Export Summary"
# # # #     t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
# # # #     t.fill      = PatternFill("solid", fgColor=NAVY)
# # # #     t.alignment = Alignment(horizontal="center")

# # # #     summary_stats = [
# # # #         ("Total Jobs",        len(all_jobs)),
# # # #         ("Total Contacts",    total_contacts),
# # # #         ("Primary Contacts",  sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "Primary")),
# # # #         ("Secondary",         sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "Secondary")),
# # # #         ("Tertiary",          sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "Tertiary")),
# # # #         ("General",           sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("priority") == "General")),
# # # #         ("With Email",        sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("email"))),
# # # #         ("With LinkedIn",     sum(1 for j in all_jobs for c in j.get("contacts", []) if c.get("linkedin_url"))),
# # # #         ("Generated",         datetime.now().strftime("%d %b %Y  %H:%M")),
# # # #     ]

# # # #     for r, (lbl, val) in enumerate(summary_stats, 2):
# # # #         lc = ws2.cell(row=r, column=1, value=lbl)
# # # #         vc = ws2.cell(row=r, column=2, value=val)
# # # #         bg = LBLUE if r % 2 == 0 else WHITE
# # # #         for cell in (lc, vc):
# # # #             cell.font   = Font(name="Arial", bold=(cell.column == 1), size=10)
# # # #             cell.fill   = PatternFill("solid", fgColor=bg)
# # # #             cell.border = border

# # # #     ws2.column_dimensions["A"].width = 20
# # # #     ws2.column_dimensions["B"].width = 30

# # # #     buf = io.BytesIO()
# # # #     wb.save(buf)
# # # #     buf.seek(0)
# # # #     return buf.getvalue()


# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # #  INDIVIDUAL JOB CONTACTS EXCEL
# # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # def build_contacts_excel(contacts: list, company: str, role: str):
# # # #     """Build a formatted Excel workbook from contacts. Returns bytes or None."""
# # # #     try:
# # # #         from openpyxl import Workbook
# # # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # # #         from openpyxl.utils import get_column_letter
# # # #     except ImportError:
# # # #         return None

# # # #     wb = Workbook()
# # # #     ws = wb.active
# # # #     ws.title = "Contacts"

# # # #     NAVY  = "1E3A5F"
# # # #     BLUE  = "2563EB"
# # # #     LBLUE = "EFF6FF"
# # # #     GREY  = "F8FAFC"
# # # #     WHITE = "FFFFFF"

# # # #     pri_colors = {
# # # #         "Primary":   ("FEF2F2", "B91C1C"),
# # # #         "Secondary": ("FFFBEB", "B45309"),
# # # #         "Tertiary":  ("EFF6FF", "1D4ED8"),
# # # #         "General":   ("F5F3FF", "6D28D9"),
# # # #     }

# # # #     thin   = Side(border_style="thin", color="D9E2EC")
# # # #     border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # # #     # Row 1: Title
# # # #     ws.merge_cells("A1:H1")
# # # #     c = ws["A1"]
# # # #     c.value     = f"Contacts Export  —  {company}  |  {role}"
# # # #     c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
# # # #     c.fill      = PatternFill("solid", fgColor=NAVY)
# # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # #     ws.row_dimensions[1].height = 30

# # # #     # Row 2: Sub-title
# # # #     ws.merge_cells("A2:H2")
# # # #     c = ws["A2"]
# # # #     c.value     = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(contacts)} contacts"
# # # #     c.font      = Font(name="Arial", size=9, color="64748B")
# # # #     c.fill      = PatternFill("solid", fgColor="F4F7FB")
# # # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # # #     ws.row_dimensions[2].height = 18

# # # #     # Row 3: Headers
# # # #     headers    = ["#", "Priority", "Name", "Title / Role", "Company", "Email", "LinkedIn URL", "Source"]
# # # #     col_widths = [4,    12,         24,     32,              22,        34,       42,              18]

# # # #     for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
# # # #         c = ws.cell(row=3, column=ci, value=h)
# # # #         c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
# # # #         c.fill      = PatternFill("solid", fgColor=BLUE)
# # # #         c.alignment = Alignment(horizontal="center", vertical="center")
# # # #         c.border    = border
# # # #         ws.column_dimensions[get_column_letter(ci)].width = w

# # # #     ws.row_dimensions[3].height = 22

# # # #     # Data rows
# # # #     for ri, ct in enumerate(contacts, start=4):
# # # #         prio           = ct.get("priority", "General")
# # # #         bg_hex, fg_hex = pri_colors.get(prio, (WHITE, "0F172A"))
# # # #         patterns       = ct.get("email_patterns", [])
# # # #         email_val      = ct.get("email") or (patterns[0] + "  (pattern)" if patterns else "")
# # # #         row_fill       = bg_hex if ri % 2 == 0 else GREY

# # # #         for ci, val in enumerate([
# # # #             ri - 3,
# # # #             prio,
# # # #             ct.get("name", ""),
# # # #             ct.get("title", ""),
# # # #             ct.get("company", ""),
# # # #             email_val,
# # # #             ct.get("linkedin_url", ""),
# # # #             ct.get("source", ""),
# # # #         ], 1):
# # # #             cell = ws.cell(row=ri, column=ci, value=val)
# # # #             cell.font      = Font(name="Arial", size=9,
# # # #                                   bold=(ci == 2),
# # # #                                   color=fg_hex if ci == 2 else "0F172A")
# # # #             cell.fill      = PatternFill("solid", fgColor=row_fill)
# # # #             cell.alignment = Alignment(vertical="center", wrap_text=(ci in [4, 7]))
# # # #             cell.border    = border

# # # #         ws.row_dimensions[ri].height = 18

# # # #     ws.freeze_panes   = "A4"
# # # #     ws.auto_filter.ref = f"A3:H{3 + len(contacts)}"

# # # #     # Summary sheet
# # # #     ws2 = wb.create_sheet("Summary")
# # # #     ws2.merge_cells("A1:C1")
# # # #     t = ws2["A1"]
# # # #     t.value     = "Export Summary"
# # # #     t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
# # # #     t.fill      = PatternFill("solid", fgColor=NAVY)
# # # #     t.alignment = Alignment(horizontal="center")

# # # #     for r, (lbl, val) in enumerate([
# # # #         ("Company",        company),
# # # #         ("Role",           role),
# # # #         ("Total Contacts", len(contacts)),
# # # #         ("Primary",        sum(1 for x in contacts if x.get("priority") == "Primary")),
# # # #         ("Secondary",      sum(1 for x in contacts if x.get("priority") == "Secondary")),
# # # #         ("Tertiary",       sum(1 for x in contacts if x.get("priority") == "Tertiary")),
# # # #         ("General",        sum(1 for x in contacts if x.get("priority") == "General")),
# # # #         ("With Email",     sum(1 for x in contacts if x.get("email"))),
# # # #         ("With LinkedIn",  sum(1 for x in contacts if x.get("linkedin_url"))),
# # # #         ("Generated",      datetime.now().strftime("%d %b %Y  %H:%M")),
# # # #     ], 2):
# # # #         lc = ws2.cell(row=r, column=1, value=lbl)
# # # #         vc = ws2.cell(row=r, column=2, value=val)
# # # #         bg = LBLUE if r % 2 == 0 else WHITE
# # # #         for cell in (lc, vc):
# # # #             cell.font   = Font(name="Arial", bold=(cell.column == 1), size=10)
# # # #             cell.fill   = PatternFill("solid", fgColor=bg)
# # # #             cell.border = border

# # # #     ws2.column_dimensions["A"].width = 20
# # # #     ws2.column_dimensions["B"].width = 30

# # # #     buf = io.BytesIO()
# # # #     wb.save(buf)
# # # #     buf.seek(0)
# # # #     return buf.getvalue()


# # # # # ── MongoDB helpers ───────────────────────────────────────────────────────────
# # # # @st.cache_resource
# # # # def get_db():
# # # #     URI = st.secrets.get("MONGO_URI", "mongodb://localhost:27017")
# # # #     DB  = st.secrets.get("MONGO_DB",  "secureitlab_job_pipeline")
# # # #     client = MongoClient(URI, serverSelectionTimeoutMS=6000)
# # # #     return client[DB]

# # # # @st.cache_data(ttl=60)
# # # # def load_all_jobs():
# # # #     db = get_db()
# # # #     return list(db.jobs.find({}, {
# # # #         "_id": 1, "company": 1, "role": 1, "job_number": 1,
# # # #         "opp_score": 1, "contacts_found": 1, "pipeline_ok": 1,
# # # #         "coverage_score": 1, "run_at": 1, "contact_domain": 1,
# # # #         "contacts": 1, "country": 1,
# # # #     }))

# # # # @st.cache_data(ttl=60)
# # # # def load_job(job_id):
# # # #     from bson import ObjectId
# # # #     return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# # # # # ── Render helpers ────────────────────────────────────────────────────────────
# # # # def badge(text, color="blue"):
# # # #     return f'<span class="badge badge-{color}">{text}</span>'

# # # # def safe_str(val, limit=300):
# # # #     if val is None: return "—"
# # # #     s = str(val)
# # # #     return s[:limit] + "…" if len(s) > limit else s

# # # # def as_dict(raw):
# # # #     if isinstance(raw, dict): return raw
# # # #     if isinstance(raw, list): return next((x for x in raw if isinstance(x, dict)), {})
# # # #     return {}

# # # # def render_json_pretty(data, title=""):
# # # #     if not data: return
# # # #     with st.expander(f"📄 Raw JSON — {title}", expanded=False):
# # # #         st.code(json.dumps(data, indent=2, default=str), language="json")

# # # # def render_qa_block(data, label):
# # # #     if not data:
# # # #         st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>', unsafe_allow_html=True)
# # # #         return
# # # #     data = as_dict(data)
# # # #     if not data: return
# # # #     passed    = data.get("passed") or data.get("Passed") or False
# # # #     rec       = data.get("recommendation") or data.get("Recommendation", "")
# # # #     issues    = data.get("issues") or data.get("Issues") or []
# # # #     checklist = data.get("checklist") or data.get("Checklist") or []
# # # #     color     = "green" if passed else "yellow"
# # # #     status    = "✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
# # # #     html = f"""
# # # #     <div class="sil-card sil-card-accent">
# # # #       <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
# # # #         <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
# # # #         {badge(status, color)}
# # # #       </div>"""
# # # #     if rec:
# # # #         html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
# # # #     if checklist:
# # # #         html += '<div style="font-size:.82rem;margin-bottom:.5rem">'
# # # #         for item in (checklist if isinstance(checklist, list) else [checklist]):
# # # #             if isinstance(item, dict):
# # # #                 ip = item.get("pass") or item.get("passed") or item.get("status","") == "pass"
# # # #                 nm = item.get("item") or item.get("name") or item.get("check","")
# # # #                 nt = item.get("reason") or item.get("note") or item.get("issue","")
# # # #                 html += f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
# # # #                 if nt: html += f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
# # # #                 html += '</div>'
# # # #         html += '</div>'
# # # #     if issues:
# # # #         html += '<div style="margin-top:.5rem">'
# # # #         for iss in (issues if isinstance(issues, list) else [issues])[:4]:
# # # #             txt = iss if isinstance(iss, str) else json.dumps(iss)
# # # #             html += f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
# # # #         html += '</div>'
# # # #     st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # def render_contacts(contacts, title="Contacts"):
# # # #     if not contacts: st.info("No contacts found for this job."); return
# # # #     pri_color = {"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
# # # #     st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>', unsafe_allow_html=True)
# # # #     cols = st.columns(2)
# # # #     for i, c in enumerate(contacts):
# # # #         col = cols[i % 2]
# # # #         prio = c.get("priority", "General")
# # # #         email = c.get("email", ""); li = c.get("linkedin_url", "")
# # # #         patterns = c.get("email_patterns", []); src = c.get("source", "")
# # # #         with col:
# # # #             html = f"""<div class="contact-card">
# # # #               <div style="display:flex;justify-content:space-between;align-items:flex-start">
# # # #                 <div>
# # # #                   <div class="contact-name">{c.get('name','—')}</div>
# # # #                   <div class="contact-title">{c.get('title','—')}</div>
# # # #                 </div>
# # # #                 {badge(prio, pri_color.get(prio,'blue'))}
# # # #               </div>"""
# # # #             if email:      html += f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
# # # #             elif patterns: html += f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
# # # #             if li:         html += f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
# # # #             if src:        html += f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
# # # #             st.markdown(html + '</div>', unsafe_allow_html=True)

# # # # def render_emails(emails_data):
# # # #     if not emails_data: st.info("No email data available."); return
# # # #     emails_data = as_dict(emails_data)
# # # #     if not emails_data: return
# # # #     variants = {}
# # # #     for k, v in emails_data.items():
# # # #         kl = k.lower().replace("_","").replace(" ","")
# # # #         if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl == "a":
# # # #             variants["Variant A — Hiring Manager"] = v
# # # #         elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl == "b":
# # # #             variants["Variant B — CISO / VP Level"] = v
# # # #         else:
# # # #             variants[k] = v
# # # #     for label, v in variants.items():
# # # #         st.markdown(f'<div class="section-label">✉️ {label}</div>', unsafe_allow_html=True)
# # # #         if isinstance(v, dict):
# # # #             subj = v.get("subject") or v.get("Subject","")
# # # #             body = v.get("body") or v.get("Body") or v.get("content","")
# # # #             if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>', unsafe_allow_html=True)
# # # #             if body: st.markdown(f'<div class="email-box">{body}</div>', unsafe_allow_html=True)
# # # #             else:    st.code(json.dumps(v, indent=2), language="json")
# # # #         elif isinstance(v, str):
# # # #             st.markdown(f'<div class="email-box">{v}</div>', unsafe_allow_html=True)
# # # #         st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

# # # # def render_service_mapping(data):
# # # #     if not data: st.info("No service mapping data."); return
# # # #     items = data if isinstance(data, list) else []
# # # #     if not items and isinstance(data, dict):
# # # #         for key in ("services","mappings","service_mapping","ServiceMapping","items"):
# # # #             if isinstance(data.get(key), list): items = data[key]; break
# # # #         if not items: items = [data]
# # # #     fit_colors = {"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
# # # #     for item in items:
# # # #         if not isinstance(item, dict): continue
# # # #         svc  = item.get("service") or item.get("Service") or item.get("name","")
# # # #         fit  = (item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
# # # #         why  = item.get("justification") or item.get("rationale") or item.get("why","")
# # # #         reqs = item.get("requirements_addressed") or item.get("requirements") or ""
# # # #         eng  = item.get("engagement_type") or item.get("engagement","")
# # # #         color = fit_colors.get(fit, "blue")
# # # #         html = f"""<div class="sil-card" style="margin-bottom:.75rem">
# # # #           <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
# # # #             <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
# # # #             {badge(fit or "MAPPED", color) if fit else ""}
# # # #           </div>"""
# # # #         if why:  html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
# # # #         if reqs:
# # # #             rs = ", ".join(reqs) if isinstance(reqs, list) else str(reqs)
# # # #             html += f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
# # # #         if eng:  html += f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
# # # #         st.markdown(html + '</div>', unsafe_allow_html=True)
# # # #     render_json_pretty(data, "Service Mapping")

# # # # def render_microplans(data):
# # # #     if not data: st.info("No micro-plan data."); return
# # # #     plans = data if isinstance(data, list) else []
# # # #     if not plans and isinstance(data, dict):
# # # #         for k in ("plans","micro_plans","microplans","top_3","improvements"):
# # # #             if isinstance(data.get(k), list): plans = data[k]; break
# # # #         if not plans: plans = [data]
# # # #     for i, plan in enumerate(plans[:3], 1):
# # # #         if not isinstance(plan, dict): continue
# # # #         title = plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
# # # #         weeks = plan.get("duration") or plan.get("timeline","")
# # # #         obj   = plan.get("objective") or plan.get("goal","")
# # # #         kpis  = plan.get("kpis") or plan.get("KPIs") or []
# # # #         tasks = plan.get("tasks") or plan.get("workstreams") or []
# # # #         with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}", expanded=(i==1)):
# # # #             if obj and obj != title: st.markdown(f"**Objective:** {obj}")
# # # #             if kpis:
# # # #                 st.markdown("**KPIs:**")
# # # #                 for kpi in (kpis if isinstance(kpis, list) else [kpis]): st.markdown(f"• {kpi}")
# # # #             if tasks:
# # # #                 st.markdown("**Tasks / Workstreams:**")
# # # #                 for t in (tasks if isinstance(tasks, list) else [tasks]):
# # # #                     if isinstance(t, dict):
# # # #                         tn = t.get("task") or t.get("name","")
# # # #                         te = t.get("effort") or t.get("duration","")
# # # #                         st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
# # # #                     else: st.markdown(f"• {t}")
# # # #             if plan:
# # # #                 st.code(json.dumps(plan, indent=2, default=str), language="json")

# # # # def render_deal_assurance(data):
# # # #     if not data: st.info("No deal assurance data."); return
# # # #     if not isinstance(data, dict): render_json_pretty(data, "Deal Assurance Pack"); return
# # # #     evp = (data.get("executive_value_proposition") or
# # # #            data.get("value_proposition") or data.get("ExecutiveValueProposition",""))
# # # #     if evp:
# # # #         st.markdown('<div class="section-label">💼 Executive Value Proposition</div>', unsafe_allow_html=True)
# # # #         st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7;color:var(--text)">{evp}</div>', unsafe_allow_html=True)
# # # #     caps = data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
# # # #     if caps:
# # # #         st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>', unsafe_allow_html=True)
# # # #         c1, c2 = st.columns(2)
# # # #         for i, cap in enumerate(caps if isinstance(caps, list) else [caps]):
# # # #             (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
# # # #     risk = data.get("risk_mitigation") or data.get("RiskMitigation","")
# # # #     if risk:
# # # #         st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>', unsafe_allow_html=True)
# # # #         if isinstance(risk, dict):
# # # #             for k, v in risk.items(): st.markdown(f"**{k}:** {v}")
# # # #         else: st.markdown(str(risk))
# # # #     render_json_pretty(data, "Deal Assurance Pack")


# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # #  LOGS PAGE (FIXED VERSION)
# # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # def render_logs_page():
# # # #     """Dedicated logs page with real-time updates."""
# # # #     st.markdown("""
# # # #     <div style="margin-bottom:2rem">
# # # #       <h1 style="font-family:'Syne',sans-serif;font-size:2.2rem;font-weight:800;color:#0f172a;margin:0">
# # # #         ⚙️ Pipeline Execution
# # # #       </h1>
# # # #       <p style="font-size:.95rem;color:#64748b;margin-top:.5rem">
# # # #         Live logs from the job scraper, deduplication, and AI agent pipeline
# # # #       </p>
# # # #     </div>
# # # #     """, unsafe_allow_html=True)

# # # #     # Status indicator
# # # #     if st.session_state.pipeline_running:
# # # #         st.markdown("""
# # # #         <div class="logs-status running">
# # # #             <div>
# # # #                 <div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.3rem">
# # # #                     <span class="pulse-dot"></span>
# # # #                     <span style="font-family:'Syne',sans-serif;font-weight:700;color:#1d4ed8">Pipeline Running</span>
# # # #                 </div>
# # # #                 <span style="font-size:.82rem;color:#64748b">Please wait while the pipeline processes jobs...</span>
# # # #             </div>
# # # #         </div>
# # # #         """, unsafe_allow_html=True)
# # # #     elif st.session_state.pipeline_done:
# # # #         result = st.session_state.pipeline_result or {}
# # # #         if result.get("success"):
# # # #             st.markdown(f"""
# # # #             <div class="logs-status success">
# # # #                 <div>
# # # #                     <div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
# # # #                         <span style="font-size:1.5rem">✅</span>
# # # #                         <span style="font-family:'Syne',sans-serif;font-weight:700;color:#15803d;font-size:1.1rem">Pipeline Completed Successfully</span>
# # # #                     </div>
# # # #                     <div style="font-size:.85rem;color:#15803d">
# # # #                         Processed {result.get('processed',0)} jobs • {result.get('new_jobs',0)} new found • {result.get('skipped_db',0)} already in database
# # # #                     </div>
# # # #                 </div>
# # # #             </div>
# # # #             """, unsafe_allow_html=True)
# # # #         else:
# # # #             error = result.get("error", "Unknown error")
# # # #             st.markdown(f"""
# # # #             <div class="logs-status error">
# # # #                 <div>
# # # #                     <div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
# # # #                         <span style="font-size:1.5rem">❌</span>
# # # #                         <span style="font-family:'Syne',sans-serif;font-weight:700;color:#b91c1c;font-size:1.1rem">Pipeline Failed</span>
# # # #                     </div>
# # # #                     <div style="font-size:.85rem;color:#b91c1c">
# # # #                         {error}
# # # #                     </div>
# # # #                 </div>
# # # #             </div>
# # # #             """, unsafe_allow_html=True)

# # # #     # Live logs container
# # # #     st.markdown('<div class="section-label" style="margin-top:2rem">📡 Live Execution Logs</div>', unsafe_allow_html=True)
    
# # # #     # Display logs
# # # #     logs = st.session_state.pipeline_logs or []
# # # #     if logs:
# # # #         log_text = "\n".join(logs[-200:])
# # # #         st.markdown(f'<div class="pipeline-log">{log_text}</div>', unsafe_allow_html=True)
# # # #     else:
# # # #         st.markdown('<div class="pipeline-log">Waiting for logs...</div>', unsafe_allow_html=True)

# # # #     # Auto-refresh if running
# # # #     if st.session_state.pipeline_running:
# # # #         st.markdown("""
# # # #         <script>
# # # #             setTimeout(function() {
# # # #                 window.location.reload();
# # # #             }, 1500);
# # # #         </script>
# # # #         """, unsafe_allow_html=True)

# # # #     # Buttons
# # # #     col1, col2, col3 = st.columns([2, 2, 1])
# # # #     with col1:
# # # #         if st.button("← Back to Dashboard", use_container_width=True):
# # # #             st.session_state.current_page = "dashboard"
# # # #             st.rerun()
# # # #     with col2:
# # # #         if st.button("🔄 Refresh Logs", use_container_width=True):
# # # #             st.rerun()
# # # #     with col3:
# # # #         if st.button("📋 Copy All", use_container_width=True):
# # # #             st.info("Logs copied (use your browser copy feature)")


# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # #  SIDEBAR
# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # with st.sidebar:
# # # #     st.markdown("""
# # # #     <div style="padding:.75rem 0 1.25rem">
# # # #       <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;
# # # #                   color:#2563eb">🛡️ SecureITLab</div>
# # # #       <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;
# # # #                   text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
# # # #     </div>
# # # #     """, unsafe_allow_html=True)

# # # #     # ── Logout button ─────────────────────────────────────────────────────────
# # # #     if st.button("🚪 Logout", use_container_width=True):
# # # #         st.session_state.logged_in = False
# # # #         st.session_state.login_error = ""
# # # #         st.rerun()

# # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # #     # ── Find Jobs button ──────────────────────────────────────────────────────
# # # #     st.markdown("**🔍 Find New Jobs**")
# # # #     st.caption(
# # # #         "Runs scraper → checks MongoDB for duplicates → "
# # # #         "runs AI pipeline only on NEW jobs → stores in MongoDB"
# # # #     )

# # # #     find_jobs_clicked = st.button(
# # # #         "🔍  Find Jobs",
# # # #         disabled=st.session_state.pipeline_running,
# # # #         use_container_width=True,
# # # #         type="primary",
# # # #     )

# # # #     if st.session_state.pipeline_running:
# # # #         st.markdown("""
# # # #         <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
# # # #                     padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;
# # # #                     font-size:.8rem;color:#1d4ed8">
# # # #           ⏳ Pipeline is running… 👉 View logs on Logs Page
# # # #         </div>""", unsafe_allow_html=True)

# # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # #     # ── Job list ──────────────────────────────────────────────────────────────
# # # #     try:
# # # #         all_jobs = load_all_jobs()
# # # #     except Exception as e:
# # # #         st.error(f"MongoDB connection failed: {e}")
# # # #         st.stop()

# # # #     if not all_jobs:
# # # #         st.warning("No jobs in MongoDB yet. Click **Find Jobs** to scrape and process.")
# # # #         st.stop()

# # # #     st.markdown(
# # # #         f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">'
# # # #         f'{len(all_jobs)} jobs in database</div>',
# # # #         unsafe_allow_html=True,
# # # #     )

# # # #     # ── MASTER CONTACTS EXPORT ────────────────────────────────────────────────
# # # #     st.markdown("**📋 Master Contacts Export**")
# # # #     st.caption("All contacts from ALL jobs in one Excel file with auto-updates")
    
# # # #     master_excel = build_master_contacts_excel(all_jobs)
# # # #     if master_excel:
# # # #         st.download_button(
# # # #             label="📥  Download All Contacts",
# # # #             data=master_excel,
# # # #             file_name=f"master_contacts_{datetime.now().strftime('%Y%m%d')}.xlsx",
# # # #             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# # # #             use_container_width=True,
# # # #         )
# # # #     else:
# # # #         st.warning("Run `pip install openpyxl` to enable Excel export.")

# # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # #     search   = st.text_input("🔍 Filter by company / role", placeholder="e.g. Bounteous")
# # # #     filtered = [j for j in all_jobs
# # # #                 if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

# # # #     def job_label(j):
# # # #         score = j.get("opp_score")
# # # #         s_str = f" [{score}/10]" if score else ""
# # # #         ok    = "✅" if j.get("pipeline_ok") else "❌"
# # # #         return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s_str}"

# # # #     if not filtered:
# # # #         st.warning("No matching jobs.")
# # # #         st.stop()

# # # #     sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
# # # #     sel_idx     = [job_label(j) for j in filtered].index(sel_label)
# # # #     selected_id = str(filtered[sel_idx]["_id"])

# # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # #     ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
# # # #     total_c  = sum(j.get("contacts_found",0) for j in all_jobs)
# # # #     st.markdown(f"""
# # # #     <div style="font-size:.75rem;color:#64748b">
# # # #       <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
# # # #       <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
# # # #     </div>""", unsafe_allow_html=True)

# # # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # #     if st.button("🔄 Refresh Data", use_container_width=True):
# # # #         st.cache_data.clear()
# # # #         st.rerun()


# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # #  FIND JOBS — background thread
# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # if find_jobs_clicked and not st.session_state.pipeline_running:
# # # #     with _thread_lock:
# # # #         _thread_logs.clear()
# # # #         _thread_result[0] = None
# # # #         _thread_done[0]   = False

# # # #     st.session_state.pipeline_running = True
# # # #     st.session_state.pipeline_done    = False
# # # #     st.session_state.pipeline_logs    = []
# # # #     st.session_state.pipeline_result  = None
# # # #     st.session_state.current_page     = "logs"
# # # #     st.cache_data.clear()

# # # #     def _run_pipeline_bg():
# # # #         try:
# # # #             import main as _main

# # # #             def _cb(msg: str):
# # # #                 with _thread_lock:
# # # #                     _thread_logs.append(msg)

# # # #             result = _main.run_pipeline(progress_callback=_cb)
# # # #             with _thread_lock:
# # # #                 _thread_result[0] = result
# # # #         except Exception as e:
# # # #             import traceback
# # # #             with _thread_lock:
# # # #                 _thread_logs.append(f"❌ Unexpected error: {e}")
# # # #                 _thread_logs.append(traceback.format_exc())
# # # #                 _thread_result[0] = {
# # # #                     "success": False, "error": str(e),
# # # #                     "scraped": 0, "new_jobs": 0, "skipped_db": 0, "processed": 0,
# # # #                 }
# # # #         finally:
# # # #             with _thread_lock:
# # # #                 _thread_done[0] = True

# # # #     threading.Thread(target=_run_pipeline_bg, daemon=True).start()
# # # #     st.rerun()

# # # # # ── Sync thread state → session_state ────────────────────────────────────────
# # # # with _thread_lock:
# # # #     if _thread_logs:
# # # #         st.session_state.pipeline_logs = list(_thread_logs)
# # # #     if _thread_done[0] and _thread_result[0] is not None:
# # # #         st.session_state.pipeline_result  = _thread_result[0]
# # # #         st.session_state.pipeline_running = False
# # # #         st.session_state.pipeline_done    = True


# # # # # ══════════════════════════════════════════════════════════════════════════════
# # # # #  PAGE ROUTING
# # # # # ══════════════════════════════════════════════════════════════════════════════

# # # # if st.session_state.current_page == "logs":
# # # #     render_logs_page()
# # # # else:
# # # #     # Original dashboard code - same as before
# # # #     with st.spinner("Loading job from MongoDB…"):
# # # #         job = load_job(selected_id)

# # # #     if not job:
# # # #         st.error("Could not load job document.")
# # # #         st.stop()

# # # #     company   = job.get("company",  "Unknown")
# # # #     role      = job.get("role",     "Unknown")
# # # #     opp_score = job.get("opp_score")
# # # #     p_ok      = job.get("pipeline_ok", False)
# # # #     p_min     = job.get("pipeline_min", "?")
# # # #     c_found   = job.get("contacts_found", 0)
# # # #     c_cov     = job.get("coverage_score")
# # # #     c_domain  = job.get("contact_domain","")
# # # #     run_at    = job.get("run_at","")

# # # #     # ── Header ────────────────────────────────────────────────────────────────────
# # # #     st.markdown(f"""
# # # #     <div style="margin-bottom:1.75rem">
# # # #       <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;
# # # #                   letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
# # # #         Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
# # # #       </div>
# # # #       <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;
# # # #                  color:#0f172a;margin:0;line-height:1.15">{role}</h1>
# # # #       <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
# # # #         @ <span style="color:#334155;font-weight:600">{company}</span>
# # # #         {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
# # # #       </div>
# # # #     </div>
# # # #     """, unsafe_allow_html=True)

# # # #     # ── Metric tiles ──────────────────────────────────────────────────────────────
# # # #     try:
# # # #         sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
# # # #         sc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # #     except Exception:
# # # #         sc = "#2563eb"

# # # #     st.markdown(f"""
# # # #     <div class="metric-row">
# # # #       <div class="metric-tile">
# # # #         <div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div>
# # # #         <div class="lbl">Opportunity Score</div>
# # # #       </div>
# # # #       <div class="metric-tile">
# # # #         <div class="val">{c_found}</div>
# # # #         <div class="lbl">Contacts Found</div>
# # # #       </div>
# # # #       <div class="metric-tile">
# # # #         <div class="val">{f"{c_cov}%" if c_cov else "—"}</div>
# # # #         <div class="lbl">Contact Coverage</div>
# # # #       </div>
# # # #       <div class="metric-tile">
# # # #         <div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">
# # # #           {'✅ OK' if p_ok else '❌ Failed'}
# # # #         </div>
# # # #         <div class="lbl">Pipeline ({p_min} min)</div>
# # # #       </div>
# # # #     </div>
# # # #     """, unsafe_allow_html=True)

# # # #     # ══════════════════════════════════════════════════════════════════════════════
# # # #     #  TABS
# # # #     # ══════════════════════════════════════════════════════════════════════════════
# # # #     tabs = st.tabs([
# # # #         "📋 Job & Enrichment",
# # # #         "🎯 Service Mapping",
# # # #         "🔍 Fit / Gap",
# # # #         "🛠️ Capability & Plans",
# # # #         "📦 Deal Assurance",
# # # #         "✉️ Outreach Emails",
# # # #         "👥 Contacts",
# # # #         "✅ QA Gates",
# # # #         "🗄️ Raw Data",
# # # #     ])

# # # #     # ── Tab 1: Job Research + Enrichment ─────────────────────────────────────
# # # #     with tabs[0]:
# # # #         col1, col2 = st.columns([1, 1])
# # # #         with col1:
# # # #             st.markdown('<div class="section-label">📄 Job Research</div>', unsafe_allow_html=True)
# # # #             jr = as_dict(job.get("agent_job_research") or {})
# # # #             if jr:
# # # #                 for label, keys in [
# # # #                     ("Job Role",        ["job_role","Job Role","role","title"]),
# # # #                     ("Company",         ["company_name","Company Name","company"]),
# # # #                     ("Location",        ["location","Location"]),
# # # #                     ("Organization URL",["organization_url","Organization URL","url"]),
# # # #                 ]:
# # # #                     val = next((jr.get(k) for k in keys if jr.get(k)), None)
# # # #                     if val: st.markdown(f"**{label}:** {val}")
# # # #                 desc = jr.get("job_description") or jr.get("Job Description","")
# # # #                 if desc:
# # # #                     st.markdown("**Job Description:**")
# # # #                     st.markdown(
# # # #                         f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;'
# # # #                         f'max-height:300px;overflow-y:auto;color:var(--text)">{desc[:2000]}</div>',
# # # #                         unsafe_allow_html=True,
# # # #                     )
# # # #                 render_json_pretty(jr, "Job Research")
# # # #             else:
# # # #                 st.info("No job research data.")
# # # #         with col2:
# # # #             st.markdown('<div class="section-label">🏢 Company Enrichment</div>', unsafe_allow_html=True)
# # # #             enr = as_dict(job.get("agent_enrichment") or {})
# # # #             if enr:
# # # #                 for label, keys in [
# # # #                     ("Industry",          ["industry","Industry"]),
# # # #                     ("Company Size",      ["company_size","size","Company Size"]),
# # # #                     ("Regulatory Env",    ["regulatory_environment","regulatory"]),
# # # #                     ("Certifications",    ["certifications","Certifications"]),
# # # #                     ("Tech Stack",        ["tech_stack","technology_stack"]),
# # # #                     ("Security Maturity", ["security_maturity","maturity"]),
# # # #                 ]:
# # # #                     val = next((enr.get(k) for k in keys if enr.get(k)), None)
# # # #                     if val:
# # # #                         if isinstance(val, list): val = ", ".join(str(v) for v in val)
# # # #                         st.markdown(f"**{label}:** {safe_str(val, 200)}")
# # # #                 render_json_pretty(enr, "Enrichment")
# # # #             else:
# # # #                 st.info("No enrichment data.")

# # # #     # ── Tab 2: Service Mapping ────────────────────────────────────────────────────
# # # #     with tabs[1]:
# # # #         st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>', unsafe_allow_html=True)
# # # #         render_service_mapping(job.get("agent_service_mapping"))

# # # #     # ── Tab 3: Fit / Gap Analysis ─────────────────────────────────────────────────
# # # #     with tabs[2]:
# # # #         fg = as_dict(job.get("agent_fit_gap") or {})
# # # #         if opp_score:
# # # #             try:
# # # #                 sn = float(str(opp_score).split("/")[0])
# # # #                 bc = "#16a34a" if sn >= 7 else "#f59e0b" if sn >= 5 else "#dc2626"
# # # #                 bp = int(sn / 10 * 100)
# # # #                 st.markdown(f"""
# # # #                 <div style="margin-bottom:1.5rem">
# # # #                   <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
# # # #                     <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">Opportunity Score</span>
# # # #                     <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
# # # #                   </div>
# # # #                   <div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%">
# # # #                     <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
# # # #                   </div>
# # # #                 </div>""", unsafe_allow_html=True)
# # # #             except Exception:
# # # #                 pass
# # # #         st.markdown('<div class="section-label">📊 Service Classifications</div>', unsafe_allow_html=True)
# # # #         services = []
# # # #         if isinstance(fg, dict):
# # # #             for k in ("services","classifications","service_classifications","items","fit_gap"):
# # # #                 v = fg.get(k)
# # # #                 if isinstance(v, list): services = v; break
# # # #             if not services and (fg.get("service") or fg.get("Service")):
# # # #                 services = [fg]
# # # #         elif isinstance(fg, list):
# # # #             services = fg
# # # #         if services:
# # # #             buckets = {"STRONG FIT": [], "PARTIAL FIT": [], "GAP": []}
# # # #             other   = []
# # # #             for s in services:
# # # #                 if not isinstance(s, dict): continue
# # # #                 fit = (s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
# # # #                 if "STRONG" in fit: buckets["STRONG FIT"].append(s)
# # # #                 elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
# # # #                 elif "GAP" in fit: buckets["GAP"].append(s)
# # # #                 else: other.append(s)
# # # #             c1, c2, c3 = st.columns(3)
# # # #             cm  = {"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
# # # #             bgm = {"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
# # # #             bdm = {"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
# # # #             for col, (fl, items) in zip([c1, c2, c3], buckets.items()):
# # # #                 col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>', unsafe_allow_html=True)
# # # #                 for s in items:
# # # #                     svc  = s.get("service") or s.get("Service") or s.get("name","")
# # # #                     just = s.get("justification") or s.get("reason","")
# # # #                     col.markdown(
# # # #                         f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem">'
# # # #                         f'<div style="font-weight:600;color:#0f172a">{svc}</div>'
# # # #                         f'<div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>',
# # # #                         unsafe_allow_html=True,
# # # #                     )
# # # #         elif fg:
# # # #             st.json(fg)
# # # #         else:
# # # #             st.info("No fit/gap data.")
# # # #         render_json_pretty(job.get("agent_fit_gap"), "Fit/Gap Analysis")

# # # #     # ── Tab 4: Capability + Micro-plans ──────────────────────────────────────
# # # #     with tabs[3]:
# # # #         col1, col2 = st.columns([1, 1])
# # # #         with col1:
# # # #             st.markdown('<div class="section-label">🔧 Capability Improvements</div>', unsafe_allow_html=True)
# # # #             cap = job.get("agent_capability") or {}
# # # #             items_cap = cap if isinstance(cap, list) else []
# # # #             if not items_cap and isinstance(cap, dict):
# # # #                 for k in ("improvements","recommendations","capabilities","items"):
# # # #                     v = cap.get(k)
# # # #                     if isinstance(v, list): items_cap = v; break
# # # #                 if not items_cap: items_cap = [cap]
# # # #             for item in items_cap:
# # # #                 if not isinstance(item, dict): continue
# # # #                 title  = item.get("title") or item.get("gap") or item.get("service","")
# # # #                 rec    = item.get("recommendation") or item.get("steps","")
# # # #                 effort = item.get("build_effort") or item.get("effort","")
# # # #                 demand = item.get("market_demand") or item.get("priority","")
# # # #                 st.markdown(f"""
# # # #                 <div class="sil-card" style="margin-bottom:.6rem">
# # # #                   <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem;color:var(--text)">{title}</div>
# # # #                   <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec, 250)}</div>
# # # #                   {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
# # # #                 </div>""", unsafe_allow_html=True)
# # # #             if not items_cap:
# # # #                 render_json_pretty(cap, "Capability Improvement")
# # # #         with col2:
# # # #             st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>', unsafe_allow_html=True)
# # # #             render_microplans(job.get("agent_microplans"))

# # # #     # ── Tab 5: Deal Assurance Pack ────────────────────────────────────────────
# # # #     with tabs[4]:
# # # #         render_deal_assurance(job.get("agent_deal_assurance"))

# # # #     # ── Tab 6: Outreach Emails ────────────────────────────────────────────────
# # # #     with tabs[5]:
# # # #         st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>', unsafe_allow_html=True)
# # # #         emails_src = job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
# # # #         oq = as_dict(job.get("agent_outreach_qa") or {})
# # # #         improved = (oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
# # # #         if improved:
# # # #             st.info("⚡ Showing QA-improved versions where available")
# # # #             render_emails(improved)
# # # #             with st.expander("📬 Original (pre-QA) versions"):
# # # #                 render_emails(emails_src)
# # # #         else:
# # # #             render_emails(emails_src)

# # # #     # ── Tab 7: Contacts ───────────────────────────────────────────────────────
# # # #     with tabs[6]:
# # # #         contacts        = job.get("contacts") or []
# # # #         contact_sources = job.get("contact_sources") or []
# # # #         pri = [c for c in contacts if c.get("priority") == "Primary"]
# # # #         sec = [c for c in contacts if c.get("priority") == "Secondary"]
# # # #         ter = [c for c in contacts if c.get("priority") == "Tertiary"]
# # # #         gen = [c for c in contacts if c.get("priority") == "General"]
# # # #         st.markdown(f"""
# # # #         <div class="metric-row" style="margin-bottom:1.5rem">
# # # #           <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
# # # #           <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
# # # #           <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
# # # #           <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
# # # #         </div>""", unsafe_allow_html=True)
# # # #         if contact_sources:
# # # #             st.markdown('Sources: ' + " ".join(badge(s,"blue") for s in contact_sources), unsafe_allow_html=True)
# # # #             st.markdown("")
# # # #         missing = job.get("missing_roles") or []
# # # #         if missing:
# # # #             st.markdown('⚠️ Missing roles: ' + " ".join(badge(r,"red") for r in missing), unsafe_allow_html=True)
# # # #             st.markdown("")
# # # #         if contacts:
# # # #             excel_bytes = build_contacts_excel(contacts, company, role)
# # # #             if excel_bytes:
# # # #                 safe_co = re.sub(r'[^a-z0-9]', '_', company.lower())[:20]
# # # #                 fname   = f"contacts_{safe_co}_{datetime.now().strftime('%Y%m%d')}.xlsx"
# # # #                 btn_col, _ = st.columns([1, 3])
# # # #                 with btn_col:
# # # #                     st.download_button(
# # # #                         label="📥  Download Contacts (.xlsx)",
# # # #                         data=excel_bytes,
# # # #                         file_name=fname,
# # # #                         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# # # #                         use_container_width=True,
# # # #                         type="primary",
# # # #                     )
# # # #             else:
# # # #                 st.warning("Run `pip install openpyxl` to enable Excel export.")

# # # #             st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # # #             pri_filter = st.multiselect(
# # # #                 "Filter by priority",
# # # #                 ["Primary","Secondary","Tertiary","General"],
# # # #                 default=["Primary","Secondary","Tertiary","General"],
# # # #             )
# # # #             shown = [c for c in contacts if c.get("priority","General") in pri_filter]
# # # #             render_contacts(shown, f"Contacts ({len(shown)} shown)")
# # # #             agent_contacts = job.get("agent_prospect_contacts") or {}
# # # #             if agent_contacts:
# # # #                 with st.expander("🤖 CrewAI Agent's Contact Search"):
# # # #                     if isinstance(agent_contacts, dict):
# # # #                         ac_list = agent_contacts.get("contacts") or []
# # # #                         if ac_list: render_contacts(ac_list, "Agent Contacts")
# # # #                         else:       st.json(agent_contacts)
# # # #                     else:           st.json(agent_contacts)
# # # #         else:
# # # #             st.info("No contacts found for this job.")

# # # #     # ── Tab 8: QA Gates ───────────────────────────────────────────────────────
# # # #     with tabs[7]:
# # # #         st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>', unsafe_allow_html=True)
# # # #         col1, col2 = st.columns(2)
# # # #         for i, (label, key) in enumerate([
# # # #             ("Research QA",       "agent_research_qa"),
# # # #             ("Service Mapping QA","agent_mapping_qa"),
# # # #             ("Deal Assurance QA", "agent_assurance_qa"),
# # # #             ("Outreach Email QA", "agent_outreach_qa"),
# # # #         ]):
# # # #             with (col1 if i % 2 == 0 else col2):
# # # #                 render_qa_block(job.get(key), label)
# # # #         st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # #         st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>', unsafe_allow_html=True)
# # # #         enf = as_dict(job.get("agent_prospect_enforcer") or {})
# # # #         if enf:
# # # #             cov  = enf.get("coverage_score","?")
# # # #             miss = enf.get("missing_roles",[])
# # # #             note = enf.get("note","")
# # # #             ec   = enf.get("contacts",[])
# # # #             x1, x2, x3 = st.columns(3)
# # # #             x1.metric("Coverage Score",    f"{cov}%")
# # # #             x2.metric("Missing Roles",     len(miss))
# # # #             x3.metric("Contacts Verified", len(ec))
# # # #             if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
# # # #             if note: st.caption(note)
# # # #         else:
# # # #             st.info("No enforcer data.")

# # # #     # ── Tab 9: Raw Data ───────────────────────────────────────────────────────
# # # #     with tabs[8]:
# # # #         st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>', unsafe_allow_html=True)
# # # #         st.caption("All fields stored in the `jobs` collection for this document.")
# # # #         rows = []
# # # #         for k, v in job.items():
# # # #             if k == "_id": continue
# # # #             rows.append({"Field": k, "Type": type(v).__name__, "Len": len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
# # # #         hc1, hc2, hc3 = st.columns([3,1,1])
# # # #         hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
# # # #         for r in rows:
# # # #             rc1, rc2, rc3 = st.columns([3,1,1])
# # # #             rc1.code(r["Field"], language=None)
# # # #             rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>', unsafe_allow_html=True)
# # # #             rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>',  unsafe_allow_html=True)
# # # #         st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # # #         for label, key in [
# # # #             ("Job Research","agent_job_research"), ("Enrichment","agent_enrichment"),
# # # #             ("Service Mapping","agent_service_mapping"), ("Fit/Gap Analysis","agent_fit_gap"),
# # # #             ("Capability","agent_capability"), ("Micro-Plans","agent_microplans"),
# # # #             ("Deal Assurance","agent_deal_assurance"), ("Outreach Emails","agent_outreach_emails"),
# # # #             ("Prospect Contacts","agent_prospect_contacts"), ("Prospect Enforcer","agent_prospect_enforcer"),
# # # #             ("Research QA","agent_research_qa"), ("Mapping QA","agent_mapping_qa"),
# # # #             ("Assurance QA","agent_assurance_qa"), ("Outreach QA","agent_outreach_qa"),
# # # #             ("Contacts (5-source)","contacts"),
# # # #         ]:
# # # #             data = job.get(key)
# # # #             if data:
# # # #                 with st.expander(f"📄 {label}"):
# # # #                     st.code(json.dumps(data, indent=2, default=str), language="json")
























# # # """
# # # ╔══════════════════════════════════════════════════════════╗
# # # ║   SecureITLab Pipeline Dashboard — Streamlit             ║
# # # ║   WITH MASTER CONTACTS AGGREGATION & AUTO-UPDATE         ║
# # # ║   WITH 12-HOUR AUTO-SCHEDULER STATUS IN SIDEBAR          ║
# # # ║   Reads from MongoDB → secureitlab_job_pipeline          ║
# # # ╠══════════════════════════════════════════════════════════╣
# # # ║  Install: pip install streamlit pymongo python-dotenv    ║
# # # ║  Run:     streamlit run streamlit_dashboard.py           ║
# # # ╚══════════════════════════════════════════════════════════╝
# # # """

# # # import io
# # # import re
# # # import streamlit as st
# # # from pymongo import MongoClient
# # # import json
# # # import threading
# # # import time
# # # import logging
# # # from datetime import datetime, timezone
# # # from io import StringIO

# # # # ── Thread-safe shared state ──────────────────────────────────────────────────
# # # _thread_logs   = []
# # # _thread_result = [None]
# # # _thread_done   = [False]
# # # _thread_lock   = threading.Lock()

# # # # Capture all logs
# # # _log_capture = StringIO()
# # # _log_handler = logging.StreamHandler(_log_capture)
# # # _log_handler.setLevel(logging.INFO)
# # # _log_formatter = logging.Formatter('%(asctime)s | %(levelname)s | %(message)s')
# # # _log_handler.setFormatter(_log_formatter)

# # # # ── Page config ───────────────────────────────────────────────────────────────
# # # st.set_page_config(
# # #     page_title="SecureITLab Pipeline",
# # #     page_icon="🛡️",
# # #     layout="wide",
# # #     initial_sidebar_state="expanded",
# # # )

# # # # ── LOGIN CREDENTIALS (change these) ─────────────────────────────────────────
# # # LOGIN_USERNAME = "admin"
# # # LOGIN_PASSWORD = "secureitlab2024"

# # # # ══════════════════════════════════════════════════════════════════════════════
# # # #  GLOBAL CSS
# # # # ══════════════════════════════════════════════════════════════════════════════
# # # st.markdown("""
# # # <style>

# # # @import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');

# # # :root {
# # #     --bg:        #f4f7fb;
# # #     --surface:   #ffffff;
# # #     --surface2:  #eef2f7;
# # #     --border:    #d9e2ec;
# # #     --accent:    #2563eb;
# # #     --accent2:   #7c3aed;
# # #     --green:     #16a34a;
# # #     --yellow:    #f59e0b;
# # #     --red:       #dc2626;
# # #     --text:      #0f172a;
# # #     --muted:     #64748b;
# # # }

# # # html, body, [class*="css"] {
# # #     background-color: var(--bg) !important;
# # #     color: var(--text) !important;
# # #     font-family: 'DM Sans', sans-serif !important;
# # # }

# # # .login-wrap { display:flex; align-items:center; justify-content:center; min-height:80vh; }
# # # .login-card {
# # #     background:var(--surface); border:1px solid var(--border);
# # #     border-radius:20px; padding:3rem 3.5rem; width:100%; max-width:420px;
# # #     box-shadow:0 20px 60px rgba(37,99,235,0.08); text-align:center;
# # # }
# # # .login-logo { font-family:'Syne',sans-serif; font-size:1.6rem; font-weight:800; color:var(--accent); margin-bottom:.25rem; }
# # # .login-subtitle { font-size:.75rem; color:var(--muted); letter-spacing:.12em; text-transform:uppercase; margin-bottom:2.5rem; }
# # # .login-error { background:#fef2f2; border:1px solid #fecaca; border-radius:8px; padding:.65rem 1rem; color:#b91c1c; font-size:.85rem; margin-top:1rem; }
# # # .login-divider { width:40px; height:3px; background:linear-gradient(90deg,var(--accent),var(--accent2)); border-radius:2px; margin:0 auto 2rem; }

# # # [data-testid="stSidebar"] { background:var(--surface) !important; border-right:1px solid var(--border) !important; }
# # # [data-testid="stSidebar"] * { color:var(--text) !important; }
# # # .main .block-container { padding:2rem 2rem 3rem !important; }
# # # h1,h2,h3,h4 { font-family:'Syne',sans-serif !important; color:var(--text) !important; }

# # # .sil-card {
# # #     background:var(--surface); border:1px solid var(--border);
# # #     border-radius:14px; padding:1.25rem 1.5rem; margin-bottom:1rem; transition:all 0.25s ease;
# # # }
# # # .sil-card:hover { transform:translateY(-2px); box-shadow:0 8px 22px rgba(0,0,0,0.05); }
# # # .sil-card-accent { border-left:4px solid var(--accent); }

# # # .metric-row { display:flex; gap:1rem; flex-wrap:wrap; margin-bottom:1.5rem; }
# # # .metric-tile {
# # #     flex:1; min-width:140px; background:var(--surface2); border:1px solid var(--border);
# # #     border-radius:12px; padding:1rem; text-align:center; transition:all .25s ease;
# # # }
# # # .metric-tile:hover { transform:translateY(-3px); box-shadow:0 10px 24px rgba(0,0,0,0.06); }
# # # .metric-tile .val { font-family:'Syne',sans-serif; font-size:2rem; font-weight:800; color:var(--accent); }
# # # .metric-tile .lbl { font-size:.72rem; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; }

# # # .badge { padding:.25rem .7rem; border-radius:20px; font-size:.72rem; font-weight:600; font-family:'DM Mono',monospace; }
# # # .badge-green  { background:#ecfdf5; color:#15803d; }
# # # .badge-yellow { background:#fffbeb; color:#b45309; }
# # # .badge-red    { background:#fef2f2; color:#b91c1c; }
# # # .badge-blue   { background:#eff6ff; color:#1d4ed8; }
# # # .badge-purple { background:#f5f3ff; color:#6d28d9; }

# # # .contact-card { background:var(--surface2); border:1px solid var(--border); border-radius:10px; padding:1rem; margin-bottom:.8rem; }
# # # .contact-name  { font-family:'Syne',sans-serif; font-weight:700; color:var(--text); }
# # # .contact-title { color:var(--muted); font-size:.85rem; }
# # # .contact-email { font-family:'DM Mono',monospace; color:var(--accent); font-size:.8rem; }

# # # .email-box {
# # #     background:#f8fafc; border:1px solid var(--border); border-radius:10px;
# # #     padding:1rem 1.25rem; font-size:.9rem; line-height:1.65; white-space:pre-wrap; color:var(--text);
# # # }
# # # .email-subject { font-family:'Syne',sans-serif; font-weight:700; color:var(--accent); margin-bottom:.5rem; }

# # # .section-label {
# # #     font-family:'DM Mono',monospace; font-size:.72rem;
# # #     text-transform:uppercase; letter-spacing:.12em; color:var(--accent); margin-bottom:.6rem;
# # # }
# # # .sil-divider { border-top:1px solid var(--border); margin:1rem 0; }

# # # [data-testid="stExpander"] { background:var(--surface) !important; border:1px solid var(--border) !important; border-radius:10px !important; }
# # # [data-testid="stSelectbox"] > div, [data-testid="stMultiSelect"] > div { background:var(--surface2) !important; border-color:var(--border) !important; }
# # # [data-testid="stTabs"] button { font-family:'Syne',sans-serif !important; font-weight:600 !important; }
# # # ::-webkit-scrollbar { width:6px; }
# # # ::-webkit-scrollbar-thumb { background:var(--border); border-radius:3px; }

# # # .pipeline-log {
# # #     background:#0f172a; color:#10b981; border-radius:10px; padding:1.5rem;
# # #     font-family:'DM Mono',monospace; font-size:.8rem; line-height:1.8; max-height:700px;
# # #     overflow-y:auto; white-space:pre-wrap; word-break:break-word; border:1px solid #1e293b;
# # # }
# # # .fit-box { border-radius:8px; padding:.75rem; margin-bottom:.5rem; font-size:.85rem; }
# # # .hide-sidebar [data-testid="stSidebar"] { display:none !important; }
# # # .hide-sidebar .main .block-container { max-width:480px; margin:0 auto; }

# # # .logs-container { display:flex; flex-direction:column; gap:1rem; }
# # # .logs-status {
# # #     display:flex; gap:1rem; justify-content:space-between; align-items:center;
# # #     margin-bottom:1.5rem; padding:1rem; background:var(--surface2);
# # #     border-radius:10px; border:1px solid var(--border);
# # # }
# # # .logs-status.running  { background:#eff6ff; border-color:#bfdbfe; }
# # # .logs-status.success  { background:#f0fdf4; border-color:#bbf7d0; }
# # # .logs-status.error    { background:#fef2f2; border-color:#fecaca; }

# # # @keyframes pulse { 0%,100% { opacity:1; } 50% { opacity:0.5; } }
# # # .pulse-dot {
# # #     display:inline-block; width:10px; height:10px; background:#2563eb;
# # #     border-radius:50%; animation:pulse 2s infinite; margin-right:8px;
# # # }

# # # /* ── Scheduler badge ──────────────────────────────────────── */
# # # .sched-on {
# # #     background:#f0fdf4; border:1px solid #bbf7d0; border-radius:8px;
# # #     padding:.65rem .9rem; margin-top:.5rem; font-size:.8rem; color:#15803d;
# # #     line-height:1.6;
# # # }
# # # .sched-paused {
# # #     background:#fef9c3; border:1px solid #fde68a; border-radius:8px;
# # #     padding:.65rem .9rem; margin-top:.5rem; font-size:.8rem; color:#92400e;
# # #     line-height:1.6;
# # # }

# # # </style>
# # # """, unsafe_allow_html=True)


# # # # ══════════════════════════════════════════════════════════════════════════════
# # # #  SESSION STATE INIT
# # # # ══════════════════════════════════════════════════════════════════════════════
# # # for _k, _v in [
# # #     ("logged_in",        False),
# # #     ("login_error",      ""),
# # #     ("pipeline_running", False),
# # #     ("pipeline_logs",    []),
# # #     ("pipeline_result",  None),
# # #     ("pipeline_done",    False),
# # #     ("current_page",     "dashboard"),
# # # ]:
# # #     if _k not in st.session_state:
# # #         st.session_state[_k] = _v


# # # # ══════════════════════════════════════════════════════════════════════════════
# # # #  LOGIN PAGE
# # # # ══════════════════════════════════════════════════════════════════════════════
# # # if not st.session_state.logged_in:

# # #     st.markdown("""
# # #     <script>
# # #     document.querySelector('[data-testid="stSidebar"]') &&
# # #     (document.querySelector('[data-testid="stSidebar"]').style.display = 'none');
# # #     </script>
# # #     """, unsafe_allow_html=True)

# # #     _, col, _ = st.columns([1, 1.2, 1])
# # #     with col:
# # #         st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)
# # #         st.markdown("""
# # #         <div class="login-card">
# # #           <div class="login-logo">🛡️ SecureITLab</div>
# # #           <div class="login-subtitle">Pipeline Intelligence</div>
# # #           <div class="login-divider"></div>
# # #         </div>
# # #         """, unsafe_allow_html=True)
# # #         st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

# # #         with st.container():
# # #             username = st.text_input("Username", placeholder="Enter username", key="login_username")
# # #             password = st.text_input("Password", placeholder="Enter password", type="password", key="login_password")
# # #             st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)
# # #             login_btn = st.button("Sign In →", use_container_width=True, type="primary")

# # #             if login_btn:
# # #                 if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
# # #                     st.session_state.logged_in   = True
# # #                     st.session_state.login_error = ""
# # #                     st.rerun()
# # #                 else:
# # #                     st.session_state.login_error = "Incorrect username or password."

# # #             if st.session_state.login_error:
# # #                 st.markdown(f'<div class="login-error">⚠️ {st.session_state.login_error}</div>', unsafe_allow_html=True)

# # #         st.markdown("<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>SecureITLab · Confidential</div>", unsafe_allow_html=True)

# # #     st.stop()


# # # # ══════════════════════════════════════════════════════════════════════════════
# # # #  MASTER CONTACTS EXCEL BUILDER
# # # # ══════════════════════════════════════════════════════════════════════════════

# # # def build_master_contacts_excel(all_jobs: list):
# # #     try:
# # #         from openpyxl import Workbook
# # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # #         from openpyxl.utils import get_column_letter
# # #     except ImportError:
# # #         return None

# # #     wb = Workbook()
# # #     ws = wb.active
# # #     ws.title = "All Contacts"

# # #     NAVY  = "1E3A5F"; BLUE = "2563EB"; LBLUE = "EFF6FF"; GREY = "F8FAFC"; WHITE = "FFFFFF"
# # #     pri_colors = {
# # #         "Primary":   ("FEF2F2","B91C1C"), "Secondary": ("FFFBEB","B45309"),
# # #         "Tertiary":  ("EFF6FF","1D4ED8"), "General":   ("F5F3FF","6D28D9"),
# # #     }
# # #     thin = Side(border_style="thin", color="D9E2EC")
# # #     border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # #     ws.merge_cells("A1:K1")
# # #     c = ws["A1"]; c.value = "Master Contacts Export — All Jobs"
# # #     c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
# # #     c.fill = PatternFill("solid", fgColor=NAVY)
# # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # #     ws.row_dimensions[1].height = 30

# # #     ws.merge_cells("A2:K2")
# # #     c = ws["A2"]
# # #     total_contacts = sum(len(j.get("contacts", [])) for j in all_jobs)
# # #     c.value = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(all_jobs)} jobs   ·   {total_contacts} total contacts"
# # #     c.font = Font(name="Arial", size=9, color="64748B")
# # #     c.fill = PatternFill("solid", fgColor="F4F7FB")
# # #     c.alignment = Alignment(horizontal="center", vertical="center")
# # #     ws.row_dimensions[2].height = 18

# # #     headers    = ["#","Job","Company","Country","Priority","Name","Title / Role","Email","LinkedIn URL","Source","Job Score"]
# # #     col_widths = [4,  28,   18,       12,       12,       24,    32,            34,     42,            18,      10]
# # #     for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
# # #         c = ws.cell(row=3, column=ci, value=h)
# # #         c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
# # #         c.fill = PatternFill("solid", fgColor=BLUE)
# # #         c.alignment = Alignment(horizontal="center", vertical="center")
# # #         c.border = border
# # #         ws.column_dimensions[get_column_letter(ci)].width = w
# # #     ws.row_dimensions[3].height = 22

# # #     ri = 4
# # #     for job in all_jobs:
# # #         company = job.get("company",""); role = job.get("role","")
# # #         country = job.get("country","?"); contacts = job.get("contacts",[])
# # #         job_score = job.get("opp_score","—")
# # #         for ci, contact in enumerate(contacts):
# # #             prio = contact.get("priority","General")
# # #             bg_hex, fg_hex = pri_colors.get(prio,(WHITE,"0F172A"))
# # #             patterns = contact.get("email_patterns",[])
# # #             email_val = contact.get("email") or (patterns[0]+"  (pattern)" if patterns else "")
# # #             row_fill = bg_hex if ri%2==0 else GREY
# # #             for col_idx, val in enumerate([
# # #                 ri-3, role if ci==0 else "", company if ci==0 else "",
# # #                 country if ci==0 else "", prio, contact.get("name",""),
# # #                 contact.get("title",""), email_val, contact.get("linkedin_url",""),
# # #                 contact.get("source",""), str(job_score) if ci==0 else "",
# # #             ], 1):
# # #                 cell = ws.cell(row=ri, column=col_idx, value=val)
# # #                 cell.font = Font(name="Arial", size=9, bold=(col_idx==5), color=fg_hex if col_idx==5 else "0F172A")
# # #                 cell.fill = PatternFill("solid", fgColor=row_fill)
# # #                 cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [4,7]))
# # #                 cell.border = border
# # #             ws.row_dimensions[ri].height = 18
# # #             ri += 1

# # #     ws.freeze_panes = "A4"
# # #     ws.auto_filter.ref = f"A3:K{ri-1}"

# # #     ws2 = wb.create_sheet("Summary")
# # #     ws2.merge_cells("A1:C1"); t = ws2["A1"]
# # #     t.value = "Master Export Summary"
# # #     t.font = Font(name="Arial", bold=True, size=12, color=WHITE)
# # #     t.fill = PatternFill("solid", fgColor=NAVY); t.alignment = Alignment(horizontal="center")
# # #     for r, (lbl, val) in enumerate([
# # #         ("Total Jobs",len(all_jobs)), ("Total Contacts",total_contacts),
# # #         ("Primary",  sum(1 for j in all_jobs for c in j.get("contacts",[]) if c.get("priority")=="Primary")),
# # #         ("Secondary",sum(1 for j in all_jobs for c in j.get("contacts",[]) if c.get("priority")=="Secondary")),
# # #         ("Tertiary", sum(1 for j in all_jobs for c in j.get("contacts",[]) if c.get("priority")=="Tertiary")),
# # #         ("General",  sum(1 for j in all_jobs for c in j.get("contacts",[]) if c.get("priority")=="General")),
# # #         ("With Email",   sum(1 for j in all_jobs for c in j.get("contacts",[]) if c.get("email"))),
# # #         ("With LinkedIn",sum(1 for j in all_jobs for c in j.get("contacts",[]) if c.get("linkedin_url"))),
# # #         ("Generated",datetime.now().strftime("%d %b %Y  %H:%M")),
# # #     ], 2):
# # #         lc = ws2.cell(row=r,column=1,value=lbl); vc = ws2.cell(row=r,column=2,value=val)
# # #         bg = LBLUE if r%2==0 else WHITE
# # #         for cell in (lc,vc):
# # #             cell.font=Font(name="Arial",bold=(cell.column==1),size=10)
# # #             cell.fill=PatternFill("solid",fgColor=bg); cell.border=border
# # #     ws2.column_dimensions["A"].width=20; ws2.column_dimensions["B"].width=30

# # #     buf = io.BytesIO(); wb.save(buf); buf.seek(0)
# # #     return buf.getvalue()


# # # # ══════════════════════════════════════════════════════════════════════════════
# # # #  INDIVIDUAL JOB CONTACTS EXCEL
# # # # ══════════════════════════════════════════════════════════════════════════════

# # # def build_contacts_excel(contacts: list, company: str, role: str):
# # #     try:
# # #         from openpyxl import Workbook
# # #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # #         from openpyxl.utils import get_column_letter
# # #     except ImportError:
# # #         return None

# # #     wb = Workbook(); ws = wb.active; ws.title = "Contacts"
# # #     NAVY="1E3A5F"; BLUE="2563EB"; LBLUE="EFF6FF"; GREY="F8FAFC"; WHITE="FFFFFF"
# # #     pri_colors = {
# # #         "Primary":("FEF2F2","B91C1C"),"Secondary":("FFFBEB","B45309"),
# # #         "Tertiary":("EFF6FF","1D4ED8"),"General":("F5F3FF","6D28D9"),
# # #     }
# # #     thin=Side(border_style="thin",color="D9E2EC"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

# # #     ws.merge_cells("A1:H1"); c=ws["A1"]
# # #     c.value=f"Contacts Export  —  {company}  |  {role}"
# # #     c.font=Font(name="Arial",bold=True,size=13,color=WHITE)
# # #     c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment(horizontal="center",vertical="center")
# # #     ws.row_dimensions[1].height=30

# # #     ws.merge_cells("A2:H2"); c=ws["A2"]
# # #     c.value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(contacts)} contacts"
# # #     c.font=Font(name="Arial",size=9,color="64748B"); c.fill=PatternFill("solid",fgColor="F4F7FB")
# # #     c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=18

# # #     headers=["#","Priority","Name","Title / Role","Company","Email","LinkedIn URL","Source"]
# # #     col_widths=[4,12,24,32,22,34,42,18]
# # #     for ci,(h,w) in enumerate(zip(headers,col_widths),1):
# # #         c=ws.cell(row=3,column=ci,value=h)
# # #         c.font=Font(name="Arial",bold=True,size=10,color=WHITE); c.fill=PatternFill("solid",fgColor=BLUE)
# # #         c.alignment=Alignment(horizontal="center",vertical="center"); c.border=border
# # #         ws.column_dimensions[get_column_letter(ci)].width=w
# # #     ws.row_dimensions[3].height=22

# # #     for ri,ct in enumerate(contacts,start=4):
# # #         prio=ct.get("priority","General"); bg_hex,fg_hex=pri_colors.get(prio,(WHITE,"0F172A"))
# # #         patterns=ct.get("email_patterns",[]); email_val=ct.get("email") or (patterns[0]+"  (pattern)" if patterns else "")
# # #         row_fill=bg_hex if ri%2==0 else GREY
# # #         for ci,val in enumerate([ri-3,prio,ct.get("name",""),ct.get("title",""),ct.get("company",""),email_val,ct.get("linkedin_url",""),ct.get("source","")],1):
# # #             cell=ws.cell(row=ri,column=ci,value=val)
# # #             cell.font=Font(name="Arial",size=9,bold=(ci==2),color=fg_hex if ci==2 else "0F172A")
# # #             cell.fill=PatternFill("solid",fgColor=row_fill)
# # #             cell.alignment=Alignment(vertical="center",wrap_text=(ci in [4,7])); cell.border=border
# # #         ws.row_dimensions[ri].height=18

# # #     ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:H{3+len(contacts)}"
# # #     ws2=wb.create_sheet("Summary"); ws2.merge_cells("A1:C1"); t=ws2["A1"]
# # #     t.value="Export Summary"; t.font=Font(name="Arial",bold=True,size=12,color=WHITE)
# # #     t.fill=PatternFill("solid",fgColor=NAVY); t.alignment=Alignment(horizontal="center")
# # #     for r,(lbl,val) in enumerate([
# # #         ("Company",company),("Role",role),("Total Contacts",len(contacts)),
# # #         ("Primary",  sum(1 for x in contacts if x.get("priority")=="Primary")),
# # #         ("Secondary",sum(1 for x in contacts if x.get("priority")=="Secondary")),
# # #         ("Tertiary", sum(1 for x in contacts if x.get("priority")=="Tertiary")),
# # #         ("General",  sum(1 for x in contacts if x.get("priority")=="General")),
# # #         ("With Email",   sum(1 for x in contacts if x.get("email"))),
# # #         ("With LinkedIn",sum(1 for x in contacts if x.get("linkedin_url"))),
# # #         ("Generated",datetime.now().strftime("%d %b %Y  %H:%M")),
# # #     ],2):
# # #         lc=ws2.cell(row=r,column=1,value=lbl); vc=ws2.cell(row=r,column=2,value=val)
# # #         bg=LBLUE if r%2==0 else WHITE
# # #         for cell in (lc,vc):
# # #             cell.font=Font(name="Arial",bold=(cell.column==1),size=10)
# # #             cell.fill=PatternFill("solid",fgColor=bg); cell.border=border
# # #     ws2.column_dimensions["A"].width=20; ws2.column_dimensions["B"].width=30
# # #     buf=io.BytesIO(); wb.save(buf); buf.seek(0)
# # #     return buf.getvalue()


# # # # ── MongoDB helpers ───────────────────────────────────────────────────────────
# # # @st.cache_resource
# # # def get_db():
# # #     URI = st.secrets.get("MONGO_URI","mongodb://localhost:27017")
# # #     DB  = st.secrets.get("MONGO_DB", "secureitlab_job_pipeline")
# # #     client = MongoClient(URI, serverSelectionTimeoutMS=6000)
# # #     return client[DB]

# # # @st.cache_data(ttl=60)
# # # def load_all_jobs():
# # #     db = get_db()
# # #     return list(db.jobs.find({}, {
# # #         "_id":1,"company":1,"role":1,"job_number":1,
# # #         "opp_score":1,"contacts_found":1,"pipeline_ok":1,
# # #         "coverage_score":1,"run_at":1,"contact_domain":1,
# # #         "contacts":1,"country":1,
# # #     }))

# # # @st.cache_data(ttl=60)
# # # def load_job(job_id):
# # #     from bson import ObjectId
# # #     return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# # # # ── Render helpers ────────────────────────────────────────────────────────────
# # # def badge(text, color="blue"):
# # #     return f'<span class="badge badge-{color}">{text}</span>'

# # # def safe_str(val, limit=300):
# # #     if val is None: return "—"
# # #     s = str(val)
# # #     return s[:limit]+"…" if len(s)>limit else s

# # # def as_dict(raw):
# # #     if isinstance(raw,dict): return raw
# # #     if isinstance(raw,list): return next((x for x in raw if isinstance(x,dict)),{})
# # #     return {}

# # # def render_json_pretty(data, title=""):
# # #     if not data: return
# # #     with st.expander(f"📄 Raw JSON — {title}", expanded=False):
# # #         st.code(json.dumps(data,indent=2,default=str),language="json")

# # # def render_qa_block(data, label):
# # #     if not data:
# # #         st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>',unsafe_allow_html=True); return
# # #     data=as_dict(data)
# # #     if not data: return
# # #     passed=data.get("passed") or data.get("Passed") or False
# # #     rec=data.get("recommendation") or data.get("Recommendation","")
# # #     issues=data.get("issues") or data.get("Issues") or []
# # #     checklist=data.get("checklist") or data.get("Checklist") or []
# # #     color="green" if passed else "yellow"
# # #     status="✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
# # #     html=f"""<div class="sil-card sil-card-accent">
# # #       <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
# # #         <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
# # #         {badge(status,color)}
# # #       </div>"""
# # #     if rec: html+=f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
# # #     if checklist:
# # #         html+='<div style="font-size:.82rem;margin-bottom:.5rem">'
# # #         for item in (checklist if isinstance(checklist,list) else [checklist]):
# # #             if isinstance(item,dict):
# # #                 ip=item.get("pass") or item.get("passed") or item.get("status","")=="pass"
# # #                 nm=item.get("item") or item.get("name") or item.get("check","")
# # #                 nt=item.get("reason") or item.get("note") or item.get("issue","")
# # #                 html+=f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
# # #                 if nt: html+=f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
# # #                 html+='</div>'
# # #         html+='</div>'
# # #     if issues:
# # #         html+='<div style="margin-top:.5rem">'
# # #         for iss in (issues if isinstance(issues,list) else [issues])[:4]:
# # #             txt=iss if isinstance(iss,str) else json.dumps(iss)
# # #             html+=f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
# # #         html+='</div>'
# # #     st.markdown(html+'</div>',unsafe_allow_html=True)

# # # def render_contacts(contacts, title="Contacts"):
# # #     if not contacts: st.info("No contacts found for this job."); return
# # #     pri_color={"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
# # #     st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>',unsafe_allow_html=True)
# # #     cols=st.columns(2)
# # #     for i,c in enumerate(contacts):
# # #         col=cols[i%2]; prio=c.get("priority","General")
# # #         email=c.get("email",""); li=c.get("linkedin_url","")
# # #         patterns=c.get("email_patterns",[]); src=c.get("source","")
# # #         with col:
# # #             html=f"""<div class="contact-card">
# # #               <div style="display:flex;justify-content:space-between;align-items:flex-start">
# # #                 <div>
# # #                   <div class="contact-name">{c.get('name','—')}</div>
# # #                   <div class="contact-title">{c.get('title','—')}</div>
# # #                 </div>
# # #                 {badge(prio,pri_color.get(prio,'blue'))}
# # #               </div>"""
# # #             if email:      html+=f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
# # #             elif patterns: html+=f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
# # #             if li:         html+=f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
# # #             if src:        html+=f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
# # #             st.markdown(html+'</div>',unsafe_allow_html=True)

# # # def render_emails(emails_data):
# # #     if not emails_data: st.info("No email data available."); return
# # #     emails_data=as_dict(emails_data)
# # #     if not emails_data: return
# # #     variants={}
# # #     for k,v in emails_data.items():
# # #         kl=k.lower().replace("_","").replace(" ","")
# # #         if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl=="a":
# # #             variants["Variant A — Hiring Manager"]=v
# # #         elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl=="b":
# # #             variants["Variant B — CISO / VP Level"]=v
# # #         else: variants[k]=v
# # #     for label,v in variants.items():
# # #         st.markdown(f'<div class="section-label">✉️ {label}</div>',unsafe_allow_html=True)
# # #         if isinstance(v,dict):
# # #             subj=v.get("subject") or v.get("Subject","")
# # #             body=v.get("body") or v.get("Body") or v.get("content","")
# # #             if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>',unsafe_allow_html=True)
# # #             if body: st.markdown(f'<div class="email-box">{body}</div>',unsafe_allow_html=True)
# # #             else:    st.code(json.dumps(v,indent=2),language="json")
# # #         elif isinstance(v,str):
# # #             st.markdown(f'<div class="email-box">{v}</div>',unsafe_allow_html=True)
# # #         st.markdown('<div style="height:1rem"></div>',unsafe_allow_html=True)

# # # def render_service_mapping(data):
# # #     if not data: st.info("No service mapping data."); return
# # #     items=data if isinstance(data,list) else []
# # #     if not items and isinstance(data,dict):
# # #         for key in ("services","mappings","service_mapping","ServiceMapping","items"):
# # #             if isinstance(data.get(key),list): items=data[key]; break
# # #         if not items: items=[data]
# # #     fit_colors={"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
# # #     for item in items:
# # #         if not isinstance(item,dict): continue
# # #         svc=item.get("service") or item.get("Service") or item.get("name","")
# # #         fit=(item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
# # #         why=item.get("justification") or item.get("rationale") or item.get("why","")
# # #         reqs=item.get("requirements_addressed") or item.get("requirements") or ""
# # #         eng=item.get("engagement_type") or item.get("engagement","")
# # #         color=fit_colors.get(fit,"blue")
# # #         html=f"""<div class="sil-card" style="margin-bottom:.75rem">
# # #           <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
# # #             <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
# # #             {badge(fit or "MAPPED",color) if fit else ""}
# # #           </div>"""
# # #         if why:  html+=f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
# # #         if reqs:
# # #             rs=", ".join(reqs) if isinstance(reqs,list) else str(reqs)
# # #             html+=f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
# # #         if eng:  html+=f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
# # #         st.markdown(html+'</div>',unsafe_allow_html=True)
# # #     render_json_pretty(data,"Service Mapping")

# # # def render_microplans(data):
# # #     if not data: st.info("No micro-plan data."); return
# # #     plans=data if isinstance(data,list) else []
# # #     if not plans and isinstance(data,dict):
# # #         for k in ("plans","micro_plans","microplans","top_3","improvements"):
# # #             if isinstance(data.get(k),list): plans=data[k]; break
# # #         if not plans: plans=[data]
# # #     for i,plan in enumerate(plans[:3],1):
# # #         if not isinstance(plan,dict): continue
# # #         title=plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
# # #         weeks=plan.get("duration") or plan.get("timeline","")
# # #         obj=plan.get("objective") or plan.get("goal","")
# # #         kpis=plan.get("kpis") or plan.get("KPIs") or []
# # #         tasks=plan.get("tasks") or plan.get("workstreams") or []
# # #         with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}",expanded=(i==1)):
# # #             if obj and obj!=title: st.markdown(f"**Objective:** {obj}")
# # #             if kpis:
# # #                 st.markdown("**KPIs:**")
# # #                 for kpi in (kpis if isinstance(kpis,list) else [kpis]): st.markdown(f"• {kpi}")
# # #             if tasks:
# # #                 st.markdown("**Tasks / Workstreams:**")
# # #                 for t in (tasks if isinstance(tasks,list) else [tasks]):
# # #                     if isinstance(t,dict):
# # #                         tn=t.get("task") or t.get("name",""); te=t.get("effort") or t.get("duration","")
# # #                         st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
# # #                     else: st.markdown(f"• {t}")
# # #             if plan: st.code(json.dumps(plan,indent=2,default=str),language="json")

# # # def render_deal_assurance(data):
# # #     if not data: st.info("No deal assurance data."); return
# # #     if not isinstance(data,dict): render_json_pretty(data,"Deal Assurance Pack"); return
# # #     evp=(data.get("executive_value_proposition") or data.get("value_proposition") or data.get("ExecutiveValueProposition",""))
# # #     if evp:
# # #         st.markdown('<div class="section-label">💼 Executive Value Proposition</div>',unsafe_allow_html=True)
# # #         st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7;color:var(--text)">{evp}</div>',unsafe_allow_html=True)
# # #     caps=data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
# # #     if caps:
# # #         st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>',unsafe_allow_html=True)
# # #         c1,c2=st.columns(2)
# # #         for i,cap in enumerate(caps if isinstance(caps,list) else [caps]):
# # #             (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
# # #     risk=data.get("risk_mitigation") or data.get("RiskMitigation","")
# # #     if risk:
# # #         st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>',unsafe_allow_html=True)
# # #         if isinstance(risk,dict):
# # #             for k,v in risk.items(): st.markdown(f"**{k}:** {v}")
# # #         else: st.markdown(str(risk))
# # #     render_json_pretty(data,"Deal Assurance Pack")


# # # # ══════════════════════════════════════════════════════════════════════════════
# # # #  LOGS PAGE
# # # # ══════════════════════════════════════════════════════════════════════════════

# # # def render_logs_page():
# # #     st.markdown("""
# # #     <div style="margin-bottom:2rem">
# # #       <h1 style="font-family:'Syne',sans-serif;font-size:2.2rem;font-weight:800;color:#0f172a;margin:0">
# # #         ⚙️ Pipeline Execution
# # #       </h1>
# # #       <p style="font-size:.95rem;color:#64748b;margin-top:.5rem">
# # #         Live logs from the job scraper, deduplication, and AI agent pipeline
# # #       </p>
# # #     </div>
# # #     """, unsafe_allow_html=True)

# # #     if st.session_state.pipeline_running:
# # #         st.markdown("""
# # #         <div class="logs-status running">
# # #           <div>
# # #             <div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.3rem">
# # #               <span class="pulse-dot"></span>
# # #               <span style="font-family:'Syne',sans-serif;font-weight:700;color:#1d4ed8">Pipeline Running</span>
# # #             </div>
# # #             <span style="font-size:.82rem;color:#64748b">Please wait while the pipeline processes jobs...</span>
# # #           </div>
# # #         </div>""", unsafe_allow_html=True)
# # #     elif st.session_state.pipeline_done:
# # #         result = st.session_state.pipeline_result or {}
# # #         if result.get("success"):
# # #             st.markdown(f"""
# # #             <div class="logs-status success">
# # #               <div>
# # #                 <div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
# # #                   <span style="font-size:1.5rem">✅</span>
# # #                   <span style="font-family:'Syne',sans-serif;font-weight:700;color:#15803d;font-size:1.1rem">Pipeline Completed Successfully</span>
# # #                 </div>
# # #                 <div style="font-size:.85rem;color:#15803d">
# # #                   Processed {result.get('processed',0)} jobs · {result.get('new_jobs',0)} new found · {result.get('skipped_db',0)} already in database
# # #                 </div>
# # #               </div>
# # #             </div>""", unsafe_allow_html=True)
# # #         else:
# # #             st.markdown(f"""
# # #             <div class="logs-status error">
# # #               <div>
# # #                 <div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
# # #                   <span style="font-size:1.5rem">❌</span>
# # #                   <span style="font-family:'Syne',sans-serif;font-weight:700;color:#b91c1c;font-size:1.1rem">Pipeline Failed</span>
# # #                 </div>
# # #                 <div style="font-size:.85rem;color:#b91c1c">{result.get('error','Unknown error')}</div>
# # #               </div>
# # #             </div>""", unsafe_allow_html=True)

# # #     st.markdown('<div class="section-label" style="margin-top:2rem">📡 Live Execution Logs</div>', unsafe_allow_html=True)
# # #     logs = st.session_state.pipeline_logs or []
# # #     if logs:
# # #         st.markdown(f'<div class="pipeline-log">{chr(10).join(logs[-200:])}</div>', unsafe_allow_html=True)
# # #     else:
# # #         st.markdown('<div class="pipeline-log">Waiting for logs...</div>', unsafe_allow_html=True)

# # #     if st.session_state.pipeline_running:
# # #         st.markdown("""<script>setTimeout(function(){window.location.reload();},1500);</script>""", unsafe_allow_html=True)

# # #     col1, col2, col3 = st.columns([2,2,1])
# # #     with col1:
# # #         if st.button("← Back to Dashboard", use_container_width=True):
# # #             st.session_state.current_page = "dashboard"; st.rerun()
# # #     with col2:
# # #         if st.button("🔄 Refresh Logs", use_container_width=True): st.rerun()
# # #     with col3:
# # #         if st.button("📋 Copy All", use_container_width=True): st.info("Use your browser copy feature")


# # # # ══════════════════════════════════════════════════════════════════════════════
# # # #  SIDEBAR
# # # # ══════════════════════════════════════════════════════════════════════════════
# # # with st.sidebar:
# # #     st.markdown("""
# # #     <div style="padding:.75rem 0 1.25rem">
# # #       <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;color:#2563eb">🛡️ SecureITLab</div>
# # #       <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
# # #     </div>
# # #     """, unsafe_allow_html=True)

# # #     if st.button("🚪 Logout", use_container_width=True):
# # #         st.session_state.logged_in = False; st.session_state.login_error = ""; st.rerun()

# # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # #     # ── Find Jobs button ──────────────────────────────────────────────────────
# # #     st.markdown("**🔍 Find New Jobs**")
# # #     st.caption("Runs scraper → checks MongoDB for duplicates → runs AI pipeline only on NEW jobs → stores in MongoDB · Auto-repeats every 12h")

# # #     find_jobs_clicked = st.button(
# # #         "🔍  Find Jobs",
# # #         disabled=st.session_state.pipeline_running,
# # #         use_container_width=True,
# # #         type="primary",
# # #     )

# # #     if st.session_state.pipeline_running:
# # #         st.markdown("""
# # #         <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
# # #                     padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;
# # #                     font-size:.8rem;color:#1d4ed8">
# # #           ⏳ Pipeline is running… 👉 View logs on Logs Page
# # #         </div>""", unsafe_allow_html=True)

# # #     # ── ✅ SCHEDULER STATUS BADGE (NEW) ───────────────────────────────────────
# # #     try:
# # #         import main as _main_module
# # #         sched = _main_module.get_scheduler_status()

# # #         if sched.get("active"):
# # #             secs = sched.get("seconds_until_next")
# # #             if secs is not None:
# # #                 hrs  = secs // 3600
# # #                 mins = (secs % 3600) // 60
# # #                 countdown = f"{hrs}h {mins}m" if hrs else f"{mins}m {secs % 60}s"
# # #                 next_label = f"Next run in: <b>{countdown}</b>"
# # #             else:
# # #                 next_label = "Next run: calculating…"

# # #             last = (sched.get("last_run") or "")[:19]
# # #             st.markdown(f"""
# # #             <div class="sched-on">
# # #               🟢 <b>Auto-scheduler ON</b><br>
# # #               <span style="color:#64748b">{next_label}</span><br>
# # #               <span style="color:#64748b">Runs every 12h · #{sched.get('run_count',0)} runs so far</span><br>
# # #               <span style="color:#64748b;font-size:.75rem">Last: {last} UTC</span>
# # #             </div>""", unsafe_allow_html=True)

# # #             if st.button("⏹️ Stop Auto-Scheduler", use_container_width=True):
# # #                 _main_module.stop_scheduler()
# # #                 st.rerun()

# # #         elif sched.get("run_count", 0) > 0:
# # #             last = (sched.get("last_run") or "")[:19]
# # #             st.markdown(f"""
# # #             <div class="sched-paused">
# # #               ⏸️ <b>Scheduler paused</b><br>
# # #               <span style="color:#64748b">Last ran: {last} UTC</span><br>
# # #               <span style="color:#64748b">Click Find Jobs to restart</span>
# # #             </div>""", unsafe_allow_html=True)

# # #     except Exception:
# # #         pass   # silently skip if main.py not importable yet
# # #     # ── END SCHEDULER STATUS ──────────────────────────────────────────────────

# # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # #     # ── Job list ──────────────────────────────────────────────────────────────
# # #     try:
# # #         all_jobs = load_all_jobs()
# # #     except Exception as e:
# # #         st.error(f"MongoDB connection failed: {e}"); st.stop()

# # #     if not all_jobs:
# # #         st.warning("No jobs in MongoDB yet. Click **Find Jobs** to scrape and process.")
# # #         st.stop()

# # #     st.markdown(f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">{len(all_jobs)} jobs in database</div>', unsafe_allow_html=True)

# # #     # ── Master contacts export ────────────────────────────────────────────────
# # #     st.markdown("**📋 Master Contacts Export**")
# # #     st.caption("All contacts from ALL jobs in one Excel file with auto-updates")
# # #     master_excel = build_master_contacts_excel(all_jobs)
# # #     if master_excel:
# # #         st.download_button(
# # #             label="📥  Download All Contacts",
# # #             data=master_excel,
# # #             file_name=f"master_contacts_{datetime.now().strftime('%Y%m%d')}.xlsx",
# # #             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# # #             use_container_width=True,
# # #         )
# # #     else:
# # #         st.warning("Run `pip install openpyxl` to enable Excel export.")

# # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# # #     search   = st.text_input("🔍 Filter by company / role", placeholder="e.g. Bounteous")
# # #     filtered = [j for j in all_jobs if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

# # #     def job_label(j):
# # #         score = j.get("opp_score")
# # #         s_str = f" [{score}/10]" if score else ""
# # #         ok    = "✅" if j.get("pipeline_ok") else "❌"
# # #         return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s_str}"

# # #     if not filtered:
# # #         st.warning("No matching jobs."); st.stop()

# # #     sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
# # #     sel_idx     = [job_label(j) for j in filtered].index(sel_label)
# # #     selected_id = str(filtered[sel_idx]["_id"])

# # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # #     ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
# # #     total_c  = sum(j.get("contacts_found",0) for j in all_jobs)
# # #     st.markdown(f"""
# # #     <div style="font-size:.75rem;color:#64748b">
# # #       <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
# # #       <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
# # #     </div>""", unsafe_allow_html=True)

# # #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# # #     if st.button("🔄 Refresh Data", use_container_width=True):
# # #         st.cache_data.clear(); st.rerun()


# # # # ══════════════════════════════════════════════════════════════════════════════
# # # #  FIND JOBS — background thread
# # # # ══════════════════════════════════════════════════════════════════════════════
# # # if find_jobs_clicked and not st.session_state.pipeline_running:
# # #     with _thread_lock:
# # #         _thread_logs.clear(); _thread_result[0]=None; _thread_done[0]=False

# # #     st.session_state.pipeline_running = True
# # #     st.session_state.pipeline_done    = False
# # #     st.session_state.pipeline_logs    = []
# # #     st.session_state.pipeline_result  = None
# # #     st.session_state.current_page     = "logs"
# # #     st.cache_data.clear()

# # #     def _run_pipeline_bg():
# # #         try:
# # #             import main as _main

# # #             def _cb(msg: str):
# # #                 with _thread_lock:
# # #                     _thread_logs.append(msg)

# # #             result = _main.run_pipeline(progress_callback=_cb)
# # #             with _thread_lock:
# # #                 _thread_result[0] = result
# # #         except Exception as e:
# # #             import traceback
# # #             with _thread_lock:
# # #                 _thread_logs.append(f"❌ Unexpected error: {e}")
# # #                 _thread_logs.append(traceback.format_exc())
# # #                 _thread_result[0] = {"success":False,"error":str(e),"scraped":0,"new_jobs":0,"skipped_db":0,"processed":0}
# # #         finally:
# # #             with _thread_lock:
# # #                 _thread_done[0] = True

# # #     threading.Thread(target=_run_pipeline_bg, daemon=True).start()
# # #     st.rerun()

# # # # ── Sync thread state → session_state ────────────────────────────────────────
# # # with _thread_lock:
# # #     if _thread_logs:
# # #         st.session_state.pipeline_logs = list(_thread_logs)
# # #     if _thread_done[0] and _thread_result[0] is not None:
# # #         st.session_state.pipeline_result  = _thread_result[0]
# # #         st.session_state.pipeline_running = False
# # #         st.session_state.pipeline_done    = True


# # # # ══════════════════════════════════════════════════════════════════════════════
# # # #  PAGE ROUTING
# # # # ══════════════════════════════════════════════════════════════════════════════

# # # if st.session_state.current_page == "logs":
# # #     render_logs_page()
# # # else:
# # #     with st.spinner("Loading job from MongoDB…"):
# # #         job = load_job(selected_id)

# # #     if not job:
# # #         st.error("Could not load job document."); st.stop()

# # #     company   = job.get("company",  "Unknown")
# # #     role      = job.get("role",     "Unknown")
# # #     opp_score = job.get("opp_score")
# # #     p_ok      = job.get("pipeline_ok", False)
# # #     p_min     = job.get("pipeline_min", "?")
# # #     c_found   = job.get("contacts_found", 0)
# # #     c_cov     = job.get("coverage_score")
# # #     c_domain  = job.get("contact_domain","")
# # #     run_at    = job.get("run_at","")

# # #     st.markdown(f"""
# # #     <div style="margin-bottom:1.75rem">
# # #       <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;
# # #                   letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
# # #         Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
# # #       </div>
# # #       <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;
# # #                  color:#0f172a;margin:0;line-height:1.15">{role}</h1>
# # #       <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
# # #         @ <span style="color:#334155;font-weight:600">{company}</span>
# # #         {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
# # #       </div>
# # #     </div>""", unsafe_allow_html=True)

# # #     try:
# # #         sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
# # #         sc = "#16a34a" if sn>=7 else "#f59e0b" if sn>=5 else "#dc2626"
# # #     except Exception:
# # #         sc = "#2563eb"

# # #     st.markdown(f"""
# # #     <div class="metric-row">
# # #       <div class="metric-tile"><div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div><div class="lbl">Opportunity Score</div></div>
# # #       <div class="metric-tile"><div class="val">{c_found}</div><div class="lbl">Contacts Found</div></div>
# # #       <div class="metric-tile"><div class="val">{f"{c_cov}%" if c_cov else "—"}</div><div class="lbl">Contact Coverage</div></div>
# # #       <div class="metric-tile"><div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">{'✅ OK' if p_ok else '❌ Failed'}</div><div class="lbl">Pipeline ({p_min} min)</div></div>
# # #     </div>""", unsafe_allow_html=True)

# # #     tabs = st.tabs([
# # #         "📋 Job & Enrichment","🎯 Service Mapping","🔍 Fit / Gap",
# # #         "🛠️ Capability & Plans","📦 Deal Assurance","✉️ Outreach Emails",
# # #         "👥 Contacts","✅ QA Gates","🗄️ Raw Data",
# # #     ])

# # #     with tabs[0]:
# # #         col1,col2=st.columns([1,1])
# # #         with col1:
# # #             st.markdown('<div class="section-label">📄 Job Research</div>',unsafe_allow_html=True)
# # #             jr=as_dict(job.get("agent_job_research") or {})
# # #             if jr:
# # #                 for label,keys in [
# # #                     ("Job Role",["job_role","Job Role","role","title"]),
# # #                     ("Company",["company_name","Company Name","company"]),
# # #                     ("Location",["location","Location"]),
# # #                     ("Organization URL",["organization_url","Organization URL","url"]),
# # #                 ]:
# # #                     val=next((jr.get(k) for k in keys if jr.get(k)),None)
# # #                     if val: st.markdown(f"**{label}:** {val}")
# # #                 desc=jr.get("job_description") or jr.get("Job Description","")
# # #                 if desc:
# # #                     st.markdown("**Job Description:**")
# # #                     st.markdown(f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;max-height:300px;overflow-y:auto;color:var(--text)">{desc[:2000]}</div>',unsafe_allow_html=True)
# # #                 render_json_pretty(jr,"Job Research")
# # #             else: st.info("No job research data.")
# # #         with col2:
# # #             st.markdown('<div class="section-label">🏢 Company Enrichment</div>',unsafe_allow_html=True)
# # #             enr=as_dict(job.get("agent_enrichment") or {})
# # #             if enr:
# # #                 for label,keys in [
# # #                     ("Industry",["industry","Industry"]),("Company Size",["company_size","size","Company Size"]),
# # #                     ("Regulatory Env",["regulatory_environment","regulatory"]),("Certifications",["certifications","Certifications"]),
# # #                     ("Tech Stack",["tech_stack","technology_stack"]),("Security Maturity",["security_maturity","maturity"]),
# # #                 ]:
# # #                     val=next((enr.get(k) for k in keys if enr.get(k)),None)
# # #                     if val:
# # #                         if isinstance(val,list): val=", ".join(str(v) for v in val)
# # #                         st.markdown(f"**{label}:** {safe_str(val,200)}")
# # #                 render_json_pretty(enr,"Enrichment")
# # #             else: st.info("No enrichment data.")

# # #     with tabs[1]:
# # #         st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>',unsafe_allow_html=True)
# # #         render_service_mapping(job.get("agent_service_mapping"))

# # #     with tabs[2]:
# # #         fg=as_dict(job.get("agent_fit_gap") or {})
# # #         if opp_score:
# # #             try:
# # #                 sn=float(str(opp_score).split("/")[0]); bc="#16a34a" if sn>=7 else "#f59e0b" if sn>=5 else "#dc2626"; bp=int(sn/10*100)
# # #                 st.markdown(f"""<div style="margin-bottom:1.5rem">
# # #                   <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
# # #                     <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">Opportunity Score</span>
# # #                     <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
# # #                   </div>
# # #                   <div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%">
# # #                     <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
# # #                   </div>
# # #                 </div>""", unsafe_allow_html=True)
# # #             except Exception: pass
# # #         st.markdown('<div class="section-label">📊 Service Classifications</div>',unsafe_allow_html=True)
# # #         services=[]
# # #         if isinstance(fg,dict):
# # #             for k in ("services","classifications","service_classifications","items","fit_gap"):
# # #                 v=fg.get(k)
# # #                 if isinstance(v,list): services=v; break
# # #             if not services and (fg.get("service") or fg.get("Service")): services=[fg]
# # #         elif isinstance(fg,list): services=fg
# # #         if services:
# # #             buckets={"STRONG FIT":[],"PARTIAL FIT":[],"GAP":[]}; other=[]
# # #             for s in services:
# # #                 if not isinstance(s,dict): continue
# # #                 fit=(s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
# # #                 if "STRONG" in fit: buckets["STRONG FIT"].append(s)
# # #                 elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
# # #                 elif "GAP" in fit: buckets["GAP"].append(s)
# # #                 else: other.append(s)
# # #             c1,c2,c3=st.columns(3)
# # #             cm={"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
# # #             bgm={"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
# # #             bdm={"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
# # #             for col,(fl,items) in zip([c1,c2,c3],buckets.items()):
# # #                 col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>',unsafe_allow_html=True)
# # #                 for s in items:
# # #                     svc=s.get("service") or s.get("Service") or s.get("name",""); just=s.get("justification") or s.get("reason","")
# # #                     col.markdown(f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem"><div style="font-weight:600;color:#0f172a">{svc}</div><div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>',unsafe_allow_html=True)
# # #         elif fg: st.json(fg)
# # #         else: st.info("No fit/gap data.")
# # #         render_json_pretty(job.get("agent_fit_gap"),"Fit/Gap Analysis")

# # #     with tabs[3]:
# # #         col1,col2=st.columns([1,1])
# # #         with col1:
# # #             st.markdown('<div class="section-label">🔧 Capability Improvements</div>',unsafe_allow_html=True)
# # #             cap=job.get("agent_capability") or {}
# # #             items_cap=cap if isinstance(cap,list) else []
# # #             if not items_cap and isinstance(cap,dict):
# # #                 for k in ("improvements","recommendations","capabilities","items"):
# # #                     v=cap.get(k)
# # #                     if isinstance(v,list): items_cap=v; break
# # #                 if not items_cap: items_cap=[cap]
# # #             for item in items_cap:
# # #                 if not isinstance(item,dict): continue
# # #                 title=item.get("title") or item.get("gap") or item.get("service","")
# # #                 rec=item.get("recommendation") or item.get("steps","")
# # #                 effort=item.get("build_effort") or item.get("effort",""); demand=item.get("market_demand") or item.get("priority","")
# # #                 st.markdown(f"""<div class="sil-card" style="margin-bottom:.6rem">
# # #                   <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem;color:var(--text)">{title}</div>
# # #                   <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec,250)}</div>
# # #                   {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
# # #                 </div>""", unsafe_allow_html=True)
# # #             if not items_cap: render_json_pretty(cap,"Capability Improvement")
# # #         with col2:
# # #             st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>',unsafe_allow_html=True)
# # #             render_microplans(job.get("agent_microplans"))

# # #     with tabs[4]:
# # #         render_deal_assurance(job.get("agent_deal_assurance"))

# # #     with tabs[5]:
# # #         st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>',unsafe_allow_html=True)
# # #         emails_src=job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
# # #         oq=as_dict(job.get("agent_outreach_qa") or {})
# # #         improved=(oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
# # #         if improved:
# # #             st.info("⚡ Showing QA-improved versions where available")
# # #             render_emails(improved)
# # #             with st.expander("📬 Original (pre-QA) versions"): render_emails(emails_src)
# # #         else: render_emails(emails_src)

# # #     with tabs[6]:
# # #         contacts=job.get("contacts") or []; contact_sources=job.get("contact_sources") or []
# # #         pri=[c for c in contacts if c.get("priority")=="Primary"]
# # #         sec=[c for c in contacts if c.get("priority")=="Secondary"]
# # #         ter=[c for c in contacts if c.get("priority")=="Tertiary"]
# # #         gen=[c for c in contacts if c.get("priority")=="General"]
# # #         st.markdown(f"""<div class="metric-row" style="margin-bottom:1.5rem">
# # #           <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
# # #           <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
# # #           <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
# # #           <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
# # #         </div>""", unsafe_allow_html=True)
# # #         if contact_sources:
# # #             st.markdown('Sources: '+" ".join(badge(s,"blue") for s in contact_sources),unsafe_allow_html=True); st.markdown("")
# # #         missing=job.get("missing_roles") or []
# # #         if missing:
# # #             st.markdown('⚠️ Missing roles: '+" ".join(badge(r,"red") for r in missing),unsafe_allow_html=True); st.markdown("")
# # #         if contacts:
# # #             excel_bytes=build_contacts_excel(contacts,company,role)
# # #             if excel_bytes:
# # #                 safe_co=re.sub(r'[^a-z0-9]','_',company.lower())[:20]
# # #                 fname=f"contacts_{safe_co}_{datetime.now().strftime('%Y%m%d')}.xlsx"
# # #                 btn_col,_=st.columns([1,3])
# # #                 with btn_col:
# # #                     st.download_button(label="📥  Download Contacts (.xlsx)",data=excel_bytes,file_name=fname,
# # #                         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,type="primary")
# # #             else: st.warning("Run `pip install openpyxl` to enable Excel export.")
# # #             st.markdown('<div class="sil-divider"></div>',unsafe_allow_html=True)
# # #             pri_filter=st.multiselect("Filter by priority",["Primary","Secondary","Tertiary","General"],default=["Primary","Secondary","Tertiary","General"])
# # #             shown=[c for c in contacts if c.get("priority","General") in pri_filter]
# # #             render_contacts(shown,f"Contacts ({len(shown)} shown)")
# # #             agent_contacts=job.get("agent_prospect_contacts") or {}
# # #             if agent_contacts:
# # #                 with st.expander("🤖 CrewAI Agent's Contact Search"):
# # #                     if isinstance(agent_contacts,dict):
# # #                         ac_list=agent_contacts.get("contacts") or []
# # #                         if ac_list: render_contacts(ac_list,"Agent Contacts")
# # #                         else: st.json(agent_contacts)
# # #                     else: st.json(agent_contacts)
# # #         else: st.info("No contacts found for this job.")

# # #     with tabs[7]:
# # #         st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>',unsafe_allow_html=True)
# # #         col1,col2=st.columns(2)
# # #         for i,(label,key) in enumerate([
# # #             ("Research QA","agent_research_qa"),("Service Mapping QA","agent_mapping_qa"),
# # #             ("Deal Assurance QA","agent_assurance_qa"),("Outreach Email QA","agent_outreach_qa"),
# # #         ]):
# # #             with (col1 if i%2==0 else col2): render_qa_block(job.get(key),label)
# # #         st.markdown('<div class="sil-divider"></div>',unsafe_allow_html=True)
# # #         st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>',unsafe_allow_html=True)
# # #         enf=as_dict(job.get("agent_prospect_enforcer") or {})
# # #         if enf:
# # #             cov=enf.get("coverage_score","?"); miss=enf.get("missing_roles",[]); note=enf.get("note",""); ec=enf.get("contacts",[])
# # #             x1,x2,x3=st.columns(3)
# # #             x1.metric("Coverage Score",f"{cov}%"); x2.metric("Missing Roles",len(miss)); x3.metric("Contacts Verified",len(ec))
# # #             if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
# # #             if note: st.caption(note)
# # #         else: st.info("No enforcer data.")

# # #     with tabs[8]:
# # #         st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>',unsafe_allow_html=True)
# # #         st.caption("All fields stored in the `jobs` collection for this document.")
# # #         rows=[]
# # #         for k,v in job.items():
# # #             if k=="_id": continue
# # #             rows.append({"Field":k,"Type":type(v).__name__,"Len":len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
# # #         hc1,hc2,hc3=st.columns([3,1,1])
# # #         hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
# # #         for r in rows:
# # #             rc1,rc2,rc3=st.columns([3,1,1])
# # #             rc1.code(r["Field"],language=None)
# # #             rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>',unsafe_allow_html=True)
# # #             rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>',unsafe_allow_html=True)
# # #         st.markdown('<div class="sil-divider"></div>',unsafe_allow_html=True)
# # #         for label,key in [
# # #             ("Job Research","agent_job_research"),("Enrichment","agent_enrichment"),
# # #             ("Service Mapping","agent_service_mapping"),("Fit/Gap Analysis","agent_fit_gap"),
# # #             ("Capability","agent_capability"),("Micro-Plans","agent_microplans"),
# # #             ("Deal Assurance","agent_deal_assurance"),("Outreach Emails","agent_outreach_emails"),
# # #             ("Prospect Contacts","agent_prospect_contacts"),("Prospect Enforcer","agent_prospect_enforcer"),
# # #             ("Research QA","agent_research_qa"),("Mapping QA","agent_mapping_qa"),
# # #             ("Assurance QA","agent_assurance_qa"),("Outreach QA","agent_outreach_qa"),
# # #             ("Contacts (5-source)","contacts"),
# # #         ]:
# # #             data=job.get(key)
# # #             if data:
# # #                 with st.expander(f"📄 {label}"):
# # #                     st.code(json.dumps(data,indent=2,default=str),language="json")


























# # """
# # ╔══════════════════════════════════════════════════════════╗
# # ║   SecureITLab Pipeline Dashboard — Streamlit             ║
# # ║   WITH GOOGLE SHEETS AUTO-SYNC (Master Contacts)         ║
# # ║   WITH 12-HOUR AUTO-SCHEDULER STATUS IN SIDEBAR          ║
# # ║   Reads from MongoDB → secureitlab_job_pipeline          ║
# # ╠══════════════════════════════════════════════════════════╣
# # ║  Install: pip install streamlit pymongo python-dotenv    ║
# # ║           gspread google-auth openpyxl                   ║
# # ║  Run:     streamlit run streamlit_dashboard.py           ║
# # ╚══════════════════════════════════════════════════════════╝
# # """

# # import io
# # import re
# # import streamlit as st
# # from pymongo import MongoClient
# # import json
# # import threading
# # import time
# # import logging
# # from datetime import datetime, timezone
# # from io import StringIO
# # from pathlib import Path

# # # ── Thread-safe shared state ──────────────────────────────────────────────────
# # _thread_logs   = []
# # _thread_result = [None]
# # _thread_done   = [False]
# # _thread_lock   = threading.Lock()

# # # ── Google Sheets config ──────────────────────────────────────────────────────
# # GSHEET_URL        = "https://docs.google.com/spreadsheets/d/1u9_Fqy8a8Dj-yX8FwZtkFfw_3BLdpmBbtagH2IjhyV4/edit"
# # GSHEET_ID         = "1u9_Fqy8a8Dj-yX8FwZtkFfw_3BLdpmBbtagH2IjhyV4"
# # GSHEET_TAB        = "master_contacts"
# # GCREDS_FILE       = "google_credentials.json"   # service account JSON in project folder
# # GSHEET_SYNC_STATE = "gsheet_sync_state.json"    # persists last sync info across reruns

# # # ── Log capture ───────────────────────────────────────────────────────────────
# # _log_capture  = StringIO()
# # _log_handler  = logging.StreamHandler(_log_capture)
# # _log_handler.setLevel(logging.INFO)
# # _log_handler.setFormatter(logging.Formatter('%(asctime)s | %(levelname)s | %(message)s'))

# # # ── Page config ───────────────────────────────────────────────────────────────
# # st.set_page_config(
# #     page_title="SecureITLab Pipeline",
# #     page_icon="🛡️",
# #     layout="wide",
# #     initial_sidebar_state="expanded",
# # )

# # LOGIN_USERNAME = "admin"
# # LOGIN_PASSWORD = "secureitlab2024"

# # # ══════════════════════════════════════════════════════════════════════════════
# # #  GLOBAL CSS
# # # ══════════════════════════════════════════════════════════════════════════════
# # st.markdown("""
# # <style>
# # @import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');
# # :root{--bg:#f4f7fb;--surface:#ffffff;--surface2:#eef2f7;--border:#d9e2ec;--accent:#2563eb;--accent2:#7c3aed;--green:#16a34a;--yellow:#f59e0b;--red:#dc2626;--text:#0f172a;--muted:#64748b;}
# # html,body,[class*="css"]{background-color:var(--bg)!important;color:var(--text)!important;font-family:'DM Sans',sans-serif!important;}
# # .login-card{background:var(--surface);border:1px solid var(--border);border-radius:20px;padding:3rem 3.5rem;width:100%;max-width:420px;box-shadow:0 20px 60px rgba(37,99,235,0.08);text-align:center;}
# # .login-logo{font-family:'Syne',sans-serif;font-size:1.6rem;font-weight:800;color:var(--accent);margin-bottom:.25rem;}
# # .login-subtitle{font-size:.75rem;color:var(--muted);letter-spacing:.12em;text-transform:uppercase;margin-bottom:2.5rem;}
# # .login-error{background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:.65rem 1rem;color:#b91c1c;font-size:.85rem;margin-top:1rem;}
# # .login-divider{width:40px;height:3px;background:linear-gradient(90deg,var(--accent),var(--accent2));border-radius:2px;margin:0 auto 2rem;}
# # [data-testid="stSidebar"]{background:var(--surface)!important;border-right:1px solid var(--border)!important;}
# # [data-testid="stSidebar"] *{color:var(--text)!important;}
# # .main .block-container{padding:2rem 2rem 3rem!important;}
# # h1,h2,h3,h4{font-family:'Syne',sans-serif!important;color:var(--text)!important;}
# # .sil-card{background:var(--surface);border:1px solid var(--border);border-radius:14px;padding:1.25rem 1.5rem;margin-bottom:1rem;transition:all 0.25s ease;}
# # .sil-card:hover{transform:translateY(-2px);box-shadow:0 8px 22px rgba(0,0,0,0.05);}
# # .sil-card-accent{border-left:4px solid var(--accent);}
# # .metric-row{display:flex;gap:1rem;flex-wrap:wrap;margin-bottom:1.5rem;}
# # .metric-tile{flex:1;min-width:140px;background:var(--surface2);border:1px solid var(--border);border-radius:12px;padding:1rem;text-align:center;transition:all .25s ease;}
# # .metric-tile:hover{transform:translateY(-3px);box-shadow:0 10px 24px rgba(0,0,0,0.06);}
# # .metric-tile .val{font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;color:var(--accent);}
# # .metric-tile .lbl{font-size:.72rem;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;}
# # .badge{padding:.25rem .7rem;border-radius:20px;font-size:.72rem;font-weight:600;font-family:'DM Mono',monospace;}
# # .badge-green{background:#ecfdf5;color:#15803d;}
# # .badge-yellow{background:#fffbeb;color:#b45309;}
# # .badge-red{background:#fef2f2;color:#b91c1c;}
# # .badge-blue{background:#eff6ff;color:#1d4ed8;}
# # .badge-purple{background:#f5f3ff;color:#6d28d9;}
# # .contact-card{background:var(--surface2);border:1px solid var(--border);border-radius:10px;padding:1rem;margin-bottom:.8rem;}
# # .contact-name{font-family:'Syne',sans-serif;font-weight:700;color:var(--text);}
# # .contact-title{color:var(--muted);font-size:.85rem;}
# # .contact-email{font-family:'DM Mono',monospace;color:var(--accent);font-size:.8rem;}
# # .email-box{background:#f8fafc;border:1px solid var(--border);border-radius:10px;padding:1rem 1.25rem;font-size:.9rem;line-height:1.65;white-space:pre-wrap;color:var(--text);}
# # .email-subject{font-family:'Syne',sans-serif;font-weight:700;color:var(--accent);margin-bottom:.5rem;}
# # .section-label{font-family:'DM Mono',monospace;font-size:.72rem;text-transform:uppercase;letter-spacing:.12em;color:var(--accent);margin-bottom:.6rem;}
# # .sil-divider{border-top:1px solid var(--border);margin:1rem 0;}
# # [data-testid="stExpander"]{background:var(--surface)!important;border:1px solid var(--border)!important;border-radius:10px!important;}
# # [data-testid="stTabs"] button{font-family:'Syne',sans-serif!important;font-weight:600!important;}
# # ::-webkit-scrollbar{width:6px;}
# # ::-webkit-scrollbar-thumb{background:var(--border);border-radius:3px;}
# # .pipeline-log{background:#0f172a;color:#10b981;border-radius:10px;padding:1.5rem;font-family:'DM Mono',monospace;font-size:.8rem;line-height:1.8;max-height:700px;overflow-y:auto;white-space:pre-wrap;word-break:break-word;border:1px solid #1e293b;}
# # .logs-status{display:flex;gap:1rem;justify-content:space-between;align-items:center;margin-bottom:1.5rem;padding:1rem;background:var(--surface2);border-radius:10px;border:1px solid var(--border);}
# # .logs-status.running{background:#eff6ff;border-color:#bfdbfe;}
# # .logs-status.success{background:#f0fdf4;border-color:#bbf7d0;}
# # .logs-status.error{background:#fef2f2;border-color:#fecaca;}
# # @keyframes pulse{0%,100%{opacity:1;}50%{opacity:0.5;}}
# # .pulse-dot{display:inline-block;width:10px;height:10px;background:#2563eb;border-radius:50%;animation:pulse 2s infinite;margin-right:8px;}
# # .sched-on{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:.65rem .9rem;margin-top:.5rem;font-size:.8rem;color:#15803d;line-height:1.6;}
# # .sched-paused{background:#fef9c3;border:1px solid #fde68a;border-radius:8px;padding:.65rem .9rem;margin-top:.5rem;font-size:.8rem;color:#92400e;line-height:1.6;}

# # /* ── Google Sheet box ── */
# # .gsheet-box{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:.9rem 1rem;margin-top:.4rem;font-size:.82rem;color:#15803d;line-height:1.8;}
# # .gsheet-box a{color:#1d4ed8!important;font-weight:700;text-decoration:none;}
# # .gsheet-box a:hover{text-decoration:underline;}
# # .gsheet-syncing{background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;padding:.9rem 1rem;margin-top:.4rem;font-size:.82rem;color:#1d4ed8;line-height:1.8;}
# # .gsheet-error{background:#fef2f2;border:1px solid #fecaca;border-radius:10px;padding:.75rem 1rem;font-size:.8rem;color:#b91c1c;margin-top:.4rem;line-height:1.6;}
# # </style>
# # """, unsafe_allow_html=True)


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  SESSION STATE
# # # ══════════════════════════════════════════════════════════════════════════════
# # for _k, _v in [
# #     ("logged_in",         False),
# #     ("login_error",       ""),
# #     ("pipeline_running",  False),
# #     ("pipeline_logs",     []),
# #     ("pipeline_result",   None),
# #     ("pipeline_done",     False),
# #     ("current_page",      "dashboard"),
# #     ("gsheet_sync_error", ""),
# #     ("gsheet_last_sync",  None),
# #     ("gsheet_appended",   None),
# # ]:
# #     if _k not in st.session_state:
# #         st.session_state[_k] = _v


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  GOOGLE SHEETS HELPERS
# # # ══════════════════════════════════════════════════════════════════════════════

# # def _gsheet_client():
# #     import gspread
# #     from google.oauth2.service_account import Credentials
# #     scopes = [
# #         "https://www.googleapis.com/auth/spreadsheets",
# #         "https://www.googleapis.com/auth/drive",
# #     ]
# #     creds = Credentials.from_service_account_file(GCREDS_FILE, scopes=scopes)
# #     return gspread.authorize(creds)


# # def _existing_contact_keys(ws) -> set:
# #     """Return set of (name_lower, company_lower) already in the sheet."""
# #     try:
# #         rows = ws.get_all_values()
# #         keys = set()
# #         for row in rows[3:]:        # skip 3 header rows
# #             name    = row[5].strip().lower() if len(row) > 5 else ""
# #             company = row[2].strip().lower() if len(row) > 2 else ""
# #             if name:
# #                 keys.add((name, company))
# #         return keys
# #     except Exception:
# #         return set()


# # def _write_sync_state(appended: int, skipped: int):
# #     try:
# #         Path(GSHEET_SYNC_STATE).write_text(json.dumps({
# #             "last_sync": datetime.now(timezone.utc).isoformat(),
# #             "appended":  appended,
# #             "skipped":   skipped,
# #         }), encoding="utf-8")
# #     except Exception:
# #         pass


# # def _read_sync_state() -> dict:
# #     try:
# #         if Path(GSHEET_SYNC_STATE).exists():
# #             return json.loads(Path(GSHEET_SYNC_STATE).read_text(encoding="utf-8"))
# #     except Exception:
# #         pass
# #     return {}


# # def sync_contacts_to_gsheet(all_jobs: list) -> dict:
# #     """
# #     Append NEW contacts from MongoDB jobs list to the Google Sheet.
# #     Never deletes existing rows — skips any contact already present.
# #     Returns: {"appended": int, "skipped": int, "error": str|None}
# #     """
# #     result = {"appended": 0, "skipped": 0, "error": None}

# #     if not Path(GCREDS_FILE).exists():
# #         result["error"] = (
# #             f"'{GCREDS_FILE}' not found in project folder. "
# #             "Create a service account on Google Cloud, download the JSON key, "
# #             "save it as google_credentials.json, then share your sheet with the service account email."
# #         )
# #         return result

# #     try:
# #         gc = _gsheet_client()
# #         sh = gc.open_by_key(GSHEET_ID)

# #         # Get or create the worksheet tab
# #         try:
# #             ws = sh.worksheet(GSHEET_TAB)
# #         except Exception:
# #             ws = sh.add_worksheet(title=GSHEET_TAB, rows=10000, cols=11)

# #         # Write headers if sheet is completely empty
# #         all_vals = ws.get_all_values()
# #         if not all_vals or len(all_vals) < 3:
# #             header_rows = [
# #                 ["Master Contacts — SecureITLab Pipeline Auto-Sync"] + [""] * 10,
# #                 [f"Last synced: {datetime.now().strftime('%d %b %Y  %H:%M')} UTC  ·  Append-only, duplicates skipped"] + [""] * 10,
# #                 ["#", "Job Role", "Company", "Country", "Priority",
# #                  "Name", "Title / Role", "Email", "LinkedIn URL", "Source", "Job Score"],
# #             ]
# #             ws.update("A1", header_rows)
# #         else:
# #             # Always refresh the subtitle row with latest sync time
# #             ws.update("A2", [[
# #                 f"Last synced: {datetime.now().strftime('%d %b %Y  %H:%M')} UTC  ·  Append-only, duplicates skipped"
# #             ]])

# #         # Collect keys already in the sheet
# #         existing_keys = _existing_contact_keys(ws)
# #         current_row_count = max(0, len(ws.get_all_values()) - 3)  # data rows only

# #         rows_to_add = []
# #         for job in all_jobs:
# #             company   = job.get("company", "")
# #             role      = job.get("role", "")
# #             country   = job.get("country", "?")
# #             job_score = str(job.get("opp_score", ""))
# #             contacts  = job.get("contacts", [])

# #             for ci, contact in enumerate(contacts):
# #                 name     = contact.get("name", "").strip()
# #                 prio     = contact.get("priority", "General")
# #                 title    = contact.get("title", "")
# #                 email    = contact.get("email", "")
# #                 li       = contact.get("linkedin_url", "")
# #                 source   = contact.get("source", "")
# #                 patterns = contact.get("email_patterns", [])

# #                 if not email and patterns:
# #                     email = patterns[0] + "  (pattern)"

# #                 key = (name.lower(), company.strip().lower())
# #                 if not name or key in existing_keys:
# #                     result["skipped"] += 1
# #                     continue

# #                 rows_to_add.append([
# #                     current_row_count + len(rows_to_add) + 1,   # row #
# #                     role    if ci == 0 else "",
# #                     company if ci == 0 else "",
# #                     country if ci == 0 else "",
# #                     prio, name, title, email, li, source,
# #                     job_score if ci == 0 else "",
# #                 ])
# #                 existing_keys.add(key)  # dedup within this batch too

# #         if rows_to_add:
# #             ws.append_rows(rows_to_add, value_input_option="USER_ENTERED")
# #             result["appended"] = len(rows_to_add)

# #         _write_sync_state(result["appended"], result["skipped"])

# #     except Exception as e:
# #         result["error"] = str(e)

# #     return result


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  LOGIN
# # # ══════════════════════════════════════════════════════════════════════════════
# # if not st.session_state.logged_in:
# #     _, col, _ = st.columns([1, 1.2, 1])
# #     with col:
# #         st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)
# #         st.markdown("""
# #         <div class="login-card">
# #           <div class="login-logo">🛡️ SecureITLab</div>
# #           <div class="login-subtitle">Pipeline Intelligence</div>
# #           <div class="login-divider"></div>
# #         </div>""", unsafe_allow_html=True)
# #         username = st.text_input("Username", placeholder="Enter username", key="lu")
# #         password = st.text_input("Password", placeholder="Enter password", type="password", key="lp")
# #         if st.button("Sign In →", use_container_width=True, type="primary"):
# #             if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
# #                 st.session_state.logged_in = True; st.session_state.login_error = ""; st.rerun()
# #             else:
# #                 st.session_state.login_error = "Incorrect username or password."
# #         if st.session_state.login_error:
# #             st.markdown(f'<div class="login-error">⚠️ {st.session_state.login_error}</div>', unsafe_allow_html=True)
# #         st.markdown("<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>SecureITLab · Confidential</div>", unsafe_allow_html=True)
# #     st.stop()


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  PER-JOB CONTACTS EXCEL  (individual job download — kept)
# # # ══════════════════════════════════════════════════════════════════════════════
# # def build_contacts_excel(contacts: list, company: str, role: str):
# #     try:
# #         from openpyxl import Workbook
# #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# #         from openpyxl.utils import get_column_letter
# #     except ImportError:
# #         return None
# #     wb = Workbook(); ws = wb.active; ws.title = "Contacts"
# #     NAVY="1E3A5F"; BLUE="2563EB"; GREY="F8FAFC"; WHITE="FFFFFF"
# #     pri_colors={"Primary":("FEF2F2","B91C1C"),"Secondary":("FFFBEB","B45309"),"Tertiary":("EFF6FF","1D4ED8"),"General":("F5F3FF","6D28D9")}
# #     thin=Side(border_style="thin",color="D9E2EC"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
# #     ws.merge_cells("A1:H1"); c=ws["A1"]
# #     c.value=f"Contacts Export  —  {company}  |  {role}"
# #     c.font=Font(name="Arial",bold=True,size=13,color=WHITE); c.fill=PatternFill("solid",fgColor=NAVY)
# #     c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=30
# #     ws.merge_cells("A2:H2"); c=ws["A2"]
# #     c.value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(contacts)} contacts"
# #     c.font=Font(name="Arial",size=9,color="64748B"); c.fill=PatternFill("solid",fgColor="F4F7FB")
# #     c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=18
# #     headers=["#","Priority","Name","Title / Role","Company","Email","LinkedIn URL","Source"]; col_widths=[4,12,24,32,22,34,42,18]
# #     for ci,(h,w) in enumerate(zip(headers,col_widths),1):
# #         c=ws.cell(row=3,column=ci,value=h)
# #         c.font=Font(name="Arial",bold=True,size=10,color=WHITE); c.fill=PatternFill("solid",fgColor=BLUE)
# #         c.alignment=Alignment(horizontal="center",vertical="center"); c.border=border
# #         ws.column_dimensions[get_column_letter(ci)].width=w
# #     ws.row_dimensions[3].height=22
# #     for ri,ct in enumerate(contacts,start=4):
# #         prio=ct.get("priority","General"); bg_hex,fg_hex=pri_colors.get(prio,(WHITE,"0F172A"))
# #         patterns=ct.get("email_patterns",[]); email_val=ct.get("email") or (patterns[0]+"  (pattern)" if patterns else "")
# #         row_fill=bg_hex if ri%2==0 else GREY
# #         for ci,val in enumerate([ri-3,prio,ct.get("name",""),ct.get("title",""),ct.get("company",""),email_val,ct.get("linkedin_url",""),ct.get("source","")],1):
# #             cell=ws.cell(row=ri,column=ci,value=val)
# #             cell.font=Font(name="Arial",size=9,bold=(ci==2),color=fg_hex if ci==2 else "0F172A")
# #             cell.fill=PatternFill("solid",fgColor=row_fill)
# #             cell.alignment=Alignment(vertical="center",wrap_text=(ci in [4,7])); cell.border=border
# #         ws.row_dimensions[ri].height=18
# #     ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:H{3+len(contacts)}"
# #     buf=io.BytesIO(); wb.save(buf); buf.seek(0)
# #     return buf.getvalue()


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  MONGODB
# # # ══════════════════════════════════════════════════════════════════════════════
# # @st.cache_resource
# # def get_db():
# #     URI = st.secrets.get("MONGO_URI", "mongodb://localhost:27017")
# #     DB  = st.secrets.get("MONGO_DB",  "secureitlab_job_pipeline")
# #     return MongoClient(URI, serverSelectionTimeoutMS=6000)[DB]

# # @st.cache_data(ttl=60)
# # def load_all_jobs():
# #     return list(get_db().jobs.find({}, {
# #         "_id":1,"company":1,"role":1,"job_number":1,"opp_score":1,
# #         "contacts_found":1,"pipeline_ok":1,"coverage_score":1,
# #         "run_at":1,"contact_domain":1,"contacts":1,"country":1,
# #     }))

# # @st.cache_data(ttl=60)
# # def load_job(job_id):
# #     from bson import ObjectId
# #     return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  RENDER HELPERS
# # # ══════════════════════════════════════════════════════════════════════════════
# # def badge(text, color="blue"):
# #     return f'<span class="badge badge-{color}">{text}</span>'

# # def safe_str(val, limit=300):
# #     if val is None: return "—"
# #     s = str(val); return s[:limit]+"…" if len(s) > limit else s

# # def as_dict(raw):
# #     if isinstance(raw, dict): return raw
# #     if isinstance(raw, list): return next((x for x in raw if isinstance(x, dict)), {})
# #     return {}

# # def render_json_pretty(data, title=""):
# #     if not data: return
# #     with st.expander(f"📄 Raw JSON — {title}", expanded=False):
# #         st.code(json.dumps(data, indent=2, default=str), language="json")

# # def render_qa_block(data, label):
# #     if not data:
# #         st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>', unsafe_allow_html=True); return
# #     data = as_dict(data)
# #     if not data: return
# #     passed   = data.get("passed") or data.get("Passed") or False
# #     rec      = data.get("recommendation") or data.get("Recommendation", "")
# #     issues   = data.get("issues") or data.get("Issues") or []
# #     checklist= data.get("checklist") or data.get("Checklist") or []
# #     color = "green" if passed else "yellow"; status = "✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
# #     html = f"""<div class="sil-card sil-card-accent">
# #       <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
# #         <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
# #         {badge(status, color)}
# #       </div>"""
# #     if rec: html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
# #     if checklist:
# #         html += '<div style="font-size:.82rem;margin-bottom:.5rem">'
# #         for item in (checklist if isinstance(checklist, list) else [checklist]):
# #             if isinstance(item, dict):
# #                 ip = item.get("pass") or item.get("passed") or item.get("status","") == "pass"
# #                 nm = item.get("item") or item.get("name") or item.get("check","")
# #                 nt = item.get("reason") or item.get("note") or item.get("issue","")
# #                 html += f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
# #                 if nt: html += f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
# #                 html += '</div>'
# #         html += '</div>'
# #     if issues:
# #         html += '<div style="margin-top:.5rem">'
# #         for iss in (issues if isinstance(issues, list) else [issues])[:4]:
# #             txt = iss if isinstance(iss, str) else json.dumps(iss)
# #             html += f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
# #         html += '</div>'
# #     st.markdown(html + '</div>', unsafe_allow_html=True)

# # def render_contacts(contacts, title="Contacts"):
# #     if not contacts: st.info("No contacts found for this job."); return
# #     pri_color = {"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
# #     st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>', unsafe_allow_html=True)
# #     cols = st.columns(2)
# #     for i, c in enumerate(contacts):
# #         col = cols[i % 2]; prio = c.get("priority","General")
# #         email = c.get("email",""); li = c.get("linkedin_url","")
# #         patterns = c.get("email_patterns",[]); src = c.get("source","")
# #         with col:
# #             html = f"""<div class="contact-card">
# #               <div style="display:flex;justify-content:space-between;align-items:flex-start">
# #                 <div><div class="contact-name">{c.get('name','—')}</div>
# #                 <div class="contact-title">{c.get('title','—')}</div></div>
# #                 {badge(prio, pri_color.get(prio,'blue'))}
# #               </div>"""
# #             if email:      html += f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
# #             elif patterns: html += f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
# #             if li:         html += f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
# #             if src:        html += f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
# #             st.markdown(html + '</div>', unsafe_allow_html=True)

# # def render_emails(emails_data):
# #     if not emails_data: st.info("No email data available."); return
# #     emails_data = as_dict(emails_data)
# #     if not emails_data: return
# #     variants = {}
# #     for k, v in emails_data.items():
# #         kl = k.lower().replace("_","").replace(" ","")
# #         if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl=="a": variants["Variant A — Hiring Manager"] = v
# #         elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl=="b": variants["Variant B — CISO / VP Level"] = v
# #         else: variants[k] = v
# #     for label, v in variants.items():
# #         st.markdown(f'<div class="section-label">✉️ {label}</div>', unsafe_allow_html=True)
# #         if isinstance(v, dict):
# #             subj = v.get("subject") or v.get("Subject",""); body = v.get("body") or v.get("Body") or v.get("content","")
# #             if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>', unsafe_allow_html=True)
# #             if body: st.markdown(f'<div class="email-box">{body}</div>', unsafe_allow_html=True)
# #             else:    st.code(json.dumps(v, indent=2), language="json")
# #         elif isinstance(v, str):
# #             st.markdown(f'<div class="email-box">{v}</div>', unsafe_allow_html=True)
# #         st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

# # def render_service_mapping(data):
# #     if not data: st.info("No service mapping data."); return
# #     items = data if isinstance(data, list) else []
# #     if not items and isinstance(data, dict):
# #         for key in ("services","mappings","service_mapping","ServiceMapping","items"):
# #             if isinstance(data.get(key), list): items = data[key]; break
# #         if not items: items = [data]
# #     fit_colors = {"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
# #     for item in items:
# #         if not isinstance(item, dict): continue
# #         svc  = item.get("service") or item.get("Service") or item.get("name","")
# #         fit  = (item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
# #         why  = item.get("justification") or item.get("rationale") or item.get("why","")
# #         reqs = item.get("requirements_addressed") or item.get("requirements") or ""
# #         eng  = item.get("engagement_type") or item.get("engagement","")
# #         color = fit_colors.get(fit,"blue")
# #         html = f"""<div class="sil-card" style="margin-bottom:.75rem">
# #           <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
# #             <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
# #             {badge(fit or "MAPPED", color) if fit else ""}
# #           </div>"""
# #         if why:  html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
# #         if reqs:
# #             rs = ", ".join(reqs) if isinstance(reqs, list) else str(reqs)
# #             html += f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
# #         if eng:  html += f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
# #         st.markdown(html + '</div>', unsafe_allow_html=True)
# #     render_json_pretty(data, "Service Mapping")

# # def render_microplans(data):
# #     if not data: st.info("No micro-plan data."); return
# #     plans = data if isinstance(data, list) else []
# #     if not plans and isinstance(data, dict):
# #         for k in ("plans","micro_plans","microplans","top_3","improvements"):
# #             if isinstance(data.get(k), list): plans = data[k]; break
# #         if not plans: plans = [data]
# #     for i, plan in enumerate(plans[:3], 1):
# #         if not isinstance(plan, dict): continue
# #         title = plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
# #         weeks = plan.get("duration") or plan.get("timeline","")
# #         obj   = plan.get("objective") or plan.get("goal","")
# #         kpis  = plan.get("kpis") or plan.get("KPIs") or []
# #         tasks = plan.get("tasks") or plan.get("workstreams") or []
# #         with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}", expanded=(i==1)):
# #             if obj and obj != title: st.markdown(f"**Objective:** {obj}")
# #             if kpis:
# #                 st.markdown("**KPIs:**")
# #                 for kpi in (kpis if isinstance(kpis, list) else [kpis]): st.markdown(f"• {kpi}")
# #             if tasks:
# #                 st.markdown("**Tasks:**")
# #                 for t in (tasks if isinstance(tasks, list) else [tasks]):
# #                     if isinstance(t, dict):
# #                         tn = t.get("task") or t.get("name",""); te = t.get("effort") or t.get("duration","")
# #                         st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
# #                     else: st.markdown(f"• {t}")
# #             st.code(json.dumps(plan, indent=2, default=str), language="json")

# # def render_deal_assurance(data):
# #     if not data: st.info("No deal assurance data."); return
# #     if not isinstance(data, dict): render_json_pretty(data, "Deal Assurance Pack"); return
# #     evp = data.get("executive_value_proposition") or data.get("value_proposition") or data.get("ExecutiveValueProposition","")
# #     if evp:
# #         st.markdown('<div class="section-label">💼 Executive Value Proposition</div>', unsafe_allow_html=True)
# #         st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7">{evp}</div>', unsafe_allow_html=True)
# #     caps = data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
# #     if caps:
# #         st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>', unsafe_allow_html=True)
# #         c1, c2 = st.columns(2)
# #         for i, cap in enumerate(caps if isinstance(caps, list) else [caps]):
# #             (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
# #     risk = data.get("risk_mitigation") or data.get("RiskMitigation","")
# #     if risk:
# #         st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>', unsafe_allow_html=True)
# #         if isinstance(risk, dict):
# #             for k, v in risk.items(): st.markdown(f"**{k}:** {v}")
# #         else: st.markdown(str(risk))
# #     render_json_pretty(data, "Deal Assurance Pack")


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  LOGS PAGE
# # # ══════════════════════════════════════════════════════════════════════════════
# # def render_logs_page():
# #     st.markdown("""<div style="margin-bottom:2rem">
# #       <h1 style="font-family:'Syne',sans-serif;font-size:2.2rem;font-weight:800;color:#0f172a;margin:0">⚙️ Pipeline Execution</h1>
# #       <p style="font-size:.95rem;color:#64748b;margin-top:.5rem">Live logs from the scraper, deduplication, AI agents, and Google Sheet sync</p>
# #     </div>""", unsafe_allow_html=True)

# #     if st.session_state.pipeline_running:
# #         st.markdown("""<div class="logs-status running">
# #           <div><div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.3rem">
# #             <span class="pulse-dot"></span>
# #             <span style="font-family:'Syne',sans-serif;font-weight:700;color:#1d4ed8">Pipeline Running</span>
# #           </div>
# #           <span style="font-size:.82rem;color:#64748b">Processing jobs... Google Sheet will auto-sync when complete.</span></div>
# #         </div>""", unsafe_allow_html=True)
# #     elif st.session_state.pipeline_done:
# #         result = st.session_state.pipeline_result or {}
# #         if result.get("success"):
# #             st.markdown(f"""<div class="logs-status success">
# #               <div><div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
# #                 <span style="font-size:1.5rem">✅</span>
# #                 <span style="font-family:'Syne',sans-serif;font-weight:700;color:#15803d;font-size:1.1rem">Pipeline Completed</span>
# #               </div>
# #               <div style="font-size:.85rem;color:#15803d">
# #                 {result.get('processed',0)} processed · {result.get('new_jobs',0)} new · {result.get('skipped_db',0)} already in DB
# #               </div></div></div>""", unsafe_allow_html=True)
# #         else:
# #             st.markdown(f"""<div class="logs-status error">
# #               <div><div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
# #                 <span style="font-size:1.5rem">❌</span>
# #                 <span style="font-family:'Syne',sans-serif;font-weight:700;color:#b91c1c;font-size:1.1rem">Pipeline Failed</span>
# #               </div>
# #               <div style="font-size:.85rem;color:#b91c1c">{result.get('error','Unknown error')}</div></div></div>""", unsafe_allow_html=True)

# #     st.markdown('<div class="section-label" style="margin-top:2rem">📡 Live Execution Logs</div>', unsafe_allow_html=True)
# #     logs = st.session_state.pipeline_logs or []
# #     log_text = "\n".join(logs[-200:]) if logs else "Waiting for logs..."
# #     st.markdown(f'<div class="pipeline-log">{log_text}</div>', unsafe_allow_html=True)

# #     if st.session_state.pipeline_running:
# #         st.markdown("<script>setTimeout(function(){window.location.reload();},1500);</script>", unsafe_allow_html=True)

# #     c1, c2, _ = st.columns([2, 2, 1])
# #     with c1:
# #         if st.button("← Back to Dashboard", use_container_width=True):
# #             st.session_state.current_page = "dashboard"; st.rerun()
# #     with c2:
# #         if st.button("🔄 Refresh Logs", use_container_width=True): st.rerun()


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  SIDEBAR
# # # ══════════════════════════════════════════════════════════════════════════════
# # with st.sidebar:
# #     st.markdown("""<div style="padding:.75rem 0 1.25rem">
# #       <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;color:#2563eb">🛡️ SecureITLab</div>
# #       <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
# #     </div>""", unsafe_allow_html=True)

# #     if st.button("🚪 Logout", use_container_width=True):
# #         st.session_state.logged_in = False; st.session_state.login_error = ""; st.rerun()

# #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# #     # ── Find Jobs ─────────────────────────────────────────────────────────────
# #     st.markdown("**🔍 Find New Jobs**")
# #     st.caption("Scrapes → deduplicates → AI agents → MongoDB → auto-syncs to Google Sheet · Repeats every 12h")

# #     find_jobs_clicked = st.button(
# #         "🔍  Find Jobs",
# #         disabled=st.session_state.pipeline_running,
# #         use_container_width=True,
# #         type="primary",
# #     )

# #     if st.session_state.pipeline_running:
# #         st.markdown("""<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
# #             padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;font-size:.8rem;color:#1d4ed8">
# #           ⏳ Pipeline running… 👉 View Logs Page
# #         </div>""", unsafe_allow_html=True)

# #     # ── Scheduler badge ───────────────────────────────────────────────────────
# #     try:
# #         import main as _main_module
# #         sched = _main_module.get_scheduler_status()
# #         if sched.get("active"):
# #             secs = sched.get("seconds_until_next")
# #             if secs is not None:
# #                 hrs=secs//3600; mins=(secs%3600)//60
# #                 countdown=f"{hrs}h {mins}m" if hrs else f"{mins}m {secs%60}s"
# #                 next_label=f"Next run in: <b>{countdown}</b>"
# #             else:
# #                 next_label="Next run: calculating…"
# #             last=(sched.get("last_run") or "")[:19]
# #             st.markdown(f"""<div class="sched-on">
# #               🟢 <b>Auto-scheduler ON</b><br>
# #               <span style="color:#64748b">{next_label}</span><br>
# #               <span style="color:#64748b">Runs every 12h · #{sched.get('run_count',0)} so far</span><br>
# #               <span style="color:#64748b;font-size:.75rem">Last: {last} UTC</span>
# #             </div>""", unsafe_allow_html=True)
# #             if st.button("⏹️ Stop Auto-Scheduler", use_container_width=True):
# #                 _main_module.stop_scheduler(); st.rerun()
# #         elif sched.get("run_count", 0) > 0:
# #             last=(sched.get("last_run") or "")[:19]
# #             st.markdown(f"""<div class="sched-paused">
# #               ⏸️ <b>Scheduler paused</b><br>
# #               <span style="color:#64748b">Last ran: {last} UTC</span><br>
# #               <span style="color:#64748b">Click Find Jobs to restart</span>
# #             </div>""", unsafe_allow_html=True)
# #     except Exception:
# #         pass

# #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# #     # ══════════════════════════════════════════════════════════════════════════
# #     #  ✅ GOOGLE SHEETS MASTER CONTACTS  (replaces Excel download button)
# #     # ══════════════════════════════════════════════════════════════════════════
# #     st.markdown("**📊 Master Contacts — Google Sheet**")
# #     st.caption("All contacts auto-synced after every pipeline run. New rows only — existing data is never deleted.")

# #     # Manual sync button
# #     sync_col1, sync_col2 = st.columns([3, 1])
# #     with sync_col1:
# #         sync_clicked = st.button("🔄  Sync Now", use_container_width=True)
# #     with sync_col2:
# #         open_clicked = st.button("↗", use_container_width=True, help="Open Google Sheet")

# #     if open_clicked:
# #         st.markdown(f'<meta http-equiv="refresh" content="0; url={GSHEET_URL}">', unsafe_allow_html=True)

# #     if sync_clicked:
# #         try:
# #             jobs_for_sync = load_all_jobs()
# #         except Exception:
# #             jobs_for_sync = []

# #         if not jobs_for_sync:
# #             st.warning("No jobs in MongoDB to sync yet.")
# #         else:
# #             with st.spinner("Syncing to Google Sheet…"):
# #                 res = sync_contacts_to_gsheet(jobs_for_sync)

# #             if res["error"]:
# #                 st.session_state.gsheet_sync_error = res["error"]
# #                 st.session_state.gsheet_last_sync  = None
# #                 st.session_state.gsheet_appended   = None
# #             else:
# #                 st.session_state.gsheet_sync_error = ""
# #                 st.session_state.gsheet_last_sync  = datetime.now().strftime("%d %b %Y  %H:%M")
# #                 st.session_state.gsheet_appended   = res["appended"]

# #     # Read persisted sync state from disk
# #     disk_state = _read_sync_state()
# #     last_sync  = (
# #         st.session_state.gsheet_last_sync
# #         or (disk_state.get("last_sync","")[:16].replace("T"," ") if disk_state.get("last_sync") else None)
# #     )
# #     appended = st.session_state.gsheet_appended if st.session_state.gsheet_appended is not None else disk_state.get("appended","")

# #     if st.session_state.gsheet_sync_error:
# #         # Show error + setup hint
# #         short_err = st.session_state.gsheet_sync_error[:180]
# #         creds_hint = ""
# #         if "not found" in short_err or "No such file" in short_err:
# #             creds_hint = "<br><span style='font-size:.74rem'>💡 Save your service account JSON as <code>google_credentials.json</code> in the project folder, then share the sheet with the service account email.</span>"
# #         st.markdown(f"""<div class="gsheet-error">
# #           ❌ <b>Sync failed</b><br>{short_err}{creds_hint}
# #         </div>""", unsafe_allow_html=True)
# #     else:
# #         last_line  = f"Last synced: <b>{last_sync}</b>" if last_sync else "Not synced yet — click Sync Now above"
# #         added_line = f" · <b>+{appended}</b> rows added" if appended != "" and appended != 0 else (" · no new rows" if appended == 0 and last_sync else "")
# #         st.markdown(f"""<div class="gsheet-box">
# #           🟢 <b>Google Sheet connected</b><br>
# #           <a href="{GSHEET_URL}" target="_blank">📋 Open Master Contacts Sheet ↗</a><br>
# #           <span style="font-size:.76rem;color:#64748b">{last_line}{added_line}</span><br>
# #           <span style="font-size:.73rem;color:#94a3b8">Append-only · duplicates skipped automatically</span>
# #         </div>""", unsafe_allow_html=True)

# #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

# #     # ── Job list ──────────────────────────────────────────────────────────────
# #     try:
# #         all_jobs = load_all_jobs()
# #     except Exception as e:
# #         st.error(f"MongoDB error: {e}"); st.stop()

# #     if not all_jobs:
# #         st.warning("No jobs in MongoDB yet. Click **Find Jobs** to start."); st.stop()

# #     st.markdown(f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">{len(all_jobs)} jobs in database</div>', unsafe_allow_html=True)

# #     search   = st.text_input("🔍 Filter jobs", placeholder="e.g. Bounteous")
# #     filtered = [j for j in all_jobs if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

# #     def job_label(j):
# #         score = j.get("opp_score"); s = f" [{score}/10]" if score else ""
# #         ok = "✅" if j.get("pipeline_ok") else "❌"
# #         return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s}"

# #     if not filtered:
# #         st.warning("No matching jobs."); st.stop()

# #     sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
# #     selected_id = str(filtered[[job_label(j) for j in filtered].index(sel_label)]["_id"])

# #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# #     ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
# #     total_c  = sum(j.get("contacts_found", 0) for j in all_jobs)
# #     st.markdown(f"""<div style="font-size:.75rem;color:#64748b">
# #       <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
# #       <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
# #     </div>""", unsafe_allow_html=True)

# #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# #     if st.button("🔄 Refresh Data", use_container_width=True):
# #         st.cache_data.clear(); st.rerun()


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  FIND JOBS — background thread
# # # ══════════════════════════════════════════════════════════════════════════════
# # if find_jobs_clicked and not st.session_state.pipeline_running:
# #     with _thread_lock:
# #         _thread_logs.clear(); _thread_result[0] = None; _thread_done[0] = False

# #     st.session_state.pipeline_running = True
# #     st.session_state.pipeline_done    = False
# #     st.session_state.pipeline_logs    = []
# #     st.session_state.pipeline_result  = None
# #     st.session_state.current_page     = "logs"
# #     st.cache_data.clear()

# #     def _run_pipeline_bg():
# #         try:
# #             import main as _main

# #             def _cb(msg: str):
# #                 with _thread_lock: _thread_logs.append(msg)

# #             result = _main.run_pipeline(progress_callback=_cb)

# #             # ── Auto-sync Google Sheet after every successful run ────────────
# #             if result.get("success"):
# #                 try:
# #                     _cb("🔄 [GSheet] Syncing new contacts to Google Sheet…")
# #                     sync_res = sync_contacts_to_gsheet(load_all_jobs())
# #                     if sync_res["error"]:
# #                         _cb(f"⚠️  [GSheet] Sync error: {sync_res['error']}")
# #                     else:
# #                         _cb(f"✅ [GSheet] Done — +{sync_res['appended']} new rows · {sync_res['skipped']} already existed")
# #                 except Exception as se:
# #                     _cb(f"⚠️  [GSheet] Sync failed: {se}")

# #             with _thread_lock: _thread_result[0] = result

# #         except Exception as e:
# #             import traceback
# #             with _thread_lock:
# #                 _thread_logs.append(f"❌ Unexpected error: {e}")
# #                 _thread_logs.append(traceback.format_exc())
# #                 _thread_result[0] = {"success":False,"error":str(e),"scraped":0,"new_jobs":0,"skipped_db":0,"processed":0}
# #         finally:
# #             with _thread_lock: _thread_done[0] = True

# #     threading.Thread(target=_run_pipeline_bg, daemon=True).start()
# #     st.rerun()

# # # ── Sync thread state → session_state ────────────────────────────────────────
# # with _thread_lock:
# #     if _thread_logs:
# #         st.session_state.pipeline_logs = list(_thread_logs)
# #     if _thread_done[0] and _thread_result[0] is not None:
# #         st.session_state.pipeline_result  = _thread_result[0]
# #         st.session_state.pipeline_running = False
# #         st.session_state.pipeline_done    = True


# # # ══════════════════════════════════════════════════════════════════════════════
# # #  PAGE ROUTING
# # # ══════════════════════════════════════════════════════════════════════════════
# # if st.session_state.current_page == "logs":
# #     render_logs_page()
# #     st.stop()

# # with st.spinner("Loading job…"):
# #     job = load_job(selected_id)

# # if not job:
# #     st.error("Could not load job document."); st.stop()

# # company   = job.get("company","Unknown"); role  = job.get("role","Unknown")
# # opp_score = job.get("opp_score");         p_ok  = job.get("pipeline_ok",False)
# # p_min     = job.get("pipeline_min","?");  c_found = job.get("contacts_found",0)
# # c_cov     = job.get("coverage_score");    c_domain = job.get("contact_domain","")
# # run_at    = job.get("run_at","")

# # st.markdown(f"""<div style="margin-bottom:1.75rem">
# #   <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
# #     Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
# #   </div>
# #   <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;color:#0f172a;margin:0;line-height:1.15">{role}</h1>
# #   <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
# #     @ <span style="color:#334155;font-weight:600">{company}</span>
# #     {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
# #   </div>
# # </div>""", unsafe_allow_html=True)

# # try:
# #     sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
# #     sc = "#16a34a" if sn>=7 else "#f59e0b" if sn>=5 else "#dc2626"
# # except Exception:
# #     sc = "#2563eb"

# # st.markdown(f"""<div class="metric-row">
# #   <div class="metric-tile"><div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div><div class="lbl">Opportunity Score</div></div>
# #   <div class="metric-tile"><div class="val">{c_found}</div><div class="lbl">Contacts Found</div></div>
# #   <div class="metric-tile"><div class="val">{f"{c_cov}%" if c_cov else "—"}</div><div class="lbl">Contact Coverage</div></div>
# #   <div class="metric-tile"><div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">{'✅ OK' if p_ok else '❌ Failed'}</div><div class="lbl">Pipeline ({p_min} min)</div></div>
# # </div>""", unsafe_allow_html=True)

# # tabs = st.tabs(["📋 Job & Enrichment","🎯 Service Mapping","🔍 Fit / Gap",
# #                 "🛠️ Capability & Plans","📦 Deal Assurance","✉️ Outreach Emails",
# #                 "👥 Contacts","✅ QA Gates","🗄️ Raw Data"])

# # with tabs[0]:
# #     c1, c2 = st.columns(2)
# #     with c1:
# #         st.markdown('<div class="section-label">📄 Job Research</div>', unsafe_allow_html=True)
# #         jr = as_dict(job.get("agent_job_research") or {})
# #         if jr:
# #             for lbl, keys in [("Job Role",["job_role","Job Role","role","title"]),("Company",["company_name","Company Name","company"]),("Location",["location","Location"]),("URL",["organization_url","Organization URL","url"])]:
# #                 val = next((jr.get(k) for k in keys if jr.get(k)), None)
# #                 if val: st.markdown(f"**{lbl}:** {val}")
# #             desc = jr.get("job_description") or jr.get("Job Description","")
# #             if desc:
# #                 st.markdown("**Job Description:**")
# #                 st.markdown(f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;max-height:300px;overflow-y:auto">{desc[:2000]}</div>', unsafe_allow_html=True)
# #             render_json_pretty(jr,"Job Research")
# #         else: st.info("No job research data.")
# #     with c2:
# #         st.markdown('<div class="section-label">🏢 Company Enrichment</div>', unsafe_allow_html=True)
# #         enr = as_dict(job.get("agent_enrichment") or {})
# #         if enr:
# #             for lbl, keys in [("Industry",["industry","Industry"]),("Company Size",["company_size","size"]),("Regulatory Env",["regulatory_environment","regulatory"]),("Certifications",["certifications","Certifications"]),("Tech Stack",["tech_stack","technology_stack"]),("Security Maturity",["security_maturity","maturity"])]:
# #                 val = next((enr.get(k) for k in keys if enr.get(k)), None)
# #                 if val:
# #                     if isinstance(val, list): val = ", ".join(str(v) for v in val)
# #                     st.markdown(f"**{lbl}:** {safe_str(val,200)}")
# #             render_json_pretty(enr,"Enrichment")
# #         else: st.info("No enrichment data.")

# # with tabs[1]:
# #     st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>', unsafe_allow_html=True)
# #     render_service_mapping(job.get("agent_service_mapping"))

# # with tabs[2]:
# #     fg = as_dict(job.get("agent_fit_gap") or {})
# #     if opp_score:
# #         try:
# #             sn=float(str(opp_score).split("/")[0]); bc="#16a34a" if sn>=7 else "#f59e0b" if sn>=5 else "#dc2626"; bp=int(sn/10*100)
# #             st.markdown(f"""<div style="margin-bottom:1.5rem">
# #               <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
# #                 <span style="font-family:'Syne',sans-serif;font-weight:700">Opportunity Score</span>
# #                 <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
# #               </div>
# #               <div style="background:#e2e8f0;border-radius:4px;height:8px">
# #                 <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
# #               </div></div>""", unsafe_allow_html=True)
# #         except Exception: pass
# #     st.markdown('<div class="section-label">📊 Service Classifications</div>', unsafe_allow_html=True)
# #     services = []
# #     if isinstance(fg, dict):
# #         for k in ("services","classifications","service_classifications","items","fit_gap"):
# #             v = fg.get(k)
# #             if isinstance(v, list): services = v; break
# #         if not services and (fg.get("service") or fg.get("Service")): services = [fg]
# #     elif isinstance(fg, list): services = fg
# #     if services:
# #         buckets = {"STRONG FIT":[],"PARTIAL FIT":[],"GAP":[]}
# #         for s in services:
# #             if not isinstance(s, dict): continue
# #             fit = (s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
# #             if "STRONG" in fit: buckets["STRONG FIT"].append(s)
# #             elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
# #             elif "GAP" in fit: buckets["GAP"].append(s)
# #         c1,c2,c3=st.columns(3)
# #         cm={"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
# #         bgm={"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
# #         bdm={"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
# #         for col,(fl,items) in zip([c1,c2,c3],buckets.items()):
# #             col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>', unsafe_allow_html=True)
# #             for s in items:
# #                 svc=s.get("service") or s.get("Service") or s.get("name",""); just=s.get("justification") or s.get("reason","")
# #                 col.markdown(f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem"><div style="font-weight:600">{svc}</div><div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>', unsafe_allow_html=True)
# #     elif fg: st.json(fg)
# #     else: st.info("No fit/gap data.")
# #     render_json_pretty(job.get("agent_fit_gap"),"Fit/Gap Analysis")

# # with tabs[3]:
# #     c1, c2 = st.columns(2)
# #     with c1:
# #         st.markdown('<div class="section-label">🔧 Capability Improvements</div>', unsafe_allow_html=True)
# #         cap = job.get("agent_capability") or {}
# #         items_cap = cap if isinstance(cap,list) else []
# #         if not items_cap and isinstance(cap,dict):
# #             for k in ("improvements","recommendations","capabilities","items"):
# #                 v=cap.get(k)
# #                 if isinstance(v,list): items_cap=v; break
# #             if not items_cap: items_cap=[cap]
# #         for item in items_cap:
# #             if not isinstance(item,dict): continue
# #             title=item.get("title") or item.get("gap") or item.get("service","")
# #             rec=item.get("recommendation") or item.get("steps","")
# #             effort=item.get("build_effort") or item.get("effort",""); demand=item.get("market_demand") or item.get("priority","")
# #             st.markdown(f"""<div class="sil-card" style="margin-bottom:.6rem">
# #               <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem">{title}</div>
# #               <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec,250)}</div>
# #               {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
# #             </div>""", unsafe_allow_html=True)
# #         if not items_cap: render_json_pretty(cap,"Capability Improvement")
# #     with c2:
# #         st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>', unsafe_allow_html=True)
# #         render_microplans(job.get("agent_microplans"))

# # with tabs[4]:
# #     render_deal_assurance(job.get("agent_deal_assurance"))

# # with tabs[5]:
# #     st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>', unsafe_allow_html=True)
# #     emails_src = job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
# #     oq = as_dict(job.get("agent_outreach_qa") or {})
# #     improved = (oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
# #     if improved:
# #         st.info("⚡ Showing QA-improved versions")
# #         render_emails(improved)
# #         with st.expander("📬 Original (pre-QA) versions"): render_emails(emails_src)
# #     else: render_emails(emails_src)

# # with tabs[6]:
# #     contacts = job.get("contacts") or []; contact_sources = job.get("contact_sources") or []
# #     pri=[c for c in contacts if c.get("priority")=="Primary"]
# #     sec=[c for c in contacts if c.get("priority")=="Secondary"]
# #     ter=[c for c in contacts if c.get("priority")=="Tertiary"]
# #     gen=[c for c in contacts if c.get("priority")=="General"]
# #     st.markdown(f"""<div class="metric-row" style="margin-bottom:1.5rem">
# #       <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
# #       <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
# #       <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
# #       <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
# #     </div>""", unsafe_allow_html=True)
# #     if contact_sources:
# #         st.markdown('Sources: '+" ".join(badge(s,"blue") for s in contact_sources), unsafe_allow_html=True); st.markdown("")
# #     missing = job.get("missing_roles") or []
# #     if missing:
# #         st.markdown('⚠️ Missing roles: '+" ".join(badge(r,"red") for r in missing), unsafe_allow_html=True); st.markdown("")
# #     if contacts:
# #         excel_bytes = build_contacts_excel(contacts, company, role)
# #         if excel_bytes:
# #             safe_co = re.sub(r'[^a-z0-9]','_',company.lower())[:20]
# #             fname = f"contacts_{safe_co}_{datetime.now().strftime('%Y%m%d')}.xlsx"
# #             btn_col, _ = st.columns([1,3])
# #             with btn_col:
# #                 st.download_button("📥  Download Contacts (.xlsx)", data=excel_bytes, file_name=fname,
# #                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")
# #         st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# #         pri_filter = st.multiselect("Filter by priority",["Primary","Secondary","Tertiary","General"],default=["Primary","Secondary","Tertiary","General"])
# #         shown = [c for c in contacts if c.get("priority","General") in pri_filter]
# #         render_contacts(shown, f"Contacts ({len(shown)} shown)")
# #         agent_contacts = job.get("agent_prospect_contacts") or {}
# #         if agent_contacts:
# #             with st.expander("🤖 CrewAI Agent's Contact Search"):
# #                 if isinstance(agent_contacts,dict):
# #                     ac_list=agent_contacts.get("contacts") or []
# #                     if ac_list: render_contacts(ac_list,"Agent Contacts")
# #                     else: st.json(agent_contacts)
# #                 else: st.json(agent_contacts)
# #     else: st.info("No contacts found for this job.")

# # with tabs[7]:
# #     st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>', unsafe_allow_html=True)
# #     c1, c2 = st.columns(2)
# #     for i,(lbl,key) in enumerate([("Research QA","agent_research_qa"),("Service Mapping QA","agent_mapping_qa"),("Deal Assurance QA","agent_assurance_qa"),("Outreach Email QA","agent_outreach_qa")]):
# #         with (c1 if i%2==0 else c2): render_qa_block(job.get(key), lbl)
# #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# #     st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>', unsafe_allow_html=True)
# #     enf = as_dict(job.get("agent_prospect_enforcer") or {})
# #     if enf:
# #         cov=enf.get("coverage_score","?"); miss=enf.get("missing_roles",[]); note=enf.get("note",""); ec=enf.get("contacts",[])
# #         x1,x2,x3=st.columns(3)
# #         x1.metric("Coverage Score",f"{cov}%"); x2.metric("Missing Roles",len(miss)); x3.metric("Contacts Verified",len(ec))
# #         if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
# #         if note: st.caption(note)
# #     else: st.info("No enforcer data.")

# # with tabs[8]:
# #     st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>', unsafe_allow_html=True)
# #     rows=[]
# #     for k,v in job.items():
# #         if k=="_id": continue
# #         rows.append({"Field":k,"Type":type(v).__name__,"Len":len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
# #     hc1,hc2,hc3=st.columns([3,1,1])
# #     hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
# #     for r in rows:
# #         rc1,rc2,rc3=st.columns([3,1,1])
# #         rc1.code(r["Field"],language=None)
# #         rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>', unsafe_allow_html=True)
# #         rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>', unsafe_allow_html=True)
# #     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
# #     for lbl,key in [("Job Research","agent_job_research"),("Enrichment","agent_enrichment"),("Service Mapping","agent_service_mapping"),("Fit/Gap","agent_fit_gap"),("Capability","agent_capability"),("Micro-Plans","agent_microplans"),("Deal Assurance","agent_deal_assurance"),("Outreach Emails","agent_outreach_emails"),("Prospect Contacts","agent_prospect_contacts"),("Prospect Enforcer","agent_prospect_enforcer"),("Research QA","agent_research_qa"),("Mapping QA","agent_mapping_qa"),("Assurance QA","agent_assurance_qa"),("Outreach QA","agent_outreach_qa"),("Contacts (5-source)","contacts")]:
# #         data=job.get(key)
# #         if data:
# #             with st.expander(f"📄 {lbl}"):
# #                 st.code(json.dumps(data,indent=2,default=str),language="json")








































# """
# ╔══════════════════════════════════════════════════════════╗
# ║   SecureITLab Pipeline Dashboard — Streamlit             ║
# ║   WITH GOOGLE SHEETS AUTO-SYNC (Master Contacts)         ║
# ║   WITH 12-HOUR AUTO-SCHEDULER STATUS IN SIDEBAR          ║
# ║   Reads from MongoDB → secureitlab_job_pipeline          ║
# ╠══════════════════════════════════════════════════════════╣
# ║  Install: pip install streamlit pymongo python-dotenv    ║
# ║           gspread google-auth openpyxl                   ║
# ║  Run:     streamlit run streamlit_dashboard.py           ║
# ╚══════════════════════════════════════════════════════════╝
# """

# import io
# import re
# import streamlit as st
# from pymongo import MongoClient
# import json
# import threading
# import time
# import logging
# from datetime import datetime, timezone
# from io import StringIO
# from pathlib import Path

# # ── Thread-safe shared state ──────────────────────────────────────────────────
# _thread_logs   = []
# _thread_result = [None]
# _thread_done   = [False]
# _thread_lock   = threading.Lock()

# # ── Google Sheets config ──────────────────────────────────────────────────────
# GSHEET_URL        = "https://docs.google.com/spreadsheets/d/1u9_Fqy8a8Dj-yX8FwZtkFfw_3BLdpmBbtagH2IjhyV4/edit"
# GSHEET_ID         = "1u9_Fqy8a8Dj-yX8FwZtkFfw_3BLdpmBbtagH2IjhyV4"
# GSHEET_TAB        = "master_contacts"
# GCREDS_FILE       = "google_credentials.json"   # service account JSON in project folder
# GSHEET_SYNC_STATE = "gsheet_sync_state.json"    # persists last sync info across reruns

# # ── Log capture ───────────────────────────────────────────────────────────────
# _log_capture  = StringIO()
# _log_handler  = logging.StreamHandler(_log_capture)
# _log_handler.setLevel(logging.INFO)
# _log_handler.setFormatter(logging.Formatter('%(asctime)s | %(levelname)s | %(message)s'))

# # ── Page config ───────────────────────────────────────────────────────────────
# st.set_page_config(
#     page_title="SecureITLab Pipeline",
#     page_icon="🛡️",
#     layout="wide",
#     initial_sidebar_state="expanded",
# )

# LOGIN_USERNAME = "admin"
# LOGIN_PASSWORD = "secureitlab2024"

# # ══════════════════════════════════════════════════════════════════════════════
# #  GLOBAL CSS
# # ══════════════════════════════════════════════════════════════════════════════
# st.markdown("""
# <style>
# @import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');
# :root{--bg:#f4f7fb;--surface:#ffffff;--surface2:#eef2f7;--border:#d9e2ec;--accent:#2563eb;--accent2:#7c3aed;--green:#16a34a;--yellow:#f59e0b;--red:#dc2626;--text:#0f172a;--muted:#64748b;}
# html,body,[class*="css"]{background-color:var(--bg)!important;color:var(--text)!important;font-family:'DM Sans',sans-serif!important;}
# .login-card{background:var(--surface);border:1px solid var(--border);border-radius:20px;padding:3rem 3.5rem;width:100%;max-width:420px;box-shadow:0 20px 60px rgba(37,99,235,0.08);text-align:center;}
# .login-logo{font-family:'Syne',sans-serif;font-size:1.6rem;font-weight:800;color:var(--accent);margin-bottom:.25rem;}
# .login-subtitle{font-size:.75rem;color:var(--muted);letter-spacing:.12em;text-transform:uppercase;margin-bottom:2.5rem;}
# .login-error{background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:.65rem 1rem;color:#b91c1c;font-size:.85rem;margin-top:1rem;}
# .login-divider{width:40px;height:3px;background:linear-gradient(90deg,var(--accent),var(--accent2));border-radius:2px;margin:0 auto 2rem;}
# [data-testid="stSidebar"]{background:var(--surface)!important;border-right:1px solid var(--border)!important;}
# [data-testid="stSidebar"] *{color:var(--text)!important;}
# .main .block-container{padding:2rem 2rem 3rem!important;}
# h1,h2,h3,h4{font-family:'Syne',sans-serif!important;color:var(--text)!important;}
# .sil-card{background:var(--surface);border:1px solid var(--border);border-radius:14px;padding:1.25rem 1.5rem;margin-bottom:1rem;transition:all 0.25s ease;}
# .sil-card:hover{transform:translateY(-2px);box-shadow:0 8px 22px rgba(0,0,0,0.05);}
# .sil-card-accent{border-left:4px solid var(--accent);}
# .metric-row{display:flex;gap:1rem;flex-wrap:wrap;margin-bottom:1.5rem;}
# .metric-tile{flex:1;min-width:140px;background:var(--surface2);border:1px solid var(--border);border-radius:12px;padding:1rem;text-align:center;transition:all .25s ease;}
# .metric-tile:hover{transform:translateY(-3px);box-shadow:0 10px 24px rgba(0,0,0,0.06);}
# .metric-tile .val{font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;color:var(--accent);}
# .metric-tile .lbl{font-size:.72rem;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;}
# .badge{padding:.25rem .7rem;border-radius:20px;font-size:.72rem;font-weight:600;font-family:'DM Mono',monospace;}
# .badge-green{background:#ecfdf5;color:#15803d;}
# .badge-yellow{background:#fffbeb;color:#b45309;}
# .badge-red{background:#fef2f2;color:#b91c1c;}
# .badge-blue{background:#eff6ff;color:#1d4ed8;}
# .badge-purple{background:#f5f3ff;color:#6d28d9;}
# .contact-card{background:var(--surface2);border:1px solid var(--border);border-radius:10px;padding:1rem;margin-bottom:.8rem;}
# .contact-name{font-family:'Syne',sans-serif;font-weight:700;color:var(--text);}
# .contact-title{color:var(--muted);font-size:.85rem;}
# .contact-email{font-family:'DM Mono',monospace;color:var(--accent);font-size:.8rem;}
# .email-box{background:#f8fafc;border:1px solid var(--border);border-radius:10px;padding:1rem 1.25rem;font-size:.9rem;line-height:1.65;white-space:pre-wrap;color:var(--text);}
# .email-subject{font-family:'Syne',sans-serif;font-weight:700;color:var(--accent);margin-bottom:.5rem;}
# .section-label{font-family:'DM Mono',monospace;font-size:.72rem;text-transform:uppercase;letter-spacing:.12em;color:var(--accent);margin-bottom:.6rem;}
# .sil-divider{border-top:1px solid var(--border);margin:1rem 0;}
# [data-testid="stExpander"]{background:var(--surface)!important;border:1px solid var(--border)!important;border-radius:10px!important;}
# [data-testid="stTabs"] button{font-family:'Syne',sans-serif!important;font-weight:600!important;}
# ::-webkit-scrollbar{width:6px;}
# ::-webkit-scrollbar-thumb{background:var(--border);border-radius:3px;}
# .pipeline-log{background:#0f172a;color:#10b981;border-radius:10px;padding:1.5rem;font-family:'DM Mono',monospace;font-size:.8rem;line-height:1.8;max-height:700px;overflow-y:auto;white-space:pre-wrap;word-break:break-word;border:1px solid #1e293b;}
# .logs-status{display:flex;gap:1rem;justify-content:space-between;align-items:center;margin-bottom:1.5rem;padding:1rem;background:var(--surface2);border-radius:10px;border:1px solid var(--border);}
# .logs-status.running{background:#eff6ff;border-color:#bfdbfe;}
# .logs-status.success{background:#f0fdf4;border-color:#bbf7d0;}
# .logs-status.error{background:#fef2f2;border-color:#fecaca;}
# @keyframes pulse{0%,100%{opacity:1;}50%{opacity:0.5;}}
# .pulse-dot{display:inline-block;width:10px;height:10px;background:#2563eb;border-radius:50%;animation:pulse 2s infinite;margin-right:8px;}
# .sched-on{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:.65rem .9rem;margin-top:.5rem;font-size:.8rem;color:#15803d;line-height:1.6;}
# .sched-paused{background:#fef9c3;border:1px solid #fde68a;border-radius:8px;padding:.65rem .9rem;margin-top:.5rem;font-size:.8rem;color:#92400e;line-height:1.6;}

# /* ── Google Sheet box ── */
# .gsheet-box{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:.9rem 1rem;margin-top:.4rem;font-size:.82rem;color:#15803d;line-height:1.8;}
# .gsheet-box a{color:#1d4ed8!important;font-weight:700;text-decoration:none;}
# .gsheet-box a:hover{text-decoration:underline;}
# .gsheet-syncing{background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;padding:.9rem 1rem;margin-top:.4rem;font-size:.82rem;color:#1d4ed8;line-height:1.8;}
# .gsheet-error{background:#fef2f2;border:1px solid #fecaca;border-radius:10px;padding:.75rem 1rem;font-size:.8rem;color:#b91c1c;margin-top:.4rem;line-height:1.6;}
# </style>
# """, unsafe_allow_html=True)


# # ══════════════════════════════════════════════════════════════════════════════
# #  SESSION STATE
# # ══════════════════════════════════════════════════════════════════════════════
# for _k, _v in [
#     ("logged_in",         False),
#     ("login_error",       ""),
#     ("pipeline_running",  False),
#     ("pipeline_logs",     []),
#     ("pipeline_result",   None),
#     ("pipeline_done",     False),
#     ("current_page",      "dashboard"),
#     ("gsheet_sync_error", ""),
#     ("gsheet_last_sync",  None),
#     ("gsheet_appended",   None),
# ]:
#     if _k not in st.session_state:
#         st.session_state[_k] = _v


# # ══════════════════════════════════════════════════════════════════════════════
# #  GOOGLE SHEETS HELPERS
# # ══════════════════════════════════════════════════════════════════════════════

# def _gsheet_client():
#     import gspread
#     from google.oauth2.service_account import Credentials
#     scopes = [
#         "https://www.googleapis.com/auth/spreadsheets",
#         "https://www.googleapis.com/auth/drive",
#     ]
#     creds = Credentials.from_service_account_file(GCREDS_FILE, scopes=scopes)
#     return gspread.authorize(creds)


# def _existing_contact_keys(ws) -> set:
#     """Return set of (name_lower, company_lower) already in the sheet."""
#     try:
#         rows = ws.get_all_values()
#         keys = set()
#         for row in rows[3:]:        # skip 3 header rows
#             name    = row[5].strip().lower() if len(row) > 5 else ""
#             company = row[2].strip().lower() if len(row) > 2 else ""
#             if name:
#                 keys.add((name, company))
#         return keys
#     except Exception:
#         return set()


# def _write_sync_state(appended: int, skipped: int):
#     try:
#         Path(GSHEET_SYNC_STATE).write_text(json.dumps({
#             "last_sync": datetime.now(timezone.utc).isoformat(),
#             "appended":  appended,
#             "skipped":   skipped,
#         }), encoding="utf-8")
#     except Exception:
#         pass


# def _read_sync_state() -> dict:
#     try:
#         if Path(GSHEET_SYNC_STATE).exists():
#             return json.loads(Path(GSHEET_SYNC_STATE).read_text(encoding="utf-8"))
#     except Exception:
#         pass
#     return {}


# def sync_contacts_to_gsheet(all_jobs: list) -> dict:
#     """
#     Append NEW contacts from MongoDB jobs list to the Google Sheet.
#     Never deletes existing rows — skips any contact already present.
#     Returns: {"appended": int, "skipped": int, "error": str|None}
#     """
#     result = {"appended": 0, "skipped": 0, "error": None}

#     if not Path(GCREDS_FILE).exists():
#         result["error"] = (
#             f"'{GCREDS_FILE}' not found in project folder. "
#             "Create a service account on Google Cloud, download the JSON key, "
#             "save it as google_credentials.json, then share your sheet with the service account email."
#         )
#         return result

#     try:
#         gc = _gsheet_client()
#         sh = gc.open_by_key(GSHEET_ID)

#         # Get or create the worksheet tab
#         try:
#             ws = sh.worksheet(GSHEET_TAB)
#         except Exception:
#             ws = sh.add_worksheet(title=GSHEET_TAB, rows=10000, cols=13)

#         # Write headers if sheet is completely empty
#         all_vals = ws.get_all_values()
#         if not all_vals or len(all_vals) < 3:
#             header_rows = [
#                 ["Master Contacts — SecureITLab Pipeline Auto-Sync"] + [""] * 12,
#                 [f"Last synced: {datetime.now().strftime('%d %b %Y  %H:%M')} UTC  ·  Append-only, duplicates skipped"] + [""] * 12,
#                 ["#", "Job Role", "Company", "Country", "Priority",
#                  "Name", "Title / Role", "Contact Email", "LinkedIn URL", "Source", "Job Score",
#                  "Outreach Email (Agent)", "LinkedIn Message (Agent)"],
#             ]
#             ws.update("A1", header_rows)
#         else:
#             # Upgrade headers if new columns are missing (existing 11-col sheet)
#             header_row = all_vals[2] if len(all_vals) >= 3 else []
#             if len(header_row) < 13:
#                 ws.update("L3", [["Outreach Email (Agent)", "LinkedIn Message (Agent)"]])
#             # Always refresh the subtitle row with latest sync time
#             ws.update("A2", [[
#                 f"Last synced: {datetime.now().strftime('%d %b %Y  %H:%M')} UTC  ·  Append-only, duplicates skipped"
#             ]])

#         # Collect keys already in the sheet
#         existing_keys = _existing_contact_keys(ws)
#         current_row_count = max(0, len(ws.get_all_values()) - 3)  # data rows only

#         rows_to_add = []
#         for job in all_jobs:
#             company   = job.get("company", "")
#             role      = job.get("role", "")
#             country   = job.get("country", "?")
#             job_score = str(job.get("opp_score", ""))
#             contacts  = job.get("contacts", [])

#             # ── Extract outreach email from agent (Variant A preferred) ───────
#             outreach_email_text = ""
#             emails_data = job.get("agent_outreach_emails") or {}
#             # Also try QA-improved version first
#             oq = job.get("agent_outreach_qa") or {}
#             if isinstance(oq, dict):
#                 improved = oq.get("improved_emails") or oq.get("ImprovedEmails")
#                 if improved and isinstance(improved, dict):
#                     emails_data = improved
#             if isinstance(emails_data, dict):
#                 # Try Variant A first, then B, then first available
#                 for k, v in emails_data.items():
#                     kl = k.lower().replace("_","").replace(" ","")
#                     if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl == "a":
#                         if isinstance(v, dict):
#                             subj = v.get("subject") or v.get("Subject","")
#                             body = v.get("body") or v.get("Body") or v.get("content","")
#                             outreach_email_text = f"Subject: {subj}\n\n{body}" if subj else body
#                         elif isinstance(v, str):
#                             outreach_email_text = v
#                         break
#                 if not outreach_email_text:
#                     # fallback: first key
#                     for v in emails_data.values():
#                         if isinstance(v, dict):
#                             subj = v.get("subject") or v.get("Subject","")
#                             body = v.get("body") or v.get("Body") or v.get("content","")
#                             outreach_email_text = f"Subject: {subj}\n\n{body}" if subj else body
#                         elif isinstance(v, str):
#                             outreach_email_text = v
#                         if outreach_email_text:
#                             break

#             # ── Extract LinkedIn message from agent ───────────────────────────
#             linkedin_msg_text = ""
#             li_sequences = job.get("agent_linkedin_sequences") or job.get("agent_linkedin") or {}
#             if isinstance(li_sequences, dict):
#                 # Try to get connection request message or first message
#                 for k in ("connection_request","connection_message","message_1","first_message","intro","sequence_1","message"):
#                     v = li_sequences.get(k) or li_sequences.get(k.title()) or li_sequences.get(k.upper())
#                     if v and isinstance(v, str):
#                         linkedin_msg_text = v; break
#                     elif v and isinstance(v, dict):
#                         linkedin_msg_text = v.get("message") or v.get("text") or v.get("content","")
#                         if linkedin_msg_text: break
#                 if not linkedin_msg_text:
#                     # Try sequences list
#                     for k in ("sequences","messages","steps"):
#                         seq = li_sequences.get(k)
#                         if isinstance(seq, list) and seq:
#                             first = seq[0]
#                             if isinstance(first, dict):
#                                 linkedin_msg_text = first.get("message") or first.get("text") or first.get("content","")
#                             elif isinstance(first, str):
#                                 linkedin_msg_text = first
#                             if linkedin_msg_text: break
#                 if not linkedin_msg_text:
#                     # Last resort: dump first string value found
#                     for v in li_sequences.values():
#                         if isinstance(v, str) and len(v) > 20:
#                             linkedin_msg_text = v; break

#             elif isinstance(li_sequences, list) and li_sequences:
#                 first = li_sequences[0]
#                 if isinstance(first, dict):
#                     linkedin_msg_text = (first.get("message") or first.get("connection_request")
#                                          or first.get("text") or first.get("content",""))
#                 elif isinstance(first, str):
#                     linkedin_msg_text = first

#             for ci, contact in enumerate(contacts):
#                 name     = contact.get("name", "").strip()
#                 prio     = contact.get("priority", "General")
#                 title    = contact.get("title", "")
#                 email    = contact.get("email", "")
#                 li       = contact.get("linkedin_url", "")
#                 source   = contact.get("source", "")
#                 patterns = contact.get("email_patterns", [])

#                 if not email and patterns:
#                     email = patterns[0] + "  (pattern)"

#                 key = (name.lower(), company.strip().lower())
#                 if not name or key in existing_keys:
#                     result["skipped"] += 1
#                     continue

#                 rows_to_add.append([
#                     current_row_count + len(rows_to_add) + 1,   # row #
#                     role    if ci == 0 else "",
#                     company if ci == 0 else "",
#                     country if ci == 0 else "",
#                     prio, name, title, email, li, source,
#                     job_score if ci == 0 else "",
#                     outreach_email_text if ci == 0 else "",   # col L — Outreach Email
#                     linkedin_msg_text   if ci == 0 else "",   # col M — LinkedIn Message
#                 ])
#                 existing_keys.add(key)  # dedup within this batch too

#         if rows_to_add:
#             ws.append_rows(rows_to_add, value_input_option="USER_ENTERED")
#             result["appended"] = len(rows_to_add)

#         _write_sync_state(result["appended"], result["skipped"])

#     except Exception as e:
#         result["error"] = str(e)

#     return result


# # ══════════════════════════════════════════════════════════════════════════════
# #  LOGIN
# # ══════════════════════════════════════════════════════════════════════════════
# if not st.session_state.logged_in:
#     _, col, _ = st.columns([1, 1.2, 1])
#     with col:
#         st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)
#         st.markdown("""
#         <div class="login-card">
#           <div class="login-logo">🛡️ SecureITLab</div>
#           <div class="login-subtitle">Pipeline Intelligence</div>
#           <div class="login-divider"></div>
#         </div>""", unsafe_allow_html=True)
#         username = st.text_input("Username", placeholder="Enter username", key="lu")
#         password = st.text_input("Password", placeholder="Enter password", type="password", key="lp")
#         if st.button("Sign In →", use_container_width=True, type="primary"):
#             if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
#                 st.session_state.logged_in = True; st.session_state.login_error = ""; st.rerun()
#             else:
#                 st.session_state.login_error = "Incorrect username or password."
#         if st.session_state.login_error:
#             st.markdown(f'<div class="login-error">⚠️ {st.session_state.login_error}</div>', unsafe_allow_html=True)
#         st.markdown("<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>SecureITLab · Confidential</div>", unsafe_allow_html=True)
#     st.stop()


# # ══════════════════════════════════════════════════════════════════════════════
# #  PER-JOB CONTACTS EXCEL  (individual job download — kept)
# # ══════════════════════════════════════════════════════════════════════════════
# def build_contacts_excel(contacts: list, company: str, role: str):
#     try:
#         from openpyxl import Workbook
#         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
#         from openpyxl.utils import get_column_letter
#     except ImportError:
#         return None
#     wb = Workbook(); ws = wb.active; ws.title = "Contacts"
#     NAVY="1E3A5F"; BLUE="2563EB"; GREY="F8FAFC"; WHITE="FFFFFF"
#     pri_colors={"Primary":("FEF2F2","B91C1C"),"Secondary":("FFFBEB","B45309"),"Tertiary":("EFF6FF","1D4ED8"),"General":("F5F3FF","6D28D9")}
#     thin=Side(border_style="thin",color="D9E2EC"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
#     ws.merge_cells("A1:H1"); c=ws["A1"]
#     c.value=f"Contacts Export  —  {company}  |  {role}"
#     c.font=Font(name="Arial",bold=True,size=13,color=WHITE); c.fill=PatternFill("solid",fgColor=NAVY)
#     c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=30
#     ws.merge_cells("A2:H2"); c=ws["A2"]
#     c.value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(contacts)} contacts"
#     c.font=Font(name="Arial",size=9,color="64748B"); c.fill=PatternFill("solid",fgColor="F4F7FB")
#     c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=18
#     headers=["#","Priority","Name","Title / Role","Company","Email","LinkedIn URL","Source"]; col_widths=[4,12,24,32,22,34,42,18]
#     for ci,(h,w) in enumerate(zip(headers,col_widths),1):
#         c=ws.cell(row=3,column=ci,value=h)
#         c.font=Font(name="Arial",bold=True,size=10,color=WHITE); c.fill=PatternFill("solid",fgColor=BLUE)
#         c.alignment=Alignment(horizontal="center",vertical="center"); c.border=border
#         ws.column_dimensions[get_column_letter(ci)].width=w
#     ws.row_dimensions[3].height=22
#     for ri,ct in enumerate(contacts,start=4):
#         prio=ct.get("priority","General"); bg_hex,fg_hex=pri_colors.get(prio,(WHITE,"0F172A"))
#         patterns=ct.get("email_patterns",[]); email_val=ct.get("email") or (patterns[0]+"  (pattern)" if patterns else "")
#         row_fill=bg_hex if ri%2==0 else GREY
#         for ci,val in enumerate([ri-3,prio,ct.get("name",""),ct.get("title",""),ct.get("company",""),email_val,ct.get("linkedin_url",""),ct.get("source","")],1):
#             cell=ws.cell(row=ri,column=ci,value=val)
#             cell.font=Font(name="Arial",size=9,bold=(ci==2),color=fg_hex if ci==2 else "0F172A")
#             cell.fill=PatternFill("solid",fgColor=row_fill)
#             cell.alignment=Alignment(vertical="center",wrap_text=(ci in [4,7])); cell.border=border
#         ws.row_dimensions[ri].height=18
#     ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:H{3+len(contacts)}"
#     buf=io.BytesIO(); wb.save(buf); buf.seek(0)
#     return buf.getvalue()


# # ══════════════════════════════════════════════════════════════════════════════
# #  MONGODB
# # ══════════════════════════════════════════════════════════════════════════════
# @st.cache_resource
# def get_db():
#     URI = st.secrets.get("MONGO_URI", "mongodb://localhost:27017")
#     DB  = st.secrets.get("MONGO_DB",  "secureitlab_job_pipeline")
#     return MongoClient(URI, serverSelectionTimeoutMS=6000)[DB]

# @st.cache_data(ttl=60)
# def load_all_jobs():
#     return list(get_db().jobs.find({}, {
#         "_id":1,"company":1,"role":1,"job_number":1,"opp_score":1,
#         "contacts_found":1,"pipeline_ok":1,"coverage_score":1,
#         "run_at":1,"contact_domain":1,"contacts":1,"country":1,
#     }))

# @st.cache_data(ttl=60)
# def load_job(job_id):
#     from bson import ObjectId
#     return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# # ══════════════════════════════════════════════════════════════════════════════
# #  RENDER HELPERS
# # ══════════════════════════════════════════════════════════════════════════════
# def badge(text, color="blue"):
#     return f'<span class="badge badge-{color}">{text}</span>'

# def safe_str(val, limit=300):
#     if val is None: return "—"
#     s = str(val); return s[:limit]+"…" if len(s) > limit else s

# def as_dict(raw):
#     if isinstance(raw, dict): return raw
#     if isinstance(raw, list): return next((x for x in raw if isinstance(x, dict)), {})
#     return {}

# def render_json_pretty(data, title=""):
#     if not data: return
#     with st.expander(f"📄 Raw JSON — {title}", expanded=False):
#         st.code(json.dumps(data, indent=2, default=str), language="json")

# def render_qa_block(data, label):
#     if not data:
#         st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>', unsafe_allow_html=True); return
#     data = as_dict(data)
#     if not data: return
#     passed   = data.get("passed") or data.get("Passed") or False
#     rec      = data.get("recommendation") or data.get("Recommendation", "")
#     issues   = data.get("issues") or data.get("Issues") or []
#     checklist= data.get("checklist") or data.get("Checklist") or []
#     color = "green" if passed else "yellow"; status = "✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
#     html = f"""<div class="sil-card sil-card-accent">
#       <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
#         <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
#         {badge(status, color)}
#       </div>"""
#     if rec: html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
#     if checklist:
#         html += '<div style="font-size:.82rem;margin-bottom:.5rem">'
#         for item in (checklist if isinstance(checklist, list) else [checklist]):
#             if isinstance(item, dict):
#                 ip = item.get("pass") or item.get("passed") or item.get("status","") == "pass"
#                 nm = item.get("item") or item.get("name") or item.get("check","")
#                 nt = item.get("reason") or item.get("note") or item.get("issue","")
#                 html += f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
#                 if nt: html += f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
#                 html += '</div>'
#         html += '</div>'
#     if issues:
#         html += '<div style="margin-top:.5rem">'
#         for iss in (issues if isinstance(issues, list) else [issues])[:4]:
#             txt = iss if isinstance(iss, str) else json.dumps(iss)
#             html += f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
#         html += '</div>'
#     st.markdown(html + '</div>', unsafe_allow_html=True)

# def render_contacts(contacts, title="Contacts"):
#     if not contacts: st.info("No contacts found for this job."); return
#     pri_color = {"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
#     st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>', unsafe_allow_html=True)
#     cols = st.columns(2)
#     for i, c in enumerate(contacts):
#         col = cols[i % 2]; prio = c.get("priority","General")
#         email = c.get("email",""); li = c.get("linkedin_url","")
#         patterns = c.get("email_patterns",[]); src = c.get("source","")
#         with col:
#             html = f"""<div class="contact-card">
#               <div style="display:flex;justify-content:space-between;align-items:flex-start">
#                 <div><div class="contact-name">{c.get('name','—')}</div>
#                 <div class="contact-title">{c.get('title','—')}</div></div>
#                 {badge(prio, pri_color.get(prio,'blue'))}
#               </div>"""
#             if email:      html += f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
#             elif patterns: html += f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
#             if li:         html += f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
#             if src:        html += f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
#             st.markdown(html + '</div>', unsafe_allow_html=True)

# def render_emails(emails_data):
#     if not emails_data: st.info("No email data available."); return
#     emails_data = as_dict(emails_data)
#     if not emails_data: return
#     variants = {}
#     for k, v in emails_data.items():
#         kl = k.lower().replace("_","").replace(" ","")
#         if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl=="a": variants["Variant A — Hiring Manager"] = v
#         elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl=="b": variants["Variant B — CISO / VP Level"] = v
#         else: variants[k] = v
#     for label, v in variants.items():
#         st.markdown(f'<div class="section-label">✉️ {label}</div>', unsafe_allow_html=True)
#         if isinstance(v, dict):
#             subj = v.get("subject") or v.get("Subject",""); body = v.get("body") or v.get("Body") or v.get("content","")
#             if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>', unsafe_allow_html=True)
#             if body: st.markdown(f'<div class="email-box">{body}</div>', unsafe_allow_html=True)
#             else:    st.code(json.dumps(v, indent=2), language="json")
#         elif isinstance(v, str):
#             st.markdown(f'<div class="email-box">{v}</div>', unsafe_allow_html=True)
#         st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

# def render_service_mapping(data):
#     if not data: st.info("No service mapping data."); return
#     items = data if isinstance(data, list) else []
#     if not items and isinstance(data, dict):
#         for key in ("services","mappings","service_mapping","ServiceMapping","items"):
#             if isinstance(data.get(key), list): items = data[key]; break
#         if not items: items = [data]
#     fit_colors = {"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
#     for item in items:
#         if not isinstance(item, dict): continue
#         svc  = item.get("service") or item.get("Service") or item.get("name","")
#         fit  = (item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
#         why  = item.get("justification") or item.get("rationale") or item.get("why","")
#         reqs = item.get("requirements_addressed") or item.get("requirements") or ""
#         eng  = item.get("engagement_type") or item.get("engagement","")
#         color = fit_colors.get(fit,"blue")
#         html = f"""<div class="sil-card" style="margin-bottom:.75rem">
#           <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
#             <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
#             {badge(fit or "MAPPED", color) if fit else ""}
#           </div>"""
#         if why:  html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
#         if reqs:
#             rs = ", ".join(reqs) if isinstance(reqs, list) else str(reqs)
#             html += f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
#         if eng:  html += f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
#         st.markdown(html + '</div>', unsafe_allow_html=True)
#     render_json_pretty(data, "Service Mapping")

# def render_microplans(data):
#     if not data: st.info("No micro-plan data."); return
#     plans = data if isinstance(data, list) else []
#     if not plans and isinstance(data, dict):
#         for k in ("plans","micro_plans","microplans","top_3","improvements"):
#             if isinstance(data.get(k), list): plans = data[k]; break
#         if not plans: plans = [data]
#     for i, plan in enumerate(plans[:3], 1):
#         if not isinstance(plan, dict): continue
#         title = plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
#         weeks = plan.get("duration") or plan.get("timeline","")
#         obj   = plan.get("objective") or plan.get("goal","")
#         kpis  = plan.get("kpis") or plan.get("KPIs") or []
#         tasks = plan.get("tasks") or plan.get("workstreams") or []
#         with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}", expanded=(i==1)):
#             if obj and obj != title: st.markdown(f"**Objective:** {obj}")
#             if kpis:
#                 st.markdown("**KPIs:**")
#                 for kpi in (kpis if isinstance(kpis, list) else [kpis]): st.markdown(f"• {kpi}")
#             if tasks:
#                 st.markdown("**Tasks:**")
#                 for t in (tasks if isinstance(tasks, list) else [tasks]):
#                     if isinstance(t, dict):
#                         tn = t.get("task") or t.get("name",""); te = t.get("effort") or t.get("duration","")
#                         st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
#                     else: st.markdown(f"• {t}")
#             st.code(json.dumps(plan, indent=2, default=str), language="json")

# def render_deal_assurance(data):
#     if not data: st.info("No deal assurance data."); return
#     if not isinstance(data, dict): render_json_pretty(data, "Deal Assurance Pack"); return
#     evp = data.get("executive_value_proposition") or data.get("value_proposition") or data.get("ExecutiveValueProposition","")
#     if evp:
#         st.markdown('<div class="section-label">💼 Executive Value Proposition</div>', unsafe_allow_html=True)
#         st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7">{evp}</div>', unsafe_allow_html=True)
#     caps = data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
#     if caps:
#         st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>', unsafe_allow_html=True)
#         c1, c2 = st.columns(2)
#         for i, cap in enumerate(caps if isinstance(caps, list) else [caps]):
#             (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
#     risk = data.get("risk_mitigation") or data.get("RiskMitigation","")
#     if risk:
#         st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>', unsafe_allow_html=True)
#         if isinstance(risk, dict):
#             for k, v in risk.items(): st.markdown(f"**{k}:** {v}")
#         else: st.markdown(str(risk))
#     render_json_pretty(data, "Deal Assurance Pack")


# # ══════════════════════════════════════════════════════════════════════════════
# #  LOGS PAGE
# # ══════════════════════════════════════════════════════════════════════════════
# def render_logs_page():
#     st.markdown("""<div style="margin-bottom:2rem">
#       <h1 style="font-family:'Syne',sans-serif;font-size:2.2rem;font-weight:800;color:#0f172a;margin:0">⚙️ Pipeline Execution</h1>
#       <p style="font-size:.95rem;color:#64748b;margin-top:.5rem">Live logs from the scraper, deduplication, AI agents, and Google Sheet sync</p>
#     </div>""", unsafe_allow_html=True)

#     if st.session_state.pipeline_running:
#         st.markdown("""<div class="logs-status running">
#           <div><div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.3rem">
#             <span class="pulse-dot"></span>
#             <span style="font-family:'Syne',sans-serif;font-weight:700;color:#1d4ed8">Pipeline Running</span>
#           </div>
#           <span style="font-size:.82rem;color:#64748b">Processing jobs... Google Sheet will auto-sync when complete.</span></div>
#         </div>""", unsafe_allow_html=True)
#     elif st.session_state.pipeline_done:
#         result = st.session_state.pipeline_result or {}
#         if result.get("success"):
#             st.markdown(f"""<div class="logs-status success">
#               <div><div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
#                 <span style="font-size:1.5rem">✅</span>
#                 <span style="font-family:'Syne',sans-serif;font-weight:700;color:#15803d;font-size:1.1rem">Pipeline Completed</span>
#               </div>
#               <div style="font-size:.85rem;color:#15803d">
#                 {result.get('processed',0)} processed · {result.get('new_jobs',0)} new · {result.get('skipped_db',0)} already in DB
#               </div></div></div>""", unsafe_allow_html=True)
#         else:
#             st.markdown(f"""<div class="logs-status error">
#               <div><div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
#                 <span style="font-size:1.5rem">❌</span>
#                 <span style="font-family:'Syne',sans-serif;font-weight:700;color:#b91c1c;font-size:1.1rem">Pipeline Failed</span>
#               </div>
#               <div style="font-size:.85rem;color:#b91c1c">{result.get('error','Unknown error')}</div></div></div>""", unsafe_allow_html=True)

#     st.markdown('<div class="section-label" style="margin-top:2rem">📡 Live Execution Logs</div>', unsafe_allow_html=True)
#     logs = st.session_state.pipeline_logs or []
#     log_text = "\n".join(logs[-200:]) if logs else "Waiting for logs..."
#     st.markdown(f'<div class="pipeline-log">{log_text}</div>', unsafe_allow_html=True)

#     if st.session_state.pipeline_running:
#         st.markdown("<script>setTimeout(function(){window.location.reload();},1500);</script>", unsafe_allow_html=True)

#     c1, c2, _ = st.columns([2, 2, 1])
#     with c1:
#         if st.button("← Back to Dashboard", use_container_width=True):
#             st.session_state.current_page = "dashboard"; st.rerun()
#     with c2:
#         if st.button("🔄 Refresh Logs", use_container_width=True): st.rerun()


# # ══════════════════════════════════════════════════════════════════════════════
# #  SIDEBAR
# # ══════════════════════════════════════════════════════════════════════════════
# with st.sidebar:
#     st.markdown("""<div style="padding:.75rem 0 1.25rem">
#       <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;color:#2563eb">🛡️ SecureITLab</div>
#       <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
#     </div>""", unsafe_allow_html=True)

#     if st.button("🚪 Logout", use_container_width=True):
#         st.session_state.logged_in = False; st.session_state.login_error = ""; st.rerun()

#     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

#     # ── Find Jobs ─────────────────────────────────────────────────────────────
#     st.markdown("**🔍 Find New Jobs**")
#     st.caption("Scrapes → deduplicates → AI agents → MongoDB → auto-syncs to Google Sheet · Repeats every 12h")

#     find_jobs_clicked = st.button(
#         "🔍  Find Jobs",
#         disabled=st.session_state.pipeline_running,
#         use_container_width=True,
#         type="primary",
#     )

#     if st.session_state.pipeline_running:
#         st.markdown("""<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
#             padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;font-size:.8rem;color:#1d4ed8">
#           ⏳ Pipeline running… 👉 View Logs Page
#         </div>""", unsafe_allow_html=True)

#     # ── Scheduler badge ───────────────────────────────────────────────────────
#     try:
#         import main as _main_module
#         sched = _main_module.get_scheduler_status()
#         if sched.get("active"):
#             secs = sched.get("seconds_until_next")
#             if secs is not None:
#                 hrs=secs//3600; mins=(secs%3600)//60
#                 countdown=f"{hrs}h {mins}m" if hrs else f"{mins}m {secs%60}s"
#                 next_label=f"Next run in: <b>{countdown}</b>"
#             else:
#                 next_label="Next run: calculating…"
#             last=(sched.get("last_run") or "")[:19]
#             st.markdown(f"""<div class="sched-on">
#               🟢 <b>Auto-scheduler ON</b><br>
#               <span style="color:#64748b">{next_label}</span><br>
#               <span style="color:#64748b">Runs every 12h · #{sched.get('run_count',0)} so far</span><br>
#               <span style="color:#64748b;font-size:.75rem">Last: {last} UTC</span>
#             </div>""", unsafe_allow_html=True)
#             if st.button("⏹️ Stop Auto-Scheduler", use_container_width=True):
#                 _main_module.stop_scheduler(); st.rerun()
#         elif sched.get("run_count", 0) > 0:
#             last=(sched.get("last_run") or "")[:19]
#             st.markdown(f"""<div class="sched-paused">
#               ⏸️ <b>Scheduler paused</b><br>
#               <span style="color:#64748b">Last ran: {last} UTC</span><br>
#               <span style="color:#64748b">Click Find Jobs to restart</span>
#             </div>""", unsafe_allow_html=True)
#     except Exception:
#         pass

#     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

#     # ══════════════════════════════════════════════════════════════════════════
#     #  ✅ GOOGLE SHEETS MASTER CONTACTS  (replaces Excel download button)
#     # ══════════════════════════════════════════════════════════════════════════
#     st.markdown("**📊 Master Contacts — Google Sheet**")
#     st.caption("All contacts synced here. Append-only — existing data is never deleted.")

#     # ── Auto-sync once per session on dashboard load ──────────────────────────
#     if not st.session_state.get("gsheet_auto_synced"):
#         st.session_state["gsheet_auto_synced"] = True
#         if Path(GCREDS_FILE).exists():
#             try:
#                 _auto_jobs = load_all_jobs()
#             except Exception:
#                 _auto_jobs = []
#             if _auto_jobs:
#                 _auto_res = sync_contacts_to_gsheet(_auto_jobs)
#                 if not _auto_res["error"]:
#                     st.session_state.gsheet_sync_error = ""
#                     st.session_state.gsheet_last_sync  = datetime.now().strftime("%d %b %Y  %H:%M")
#                     st.session_state.gsheet_appended   = _auto_res["appended"]
#                 else:
#                     st.session_state.gsheet_sync_error = _auto_res["error"]

#     # Buttons row
#     sync_col1, sync_col2 = st.columns([3, 1])
#     with sync_col1:
#         sync_clicked = st.button("🔄  Sync Now", use_container_width=True)
#     with sync_col2:
#         st.markdown(f'<a href="{GSHEET_URL}" target="_blank" style="text-decoration:none"><div style="text-align:center;padding:.42rem;border:1px solid #d9e2ec;border-radius:6px;background:#fff;font-size:.85rem;cursor:pointer">↗</div></a>', unsafe_allow_html=True)
#     if sync_clicked:
#         try:
#             jobs_for_sync = load_all_jobs()
#         except Exception:
#             jobs_for_sync = []
#         if not jobs_for_sync:
#             st.warning("No jobs in MongoDB to sync yet.")
#         else:
#             with st.spinner("Syncing to Google Sheet…"):
#                 res = sync_contacts_to_gsheet(jobs_for_sync)
#             if res["error"]:
#                 st.session_state.gsheet_sync_error = res["error"]
#                 st.session_state.gsheet_last_sync  = None
#                 st.session_state.gsheet_appended   = None
#             else:
#                 st.session_state.gsheet_sync_error = ""
#                 st.session_state.gsheet_last_sync  = datetime.now().strftime("%d %b %Y  %H:%M")
#                 st.session_state.gsheet_appended   = res["appended"]
#                 st.success(f"✅ +{res['appended']} new rows · {res['skipped']} already existed")

#     # Read persisted sync state from disk
#     disk_state = _read_sync_state()
#     last_sync  = (
#         st.session_state.gsheet_last_sync
#         or (disk_state.get("last_sync","")[:16].replace("T"," ") if disk_state.get("last_sync") else None)
#     )
#     appended = st.session_state.gsheet_appended if st.session_state.gsheet_appended is not None else disk_state.get("appended","")

#     if st.session_state.gsheet_sync_error:
#         short_err  = st.session_state.gsheet_sync_error[:180]
#         creds_hint = ""
#         if "not found" in short_err.lower() or "no such file" in short_err.lower():
#             creds_hint = "<br><span style='font-size:.74rem'>💡 Save your service account JSON as <code>google_credentials.json</code> in the same folder as this script, then share the sheet with the service account email.</span>"
#         elif "worksheet" in short_err.lower():
#             creds_hint = f"<br><span style='font-size:.74rem'>💡 Make sure the sheet tab is named <b>{GSHEET_TAB}</b> (it will be created automatically if missing).</span>"
#         st.markdown(f'<div class="gsheet-error">❌ <b>Sync failed</b><br>{short_err}{creds_hint}</div>', unsafe_allow_html=True)
#     else:
#         last_line  = f"Last synced: <b>{last_sync}</b>" if last_sync else "Not synced yet"
#         added_line = f" · <b>+{appended}</b> new rows" if appended and appended != 0 else (" · ✓ up to date" if appended == 0 and last_sync else "")
#         st.markdown(f"""<div class="gsheet-box">
#           🟢 <b>Google Sheet live</b><br>
#           <a href="{GSHEET_URL}" target="_blank">📋 Open master_contacts ↗</a><br>
#           <span style="font-size:.76rem;color:#64748b">{last_line}{added_line}</span><br>
#           <span style="font-size:.73rem;color:#94a3b8">Auto-syncs on load &amp; after every pipeline run</span>
#         </div>""", unsafe_allow_html=True)

#     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

#     # ── Job list ──────────────────────────────────────────────────────────────
#     try:
#         all_jobs = load_all_jobs()
#     except Exception as e:
#         st.error(f"MongoDB error: {e}"); st.stop()

#     if not all_jobs:
#         st.warning("No jobs in MongoDB yet. Click **Find Jobs** to start."); st.stop()

#     st.markdown(f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">{len(all_jobs)} jobs in database</div>', unsafe_allow_html=True)

#     search   = st.text_input("🔍 Filter jobs", placeholder="e.g. Bounteous")
#     filtered = [j for j in all_jobs if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

#     def job_label(j):
#         score = j.get("opp_score"); s = f" [{score}/10]" if score else ""
#         ok = "✅" if j.get("pipeline_ok") else "❌"
#         return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s}"

#     if not filtered:
#         st.warning("No matching jobs."); st.stop()

#     sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
#     selected_id = str(filtered[[job_label(j) for j in filtered].index(sel_label)]["_id"])

#     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
#     ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
#     total_c  = sum(j.get("contacts_found", 0) for j in all_jobs)
#     st.markdown(f"""<div style="font-size:.75rem;color:#64748b">
#       <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
#       <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
#     </div>""", unsafe_allow_html=True)

#     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
#     if st.button("🔄 Refresh Data", use_container_width=True):
#         st.cache_data.clear(); st.rerun()


# # ══════════════════════════════════════════════════════════════════════════════
# #  FIND JOBS — background thread
# # ══════════════════════════════════════════════════════════════════════════════
# if find_jobs_clicked and not st.session_state.pipeline_running:
#     with _thread_lock:
#         _thread_logs.clear(); _thread_result[0] = None; _thread_done[0] = False

#     st.session_state.pipeline_running = True
#     st.session_state.pipeline_done    = False
#     st.session_state.pipeline_logs    = []
#     st.session_state.pipeline_result  = None
#     st.session_state.current_page     = "logs"
#     st.cache_data.clear()

#     def _run_pipeline_bg():
#         try:
#             import main as _main

#             def _cb(msg: str):
#                 with _thread_lock: _thread_logs.append(msg)

#             result = _main.run_pipeline(progress_callback=_cb)

#             # ── Auto-sync Google Sheet after every successful run ────────────
#             if result.get("success"):
#                 try:
#                     _cb("🔄 [GSheet] Syncing new contacts to Google Sheet…")
#                     sync_res = sync_contacts_to_gsheet(load_all_jobs())
#                     if sync_res["error"]:
#                         _cb(f"⚠️  [GSheet] Sync error: {sync_res['error']}")
#                     else:
#                         _cb(f"✅ [GSheet] Done — +{sync_res['appended']} new rows · {sync_res['skipped']} already existed")
#                 except Exception as se:
#                     _cb(f"⚠️  [GSheet] Sync failed: {se}")

#             with _thread_lock: _thread_result[0] = result

#         except Exception as e:
#             import traceback
#             with _thread_lock:
#                 _thread_logs.append(f"❌ Unexpected error: {e}")
#                 _thread_logs.append(traceback.format_exc())
#                 _thread_result[0] = {"success":False,"error":str(e),"scraped":0,"new_jobs":0,"skipped_db":0,"processed":0}
#         finally:
#             with _thread_lock: _thread_done[0] = True

#     threading.Thread(target=_run_pipeline_bg, daemon=True).start()
#     st.rerun()

# # ── Sync thread state → session_state ────────────────────────────────────────
# with _thread_lock:
#     if _thread_logs:
#         st.session_state.pipeline_logs = list(_thread_logs)
#     if _thread_done[0] and _thread_result[0] is not None:
#         st.session_state.pipeline_result  = _thread_result[0]
#         st.session_state.pipeline_running = False
#         st.session_state.pipeline_done    = True


# # ══════════════════════════════════════════════════════════════════════════════
# #  PAGE ROUTING
# # ══════════════════════════════════════════════════════════════════════════════
# if st.session_state.current_page == "logs":
#     render_logs_page()
#     st.stop()

# with st.spinner("Loading job…"):
#     job = load_job(selected_id)

# if not job:
#     st.error("Could not load job document."); st.stop()

# company   = job.get("company","Unknown"); role  = job.get("role","Unknown")
# opp_score = job.get("opp_score");         p_ok  = job.get("pipeline_ok",False)
# p_min     = job.get("pipeline_min","?");  c_found = job.get("contacts_found",0)
# c_cov     = job.get("coverage_score");    c_domain = job.get("contact_domain","")
# run_at    = job.get("run_at","")

# st.markdown(f"""<div style="margin-bottom:1.75rem">
#   <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
#     Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
#   </div>
#   <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;color:#0f172a;margin:0;line-height:1.15">{role}</h1>
#   <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
#     @ <span style="color:#334155;font-weight:600">{company}</span>
#     {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
#   </div>
# </div>""", unsafe_allow_html=True)

# try:
#     sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
#     sc = "#16a34a" if sn>=7 else "#f59e0b" if sn>=5 else "#dc2626"
# except Exception:
#     sc = "#2563eb"

# st.markdown(f"""<div class="metric-row">
#   <div class="metric-tile"><div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div><div class="lbl">Opportunity Score</div></div>
#   <div class="metric-tile"><div class="val">{c_found}</div><div class="lbl">Contacts Found</div></div>
#   <div class="metric-tile"><div class="val">{f"{c_cov}%" if c_cov else "—"}</div><div class="lbl">Contact Coverage</div></div>
#   <div class="metric-tile"><div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">{'✅ OK' if p_ok else '❌ Failed'}</div><div class="lbl">Pipeline ({p_min} min)</div></div>
# </div>""", unsafe_allow_html=True)

# tabs = st.tabs(["📋 Job & Enrichment","🎯 Service Mapping","🔍 Fit / Gap",
#                 "🛠️ Capability & Plans","📦 Deal Assurance","✉️ Outreach Emails",
#                 "👥 Contacts","✅ QA Gates","🗄️ Raw Data"])

# with tabs[0]:
#     c1, c2 = st.columns(2)
#     with c1:
#         st.markdown('<div class="section-label">📄 Job Research</div>', unsafe_allow_html=True)
#         jr = as_dict(job.get("agent_job_research") or {})
#         if jr:
#             for lbl, keys in [("Job Role",["job_role","Job Role","role","title"]),("Company",["company_name","Company Name","company"]),("Location",["location","Location"]),("URL",["organization_url","Organization URL","url"])]:
#                 val = next((jr.get(k) for k in keys if jr.get(k)), None)
#                 if val: st.markdown(f"**{lbl}:** {val}")
#             desc = jr.get("job_description") or jr.get("Job Description","")
#             if desc:
#                 st.markdown("**Job Description:**")
#                 st.markdown(f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;max-height:300px;overflow-y:auto">{desc[:2000]}</div>', unsafe_allow_html=True)
#             render_json_pretty(jr,"Job Research")
#         else: st.info("No job research data.")
#     with c2:
#         st.markdown('<div class="section-label">🏢 Company Enrichment</div>', unsafe_allow_html=True)
#         enr = as_dict(job.get("agent_enrichment") or {})
#         if enr:
#             for lbl, keys in [("Industry",["industry","Industry"]),("Company Size",["company_size","size"]),("Regulatory Env",["regulatory_environment","regulatory"]),("Certifications",["certifications","Certifications"]),("Tech Stack",["tech_stack","technology_stack"]),("Security Maturity",["security_maturity","maturity"])]:
#                 val = next((enr.get(k) for k in keys if enr.get(k)), None)
#                 if val:
#                     if isinstance(val, list): val = ", ".join(str(v) for v in val)
#                     st.markdown(f"**{lbl}:** {safe_str(val,200)}")
#             render_json_pretty(enr,"Enrichment")
#         else: st.info("No enrichment data.")

# with tabs[1]:
#     st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>', unsafe_allow_html=True)
#     render_service_mapping(job.get("agent_service_mapping"))

# with tabs[2]:
#     fg = as_dict(job.get("agent_fit_gap") or {})
#     if opp_score:
#         try:
#             sn=float(str(opp_score).split("/")[0]); bc="#16a34a" if sn>=7 else "#f59e0b" if sn>=5 else "#dc2626"; bp=int(sn/10*100)
#             st.markdown(f"""<div style="margin-bottom:1.5rem">
#               <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
#                 <span style="font-family:'Syne',sans-serif;font-weight:700">Opportunity Score</span>
#                 <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
#               </div>
#               <div style="background:#e2e8f0;border-radius:4px;height:8px">
#                 <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
#               </div></div>""", unsafe_allow_html=True)
#         except Exception: pass
#     st.markdown('<div class="section-label">📊 Service Classifications</div>', unsafe_allow_html=True)
#     services = []
#     if isinstance(fg, dict):
#         for k in ("services","classifications","service_classifications","items","fit_gap"):
#             v = fg.get(k)
#             if isinstance(v, list): services = v; break
#         if not services and (fg.get("service") or fg.get("Service")): services = [fg]
#     elif isinstance(fg, list): services = fg
#     if services:
#         buckets = {"STRONG FIT":[],"PARTIAL FIT":[],"GAP":[]}
#         for s in services:
#             if not isinstance(s, dict): continue
#             fit = (s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
#             if "STRONG" in fit: buckets["STRONG FIT"].append(s)
#             elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
#             elif "GAP" in fit: buckets["GAP"].append(s)
#         c1,c2,c3=st.columns(3)
#         cm={"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
#         bgm={"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
#         bdm={"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
#         for col,(fl,items) in zip([c1,c2,c3],buckets.items()):
#             col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>', unsafe_allow_html=True)
#             for s in items:
#                 svc=s.get("service") or s.get("Service") or s.get("name",""); just=s.get("justification") or s.get("reason","")
#                 col.markdown(f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem"><div style="font-weight:600">{svc}</div><div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>', unsafe_allow_html=True)
#     elif fg: st.json(fg)
#     else: st.info("No fit/gap data.")
#     render_json_pretty(job.get("agent_fit_gap"),"Fit/Gap Analysis")

# with tabs[3]:
#     c1, c2 = st.columns(2)
#     with c1:
#         st.markdown('<div class="section-label">🔧 Capability Improvements</div>', unsafe_allow_html=True)
#         cap = job.get("agent_capability") or {}
#         items_cap = cap if isinstance(cap,list) else []
#         if not items_cap and isinstance(cap,dict):
#             for k in ("improvements","recommendations","capabilities","items"):
#                 v=cap.get(k)
#                 if isinstance(v,list): items_cap=v; break
#             if not items_cap: items_cap=[cap]
#         for item in items_cap:
#             if not isinstance(item,dict): continue
#             title=item.get("title") or item.get("gap") or item.get("service","")
#             rec=item.get("recommendation") or item.get("steps","")
#             effort=item.get("build_effort") or item.get("effort",""); demand=item.get("market_demand") or item.get("priority","")
#             st.markdown(f"""<div class="sil-card" style="margin-bottom:.6rem">
#               <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem">{title}</div>
#               <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec,250)}</div>
#               {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
#             </div>""", unsafe_allow_html=True)
#         if not items_cap: render_json_pretty(cap,"Capability Improvement")
#     with c2:
#         st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>', unsafe_allow_html=True)
#         render_microplans(job.get("agent_microplans"))

# with tabs[4]:
#     render_deal_assurance(job.get("agent_deal_assurance"))

# with tabs[5]:
#     st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>', unsafe_allow_html=True)
#     emails_src = job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
#     oq = as_dict(job.get("agent_outreach_qa") or {})
#     improved = (oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
#     if improved:
#         st.info("⚡ Showing QA-improved versions")
#         render_emails(improved)
#         with st.expander("📬 Original (pre-QA) versions"): render_emails(emails_src)
#     else: render_emails(emails_src)

# with tabs[6]:
#     contacts = job.get("contacts") or []; contact_sources = job.get("contact_sources") or []
#     pri=[c for c in contacts if c.get("priority")=="Primary"]
#     sec=[c for c in contacts if c.get("priority")=="Secondary"]
#     ter=[c for c in contacts if c.get("priority")=="Tertiary"]
#     gen=[c for c in contacts if c.get("priority")=="General"]
#     st.markdown(f"""<div class="metric-row" style="margin-bottom:1.5rem">
#       <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
#       <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
#       <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
#       <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
#     </div>""", unsafe_allow_html=True)
#     if contact_sources:
#         st.markdown('Sources: '+" ".join(badge(s,"blue") for s in contact_sources), unsafe_allow_html=True); st.markdown("")
#     missing = job.get("missing_roles") or []
#     if missing:
#         st.markdown('⚠️ Missing roles: '+" ".join(badge(r,"red") for r in missing), unsafe_allow_html=True); st.markdown("")
#     if contacts:
#         excel_bytes = build_contacts_excel(contacts, company, role)
#         if excel_bytes:
#             safe_co = re.sub(r'[^a-z0-9]','_',company.lower())[:20]
#             fname = f"contacts_{safe_co}_{datetime.now().strftime('%Y%m%d')}.xlsx"
#             btn_col, _ = st.columns([1,3])
#             with btn_col:
#                 st.download_button("📥  Download Contacts (.xlsx)", data=excel_bytes, file_name=fname,
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")
#         st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
#         pri_filter = st.multiselect("Filter by priority",["Primary","Secondary","Tertiary","General"],default=["Primary","Secondary","Tertiary","General"])
#         shown = [c for c in contacts if c.get("priority","General") in pri_filter]
#         render_contacts(shown, f"Contacts ({len(shown)} shown)")
#         agent_contacts = job.get("agent_prospect_contacts") or {}
#         if agent_contacts:
#             with st.expander("🤖 CrewAI Agent's Contact Search"):
#                 if isinstance(agent_contacts,dict):
#                     ac_list=agent_contacts.get("contacts") or []
#                     if ac_list: render_contacts(ac_list,"Agent Contacts")
#                     else: st.json(agent_contacts)
#                 else: st.json(agent_contacts)
#     else: st.info("No contacts found for this job.")

# with tabs[7]:
#     st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>', unsafe_allow_html=True)
#     c1, c2 = st.columns(2)
#     for i,(lbl,key) in enumerate([("Research QA","agent_research_qa"),("Service Mapping QA","agent_mapping_qa"),("Deal Assurance QA","agent_assurance_qa"),("Outreach Email QA","agent_outreach_qa")]):
#         with (c1 if i%2==0 else c2): render_qa_block(job.get(key), lbl)
#     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
#     st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>', unsafe_allow_html=True)
#     enf = as_dict(job.get("agent_prospect_enforcer") or {})
#     if enf:
#         cov=enf.get("coverage_score","?"); miss=enf.get("missing_roles",[]); note=enf.get("note",""); ec=enf.get("contacts",[])
#         x1,x2,x3=st.columns(3)
#         x1.metric("Coverage Score",f"{cov}%"); x2.metric("Missing Roles",len(miss)); x3.metric("Contacts Verified",len(ec))
#         if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
#         if note: st.caption(note)
#     else: st.info("No enforcer data.")

# with tabs[8]:
#     st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>', unsafe_allow_html=True)
#     rows=[]
#     for k,v in job.items():
#         if k=="_id": continue
#         rows.append({"Field":k,"Type":type(v).__name__,"Len":len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
#     hc1,hc2,hc3=st.columns([3,1,1])
#     hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
#     for r in rows:
#         rc1,rc2,rc3=st.columns([3,1,1])
#         rc1.code(r["Field"],language=None)
#         rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>', unsafe_allow_html=True)
#         rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>', unsafe_allow_html=True)
#     st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
#     for lbl,key in [("Job Research","agent_job_research"),("Enrichment","agent_enrichment"),("Service Mapping","agent_service_mapping"),("Fit/Gap","agent_fit_gap"),("Capability","agent_capability"),("Micro-Plans","agent_microplans"),("Deal Assurance","agent_deal_assurance"),("Outreach Emails","agent_outreach_emails"),("Prospect Contacts","agent_prospect_contacts"),("Prospect Enforcer","agent_prospect_enforcer"),("Research QA","agent_research_qa"),("Mapping QA","agent_mapping_qa"),("Assurance QA","agent_assurance_qa"),("Outreach QA","agent_outreach_qa"),("Contacts (5-source)","contacts")]:
#         data=job.get(key)
#         if data:
#             with st.expander(f"📄 {lbl}"):
#                 st.code(json.dumps(data,indent=2,default=str),language="json")



































"""
╔══════════════════════════════════════════════════════════╗
║   SecureITLab Pipeline Dashboard — Streamlit             ║
║   WITH GOOGLE SHEETS AUTO-SYNC (Master Contacts)         ║
║   WITH 12-HOUR AUTO-SCHEDULER STATUS IN SIDEBAR          ║
║   Reads from MongoDB → secureitlab_job_pipeline          ║
║   ✨ LinkedIn & Email Sync Now Included                   ║
╠══════════════════════════════════════════════════════════╣
║  Install: pip install streamlit pymongo python-dotenv    ║
║           gspread google-auth openpyxl                   ║
║  Run:     streamlit run streamlit_dashboard.py           ║
╚══════════════════════════════════════════════════════════╝
"""

import io
import re
import streamlit as st
from pymongo import MongoClient
import json
import threading
import time
import logging
from datetime import datetime, timezone
from io import StringIO
from pathlib import Path

# ── Thread-safe shared state ──────────────────────────────────────────────────
_thread_logs   = []
_thread_result = [None]
_thread_done   = [False]
_thread_lock   = threading.Lock()

# ── Google Sheets config ──────────────────────────────────────────────────────
GSHEET_URL        = "https://docs.google.com/spreadsheets/d/1u9_Fqy8a8Dj-yX8FwZtkFfw_3BLdpmBbtagH2IjhyV4/edit"
GSHEET_ID         = "1u9_Fqy8a8Dj-yX8FwZtkFfw_3BLdpmBbtagH2IjhyV4"
GSHEET_TAB        = "master_contacts"
GCREDS_FILE       = "google_credentials.json"   # service account JSON in project folder
GSHEET_SYNC_STATE = "gsheet_sync_state.json"    # persists last sync info across reruns

# ── Log capture ───────────────────────────────────────────────────────────────
_log_capture  = StringIO()
_log_handler  = logging.StreamHandler(_log_capture)
_log_handler.setLevel(logging.INFO)
_log_handler.setFormatter(logging.Formatter('%(asctime)s | %(levelname)s | %(message)s'))

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SecureITLab Pipeline",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

LOGIN_USERNAME = "admin"
LOGIN_PASSWORD = "secureitlab2024"

# ══════════════════════════════════════════════════════════════════════════════
#  GLOBAL CSS
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');
:root{--bg:#f4f7fb;--surface:#ffffff;--surface2:#eef2f7;--border:#d9e2ec;--accent:#2563eb;--accent2:#7c3aed;--green:#16a34a;--yellow:#f59e0b;--red:#dc2626;--text:#0f172a;--muted:#64748b;}
html,body,[class*="css"]{background-color:var(--bg)!important;color:var(--text)!important;font-family:'DM Sans',sans-serif!important;}
.login-card{background:var(--surface);border:1px solid var(--border);border-radius:20px;padding:3rem 3.5rem;width:100%;max-width:420px;box-shadow:0 20px 60px rgba(37,99,235,0.08);text-align:center;}
.login-logo{font-family:'Syne',sans-serif;font-size:1.6rem;font-weight:800;color:var(--accent);margin-bottom:.25rem;}
.login-subtitle{font-size:.75rem;color:var(--muted);letter-spacing:.12em;text-transform:uppercase;margin-bottom:2.5rem;}
.login-error{background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:.65rem 1rem;color:#b91c1c;font-size:.85rem;margin-top:1rem;}
.login-divider{width:40px;height:3px;background:linear-gradient(90deg,var(--accent),var(--accent2));border-radius:2px;margin:0 auto 2rem;}
[data-testid="stSidebar"]{background:var(--surface)!important;border-right:1px solid var(--border)!important;}
[data-testid="stSidebar"] *{color:var(--text)!important;}
.main .block-container{padding:2rem 2rem 3rem!important;}
h1,h2,h3,h4{font-family:'Syne',sans-serif!important;color:var(--text)!important;}
.sil-card{background:var(--surface);border:1px solid var(--border);border-radius:14px;padding:1.25rem 1.5rem;margin-bottom:1rem;transition:all 0.25s ease;}
.sil-card:hover{transform:translateY(-2px);box-shadow:0 8px 22px rgba(0,0,0,0.05);}
.sil-card-accent{border-left:4px solid var(--accent);}
.metric-row{display:flex;gap:1rem;flex-wrap:wrap;margin-bottom:1.5rem;}
.metric-tile{flex:1;min-width:140px;background:var(--surface2);border:1px solid var(--border);border-radius:12px;padding:1rem;text-align:center;transition:all .25s ease;}
.metric-tile:hover{transform:translateY(-3px);box-shadow:0 10px 24px rgba(0,0,0,0.06);}
.metric-tile .val{font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;color:var(--accent);}
.metric-tile .lbl{font-size:.72rem;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;}
.badge{padding:.25rem .7rem;border-radius:20px;font-size:.72rem;font-weight:600;font-family:'DM Mono',monospace;}
.badge-green{background:#ecfdf5;color:#15803d;}
.badge-yellow{background:#fffbeb;color:#b45309;}
.badge-red{background:#fef2f2;color:#b91c1c;}
.badge-blue{background:#eff6ff;color:#1d4ed8;}
.badge-purple{background:#f5f3ff;color:#6d28d9;}
.contact-card{background:var(--surface2);border:1px solid var(--border);border-radius:10px;padding:1rem;margin-bottom:.8rem;}
.contact-name{font-family:'Syne',sans-serif;font-weight:700;color:var(--text);}
.contact-title{color:var(--muted);font-size:.85rem;}
.contact-email{font-family:'DM Mono',monospace;color:var(--accent);font-size:.8rem;}
.email-box{background:#f8fafc;border:1px solid var(--border);border-radius:10px;padding:1rem 1.25rem;font-size:.9rem;line-height:1.65;white-space:pre-wrap;color:var(--text);}
.email-subject{font-family:'Syne',sans-serif;font-weight:700;color:var(--accent);margin-bottom:.5rem;}
.section-label{font-family:'DM Mono',monospace;font-size:.72rem;text-transform:uppercase;letter-spacing:.12em;color:var(--accent);margin-bottom:.6rem;}
.sil-divider{border-top:1px solid var(--border);margin:1rem 0;}
[data-testid="stExpander"]{background:var(--surface)!important;border:1px solid var(--border)!important;border-radius:10px!important;}
[data-testid="stTabs"] button{font-family:'Syne',sans-serif!important;font-weight:600!important;}
::-webkit-scrollbar{width:6px;}
::-webkit-scrollbar-thumb{background:var(--border);border-radius:3px;}
.pipeline-log{background:#0f172a;color:#10b981;border-radius:10px;padding:1.5rem;font-family:'DM Mono',monospace;font-size:.8rem;line-height:1.8;max-height:700px;overflow-y:auto;white-space:pre-wrap;word-break:break-word;border:1px solid #1e293b;}
.logs-status{display:flex;gap:1rem;justify-content:space-between;align-items:center;margin-bottom:1.5rem;padding:1rem;background:var(--surface2);border-radius:10px;border:1px solid var(--border);}
.logs-status.running{background:#eff6ff;border-color:#bfdbfe;}
.logs-status.success{background:#f0fdf4;border-color:#bbf7d0;}
.logs-status.error{background:#fef2f2;border-color:#fecaca;}
@keyframes pulse{0%,100%{opacity:1;}50%{opacity:0.5;}}
.pulse-dot{display:inline-block;width:10px;height:10px;background:#2563eb;border-radius:50%;animation:pulse 2s infinite;margin-right:8px;}
.sched-on{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:.65rem .9rem;margin-top:.5rem;font-size:.8rem;color:#15803d;line-height:1.6;}
.sched-paused{background:#fef9c3;border:1px solid #fde68a;border-radius:8px;padding:.65rem .9rem;margin-top:.5rem;font-size:.8rem;color:#92400e;line-height:1.6;}

/* ── Google Sheet box ── */
.gsheet-box{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:.9rem 1rem;margin-top:.4rem;font-size:.82rem;color:#15803d;line-height:1.8;}
.gsheet-box a{color:#1d4ed8!important;font-weight:700;text-decoration:none;}
.gsheet-box a:hover{text-decoration:underline;}
.gsheet-syncing{background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;padding:.9rem 1rem;margin-top:.4rem;font-size:.82rem;color:#1d4ed8;line-height:1.8;}
.gsheet-error{background:#fef2f2;border:1px solid #fecaca;border-radius:10px;padding:.75rem 1rem;font-size:.8rem;color:#b91c1c;margin-top:.4rem;line-height:1.6;}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for _k, _v in [
    ("logged_in",         False),
    ("login_error",       ""),
    ("pipeline_running",  False),
    ("pipeline_logs",     []),
    ("pipeline_result",   None),
    ("pipeline_done",     False),
    ("current_page",      "dashboard"),
    ("gsheet_sync_error", ""),
    ("gsheet_last_sync",  None),
    ("gsheet_appended",   None),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ══════════════════════════════════════════════════════════════════════════════
#  GOOGLE SHEETS HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _gsheet_client():
    import gspread
    from google.oauth2.service_account import Credentials
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(GCREDS_FILE, scopes=scopes)
    return gspread.authorize(creds)


def _existing_contact_keys(ws) -> set:
    """Return set of (name_lower, company_lower) already in the sheet."""
    try:
        rows = ws.get_all_values()
        keys = set()
        for row in rows[3:]:        # skip 3 header rows
            name    = row[5].strip().lower() if len(row) > 5 else ""
            company = row[2].strip().lower() if len(row) > 2 else ""
            if name:
                keys.add((name, company))
        return keys
    except Exception:
        return set()


def _write_sync_state(appended: int, skipped: int):
    try:
        Path(GSHEET_SYNC_STATE).write_text(json.dumps({
            "last_sync": datetime.now(timezone.utc).isoformat(),
            "appended":  appended,
            "skipped":   skipped,
        }), encoding="utf-8")
    except Exception:
        pass


def _read_sync_state() -> dict:
    try:
        if Path(GSHEET_SYNC_STATE).exists():
            return json.loads(Path(GSHEET_SYNC_STATE).read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}


def sync_contacts_to_gsheet(all_jobs: list) -> dict:
    """
    Append NEW contacts from MongoDB jobs list to the Google Sheet.
    Includes: Outreach Email (Variant A) and LinkedIn Message.
    Never deletes existing rows — skips any contact already present.
    Returns: {"appended": int, "skipped": int, "error": str|None}
    """
    result = {"appended": 0, "skipped": 0, "error": None}

    if not Path(GCREDS_FILE).exists():
        result["error"] = (
            f"'{GCREDS_FILE}' not found in project folder. "
            "Create a service account on Google Cloud, download the JSON key, "
            "save it as google_credentials.json, then share your sheet with the service account email."
        )
        return result

    try:
        gc = _gsheet_client()
        sh = gc.open_by_key(GSHEET_ID)

        # Get or create the worksheet tab
        try:
            ws = sh.worksheet(GSHEET_TAB)
        except Exception:
            ws = sh.add_worksheet(title=GSHEET_TAB, rows=10000, cols=13)

        # ── Ensure sheet has at least 13 columns ──
        try:
            ws.resize(rows=max(ws.row_count, 10000), cols=max(ws.col_count, 13))
        except Exception:
            pass

        # Write headers if sheet is completely empty
        all_vals = ws.get_all_values()
        if not all_vals or len(all_vals) < 3:
            header_rows = [
                ["Master Contacts — SecureITLab Pipeline Auto-Sync"] + [""] * 12,
                [f"Last synced: {datetime.now().strftime('%d %b %Y  %H:%M')} UTC  ·  Append-only, duplicates skipped"] + [""] * 12,
                ["#", "Job Role", "Company", "Country", "Priority",
                 "Name", "Title / Role", "Contact Email", "LinkedIn URL", "Source", "Job Score",
                 "Outreach Email (Agent)", "LinkedIn Message (Agent)"],
            ]
            ws.update("A1", header_rows)
        else:
            # Upgrade headers if new columns are missing (existing 11-col sheet)
            header_row = all_vals[2] if len(all_vals) >= 3 else []
            if len(header_row) < 13:
                ws.update("L3", [["Outreach Email (Agent)", "LinkedIn Message (Agent)"]])
            # Always refresh the subtitle row with latest sync time
            ws.update("A2", [[
                f"Last synced: {datetime.now().strftime('%d %b %Y  %H:%M')} UTC  ·  Append-only, duplicates skipped"
            ]])

        # Collect keys already in the sheet
        existing_keys = _existing_contact_keys(ws)
        current_row_count = max(0, len(ws.get_all_values()) - 3)  # data rows only

        rows_to_add = []
        for job in all_jobs:
            company   = job.get("company", "")
            role      = job.get("role", "")
            country   = job.get("country", "?")
            job_score = str(job.get("opp_score", ""))
            contacts  = job.get("contacts", [])

            # ── Extract outreach email from agent (Variant A preferred) ───────
            outreach_email_text = ""
            emails_data = job.get("agent_outreach_emails") or {}
            # Also try QA-improved version first
            oq = job.get("agent_outreach_qa") or {}
            if isinstance(oq, dict):
                improved = oq.get("improved_emails") or oq.get("ImprovedEmails")
                if improved and isinstance(improved, dict):
                    emails_data = improved
            if isinstance(emails_data, dict):
                # Try Variant A first, then B, then first available
                for k, v in emails_data.items():
                    kl = k.lower().replace("_","").replace(" ","")
                    if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl == "a":
                        if isinstance(v, dict):
                            subj = v.get("subject") or v.get("Subject","")
                            body = v.get("body") or v.get("Body") or v.get("content","")
                            outreach_email_text = f"Subject: {subj}\n\n{body}" if subj else body
                        elif isinstance(v, str):
                            outreach_email_text = v
                        break
                if not outreach_email_text:
                    # fallback: first key
                    for v in emails_data.values():
                        if isinstance(v, dict):
                            subj = v.get("subject") or v.get("Subject","")
                            body = v.get("body") or v.get("Body") or v.get("content","")
                            outreach_email_text = f"Subject: {subj}\n\n{body}" if subj else body
                        elif isinstance(v, str):
                            outreach_email_text = v
                        if outreach_email_text:
                            break

            # ── Extract LinkedIn message from agent ───────────────────────────
            linkedin_msg_text = ""
            li_sequences = job.get("agent_linkedin_sequences") or job.get("agent_linkedin") or {}
            if isinstance(li_sequences, dict):
                # Try to get connection request message or first message
                for k in ("connection_request","connection_message","message_1","first_message","intro","sequence_1","message"):
                    v = li_sequences.get(k) or li_sequences.get(k.title()) or li_sequences.get(k.upper())
                    if v and isinstance(v, str):
                        linkedin_msg_text = v; break
                    elif v and isinstance(v, dict):
                        linkedin_msg_text = v.get("message") or v.get("text") or v.get("content","")
                        if linkedin_msg_text: break
                if not linkedin_msg_text:
                    # Try sequences list
                    for k in ("sequences","messages","steps"):
                        seq = li_sequences.get(k)
                        if isinstance(seq, list) and seq:
                            first = seq[0]
                            if isinstance(first, dict):
                                linkedin_msg_text = first.get("message") or first.get("text") or first.get("content","")
                            elif isinstance(first, str):
                                linkedin_msg_text = first
                            if linkedin_msg_text: break
                if not linkedin_msg_text:
                    # Last resort: dump first string value found
                    for v in li_sequences.values():
                        if isinstance(v, str) and len(v) > 20:
                            linkedin_msg_text = v; break

            elif isinstance(li_sequences, list) and li_sequences:
                first = li_sequences[0]
                if isinstance(first, dict):
                    linkedin_msg_text = (first.get("message") or first.get("connection_request")
                                         or first.get("text") or first.get("content",""))
                elif isinstance(first, str):
                    linkedin_msg_text = first

            for ci, contact in enumerate(contacts):
                name     = contact.get("name", "").strip()
                prio     = contact.get("priority", "General")
                title    = contact.get("title", "")
                email    = contact.get("email", "")
                li       = contact.get("linkedin_url", "")
                source   = contact.get("source", "")
                patterns = contact.get("email_patterns", [])

                if not email and patterns:
                    email = patterns[0] + "  (pattern)"

                key = (name.lower(), company.strip().lower())
                if not name or key in existing_keys:
                    result["skipped"] += 1
                    continue

                rows_to_add.append([
                    current_row_count + len(rows_to_add) + 1,   # col A — row #
                    role    if ci == 0 else "",                  # col B — Job Role
                    company if ci == 0 else "",                  # col C — Company
                    country if ci == 0 else "",                  # col D — Country
                    prio,                                        # col E — Priority
                    name,                                        # col F — Name
                    title,                                       # col G — Title / Role
                    email,                                       # col H — Contact Email
                    li,                                          # col I — LinkedIn URL
                    source,                                      # col J — Source
                    job_score if ci == 0 else "",               # col K — Job Score
                    outreach_email_text if ci == 0 else "",    # col L — Outreach Email (Agent)
                    linkedin_msg_text   if ci == 0 else "",    # col M — LinkedIn Message (Agent)
                ])
                existing_keys.add(key)  # dedup within this batch too

        if rows_to_add:
            ws.append_rows(rows_to_add, value_input_option="USER_ENTERED")
            result["appended"] = len(rows_to_add)

        _write_sync_state(result["appended"], result["skipped"])

    except Exception as e:
        result["error"] = str(e)

    return result


# ══════════════════════════════════════════════════════════════════════════════
#  LOGIN
# ══════════════════════════════════════════════════════════════════════════════
if not st.session_state.logged_in:
    _, col, _ = st.columns([1, 1.2, 1])
    with col:
        st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)
        st.markdown("""
        <div class="login-card">
          <div class="login-logo">🛡️ SecureITLab</div>
          <div class="login-subtitle">Pipeline Intelligence</div>
          <div class="login-divider"></div>
        </div>""", unsafe_allow_html=True)
        username = st.text_input("Username", placeholder="Enter username", key="lu")
        password = st.text_input("Password", placeholder="Enter password", type="password", key="lp")
        if st.button("Sign In →", use_container_width=True, type="primary"):
            if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
                st.session_state.logged_in = True; st.session_state.login_error = ""; st.rerun()
            else:
                st.session_state.login_error = "Incorrect username or password."
        if st.session_state.login_error:
            st.markdown(f'<div class="login-error">⚠️ {st.session_state.login_error}</div>', unsafe_allow_html=True)
        st.markdown("<div style='text-align:center;font-size:.72rem;color:#94a3b8;margin-top:2rem'>SecureITLab · Confidential</div>", unsafe_allow_html=True)
    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
#  PER-JOB CONTACTS EXCEL  (individual job download — kept)
# ══════════════════════════════════════════════════════════════════════════════
def build_contacts_excel(contacts: list, company: str, role: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    wb = Workbook(); ws = wb.active; ws.title = "Contacts"
    NAVY="1E3A5F"; BLUE="2563EB"; GREY="F8FAFC"; WHITE="FFFFFF"
    pri_colors={"Primary":("FEF2F2","B91C1C"),"Secondary":("FFFBEB","B45309"),"Tertiary":("EFF6FF","1D4ED8"),"General":("F5F3FF","6D28D9")}
    thin=Side(border_style="thin",color="D9E2EC"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.merge_cells("A1:H1"); c=ws["A1"]
    c.value=f"Contacts Export  —  {company}  |  {role}"
    c.font=Font(name="Arial",bold=True,size=13,color=WHITE); c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=30
    ws.merge_cells("A2:H2"); c=ws["A2"]
    c.value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   {len(contacts)} contacts"
    c.font=Font(name="Arial",size=9,color="64748B"); c.fill=PatternFill("solid",fgColor="F4F7FB")
    c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=18
    headers=["#","Priority","Name","Title / Role","Company","Email","LinkedIn URL","Source"]; col_widths=[4,12,24,32,22,34,42,18]
    for ci,(h,w) in enumerate(zip(headers,col_widths),1):
        c=ws.cell(row=3,column=ci,value=h)
        c.font=Font(name="Arial",bold=True,size=10,color=WHITE); c.fill=PatternFill("solid",fgColor=BLUE)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=border
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=22
    for ri,ct in enumerate(contacts,start=4):
        prio=ct.get("priority","General"); bg_hex,fg_hex=pri_colors.get(prio,(WHITE,"0F172A"))
        patterns=ct.get("email_patterns",[]); email_val=ct.get("email") or (patterns[0]+"  (pattern)" if patterns else "")
        row_fill=bg_hex if ri%2==0 else GREY
        for ci,val in enumerate([ri-3,prio,ct.get("name",""),ct.get("title",""),ct.get("company",""),email_val,ct.get("linkedin_url",""),ct.get("source","")],1):
            cell=ws.cell(row=ri,column=ci,value=val)
            cell.font=Font(name="Arial",size=9,bold=(ci==2),color=fg_hex if ci==2 else "0F172A")
            cell.fill=PatternFill("solid",fgColor=row_fill)
            cell.alignment=Alignment(vertical="center",wrap_text=(ci in [4,7])); cell.border=border
        ws.row_dimensions[ri].height=18
    ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:H{3+len(contacts)}"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  MONGODB
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_resource
def get_db():
    URI = st.secrets.get("MONGO_URI", "mongodb://localhost:27017")
    DB  = st.secrets.get("MONGO_DB",  "secureitlab_job_pipeline")
    return MongoClient(URI, serverSelectionTimeoutMS=6000)[DB]

@st.cache_data(ttl=60)
def load_all_jobs():
    return list(get_db().jobs.find({}, {
        "_id":1,"company":1,"role":1,"job_number":1,"opp_score":1,
        "contacts_found":1,"pipeline_ok":1,"coverage_score":1,
        "run_at":1,"contact_domain":1,"contacts":1,"country":1,
    }))

@st.cache_data(ttl=60)
def load_job(job_id):
    from bson import ObjectId
    return get_db().jobs.find_one({"_id": ObjectId(job_id)})


# ══════════════════════════════════════════════════════════════════════════════
#  RENDER HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def badge(text, color="blue"):
    return f'<span class="badge badge-{color}">{text}</span>'

def safe_str(val, limit=300):
    if val is None: return "—"
    s = str(val); return s[:limit]+"…" if len(s) > limit else s

def as_dict(raw):
    if isinstance(raw, dict): return raw
    if isinstance(raw, list): return next((x for x in raw if isinstance(x, dict)), {})
    return {}

def render_json_pretty(data, title=""):
    if not data: return
    with st.expander(f"📄 Raw JSON — {title}", expanded=False):
        st.code(json.dumps(data, indent=2, default=str), language="json")

def render_qa_block(data, label):
    if not data:
        st.markdown(f'<div class="sil-card"><b>{label}</b> — <i>No data</i></div>', unsafe_allow_html=True); return
    data = as_dict(data)
    if not data: return
    passed   = data.get("passed") or data.get("Passed") or False
    rec      = data.get("recommendation") or data.get("Recommendation", "")
    issues   = data.get("issues") or data.get("Issues") or []
    checklist= data.get("checklist") or data.get("Checklist") or []
    color = "green" if passed else "yellow"; status = "✅ APPROVED" if passed else "⚠️ NEEDS REWORK"
    html = f"""<div class="sil-card sil-card-accent">
      <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.75rem">
        <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:1rem">{label}</span>
        {badge(status, color)}
      </div>"""
    if rec: html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.6rem">📝 {rec}</div>'
    if checklist:
        html += '<div style="font-size:.82rem;margin-bottom:.5rem">'
        for item in (checklist if isinstance(checklist, list) else [checklist]):
            if isinstance(item, dict):
                ip = item.get("pass") or item.get("passed") or item.get("status","") == "pass"
                nm = item.get("item") or item.get("name") or item.get("check","")
                nt = item.get("reason") or item.get("note") or item.get("issue","")
                html += f'<div style="margin:.25rem 0">{"✅" if ip else "❌"} <b>{nm}</b>'
                if nt: html += f' — <span style="color:var(--muted)">{str(nt)[:120]}</span>'
                html += '</div>'
        html += '</div>'
    if issues:
        html += '<div style="margin-top:.5rem">'
        for iss in (issues if isinstance(issues, list) else [issues])[:4]:
            txt = iss if isinstance(iss, str) else json.dumps(iss)
            html += f'<div style="font-size:.8rem;color:#f59e0b;margin:.2rem 0">• {txt[:200]}</div>'
        html += '</div>'
    st.markdown(html + '</div>', unsafe_allow_html=True)

def render_contacts(contacts, title="Contacts"):
    if not contacts: st.info("No contacts found for this job."); return
    pri_color = {"Primary":"red","Secondary":"yellow","Tertiary":"blue","General":"purple"}
    st.markdown(f'<div class="section-label">👥 {title} ({len(contacts)})</div>', unsafe_allow_html=True)
    cols = st.columns(2)
    for i, c in enumerate(contacts):
        col = cols[i % 2]; prio = c.get("priority","General")
        email = c.get("email",""); li = c.get("linkedin_url","")
        patterns = c.get("email_patterns",[]); src = c.get("source","")
        with col:
            html = f"""<div class="contact-card">
              <div style="display:flex;justify-content:space-between;align-items:flex-start">
                <div><div class="contact-name">{c.get('name','—')}</div>
                <div class="contact-title">{c.get('title','—')}</div></div>
                {badge(prio, pri_color.get(prio,'blue'))}
              </div>"""
            if email:      html += f'<div class="contact-email" style="margin-top:.5rem">✉️ {email}</div>'
            elif patterns: html += f'<div style="font-size:.75rem;color:var(--muted);margin-top:.4rem">📧 {patterns[0]}</div>'
            if li:         html += f'<div style="font-size:.75rem;margin-top:.3rem"><a href="{li}" target="_blank" style="color:var(--accent);text-decoration:none">🔗 LinkedIn</a></div>'
            if src:        html += f'<div style="font-size:.68rem;color:var(--muted);margin-top:.4rem">via {src}</div>'
            st.markdown(html + '</div>', unsafe_allow_html=True)

def render_emails(emails_data):
    if not emails_data: st.info("No email data available."); return
    emails_data = as_dict(emails_data)
    if not emails_data: return
    variants = {}
    for k, v in emails_data.items():
        kl = k.lower().replace("_","").replace(" ","")
        if any(x in kl for x in ["varianta","variant_a","emaila"]) or kl=="a": variants["Variant A — Hiring Manager"] = v
        elif any(x in kl for x in ["variantb","variant_b","emailb"]) or kl=="b": variants["Variant B — CISO / VP Level"] = v
        else: variants[k] = v
    for label, v in variants.items():
        st.markdown(f'<div class="section-label">✉️ {label}</div>', unsafe_allow_html=True)
        if isinstance(v, dict):
            subj = v.get("subject") or v.get("Subject",""); body = v.get("body") or v.get("Body") or v.get("content","")
            if subj: st.markdown(f'<div class="email-subject">Subject: {subj}</div>', unsafe_allow_html=True)
            if body: st.markdown(f'<div class="email-box">{body}</div>', unsafe_allow_html=True)
            else:    st.code(json.dumps(v, indent=2), language="json")
        elif isinstance(v, str):
            st.markdown(f'<div class="email-box">{v}</div>', unsafe_allow_html=True)
        st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

def render_service_mapping(data):
    if not data: st.info("No service mapping data."); return
    items = data if isinstance(data, list) else []
    if not items and isinstance(data, dict):
        for key in ("services","mappings","service_mapping","ServiceMapping","items"):
            if isinstance(data.get(key), list): items = data[key]; break
        if not items: items = [data]
    fit_colors = {"STRONG FIT":"green","PARTIAL FIT":"yellow","GAP":"red"}
    for item in items:
        if not isinstance(item, dict): continue
        svc  = item.get("service") or item.get("Service") or item.get("name","")
        fit  = (item.get("fit") or item.get("classification") or item.get("Fit") or item.get("status","")).upper()
        why  = item.get("justification") or item.get("rationale") or item.get("why","")
        reqs = item.get("requirements_addressed") or item.get("requirements") or ""
        eng  = item.get("engagement_type") or item.get("engagement","")
        color = fit_colors.get(fit,"blue")
        html = f"""<div class="sil-card" style="margin-bottom:.75rem">
          <div style="display:flex;align-items:center;gap:.75rem;margin-bottom:.5rem">
            <span style="font-family:'Syne',sans-serif;font-weight:700;color:var(--text)">{svc}</span>
            {badge(fit or "MAPPED", color) if fit else ""}
          </div>"""
        if why:  html += f'<div style="font-size:.85rem;color:var(--muted);margin-bottom:.4rem">💡 {str(why)[:250]}</div>'
        if reqs:
            rs = ", ".join(reqs) if isinstance(reqs, list) else str(reqs)
            html += f'<div style="font-size:.8rem;color:var(--muted)">📌 {rs[:200]}</div>'
        if eng:  html += f'<div style="font-size:.8rem;color:var(--accent2);margin-top:.3rem">🔧 {eng}</div>'
        st.markdown(html + '</div>', unsafe_allow_html=True)
    render_json_pretty(data, "Service Mapping")

def render_microplans(data):
    if not data: st.info("No micro-plan data."); return
    plans = data if isinstance(data, list) else []
    if not plans and isinstance(data, dict):
        for k in ("plans","micro_plans","microplans","top_3","improvements"):
            if isinstance(data.get(k), list): plans = data[k]; break
        if not plans: plans = [data]
    for i, plan in enumerate(plans[:3], 1):
        if not isinstance(plan, dict): continue
        title = plan.get("title") or plan.get("objective") or plan.get("name") or f"Plan {i}"
        weeks = plan.get("duration") or plan.get("timeline","")
        obj   = plan.get("objective") or plan.get("goal","")
        kpis  = plan.get("kpis") or plan.get("KPIs") or []
        tasks = plan.get("tasks") or plan.get("workstreams") or []
        with st.expander(f"📋 Plan {i}: {title} {f'({weeks})' if weeks else ''}", expanded=(i==1)):
            if obj and obj != title: st.markdown(f"**Objective:** {obj}")
            if kpis:
                st.markdown("**KPIs:**")
                for kpi in (kpis if isinstance(kpis, list) else [kpis]): st.markdown(f"• {kpi}")
            if tasks:
                st.markdown("**Tasks:**")
                for t in (tasks if isinstance(tasks, list) else [tasks]):
                    if isinstance(t, dict):
                        tn = t.get("task") or t.get("name",""); te = t.get("effort") or t.get("duration","")
                        st.markdown(f"• **{tn}** {f'— {te}' if te else ''}")
                    else: st.markdown(f"• {t}")
            st.code(json.dumps(plan, indent=2, default=str), language="json")

def render_deal_assurance(data):
    if not data: st.info("No deal assurance data."); return
    if not isinstance(data, dict): render_json_pretty(data, "Deal Assurance Pack"); return
    evp = data.get("executive_value_proposition") or data.get("value_proposition") or data.get("ExecutiveValueProposition","")
    if evp:
        st.markdown('<div class="section-label">💼 Executive Value Proposition</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="sil-card sil-card-accent" style="font-size:.9rem;line-height:1.7">{evp}</div>', unsafe_allow_html=True)
    caps = data.get("mandatory_capabilities") or data.get("capabilities_checklist") or []
    if caps:
        st.markdown('<div class="section-label" style="margin-top:1rem">✅ Mandatory Capabilities</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        for i, cap in enumerate(caps if isinstance(caps, list) else [caps]):
            (c1 if i%2==0 else c2).markdown(f"✅ {cap}")
    risk = data.get("risk_mitigation") or data.get("RiskMitigation","")
    if risk:
        st.markdown('<div class="section-label" style="margin-top:1rem">🛡️ Risk Mitigation</div>', unsafe_allow_html=True)
        if isinstance(risk, dict):
            for k, v in risk.items(): st.markdown(f"**{k}:** {v}")
        else: st.markdown(str(risk))
    render_json_pretty(data, "Deal Assurance Pack")


# ══════════════════════════════════════════════════════════════════════════════
#  LOGS PAGE
# ══════════════════════════════════════════════════════════════════════════════
def render_logs_page():
    st.markdown("""<div style="margin-bottom:2rem">
      <h1 style="font-family:'Syne',sans-serif;font-size:2.2rem;font-weight:800;color:#0f172a;margin:0">⚙️ Pipeline Execution</h1>
      <p style="font-size:.95rem;color:#64748b;margin-top:.5rem">Live logs from the scraper, deduplication, AI agents, and Google Sheet sync</p>
    </div>""", unsafe_allow_html=True)

    if st.session_state.pipeline_running:
        st.markdown("""<div class="logs-status running">
          <div><div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.3rem">
            <span class="pulse-dot"></span>
            <span style="font-family:'Syne',sans-serif;font-weight:700;color:#1d4ed8">Pipeline Running</span>
          </div>
          <span style="font-size:.82rem;color:#64748b">Processing jobs... Google Sheet will auto-sync when complete.</span></div>
        </div>""", unsafe_allow_html=True)
    elif st.session_state.pipeline_done:
        result = st.session_state.pipeline_result or {}
        if result.get("success"):
            st.markdown(f"""<div class="logs-status success">
              <div><div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
                <span style="font-size:1.5rem">✅</span>
                <span style="font-family:'Syne',sans-serif;font-weight:700;color:#15803d;font-size:1.1rem">Pipeline Completed</span>
              </div>
              <div style="font-size:.85rem;color:#15803d">
                {result.get('processed',0)} processed · {result.get('new_jobs',0)} new · {result.get('skipped_db',0)} already in DB
              </div></div></div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""<div class="logs-status error">
              <div><div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.5rem">
                <span style="font-size:1.5rem">❌</span>
                <span style="font-family:'Syne',sans-serif;font-weight:700;color:#b91c1c;font-size:1.1rem">Pipeline Failed</span>
              </div>
              <div style="font-size:.85rem;color:#b91c1c">{result.get('error','Unknown error')}</div></div></div>""", unsafe_allow_html=True)

    st.markdown('<div class="section-label" style="margin-top:2rem">📡 Live Execution Logs</div>', unsafe_allow_html=True)
    logs = st.session_state.pipeline_logs or []
    log_text = "\n".join(logs[-200:]) if logs else "Waiting for logs..."
    st.markdown(f'<div class="pipeline-log">{log_text}</div>', unsafe_allow_html=True)

    if st.session_state.pipeline_running:
        st.markdown("<script>setTimeout(function(){window.location.reload();},1500);</script>", unsafe_allow_html=True)

    c1, c2, _ = st.columns([2, 2, 1])
    with c1:
        if st.button("← Back to Dashboard", use_container_width=True):
            st.session_state.current_page = "dashboard"; st.rerun()
    with c2:
        if st.button("🔄 Refresh Logs", use_container_width=True): st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""<div style="padding:.75rem 0 1.25rem">
      <div style="font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:800;color:#2563eb">🛡️ SecureITLab</div>
      <div style="font-size:.72rem;color:#64748b;letter-spacing:.1em;text-transform:uppercase;margin-top:.2rem">Pipeline Intelligence</div>
    </div>""", unsafe_allow_html=True)

    if st.button("🚪 Logout", use_container_width=True):
        st.session_state.logged_in = False; st.session_state.login_error = ""; st.rerun()

    st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

    # ── Find Jobs ─────────────────────────────────────────────────────────────
    st.markdown("**🔍 Find New Jobs**")
    st.caption("Scrapes → deduplicates → AI agents → MongoDB → auto-syncs to Google Sheet · Repeats every 12h")

    find_jobs_clicked = st.button(
        "🔍  Find Jobs",
        disabled=st.session_state.pipeline_running,
        use_container_width=True,
        type="primary",
    )

    if st.session_state.pipeline_running:
        st.markdown("""<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
            padding:.6rem .9rem;margin-top:.5rem;font-family:'DM Mono',monospace;font-size:.8rem;color:#1d4ed8">
          ⏳ Pipeline running… 👉 View Logs Page
        </div>""", unsafe_allow_html=True)

    # ── Scheduler badge ───────────────────────────────────────────────────────
    try:
        import main as _main_module
        sched = _main_module.get_scheduler_status()
        if sched.get("active"):
            secs = sched.get("seconds_until_next")
            if secs is not None:
                hrs=secs//3600; mins=(secs%3600)//60
                countdown=f"{hrs}h {mins}m" if hrs else f"{mins}m {secs%60}s"
                next_label=f"Next run in: <b>{countdown}</b>"
            else:
                next_label="Next run: calculating…"
            last=(sched.get("last_run") or "")[:19]
            st.markdown(f"""<div class="sched-on">
              🟢 <b>Auto-scheduler ON</b><br>
              <span style="color:#64748b">{next_label}</span><br>
              <span style="color:#64748b">Runs every 12h · #{sched.get('run_count',0)} so far</span><br>
              <span style="color:#64748b;font-size:.75rem">Last: {last} UTC</span>
            </div>""", unsafe_allow_html=True)
            if st.button("⏹️ Stop Auto-Scheduler", use_container_width=True):
                _main_module.stop_scheduler(); st.rerun()
        elif sched.get("run_count", 0) > 0:
            last=(sched.get("last_run") or "")[:19]
            st.markdown(f"""<div class="sched-paused">
              ⏸️ <b>Scheduler paused</b><br>
              <span style="color:#64748b">Last ran: {last} UTC</span><br>
              <span style="color:#64748b">Click Find Jobs to restart</span>
            </div>""", unsafe_allow_html=True)
    except Exception:
        pass

    st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    #  ✅ GOOGLE SHEETS MASTER CONTACTS  (with LinkedIn & Email)
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("**📊 Master Contacts — Google Sheet**")
    st.caption("All contacts synced here. Includes LinkedIn messages & outreach emails. Append-only — no deletions.")

    # ── Auto-sync once per session on dashboard load ──────────────────────────
    if not st.session_state.get("gsheet_auto_synced"):
        st.session_state["gsheet_auto_synced"] = True
        if Path(GCREDS_FILE).exists():
            try:
                _auto_jobs = load_all_jobs()
            except Exception:
                _auto_jobs = []
            if _auto_jobs:
                _auto_res = sync_contacts_to_gsheet(_auto_jobs)
                if not _auto_res["error"]:
                    st.session_state.gsheet_sync_error = ""
                    st.session_state.gsheet_last_sync  = datetime.now().strftime("%d %b %Y  %H:%M")
                    st.session_state.gsheet_appended   = _auto_res["appended"]
                else:
                    st.session_state.gsheet_sync_error = _auto_res["error"]

    # Buttons row
    sync_col1, sync_col2 = st.columns([3, 1])
    with sync_col1:
        sync_clicked = st.button("🔄  Sync Now", use_container_width=True)
    with sync_col2:
        st.markdown(f'<a href="{GSHEET_URL}" target="_blank" style="text-decoration:none"><div style="text-align:center;padding:.42rem;border:1px solid #d9e2ec;border-radius:6px;background:#fff;font-size:.85rem;cursor:pointer">↗</div></a>', unsafe_allow_html=True)
    if sync_clicked:
        try:
            jobs_for_sync = load_all_jobs()
        except Exception:
            jobs_for_sync = []
        if not jobs_for_sync:
            st.warning("No jobs in MongoDB to sync yet.")
        else:
            with st.spinner("Syncing to Google Sheet…"):
                res = sync_contacts_to_gsheet(jobs_for_sync)
            if res["error"]:
                st.session_state.gsheet_sync_error = res["error"]
                st.session_state.gsheet_last_sync  = None
                st.session_state.gsheet_appended   = None
            else:
                st.session_state.gsheet_sync_error = ""
                st.session_state.gsheet_last_sync  = datetime.now().strftime("%d %b %Y  %H:%M")
                st.session_state.gsheet_appended   = res["appended"]
                st.success(f"✅ +{res['appended']} new rows · {res['skipped']} already existed")

    # Read persisted sync state from disk
    disk_state = _read_sync_state()
    last_sync  = (
        st.session_state.gsheet_last_sync
        or (disk_state.get("last_sync","")[:16].replace("T"," ") if disk_state.get("last_sync") else None)
    )
    appended = st.session_state.gsheet_appended if st.session_state.gsheet_appended is not None else disk_state.get("appended","")

    if st.session_state.gsheet_sync_error:
        short_err  = st.session_state.gsheet_sync_error[:180]
        creds_hint = ""
        if "not found" in short_err.lower() or "no such file" in short_err.lower():
            creds_hint = "<br><span style='font-size:.74rem'>💡 Save your service account JSON as <code>google_credentials.json</code> in the same folder as this script, then share the sheet with the service account email.</span>"
        elif "worksheet" in short_err.lower():
            creds_hint = f"<br><span style='font-size:.74rem'>💡 Make sure the sheet tab is named <b>{GSHEET_TAB}</b> (it will be created automatically if missing).</span>"
        st.markdown(f'<div class="gsheet-error">❌ <b>Sync failed</b><br>{short_err}{creds_hint}</div>', unsafe_allow_html=True)
    else:
        last_line  = f"Last synced: <b>{last_sync}</b>" if last_sync else "Not synced yet"
        added_line = f" · <b>+{appended}</b> new rows" if appended and appended != 0 else (" · ✓ up to date" if appended == 0 and last_sync else "")
        st.markdown(f"""<div class="gsheet-box">
          🟢 <b>Google Sheet live</b><br>
          <a href="{GSHEET_URL}" target="_blank">📋 Open master_contacts ↗</a><br>
          <span style="font-size:.76rem;color:#64748b">{last_line}{added_line}</span><br>
          <span style="font-size:.73rem;color:#94a3b8">LinkedIn messages & outreach emails synced · Auto-sync on load & after every pipeline run</span>
        </div>""", unsafe_allow_html=True)

    st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)

    # ── Job list ──────────────────────────────────────────────────────────────
    try:
        all_jobs = load_all_jobs()
    except Exception as e:
        st.error(f"MongoDB error: {e}"); st.stop()

    if not all_jobs:
        st.warning("No jobs in MongoDB yet. Click **Find Jobs** to start."); st.stop()

    st.markdown(f'<div style="font-size:.75rem;color:#64748b;margin-bottom:.75rem">{len(all_jobs)} jobs in database</div>', unsafe_allow_html=True)

    search   = st.text_input("🔍 Filter jobs", placeholder="e.g. Bounteous")
    filtered = [j for j in all_jobs if search.lower() in (j.get("company","")+" "+j.get("role","")).lower()]

    def job_label(j):
        score = j.get("opp_score"); s = f" [{score}/10]" if score else ""
        ok = "✅" if j.get("pipeline_ok") else "❌"
        return f"{ok} {j.get('company','?')} — {j.get('role','?')[:32]}{s}"

    if not filtered:
        st.warning("No matching jobs."); st.stop()

    sel_label   = st.selectbox("Select a Job", [job_label(j) for j in filtered], index=0)
    selected_id = str(filtered[[job_label(j) for j in filtered].index(sel_label)]["_id"])

    st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
    ok_count = sum(1 for j in all_jobs if j.get("pipeline_ok"))
    total_c  = sum(j.get("contacts_found", 0) for j in all_jobs)
    st.markdown(f"""<div style="font-size:.75rem;color:#64748b">
      <div>✅ Pipeline OK: <b style="color:#16a34a">{ok_count}/{len(all_jobs)}</b></div>
      <div>👥 Total Contacts: <b style="color:#2563eb">{total_c}</b></div>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
    if st.button("🔄 Refresh Data", use_container_width=True):
        st.cache_data.clear(); st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
#  FIND JOBS — background thread
# ══════════════════════════════════════════════════════════════════════════════
if find_jobs_clicked and not st.session_state.pipeline_running:
    with _thread_lock:
        _thread_logs.clear(); _thread_result[0] = None; _thread_done[0] = False

    st.session_state.pipeline_running = True
    st.session_state.pipeline_done    = False
    st.session_state.pipeline_logs    = []
    st.session_state.pipeline_result  = None
    st.session_state.current_page     = "logs"
    st.cache_data.clear()

    def _run_pipeline_bg():
        try:
            import main as _main

            def _cb(msg: str):
                with _thread_lock: _thread_logs.append(msg)

            result = _main.run_pipeline(progress_callback=_cb)

            # ── Auto-sync Google Sheet after every successful run ────────────
            if result.get("success"):
                try:
                    _cb("🔄 [GSheet] Syncing new contacts to Google Sheet…")
                    sync_res = sync_contacts_to_gsheet(load_all_jobs())
                    if sync_res["error"]:
                        _cb(f"⚠️  [GSheet] Sync error: {sync_res['error']}")
                    else:
                        _cb(f"✅ [GSheet] Done — +{sync_res['appended']} new rows · {sync_res['skipped']} already existed")
                except Exception as se:
                    _cb(f"⚠️  [GSheet] Sync failed: {se}")

            with _thread_lock: _thread_result[0] = result

        except Exception as e:
            import traceback
            with _thread_lock:
                _thread_logs.append(f"❌ Unexpected error: {e}")
                _thread_logs.append(traceback.format_exc())
                _thread_result[0] = {"success":False,"error":str(e),"scraped":0,"new_jobs":0,"skipped_db":0,"processed":0}
        finally:
            with _thread_lock: _thread_done[0] = True

    threading.Thread(target=_run_pipeline_bg, daemon=True).start()
    st.rerun()

# ── Sync thread state → session_state ────────────────────────────────────────
with _thread_lock:
    if _thread_logs:
        st.session_state.pipeline_logs = list(_thread_logs)
    if _thread_done[0] and _thread_result[0] is not None:
        st.session_state.pipeline_result  = _thread_result[0]
        st.session_state.pipeline_running = False
        st.session_state.pipeline_done    = True


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE ROUTING
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.current_page == "logs":
    render_logs_page()
    st.stop()

with st.spinner("Loading job…"):
    job = load_job(selected_id)

if not job:
    st.error("Could not load job document."); st.stop()

company   = job.get("company","Unknown"); role  = job.get("role","Unknown")
opp_score = job.get("opp_score");         p_ok  = job.get("pipeline_ok",False)
p_min     = job.get("pipeline_min","?");  c_found = job.get("contacts_found",0)
c_cov     = job.get("coverage_score");    c_domain = job.get("contact_domain","")
run_at    = job.get("run_at","")

st.markdown(f"""<div style="margin-bottom:1.75rem">
  <div style="font-family:'DM Mono',monospace;font-size:.72rem;color:#2563eb;letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem">
    Job #{job.get('job_number','?')} · {run_at[:10] if run_at else ''}
  </div>
  <h1 style="font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;color:#0f172a;margin:0;line-height:1.15">{role}</h1>
  <div style="font-size:1.05rem;color:#64748b;margin-top:.3rem">
    @ <span style="color:#334155;font-weight:600">{company}</span>
    {f'<span style="color:#cbd5e1;margin:0 .5rem">·</span><span style="font-family:DM Mono,monospace;font-size:.82rem;color:#94a3b8">{c_domain}</span>' if c_domain else ""}
  </div>
</div>""", unsafe_allow_html=True)

try:
    sn = float(str(opp_score).split("/")[0].split(".")[0]) if opp_score else 0
    sc = "#16a34a" if sn>=7 else "#f59e0b" if sn>=5 else "#dc2626"
except Exception:
    sc = "#2563eb"

st.markdown(f"""<div class="metric-row">
  <div class="metric-tile"><div class="val" style="color:{sc}">{f"{opp_score}/10" if opp_score else "—"}</div><div class="lbl">Opportunity Score</div></div>
  <div class="metric-tile"><div class="val">{c_found}</div><div class="lbl">Contacts Found</div></div>
  <div class="metric-tile"><div class="val">{f"{c_cov}%" if c_cov else "—"}</div><div class="lbl">Contact Coverage</div></div>
  <div class="metric-tile"><div class="val" style="color:{'#16a34a' if p_ok else '#dc2626'}">{'✅ OK' if p_ok else '❌ Failed'}</div><div class="lbl">Pipeline ({p_min} min)</div></div>
</div>""", unsafe_allow_html=True)

tabs = st.tabs(["📋 Job & Enrichment","🎯 Service Mapping","🔍 Fit / Gap",
                "🛠️ Capability & Plans","📦 Deal Assurance","✉️ Outreach Emails",
                "👥 Contacts","✅ QA Gates","🗄️ Raw Data"])

with tabs[0]:
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-label">📄 Job Research</div>', unsafe_allow_html=True)
        jr = as_dict(job.get("agent_job_research") or {})
        if jr:
            for lbl, keys in [("Job Role",["job_role","Job Role","role","title"]),("Company",["company_name","Company Name","company"]),("Location",["location","Location"]),("URL",["organization_url","Organization URL","url"])]:
                val = next((jr.get(k) for k in keys if jr.get(k)), None)
                if val: st.markdown(f"**{lbl}:** {val}")
            desc = jr.get("job_description") or jr.get("Job Description","")
            if desc:
                st.markdown("**Job Description:**")
                st.markdown(f'<div class="sil-card" style="font-size:.85rem;line-height:1.7;max-height:300px;overflow-y:auto">{desc[:2000]}</div>', unsafe_allow_html=True)
            render_json_pretty(jr,"Job Research")
        else: st.info("No job research data.")
    with c2:
        st.markdown('<div class="section-label">🏢 Company Enrichment</div>', unsafe_allow_html=True)
        enr = as_dict(job.get("agent_enrichment") or {})
        if enr:
            for lbl, keys in [("Industry",["industry","Industry"]),("Company Size",["company_size","size"]),("Regulatory Env",["regulatory_environment","regulatory"]),("Certifications",["certifications","Certifications"]),("Tech Stack",["tech_stack","technology_stack"]),("Security Maturity",["security_maturity","maturity"])]:
                val = next((enr.get(k) for k in keys if enr.get(k)), None)
                if val:
                    if isinstance(val, list): val = ", ".join(str(v) for v in val)
                    st.markdown(f"**{lbl}:** {safe_str(val,200)}")
            render_json_pretty(enr,"Enrichment")
        else: st.info("No enrichment data.")

with tabs[1]:
    st.markdown('<div class="section-label">🗺️ Service Mapping Matrix</div>', unsafe_allow_html=True)
    render_service_mapping(job.get("agent_service_mapping"))

with tabs[2]:
    fg = as_dict(job.get("agent_fit_gap") or {})
    if opp_score:
        try:
            sn=float(str(opp_score).split("/")[0]); bc="#16a34a" if sn>=7 else "#f59e0b" if sn>=5 else "#dc2626"; bp=int(sn/10*100)
            st.markdown(f"""<div style="margin-bottom:1.5rem">
              <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
                <span style="font-family:'Syne',sans-serif;font-weight:700">Opportunity Score</span>
                <span style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:800;color:{bc}">{opp_score}/10</span>
              </div>
              <div style="background:#e2e8f0;border-radius:4px;height:8px">
                <div style="background:{bc};width:{bp}%;height:100%;border-radius:4px"></div>
              </div></div>""", unsafe_allow_html=True)
        except Exception: pass
    st.markdown('<div class="section-label">📊 Service Classifications</div>', unsafe_allow_html=True)
    services = []
    if isinstance(fg, dict):
        for k in ("services","classifications","service_classifications","items","fit_gap"):
            v = fg.get(k)
            if isinstance(v, list): services = v; break
        if not services and (fg.get("service") or fg.get("Service")): services = [fg]
    elif isinstance(fg, list): services = fg
    if services:
        buckets = {"STRONG FIT":[],"PARTIAL FIT":[],"GAP":[]}
        for s in services:
            if not isinstance(s, dict): continue
            fit = (s.get("fit") or s.get("classification") or s.get("Fit","")).upper()
            if "STRONG" in fit: buckets["STRONG FIT"].append(s)
            elif "PARTIAL" in fit: buckets["PARTIAL FIT"].append(s)
            elif "GAP" in fit: buckets["GAP"].append(s)
        c1,c2,c3=st.columns(3)
        cm={"STRONG FIT":"#16a34a","PARTIAL FIT":"#f59e0b","GAP":"#dc2626"}
        bgm={"STRONG FIT":"#f0fdf4","PARTIAL FIT":"#fffbeb","GAP":"#fef2f2"}
        bdm={"STRONG FIT":"#bbf7d0","PARTIAL FIT":"#fde68a","GAP":"#fecaca"}
        for col,(fl,items) in zip([c1,c2,c3],buckets.items()):
            col.markdown(f'<div style="font-family:Syne,sans-serif;font-weight:700;color:{cm[fl]};margin-bottom:.5rem">{fl} ({len(items)})</div>', unsafe_allow_html=True)
            for s in items:
                svc=s.get("service") or s.get("Service") or s.get("name",""); just=s.get("justification") or s.get("reason","")
                col.markdown(f'<div style="background:{bgm[fl]};border:1px solid {bdm[fl]};border-radius:8px;padding:.75rem;margin-bottom:.5rem;font-size:.85rem"><div style="font-weight:600">{svc}</div><div style="color:#64748b;font-size:.78rem;margin-top:.2rem">{safe_str(just,150)}</div></div>', unsafe_allow_html=True)
    elif fg: st.json(fg)
    else: st.info("No fit/gap data.")
    render_json_pretty(job.get("agent_fit_gap"),"Fit/Gap Analysis")

with tabs[3]:
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-label">🔧 Capability Improvements</div>', unsafe_allow_html=True)
        cap = job.get("agent_capability") or {}
        items_cap = cap if isinstance(cap,list) else []
        if not items_cap and isinstance(cap,dict):
            for k in ("improvements","recommendations","capabilities","items"):
                v=cap.get(k)
                if isinstance(v,list): items_cap=v; break
            if not items_cap: items_cap=[cap]
        for item in items_cap:
            if not isinstance(item,dict): continue
            title=item.get("title") or item.get("gap") or item.get("service","")
            rec=item.get("recommendation") or item.get("steps","")
            effort=item.get("build_effort") or item.get("effort",""); demand=item.get("market_demand") or item.get("priority","")
            st.markdown(f"""<div class="sil-card" style="margin-bottom:.6rem">
              <div style="font-family:'Syne',sans-serif;font-weight:700;margin-bottom:.3rem">{title}</div>
              <div style="font-size:.82rem;color:var(--muted)">{safe_str(rec,250)}</div>
              {f'<div style="font-size:.75rem;color:var(--accent2);margin-top:.3rem">Priority: {demand} · Effort: {effort}</div>' if demand or effort else ""}
            </div>""", unsafe_allow_html=True)
        if not items_cap: render_json_pretty(cap,"Capability Improvement")
    with c2:
        st.markdown('<div class="section-label">📅 Maturity Micro-Plans</div>', unsafe_allow_html=True)
        render_microplans(job.get("agent_microplans"))

with tabs[4]:
    render_deal_assurance(job.get("agent_deal_assurance"))

with tabs[5]:
    st.markdown('<div class="section-label">✉️ Outreach Email Variants</div>', unsafe_allow_html=True)
    emails_src = job.get("agent_outreach_emails") or job.get("outreach_emails") or {}
    oq = as_dict(job.get("agent_outreach_qa") or {})
    improved = (oq.get("improved_emails") or oq.get("ImprovedEmails")) if oq else None
    if improved:
        st.info("⚡ Showing QA-improved versions")
        render_emails(improved)
        with st.expander("📬 Original (pre-QA) versions"): render_emails(emails_src)
    else: render_emails(emails_src)

with tabs[6]:
    contacts = job.get("contacts") or []; contact_sources = job.get("contact_sources") or []
    pri=[c for c in contacts if c.get("priority")=="Primary"]
    sec=[c for c in contacts if c.get("priority")=="Secondary"]
    ter=[c for c in contacts if c.get("priority")=="Tertiary"]
    gen=[c for c in contacts if c.get("priority")=="General"]
    st.markdown(f"""<div class="metric-row" style="margin-bottom:1.5rem">
      <div class="metric-tile"><div class="val" style="color:#dc2626">{len(pri)}</div><div class="lbl">Primary</div></div>
      <div class="metric-tile"><div class="val" style="color:#f59e0b">{len(sec)}</div><div class="lbl">Secondary</div></div>
      <div class="metric-tile"><div class="val" style="color:#2563eb">{len(ter)}</div><div class="lbl">Tertiary</div></div>
      <div class="metric-tile"><div class="val" style="color:#94a3b8">{len(gen)}</div><div class="lbl">General</div></div>
    </div>""", unsafe_allow_html=True)
    if contact_sources:
        st.markdown('Sources: '+" ".join(badge(s,"blue") for s in contact_sources), unsafe_allow_html=True); st.markdown("")
    missing = job.get("missing_roles") or []
    if missing:
        st.markdown('⚠️ Missing roles: '+" ".join(badge(r,"red") for r in missing), unsafe_allow_html=True); st.markdown("")
    if contacts:
        excel_bytes = build_contacts_excel(contacts, company, role)
        if excel_bytes:
            safe_co = re.sub(r'[^a-z0-9]','_',company.lower())[:20]
            fname = f"contacts_{safe_co}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            btn_col, _ = st.columns([1,3])
            with btn_col:
                st.download_button("📥  Download Contacts (.xlsx)", data=excel_bytes, file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")
        st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
        pri_filter = st.multiselect("Filter by priority",["Primary","Secondary","Tertiary","General"],default=["Primary","Secondary","Tertiary","General"])
        shown = [c for c in contacts if c.get("priority","General") in pri_filter]
        render_contacts(shown, f"Contacts ({len(shown)} shown)")
        agent_contacts = job.get("agent_prospect_contacts") or {}
        if agent_contacts:
            with st.expander("🤖 CrewAI Agent's Contact Search"):
                if isinstance(agent_contacts,dict):
                    ac_list=agent_contacts.get("contacts") or []
                    if ac_list: render_contacts(ac_list,"Agent Contacts")
                    else: st.json(agent_contacts)
                else: st.json(agent_contacts)
    else: st.info("No contacts found for this job.")

with tabs[7]:
    st.markdown('<div class="section-label" style="margin-bottom:1rem">🔍 All 4 QA Gate Results</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    for i,(lbl,key) in enumerate([("Research QA","agent_research_qa"),("Service Mapping QA","agent_mapping_qa"),("Deal Assurance QA","agent_assurance_qa"),("Outreach Email QA","agent_outreach_qa")]):
        with (c1 if i%2==0 else c2): render_qa_block(job.get(key), lbl)
    st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="section-label">🎯 Prospect Enforcer Result</div>', unsafe_allow_html=True)
    enf = as_dict(job.get("agent_prospect_enforcer") or {})
    if enf:
        cov=enf.get("coverage_score","?"); miss=enf.get("missing_roles",[]); note=enf.get("note",""); ec=enf.get("contacts",[])
        x1,x2,x3=st.columns(3)
        x1.metric("Coverage Score",f"{cov}%"); x2.metric("Missing Roles",len(miss)); x3.metric("Contacts Verified",len(ec))
        if miss: st.markdown(f"**Missing:** {', '.join(str(m) for m in miss)}")
        if note: st.caption(note)
    else: st.info("No enforcer data.")

with tabs[8]:
    st.markdown('<div class="section-label">🗄️ Raw MongoDB Document</div>', unsafe_allow_html=True)
    rows=[]
    for k,v in job.items():
        if k=="_id": continue
        rows.append({"Field":k,"Type":type(v).__name__,"Len":len(v) if isinstance(v,(list,dict)) else len(str(v)) if v else 0})
    hc1,hc2,hc3=st.columns([3,1,1])
    hc1.markdown("**Field**"); hc2.markdown("**Type**"); hc3.markdown("**Len**")
    for r in rows:
        rc1,rc2,rc3=st.columns([3,1,1])
        rc1.code(r["Field"],language=None)
        rc2.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Type"]}</span>', unsafe_allow_html=True)
        rc3.markdown(f'<span style="color:#64748b;font-size:.8rem">{r["Len"]}</span>', unsafe_allow_html=True)
    st.markdown('<div class="sil-divider"></div>', unsafe_allow_html=True)
    for lbl,key in [("Job Research","agent_job_research"),("Enrichment","agent_enrichment"),("Service Mapping","agent_service_mapping"),("Fit/Gap","agent_fit_gap"),("Capability","agent_capability"),("Micro-Plans","agent_microplans"),("Deal Assurance","agent_deal_assurance"),("Outreach Emails","agent_outreach_emails"),("Prospect Contacts","agent_prospect_contacts"),("Prospect Enforcer","agent_prospect_enforcer"),("Research QA","agent_research_qa"),("Mapping QA","agent_mapping_qa"),("Assurance QA","agent_assurance_qa"),("Outreach QA","agent_outreach_qa"),("Contacts (5-source)","contacts")]:
        data=job.get(key)
        if data:
            with st.expander(f"📄 {lbl}"):
                st.code(json.dumps(data,indent=2,default=str),language="json")